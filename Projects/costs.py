from pony.orm import *
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from . import materialCost, instalationCost, fabricationCost
import convert

def estimateCost(db, contract_number, file_name):
    # primero, obtenemos la comuna donde se realizara el proyecto
    file_read = file_name + ".xlsx"
    wb_read = load_workbook(file_read, data_only=True)
    #revisamos que el archivo tenga las hojas que debe tener
    try:
        ws_read_measures = wb_read["Measures"]
    except KeyError:
        print(' Formato invalido del archivo. El archivo debe tener una hoja llamada Measures.')
        return
    try:
        wb_read["Manufacturing"]
    except KeyError:
        print(' Formato invalido del archivo. El archivo debe tener una hoja llamada Measures.')
        return
    #obtenemos la comuna donde se realizara el proyecto
    comuna_to = ws_read_measures.cell(row = 7, column = 15).value
    try:
        comuna_to = convert.get_fuzzy(comuna_to)
    except:
        print('\n Formato del archivo es invalido. No ha podido leerse la comuna del proyecto.')
        return
    
    # luego, obtenemos el freight_cost de la tabla Freight_Costs
    warning = ' Aviso: el costo de flete para la comuna indicada en el archivo no se encuentra registrado en la base de datos. Se considerara como 0.'
    try:
        freight_cost = db.Freight_Costs[comuna_to].freight_cost
    except pony.orm.core.ObjectNotFound:
        print(warning)
        freight_cost = 0
    
    # segundo, obtenemos la lista de instaladores ligados al proyecto
    tasks = select(t for t in db.Tasks if t.skill == db.Skills[4] and t.project == db.Projects.get(contract_number = contract_number, finished = None))
    installers = []
    for t in tasks:
        if t.failed != True:
            installers_tasks = select(et for et in db.Employees_Tasks if et.task == t)
            for it in installers_tasks:
                installers.append(it.employee)
    
    # luego, obtenemos los datos viatic_cost y movilization_cost de las tablas comuna-comuna, para cada instalador
    # tambien contamos la cantidad de instaladores, el tema de si son externos o no queda pendiente
    viatic_cost = 0
    movilization_cost = 0
    num_installers = 0
    for i in installers:
        num_installers = num_installers + 1
        #recuperamos el costo de viáticos
        warning = ' Aviso: el costo de viatico para las comunas ' + i.zone + '-' + comuna_to + ' no se encuentra registrado en la base de datos. Se considerara como 0.'
        try:
            viatic_cost = viatic_cost + db.Viatic_Costs[(i.zone, comuna_to)].viatic_cost
        except pony.orm.core.ObjectNotFound:
            print(warning)
        #recuperamos el costo de movilización
        warning = ' Aviso: el costo de movilizacion para las comunas ' + i.zone + '-' + comuna_to + ' no se encuentra registrado en la base de datos. Se considerara como 0.'
        try:
            movilization_cost = movilization_cost + db.Movilization_Costs[(i.zone, comuna_to)].movilization_cost
        except pony.orm.core.ObjectNotFound:
            print(warning)
    
    # hacemos un query para el Projects_Costs que usaremos en adelante, si no existe, lo creamos
    project_cost = db.Projects_Costs.get(project = db.Projects.get(contract_number = contract_number, finished = None))
    if project_cost == None:
        project_cost = db.Projects_Costs(project = db.Projects.get(contract_number = contract_number, finished = None))
    
    parameters = viatic_cost, movilization_cost, num_installers, freight_cost, 1, file_name
    return standardCostCalculation(db, project_cost, parameters)
    
###################################################################
# Desde aqui los metodos auxiliares que calculan todos los costos #
###################################################################

def standardCostCalculation(db, project_cost, parameters):
    file_read = parameters[5] + ".xlsx"

    wb_read = load_workbook(file_read, data_only=True)

    #tercero, calculamos y escribimos los costos de instalacion
    instalationCost.computation(db, wb_read, project_cost, parameters[0], parameters[1], parameters[2], parameters[3], parameters[4])

    #cuarto, calculamos y escribimos los costos de fabricacion
    fabricationCost.computation(db, wb_read, project_cost)

    #segundo, calculamos los costos por materia prima
    materialCost.computation(db, wb_read, project_cost)

    #quinto, calculamos y escribimos los costos adicionales
    
    #por ultimo, actualizamos los datos que faltan en la tabla Projects_Costs
    with db_session:
        project_cost.standard_cost_additionals = 0
    
        profiles_cost = project_cost.standard_cost_profiles
        fittings_cost = project_cost.standard_cost_fittings
        crystals_cost = project_cost.standard_cost_crystals
        project_cost.standard_cost_material = profiles_cost + fittings_cost + crystals_cost
        
        additional_costs = project_cost.standard_cost_additionals
        installation_cost = project_cost.standard_cost_installation
        fabrication_cost = project_cost.standard_cost_fabrication
        project_cost.standard_cost_total = profiles_cost + fittings_cost + crystals_cost + additional_costs + installation_cost + fabrication_cost
        
        commit()
    #si el cálculo es exitoso
    return True
    
####################################################################################################
# Método auxiliar para manejar los casos en que alguno de los valores no están en la base de datos #
####################################################################################################

def getParameter(table, key, exception, warning):
    parameter = 0
    try:
        parameter = table[key].value
    except exception:
        print(warning)
        parameter = 0
    return parameter