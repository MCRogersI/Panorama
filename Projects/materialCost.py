from pony.orm import *
from openpyxl import load_workbook
from math import ceil
from copy import copy


######################################################################################
# Muy simple, se encarga de llamar a todas las funciones que hacen de verdad la pega #
######################################################################################

def computation(db, wb_read, project_cost):
    ws_read_manufacturing = wb_read["Manufacturing"]
    
    #necesitamos algunos parametros basicos para costear
    with db_session:
        from Projects.costs import getParameter
        
        warning = ' Aviso: el royalty asociado a la importacion de productos no se encuentra registrado en la base de datos. Se considerara como 0.'
        royalty_porcentaje = getParameter(db.Operating_Parameters, "Royalty", pony.orm.core.ObjectNotFound, warning)
        royalty_porcentaje = royalty_porcentaje/100.0
        
        warning = ' Aviso: el porcentaje asociado a seguros, transporte, aduana y flete de productos no se encuentra registrado en la base de datos. Se considerara como 0.'
        seguros_porcentaje = getParameter(db.Operating_Parameters, "Seguros, Transporte, Aduana, Flete", pony.orm.core.ObjectNotFound, warning)
        seguros_porcentaje = seguros_porcentaje/100.0
        
        warning = ' Aviso: el valor actual del Euro no se encuentra registrado en la base de datos. Se considerara como 700.'
        euro = getParameter(db.Operating_Parameters, "Valor del euro", pony.orm.core.ObjectNotFound, warning)
        if euro == 0:
            euro = 700
    
    #ahora seguimos escribiendo en el archivo, donde llamamos varias funciones mas
    writeInfoGlass(db, project_cost, ws_read_manufacturing)
    writeInfoProfile(db, project_cost, ws_read_manufacturing, royalty_porcentaje, seguros_porcentaje, euro)
    components_cost = writeInfoComponents(db, ws_read_manufacturing, royalty_porcentaje, seguros_porcentaje, euro)
    sealings_cost = writeInfoSealings(db, ws_read_manufacturing, royalty_porcentaje, seguros_porcentaje, euro)

    #terminamos con la escritura del archivo Resumen
    project_cost.standard_cost_fittings = components_cost + sealings_cost
    
    
    
##############################################################################
# Aca empezamos a escribir la informacion relacionada al proyecto especifico #
##############################################################################

# Partimos por los Glass
def writeInfoGlass(db, project_cost, ws_read_manufacturing):
    #necesitamos algunos parametros, en este caso, los tomamos de la base de datos
    with db_session:
        thickness = str(ws_read_manufacturing.cell(row = 7, column = 2).value)
        cristal = db.Crystals_Parameters.get(thickness = thickness)
        #recuperamos el costo de cristal por metro cuadrado
        warning = ' Aviso: el costo del cristal por metro cuadrado no se encuentra registrado en la base de datos. Se considerara como 0.'
        try:
            costo_cristal_por_m2 = cristal.square_meter_cost
        except AttributeError:
            print(warning)
            costo_cristal_por_m2 = 0
        #recuperamos el factor de perdida de cristales
        warning = ' Aviso: el factor de perdida del cristal no se encuentra registrado en la base de datos. Se considerara como 0.'
        try:
            factor_perdida = cristal.waste_factor
        except AttributeError:
            print(warning)
            factor_perdida = 0
        #lo pasamos de porcentaje a fraccion
        factor_perdida = factor_perdida/100.0

    #primero, calculamos los metros lineales
    metros_lineales = 0
    width = ws_read_manufacturing.cell(row = 7, column = 4).value
    next_row = 8
    while(width > 0): #float(width.replace(',', '.')) en caso que width sea leido como string
        metros_lineales = metros_lineales + width/1000
        width = ws_read_manufacturing.cell(row = next_row, column = 4).value
        next_row = next_row + 1
    
    #segundo, calculamos los metros cuadrados
    metros_cuadrados = 0
    height = ws_read_manufacturing.cell(row = 7, column = 3).value
    width = ws_read_manufacturing.cell(row = 7, column = 4).value
    next_row = 8
    while(width > 0): #float(width.replace(',', '.')) en caso que width sea leido como string
        metros_cuadrados = metros_cuadrados + (height*width)/1000000
        height = ws_read_manufacturing.cell(row = next_row, column = 3).value
        width = ws_read_manufacturing.cell(row = next_row, column = 4).value
        next_row = next_row + 1
        
    #calculamos el ultimo dato necesario
    costo_estandar_cristales = metros_cuadrados*costo_cristal_por_m2/(1 - factor_perdida)
    
    #escribimos efectivamente el costo estandar de cristales en la base de datos
    with db_session:
        project_cost.standard_cost_crystals = costo_estandar_cristales
    
 
# Pasamos a los Profiles
def writeInfoProfile(db, project_cost, ws_read_manufacturing, royalty_porcentaje, seguros_porcentaje, euro):
    #primero, para el Perfil Superior
    with db_session:
        #recuperamos el factor de perdida
        warning = ' Aviso: el factor de perdida para Perfil Superior no se encuentra registrado en la base de datos. Se considerara como 0.'
        try:
            factor_perdida = db.Profiles_Fittings_Parameters.get(name = "Upper").waste_factor
        except AttributeError:
            print(warning)
            factor_perdida = 0
        #pasamos de porcentaje a fraccion
        factor_perdida = factor_perdida/100.0
        length_coords = [48, 2]
        code_coords = [46, 2]
        profile1 = writeProfile(db, ws_read_manufacturing, royalty_porcentaje, seguros_porcentaje, euro, factor_perdida, length_coords, code_coords)
    
    #ahora, el Perfil Inferior
        #recuperamos el factor de perdida
        warning = ' Aviso: el factor de perdida para Perfil Inferior no se encuentra registrado en la base de datos. Se considerara como 0.'
        try:
            factor_perdida = db.Profiles_Fittings_Parameters.get(name = "Lower").waste_factor
        except AttributeError:
            print(warning)
            factor_perdida = 0
        #pasamos de porcentaje a fraccion
        factor_perdida = factor_perdida/100.0
        length_coords = [60, 2]
        code_coords = [58, 2]
        profile2 = writeProfile(db, ws_read_manufacturing, royalty_porcentaje, seguros_porcentaje, euro, factor_perdida, length_coords, code_coords)
    
    #por ultimo, el Perfil Telescopico
        #recuperamos el factor de perdida
        warning = ' Aviso: el factor de perdida para Perfil Telescopico no se encuentra registrado en la base de datos. Se considerara como 0.'
        try:
            factor_perdida = db.Profiles_Fittings_Parameters.get(name = "Teleskopic").waste_factor
        except AttributeError:
            print(warning)
            factor_perdida = 0
        #pasamos de porcentaje a fraccion
        factor_perdida = factor_perdida/100.0
        length_coords = [72, 2]
        code_coords = [70, 2]
        profile3 = writeProfile(db, ws_read_manufacturing, royalty_porcentaje, seguros_porcentaje, euro, factor_perdida, length_coords, code_coords)
    
    #ahora, los Glassing Beads
        #recuperamos el factor de perdida
        warning = ' Aviso: el factor de perdida para Glassing Beads no se encuentra registrado en la base de datos. Se considerara como 0.'
        try:
            factor_perdida = db.Profiles_Fittings_Parameters.get(name = "Glassing beads").waste_factor
        except AttributeError:
            print(warning)
            factor_perdida = 0
        #pasamos de porcentaje a fraccion
        factor_perdida = factor_perdida/100.0
        length_coords = [47, 13]
        code_coords = [78, 14]
        profile4 = writeProfile(db, ws_read_manufacturing, royalty_porcentaje, seguros_porcentaje, euro, \
                            factor_perdida, length_coords, code_coords, True)
 
    #ahora, los Glassing Beads Covers
        #recuperamos el factor de perdida
        warning = ' Aviso: el factor de perdida para Glassing Bead Cover no se encuentra registrado en la base de datos. Se considerara como 0.'
        try:
            factor_perdida = db.Profiles_Fittings_Parameters.get(name = "Glassing bead cover").waste_factor
        except AttributeError:
            print(warning)
            factor_perdida = 0
        #pasamos de porcentaje a fraccion
        factor_perdida = factor_perdida/100.0
        length_coords = [47, 13]
        code_coords = [79, 14]
        profile5 = writeProfile(db, ws_read_manufacturing, royalty_porcentaje, seguros_porcentaje, euro, \
                            factor_perdida, length_coords, code_coords, True)
                            
        project_cost.standard_cost_profiles = profile1 + profile2 + profile3 + profile4 + profile5
  
 
def writeProfile(db, ws_read_manufacturing, royalty_porcentaje, seguros_porcentaje, euro, \
                    factor_perdida, length_coords, code_coords, bead = False):
    #primero, calculamos los metros lineales
    metros_lineales = 0
    length = ws_read_manufacturing.cell(row = length_coords[0], column = length_coords[1]).value
    next_row = length_coords[0] + 1
    while(length > 0): #float(length.replace(',', '.')) en caso que length sea leido como string
        metros_lineales = metros_lineales + length/1000
        length = ws_read_manufacturing.cell(row = next_row, column = length_coords[1]).value
        next_row = next_row + 1
    if bead:
        metros_lineales = 2*metros_lineales
    metros_lineales_totales = metros_lineales/(1 - factor_perdida)
    
    #segundo, calculamos el costo unitario (pre royalty/seguros/flete)
    code = ws_read_manufacturing.cell(row = code_coords[0], column = code_coords[1]).value
    precio = getByCode(db, code)
    if precio == 0 and bead:
        precio = getByCode(db, 54220002)
    
    #ahora, el precio con el royalty y todo
    precio_royalty = precio * (1 + royalty_porcentaje) * (1 + seguros_porcentaje)
    
    #calculamos el ultimo dato necesario
    costo_estandar = metros_lineales_totales*precio_royalty*euro
    
    #escribimos finalmente la informacion obtenida
    return costo_estandar
        
        
# Ahora a los Components
def writeInfoComponents(db, ws_read_manufacturing, royalty_porcentaje, seguros_porcentaje, euro):
    with db_session:
        #partimos por Components to Glass Panes
        #recuperamos el factor de perdida
        warning = ' Aviso: el factor de perdida para Components to Glass Panes no se encuentra registrado en la base de datos. Se considerara como 0.'
        try:
            factor_perdida = db.Profiles_Fittings_Parameters.get(name = "Components to glass panes").waste_factor
        except AttributeError:
            print(warning)
            factor_perdida = 0
        #pasamos de porcentaje a fraccion
        factor_perdida = factor_perdida/100.0
        rows_read = [126, 134]
        columns_read = [1, 6]
        rows = [22, 30]
        column = 2
        component1 = writeComponent(db, ws_read_manufacturing, royalty_porcentaje, seguros_porcentaje, euro, \
                            factor_perdida, rows_read, columns_read, rows)
        
        #seguimos con Components to component box or profiles
        #recuperamos el factor de perdida
        warning = ' Aviso: el factor de perdida para Components to Component Box or Profiles no se encuentra registrado en la base de datos. Se considerara como 0.'
        try:
            factor_perdida = db.Profiles_Fittings_Parameters.get(name = "Components to component box or profiles").waste_factor
        except AttributeError:
            print(warning)
            factor_perdida = 0
        #pasamos de porcentaje a fraccion
        factor_perdida = factor_perdida/100.0
        rows_read = [140, 150]
        columns_read = [1, 6]
        rows = [36, 45]
        column = 2
        component2 = writeComponent(db, ws_read_manufacturing, royalty_porcentaje, seguros_porcentaje, euro, \
                            factor_perdida, rows_read, columns_read, rows)
        
        #ahora Components to component box
        #recuperamos el factor de perdida
        warning = ' Aviso: el factor de perdida para Components to Component Box no se encuentra registrado en la base de datos. Se considerara como 0.'
        try:
            factor_perdida = db.Profiles_Fittings_Parameters.get(name = "Components to component box").waste_factor
        except AttributeError:
            print(warning)
            factor_perdida = 0
        #pasamos de porcentaje a fraccion
        factor_perdida = factor_perdida/100.0
        rows_read = [126, 148]
        columns_read = [8, 15]
        rows = [51, 73]
        column = 2
        component3 = writeComponent(db, ws_read_manufacturing, royalty_porcentaje, seguros_porcentaje, euro, \
                            factor_perdida, rows_read, columns_read, rows)
                            
        return component1 + component2 + component3


def writeComponent(db, ws_read_manufacturing, royalty_porcentaje, seguros_porcentaje, euro, \
                        factor_perdida, rows_read, columns_read, rows):
    precio_total = 0
    for i in range(0, rows_read[1] - rows_read[0] + 1):
        #vemos primero unidades totales requeridas del producto
        units = ws_read_manufacturing.cell(column = columns_read[1], row = rows_read[0] + i).value
        if units == None:
            units = 0
        if units > 500:
            units = units/1000
        units_total = units/(1 - factor_perdida)
        
        #ahora recuperamos el codigo y precio del producto
        code = ws_read_manufacturing.cell(column = columns_read[0], row = rows_read[0] + i).value
        if i > 19:
            code = 51220110
        precio = getByCode(db, code)
        if precio == 0 and code == 54220001:
            precio, quantity = 1.19, 1
        precio_royalty = precio * (1 + royalty_porcentaje) * (1 + seguros_porcentaje)
        precio_total = precio_total + units_total * precio_royalty * euro #parece faltar algo como units_total/quantity
        
    return precio_total

        
# Ahora a los Sealings
def writeInfoSealings(db, ws_read_manufacturing, royalty_porcentaje, seguros_porcentaje, euro):
    #primero, calculamos la cantidad de hojas
    hojas = 0
    width = ws_read_manufacturing.cell(row = 7, column = 4).value
    next_row = 8
    while(width > 0): #float(width.replace(',', '.')) en caso que width sea leido como string
        hojas = hojas + 1
        width = ws_read_manufacturing.cell(row = next_row, column = 4).value
        next_row = next_row + 1
        
    #ahora recuperamos el codigo y precio del producto
    precio_total = 0
    for code in [54043034, 54043044, 54043064, 54043014, 54043024, 54043054, 54042014, 54042024, 54200204, 54200104, 11116200, 11116201]:
        if code in [54043034, 54043044, 54043064]:
            precio = getByCode(db, code)
            precio_royalty = precio * (1 + royalty_porcentaje) * (1 + seguros_porcentaje)
            precio_total = precio_total + hojas * precio_royalty * euro /(1 - 0.03)#parece faltar algo como units_total/quantity
        
            #ahora escribimos los datos que necesitamos
    return precio_total
        

    
# Funcion auxiliar muy usada, bastante self-descriptive, auto-descriptiva        
def getByCode(db, code):
    with db_session:
        if code == None:
            return 0
        sku = db.Stock.get(id = code)
        if(sku != None):
            return sku.price
        return 0