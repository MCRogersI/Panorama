from pony.orm import *
from datetime import date, datetime, timedelta
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from os import remove
import Stock.features as Sf
from Planning.features import sumDays, substractDays, doPlanning, changePriority, datesOverlap
from Planning.reports import createReport
import pandas
from tabulate import tabulate

def noneInt(x):
    if x == None:
        return 0
    else:
        return x

def createProject(db, contract_number = None, version = None, client_address = None, client_comuna = None,
                  client_name = None, client_rut = None, linear_meters = None, square_meters = None, year = None, month = None,
                  day = None, crystal_leadtime = None, sale_date = None, sale_price = None,estimated_cost = None, sale_date_year=None,sale_date_month=None,sale_date_day=None):
    import Planning.features as PLf
    with db_session:
        deadline = date(int(year), int(month), int(day))
        p = db.Projects(contract_number = contract_number, version = version, client_address = client_address, client_comuna=client_comuna, 
                            client_name = client_name, client_rut = client_rut, linear_meters = linear_meters, square_meters = square_meters, deadline = deadline, crystal_leadtime = crystal_leadtime, sale_date = sale_date, sale_price = sale_price)
        p.priority = select(p for p in db.Projects if p.finished == None).count()
        
        commit()
    PLf.doPlanning(db)
    
def printProjects(db):
    with db_session:
        print('')
        pr = db.Projects.select()
        data = [p.to_dict() for p in pr]
        df = pandas.DataFrame(data, columns = ['contract_number','version','client_address','client_comuna','client_name','client_rut','linear_meters','square_meters','deadline','priority','real_linear_meters'\
        ,'estimated_cost','real_cost','sale_price','fixed_planning','fixed_priority','crystal_leadtime','sale_date','finished'])                    
        df.columns = ['Número de Contrato','Versión','Dirección','Comuna','Nombre','RUT','Metros Lineales','Metros Cuadrados','Plazo Pactado Proyecto','Prioridad','Metros Lineales Reales'\
        ,'Costo Estimado','Costo Real','Precio de Venta','Planificación Fija','Prioridad Fijada','Tiempo Entrega Cristales','Fecha de Venta','Proyecto Finalizado']
        print( tabulate(df.drop(df.columns[[9,10,11,12,13,14,15,16,17,18]], axis = 1), headers='keys', tablefmt='psql'))
        print( tabulate(df.drop(df.columns[[0,1,2,3,4,5,6,7,8]], axis = 1), headers='keys', tablefmt='psql'))

def printCurrentProjects(db):
    with db_session:
        print('')
        pr = select(p for p in db.Projects if p.finished == None)
        data = [p.to_dict() for p in pr]
        df = pandas.DataFrame(data, columns = ['contract_number','version','client_address','client_comuna','client_name','client_rut','linear_meters','square_meters','deadline','priority','real_linear_meters'\
        ,'estimated_cost','real_cost','sale_price','fixed_planning','fixed_priority','crystal_leadtime','sale_date','finished'])                    
        df.columns = ['Número de Contrato','Versión','Dirección','Comuna','Nombre','Rut','Metros Lineales','Metros Cuadrados','Plazo Pactado Proyecto','Prioridad','Metros Lineales Reales'\
        ,'Costo Estimado','Costo Real','Precio de Venta','Planificación Fija','Prioridad Fijada','Tiempo Entrega Cristales','Fecha de Venta','Proyecto Finalizado']
        print( tabulate(df.drop(df.columns[[9,10,11,12,13,14,15,16,17,18]], axis = 1), headers='keys', tablefmt='psql'))
        print( tabulate(df.drop(df.columns[[0,1,2,3,4,5,6,7,8,18]], axis = 1), headers='keys', tablefmt='psql'))
        
def printFinishedProjects(db):
    with db_session:
        print('')
        pr = select(p for p in db.Projects if p.finished == True)
        data = [p.to_dict() for p in pr]
        df = pandas.DataFrame(data, columns = ['contract_number','version','client_address','client_comuna','client_name','client_rut','linear_meters','square_meters','deadline','priority','real_linear_meters'\
        ,'estimated_cost','real_cost','sale_price','fixed_planning','fixed_priority','crystal_leadtime','sale_date','finished'])                    
        df.columns = ['Número de Contrato','Versión','Dirección','Comuna','Nombre','Rut','Metros Lineales','Metros Cuadrados','Plazo Pactado Proyecto','Prioridad','Metros Lineales Reales'\
        ,'Costo Estimado','Costo Real','Precio de Venta','Planificación Fija','Prioridad Fijada','Tiempo Entrega Cristales','Fecha de Venta','Proyecto Finalizado']
        print( tabulate(df.drop(df.columns[[8,9,10,11,12,13,14,15,16,17,18]], axis = 1), headers='keys', tablefmt='psql'))
        print( tabulate(df.drop(df.columns[[0,1,2,3,4,5,6,7,8,18]], axis = 1), headers='keys', tablefmt='psql'))
        
        
def editProject(db, contract_number, new_client_address = None, new_client_comuna = None, new_client_name = None, new_client_rut = None , new_linear_meters = None,new_square_meters = None, new_real_linear_meters = None, new_deadline = None, new_estimated_cost = None, new_real_cost = None, new_crystal_leadtime = None):
    with db_session:
        projects = select(p for p in db.Projects if contract_number == contract_number)
        for p in projects:
            if new_client_address != None:
                p.client_addres = new_client_address
            if new_client_comuna != None:
                p.client_comuna = new_client_comuna
            if new_client_name != None:
                p.client_name = new_client_name
            if new_client_rut != None:
                p.client_rut = new_client_rut
            if new_linear_meters != None:
                p.linear_meters = new_linear_meters
            if new_square_meters != None:
                p.square_meters = new_square_meters
            if new_deadline != None:
                p.deadline = new_deadline
            if new_real_linear_meters != None:
                p.real_linear_meters = new_real_linear_meters
            if new_estimated_cost != None:
                p.estimated_cost = new_estimated_cost
            if new_real_cost != None:
                p.real_cost = new_real_cost
            if new_crystal_leadtime != None:
                p.crystal_leadtime = new_crystal_leadtime
        commit()

def deleteProject(db, contract_number):
    with db_session:
        new_priority = select(p for p in db.Projects if p.finished == None).count()
        changePriority(db, contract_number, new_priority)
        select(p for p in db.Projects if p.contract_number == contract_number and p.finished == None).delete()
        commit()

def finishProject(db, contract_number):
    with db_session:
        #recalculamos las prioridades
        new_priority = select(p for p in db.Projects if p.finished == None).count()
        changePriority(db, contract_number, new_priority)
        db.Projects.get(contract_number = contract_number, finished = None).priority = -1
        #cambiamos el Project a finished = True
        db.Projects.get(contract_number = contract_number, finished = None).finished = True
        #eliminamos las Employees_Restrictions asociadas al proyecto
        select(er for er in db.Employees_Restrictions if er.project.contract_number == contract_number).delete()
        #eliminamos las Employees_Tasks asociadas al proyecto
        select(et for et in db.Employees_Tasks if et.task.project == db.Projects.get(contract_number = contract_number, finished = None) and et.task.effective_end_date == None).delete()
        commit()

def finishFailedProject(db, contract_number):
    with db_session:
        db.Projects.get(contract_number = contract_number, finished = None).priority = -1
        #cambiamos el Project a finished = True
        db.Projects.get(contract_number = contract_number, finished = None).finished = True
        #eliminamos las Employees_Restrictions asociadas al proyecto
        select(er for er in db.Employees_Restrictions if er.project.contract_number == contract_number).delete()
        #eliminamos las Employees_Tasks asociadas al proyecto
        select(et for et in db.Employees_Tasks if et.task.project == db.Projects.get(contract_number = contract_number, finished = None) and et.task.effective_end_date == None).delete()
        commit()
def getNumberConcurrentProjects(db, contract_number, date):
    ''' Método que entrega la cantidad de proyectos que son realizados en la misma comuna,
    en la misma fecha, para calcular los costos de transporte si es que hay más de uno en un lugar
    en la misma fecha '''
    p = db.Projects.get(contract_number = contract_number, finished = None)
    candidates = select(pr for pr in db.Projects if pr.client_comuna == p.client_comuna and pr != p)
    quant = 1
    for pr in candidates:
        inst = db.Tasks.get(skill = 4, project = pr)
        et_ins = select(et for et in db.Employees_Tasks if et.task == inst)
        for et in et_ins:
            if(et.planned_initial_date == date):
                quant += 1
    return quant
        
def createTask(db, id_skill, contract_number, original_initial_date, original_end_date, effective_initial_date = None, effective_end_date = None):
    with db_session:
        t = db.Tasks(skill = id_skill, project = db.Projects.get(contract_number = contract_number, finished = None), original_initial_date = original_initial_date, original_end_date = original_end_date)
        commit()


def editTask(db , id_skill, contract_number, effective_initial_date = None, effective_end_date = None):
    with db_session:
        t = db.Tasks.get(skill = db.Skills[id_skill], project = db.Projects.get(contract_number = contract_number, finished = None), failed = None)
        if effective_initial_date != None:
            t.effective_initial_date = effective_initial_date
            #vemos si la tarea se inició con atraso, si fue así, entonces llamamos a createDelay() con la diferencia de días
            #para esto, primero recuperamos algún Employees_Task que contenga esta tarea
            et = select(et for et in db.Employees_Tasks if et.task == t).first()
            # if effective_initial_date > et.planned_initial_date:
                # print(" La fecha entregada indica un atraso respecto a lo planificado, por tanto, el programa quizás deba realizar una re-planificación.")
                # delay = (effective_initial_date - et.planned_initial_date).days
                # createDelay(db, t.project.contract_number, t.skill.id, delay)
        if effective_end_date != None:
            t.effective_end_date = effective_end_date
            #vemos si la tarea se terminó con atraso, si fue así, entonces llamamos a createDelay() con la diferencia de días
            #para esto, primero recuperamos algún Employees_Task que contenga esta tarea
            et = select(et for et in db.Employees_Tasks if et.task == t).first()
            # if effective_end_date > et.planned_end_date:
                # print(" La fecha entregada indica un atraso respecto a lo planificado, por tanto, el programa quizás deba realizar una re-planificación.")
                # delay = (effective_end_date - et.planned_end_date).days
                # createDelay(db, t.project.contract_number, t.skill.id, delay)
        commit()


def deleteTask(db, id_task):
    with db_session:
        db.Tasks[id_task].delete()
        commit()

def printTasks(db):
    with db_session:
        print('')
        ts = db.Tasks.select()
        data = [t.to_dict() for t in ts]
        project =pandas.Series([t.project.contract_number for t in ts], name = 'Proyecto')
        version =pandas.Series([t.project.version for t in ts], name = 'Versión')
        df = pandas.DataFrame(data, columns = ['id','skill','project','original_initial_date','original_end_date','effective_initial_date','effective_end_date','failed','fail_cost'])
        df.columns = ['ID','Tarea','Proyecto-Versión','Fecha de Inicio Original','Fecha de Finalización Original','Fecha de Inicio Efectiva','Fecha de Finalización Efectiva','Falló','Costo de Falla']
        df2 = pandas.concat([df,project,version], axis = 1)
        cols = df2.columns.tolist()
        cols = cols[0:3] + cols[-2:] + cols[4:-3]
        df2 = df2[cols]
        print( tabulate(df2.drop(df2.columns[2], axis = 1), headers='keys', tablefmt='psql'))

def failedTask(db, contract_number, id_skill, fail_cost):
    # import Planning.features as PLf
    with db_session:
        
        #primero, detectamos todos los Tasks asociados a ese Skill (o mayor) y contract_number
        tasks = select(t for t in db.Tasks if t.skill.id >= db.Skills[id_skill].id and t.project.contract_number == contract_number)
        
        #después, detectamos cuál es la última versión en la cuál ese Skill aparece como no fallado
        version = 1
        for t in tasks:
            if t.skill.id == db.Skills[id_skill] and t.project.version > version:
                version = t.project.version
            
        #después, asignamos los costos a la última versión del Skill que no aparece como fallada
        tasks = select(t for t in db.Tasks if t.skill.id >= db.Skills[id_skill].id and t.project.contract_number == contract_number and t.project.version >= version)
        for t in tasks:
            if t.skill.id == db.Skills[id_skill].id:
                task_responsible = t
        #por si no está definido antes y sigue siendo None
        task_responsible.fail_cost = noneInt(task_responsible.fail_cost)
        task_responsible.fail_cost = task_responsible.fail_cost + fail_cost
        for t in tasks:
            if t != task_responsible:
                task_responsible.fail_cost = task_responsible.fail_cost + noneInt(t.fail_cost)
                t.fail_cost = 0
        
        #después, marcamos que las Tasks, para ese Skill y los que lo siguen, falló en las versiones anteriores también
        tasks = select(t for t in db.Tasks if t.skill.id >= db.Skills[id_skill].id and t.project.contract_number == contract_number)
        for t in tasks:
            t.failed = True
        
        #por último, eliminamos las tareas, de la última versión, que no se habían alcanzado a terminar
        tasks = select(t for t in db.Tasks if t.skill.id > id_skill and t.project == db.Projects.get(contract_number = contract_number, finished = None) and t.effective_end_date == None)
        for t in tasks:
            t.delete()

        #terminamos version anterior del proyecto
        last_version = db.Projects.get(contract_number = contract_number, finished = None)
        priority = last_version.priority
        last_version.priority = -1
        # cambiamos el Project a finished = True
        last_version.finished = True
        # guardamos las Employees_Restrictions asociadas al proyecto
        restrictions = select(er for er in db.Employees_Restrictions if
               er.project == last_version)
        # eliminamos las Employees_Tasks asociadas al proyecto
        select(et for et in db.Employees_Tasks if
               et.task.project == last_version and et.task.effective_end_date == None).delete()
        
        #creamos nueva version del proyecto
        new_version = db.Projects(contract_number = last_version.contract_number, version = last_version.version + 1, client_address = last_version.client_address,
                        client_comuna = last_version.client_comuna, client_name = last_version.client_name, client_rut = last_version.client_rut, 
                        linear_meters = last_version.linear_meters, square_meters = last_version.square_meters, deadline = last_version.deadline, 
                        crystal_leadtime = last_version.crystal_leadtime, sale_date = last_version.sale_date, sale_price = last_version.sale_price)
        new_version.priority = priority
        for er in restrictions:
            er.project = new_version

        commit()
        doPlanning(db)

def createDelay(db, task, delay):
    '''Este método ingresa un delay en la tarea con id skill = skill_id del proyecto con id = contract_number, alargando el end date en delay días.    
    Todo está con ints porque si no, había problemas con los reverses, ver aquí: https://docs.ponyorm.com/relationships.html 
    Después de correr la planificacion en "delay" cantidad de dias, revisa si la planificacion que queda es factible. Si no lo es, 
    realiza una nueva planificacion'''
    with db_session:
        db.Tasks_Delays(task = task, delay = delay)
        emp_tasks = select(et for et in db.Employees_Tasks if et.task == task)
        #corremos la planned_end_date de los Employees_Tasks pertinentes
        if delay > 0:
            for et in emp_tasks:
                et.planned_end_date = sumDays(et.planned_end_date, delay)
        else:
            for et in emp_tasks:
                et.planned_end_date = substractDays(et.planned_end_date, -1*delay)

        commit()
        doPlanning(db)
        
        #No borrar: es forma efectiva de revisar si una planificación es factible
        # si la planificacion que queda no es factible, replanificamos, dejando fijo el proyecto en cuestion
        # if createReport(db, None, True) == False:
            # fixed_planning = project.fixed_planning
            # project.fixed_planning = True
            # doPlanning(db)
            # project.fixed_planning = fixed_planning

# métodos asociados a Employees_Activities (llamados en usuario.py de carpeta Employees)
def createEmployeeActivity(db, employee, activity, initial_year, initial_month, initial_day, end_year, end_month, end_day):
    '''
    Crea una actividad para un empleado de tipo licencia, vacaciones u otros. en caso de que el empleado tenga asignada una tarea en las fechas de la actividad, se replanifica
    '''
    import Planning.features as PLf
    initial_date = date(int(initial_year), int(initial_month), int(initial_day))
    end_date = date(int(end_year), int(end_month), int(end_day))
    with db_session:
        db.Employees_Activities(employee = employee, activity = activity, initial_date = initial_date, end_date = end_date)
        commit()
    if activitiyEmployeeOverlap(db, employee, initial_date, end_date):
        PLf.doPlanning(db)

        
def activitiyEmployeeOverlap(db, employee, initial_date, end_date):
    '''
    Este metodo revisa si un empleado tiene tareas asignadas durante las fechas impuestas y ,de ser cierto, deja móviles dichas actividades para una
    futura replanificación. Es un método auxiliar, por lo que no es recomendable usarlo directamente.
    '''
    with db_session:
        emp_tasks = select(et for et in db.Employees_Tasks if et.employee.id == employee)
        for et in emp_tasks:
            if datesOverlap(initial_date, end_date, et.planned_initial_date, et.planned_end_date):
                return True
        commit()
    return False

        
def deleteEmployeeActivity(db, id_employee_activity):
    with db_session:
        db.Employees_Activities[id_employee_activity].delete()
        commit()
        
def printEmployeesActivities(db):
    with db_session:
        print('')
        ea = db.Employees_Activities.select()
        data = [e.to_dict() for e in ea]
        df = pandas.DataFrame(data, columns = ['id','employee','activity','initial_date','end_date'])                    
        df.columns = ['ID','Empleado','Actividad','Fecha de inicio','Fecha de finalización']
        df.loc[df['Actividad'] == 1, 'Actividad'] = db.Activities[1].description
        df.loc[df['Actividad'] == 2, 'Actividad'] = db.Activities[2].description
        print( tabulate(df, headers='keys', tablefmt='psql'))
        
def createProjectActivity(db, project, activity, initial_year, initial_month, initial_day, end_year, end_month, end_day):
    import Planning.features as PLf
    initial_date = date(int(initial_year), int(initial_month), int(initial_day))
    end_date = date(int(end_year), int(end_month), int(end_day))
    with db_session:
        db.Projects_Activities(project = project, activity = activity, initial_date = initial_date, end_date = end_date)
        commit()
        p = db.Projects.get(contract_number = project.contract_number, finished = None)
    if activitiyProjectOverlap(db, p, initial_date, end_date):
        doPlanning(db)

def activitiyProjectOverlap(db, project, initial_date, end_date):
    '''
    Este metodo revisa si un proyecto tiene tareas asignadas durante las fechas impuestas y ,de ser cierto, deja móviles dichas actividades para una
    futura replanificación. Es un método auxiliar, por lo que no es recomendable usarlo directamente.
    '''
    with db_session:
        emp_tasks = select(et for et in db.Employees_Tasks if et.task.project == project)
        for et in emp_tasks:
            if datesOverlap(initial_date, end_date, et.planned_initial_date, et.planned_end_date):
                # Se desfija el proyecto para que no haga planificaciones infactibles
                et.task.project.fixed_planning = False
                return True
        commit()
    return False
        
def deleteProjectActivity(db, id_project_activity):
    with db_session:
        db.Projects_Activities[id_project_activity].delete()
        commit()
        
def printProjectsActivities(db):
    with db_session:
        print('')
        pr = db.Projects_Activities.select()
        data = [p.to_dict() for p in pr]
        df = pandas.DataFrame(data, columns = ['id','project','activity','initial_date','end_date'])                    
        df.columns = ['ID','Numero de contrato','Actividad','Fecha de inicio','Fecha de finalización']
        df.loc[df['Actividad'] == 3, 'Actividad'] = db.Activities[3].description
        print( tabulate(df, headers='keys', tablefmt='psql'))
        
        
        
        
#para especificar fechas de quiebre de stock
def createStockShortage(db, activity, initial_year, initial_month, initial_day, end_year, end_month, end_day):
    initial_date = date(int(initial_year), int(initial_month), int(initial_day))
    end_date = date(int(end_year), int(end_month), int(end_day))
    with db_session:
        db.Stock_Shortages(activity = activity, initial_date = initial_date, end_date = end_date)
        commit()
    if stockShortageOverlap(db, initial_date, end_date):
        doPlanning(db)
        
#para revisar si el quiebre de stock coincide con la planificación de fabricadores, en ese caso, se debe replanificar
def stockShortageOverlap(db, initial_date, end_date):
    '''
    Este metodo revisa si un empleado tiene tareas asignadas durante las fechas impuestas y ,de ser cierto, deja móviles dichas actividades para una
    futura replanificación. Es un método auxiliar, por lo que no es recomendable usarlo directamente.
    '''
    with db_session:
        emp_tasks = select(et for et in db.Employees_Tasks if et.task.skill.id == 3)
        for et in emp_tasks:
            if datesOverlap(initial_date, end_date, et.planned_initial_date, et.planned_end_date):
                return True
        commit()
    return False
        
#para eliminar fechas de quiebre de stock
def deleteStockShortage(db, id_stock_shortage):
    with db_session:
        db.Stock_Shortages[id_stock_shortage].delete()
        commit()

#para mostrar las fechas de quiebre de stock ingresadas
def printStockShortages(db):
    with db_session:
        print('')
        pr = db.Stock_Shortages.select()
        data = [p.to_dict() for p in pr]
        df = pandas.DataFrame(data, columns = ['id','activity','initial_date','end_date'])
        df.columns = ['ID','Actividad','Fecha de inicio','Fecha de finalización']
        df.loc[df['Actividad'] == 4, 'Actividad'] = db.Activities[4].description
        print( tabulate(df, headers='keys', tablefmt='psql'))
        
        
def getListProducts(db):
    with db_session:
        wb = load_workbook('Products.xlsx')
        ws = wb['Hoja2']
        r = 13
        while ws.cell(row = r, column = 2 ).value != None:
            stock_id = int(ws.cell(row = r, column = 2).value)
            stock_engname = ws.cell(row = r, column = 4).value
            stock_europrice = ws.cell(row = r, column = 5).value
            stock_packingquantity = int(ws.cell(row = r, column = 7).value)
            Sf.createSku(db, stock_id, stock_engname, stock_europrice, stock_packingquantity*50, stock_packingquantity*75, 0.03)
            #Asumimos que el nivel crítico es 50 veces el packing quantity, por mientras. 0 = estaba vacío en el excel. 
            #Asumimos además que la cantidad real (para hacer correr el programa, en realidad al comenzar a usar el 
            #software se debería saber las cantidades reales de todos los SKUs) es 1.5 veces el nivel crítico.
            #Asumimos que el factor de pérdida es 0.03 para todos los SKUs, eventualmente la tabla Products debería tener 
            #el factor de pérdida asociado al código del SKU.
            r += 1
        commit()


def getProjectFeatures(db, contract_number):
    ''' Método para obtener los parámetros de un proyecto desde un archivo de excel estandarizado '''
    wb = load_workbook('EjemploPropuestaProyecto '+str(contract_number)+'.xlsx')
    ws = wb['Edif A_Hoja Corte']

    with db_session:
        #Asumiremos, para fijar una fecha inicial, que los engagement se realizarán al comienzo de la fabricación
        p = db.Projects.get(contract_number, finished = None)
        task_aux = db.Tasks.get(skill = 3, project = p)
        et_fab = db.Employees_Tasks.get(task = task_aux)
        withdrawal = et_fab.planned_initial_date
        glass_id = int(ws.cell(row = 16, column = 2).value)
        glass_m2 = ws.cell(row = 50, column  = 3).value
        Sf.createEngagement(db, contract_number, [(glass_id, glass_m2)], withdrawal_date = withdrawal)
        # glass_ml = ws.cell(row = 51, column  = 3).value#no es necesario
        upper_profile_id = int(ws.cell(row = 58, column  = 2).value)
        upper_profile_ml = ws.cell(row = 70, column  = 3).value
        Sf.createEngagement(db, contract_number, [(upper_profile_id, upper_profile_ml)], withdrawal_date = withdrawal)
        lower_profile_id = int(ws.cell(row = 72, column  = 2).value)
        lower_profile_ml = ws.cell(row = 84, column  = 3).value
        Sf.createEngagement(db, contract_number, [(lower_profile_id, lower_profile_ml)], withdrawal_date = withdrawal)
        teles_profile_id = int(ws.cell(row = 86, column  = 2).value)
        teles_profile_ml = ws.cell(row = 98, column  = 3).value
        Sf.createEngagement(db, contract_number, [(teles_profile_id, teles_profile_ml)], withdrawal_date = withdrawal)
        glassing_bead_id = int(ws.cell(row = 98, column = 14).value)
        glassing_bead_ml = ws.cell(row = 100, column  = 14).value
        glassing_bead_price = ws.cell(row = 105, column  = 15).value#no sé donde podría utilizarse
        Sf.createEngagement(db, contract_number, [(glassing_bead_id, glassing_bead_ml)], withdrawal_date = withdrawal)

        #hasta acá no debería ser un problema mantener el formato.
        #Components to glass panes:
        c1 = 142
        while ws.cell(row = c1, column = 1).value != None:
            ide =  ws.cell(row = c1, column = 1).value
            quantity  =  ws.cell(row = c1, column = 6).value
            Sf.createEngagement(db, contract_number, [(ide, quantity)], withdrawal_date = withdrawal)
            c1 +=1
        # Components to component box or profiles:
        c2 = c1+1
        while  ws.cell(row = c2, column = 1).value != None:
            ide =  ws.cell(row = c2, column = 1).value
            quantity  =  ws.cell(row = c2, column = 6).value
            Sf.createEngagement(db, contract_number, [(ide, quantity)], withdrawal_date = withdrawal)
            c2 +=1
        #Components to component box
        c3 = 142
        while  ws.cell(row = c3, column = 8).value != None:
            ide =  ws.cell(row = c2, column = 8).value
            quantity  =  ws.cell(row = c2, column = 14).value
            Sf.createEngagement(db, contract_number, [(ide, quantity)], withdrawal_date = withdrawal)
            c3 +=1
        #Sealings:
        c4 = 181
        while ws.cell(row = c4, column = 1).value != None:
            ide =  ws.cell(row = c4, column = 1).value
            quantity  =  ws.cell(row = c4, column = 3).value
            Sf.createEngagement(db, contract_number, [(ide, quantity)], withdrawal_date = withdrawal)
            c4 += 1
        commit()







    

    
    
    
    
    
    
    
    
    
    
    
    
    
    
