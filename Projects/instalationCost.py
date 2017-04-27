from openpyxl.styles import Font, PatternFill, colors
from math import ceil

def computation(wb_read, wb_written, tarifa_viaticos, tarifa_movilizacion, numero_instaladores, costo_flete, tipo_instalador): #en este caso, 1 significa "interno", 0 significa "externo"
    ws_written = wb_written.create_sheet("Costo Estandar Instalacion")
    ws_read_measures = wb_read["Measures"]
    ws_read_manufacturing = wb_read["Manufacturing"]
    
    #parametros fijos, por ahora uso los que venian en el archivo de ejemplo
    rendimiento_diario = 6
    factor_errores_instalacion = 0.05
    tarifa_instalacion_interna = 8750 #esta es la interna, por metro lineal, la externa es aprox. 30000
    tarifa_instalacion_externa = 30000 #esta es la interna, por metro lineal, la externa es aprox. 30000
    if tipo_instalador == 1:
        tarifa_instalacion = tarifa_instalacion_interna
    else:
        tarifa_instalacion = tarifa_instalacion_externa
    
    #parametros que pueden cambiar por proyecto, pero que hay que especificar en este caso, preguntar al papa    
    # numero_instaladores = 1
    # costo_flete = 0
    
    #parametros que deben obtenerse del archivo en cuestion, si el formato es suficientemente estandar
    comuna = ws_read_measures.cell(row = 7, column = 15).value
    metros_lineales = linearMeters(ws_read_manufacturing)
    
    #ahora partimos escribiendo en el archivo
    writeTitles(ws_written)
    
    #ahora terminamos de escribir en el archivo
    writeInfo(ws_written, comuna, metros_lineales, numero_instaladores, rendimiento_diario, tipo_instalador, factor_errores_instalacion, \
                costo_flete, tarifa_viaticos, tarifa_movilizacion, tarifa_instalacion)
    
    
    
def linearMeters(ws_read_manufacturing):
    metros_lineales = 0
    width = ws_read_manufacturing.cell(row = 7, column = 4).value
    next_row = 8
    while(width > 0): #float(width.replace(',', '.')) en caso que width sea leido como string
        metros_lineales = metros_lineales + width/1000
        width = ws_read_manufacturing.cell(row = next_row, column = 4).value
        next_row = next_row + 1
    return metros_lineales

    
    
def writeTitles(ws_written):
    #aplicamos ciertos cambios al ancho y formato de algunas columnas y celdas
    ws_written.column_dimensions["B"].width = 59
    fondo_verde = PatternFill(start_color = colors.DARKGREEN, end_color = colors.DARKGREEN, fill_type='solid')
    ws_written['B15'].fill = fondo_verde
    ws_written['B17'].fill = fondo_verde
    
    #escribimos efectivamente los titulos de la columna
    rows = [4, 5, 6, 7, 8, 9, 11, 12, 13, 14, 15, 16, 17]
    columns = [2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2]
    texts = ["CoMUNA DE INSTALACIoN", "ML", "NUMERo DE INSTALADoRES ( EQUIPoS )", "RENDIMIENTo DIARIo", "TIPo DE INSTALADoR", \
                "TIEMPo ESTIMADo DE INSTALACIoN ( DIAS )", \
                "CoSTo DE INSTALACIoN", "CoSTo DE FLETE SISTEMA", "VIATICo INSTALADoRES", "INSTALACIoN", \
                "CoSTo ESTANDAR DE INSTALACIoN ANTES DE FALLAS CALIDAD", "FACToR DE ERRoRES DE INSTALACIoN", "CoSTo ESTANDAR DE INSTALACIoN"]
    for i in range(0, len(rows)):
        ws_written.cell(row = rows[i], column = columns[i], value = texts[i])
        
        
        
def writeInfo(ws_written, comuna, metros_lineales, numero_instaladores, rendimiento_diario, tipo_instalador, factor_errores_instalacion, \
                   costo_flete, tarifa_viaticos, tarifa_movilizacion, tarifa_instalacion):
    #aplicamos ciertos cambios al ancho y formato de algunas columnas y celdas
    ws_written.column_dimensions["C"].width = 14
    fondo_amarillo = PatternFill(start_color = colors.YELLOW, end_color = colors.YELLOW, fill_type='solid')
    fondo_verde = PatternFill(start_color = colors.DARKGREEN, end_color = colors.DARKGREEN, fill_type='solid')
    ws_written['C8'].fill = fondo_amarillo
    ws_written['C14'].fill = fondo_amarillo
    ws_written['C15'].fill = fondo_verde
    ws_written['C17'].fill = fondo_verde
    
    #conseguimos algunos de los datos necesarios para escribir la informacion, partiendo por el tipo de instalador (interno/externo)
    tipo_instalador_texto = "INTERNo"
    if tipo_instalador == 0:
        tipo_instalador_texto = "EXTERNo"
    
    #seguimos con el numero de dias necesarios para la instalacion
    tiempo_estimado = ceil(metros_lineales/(numero_instaladores * rendimiento_diario))
    
    #seguimos con el costo de flete, asumimos para este caso que ya viene dado
    
    #seguimos con el costo de viaticos, aca tengo dudas importantes de si se calcula asi o no
    viatico = tarifa_viaticos * numero_instaladores + tarifa_movilizacion * numero_instaladores * tiempo_estimado
    
    #seguimos con el costo de instalacion
    instalacion = tarifa_instalacion * metros_lineales
    
    #seguimos con el costo de instalacion antes de fallas
    instalacion_antes_fallas = costo_flete + viatico + instalacion
    
    #terminamos con instalacion despues de fallas
    instalacion_despues_fallas = instalacion_antes_fallas/(1 - factor_errores_instalacion)
    
    #escribimos efectivamente la informacion de la columna
    rows = [4, 5, 6, 7, 8, 9, 12, 13, 14, 15, 16, 17]
    columns = [3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3]
    texts = [comuna, metros_lineales, numero_instaladores, rendimiento_diario, tipo_instalador_texto, tiempo_estimado, \
                costo_flete, viatico, instalacion, instalacion_antes_fallas, \
                factor_errores_instalacion, instalacion_despues_fallas]
    for i in range(0, len(rows)):
        ws_written.cell(row = rows[i], column = columns[i], value = texts[i])