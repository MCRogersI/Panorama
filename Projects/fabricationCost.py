from pony.orm import *
from openpyxl.styles import Font, PatternFill, colors
from math import ceil

def computation(db, wb_read, project_cost):
    ws_read_measures = wb_read["Measures"]
    ws_read_manufacturing = wb_read["Manufacturing"]
    
    #parametros fijos, se toman de la base de datos
    with db_session:
        from Projects.costs import getParameter
        warning = ' Aviso: el monto de remuneraciones fijas asociadas a fabricacion no se encuentra registrado en la base de datos. Se considerara como 0.'
        remuneraciones_fijas = getParameter(db.Operating_Parameters, "Remuneraciones fijas fabrica", pony.orm.core.ObjectNotFound, warning)
        
        warning = ' Aviso: el monto de remuneraciones fijas asociadas a fabricacion no se encuentra registrado en la base de datos. Se considerara como 0.'
        remuneraciones_variables = getParameter(db.Operating_Parameters, "Remuneraciones variables fabrica", pony.orm.core.ObjectNotFound, warning)
        
        warning = ' Aviso: el porcentaje de la venta correspondiente a materiales de fabricacion no se encuentra registrado en la base de datos. Se considerara como 0.'
        porcentaje_venta = getParameter(db.Operating_Parameters, "Porcentaje de la venta correspondiente a materiales de fabricacion", pony.orm.core.ObjectNotFound, warning)
        #lo pasamos de porcentaje a fraccion
        porcentaje_venta = porcentaje_venta/100.0
        
        warning = ' Aviso: el monto de arriendo de la fabrica no se encuentra registrado en la base de datos. Se considerara como 0.'
        arriendo_fabrica = getParameter(db.Operating_Parameters, "Arriendo de fabrica", pony.orm.core.ObjectNotFound, warning)
        
        warning = ' Aviso: el monto de arriendo de la fabrica no se encuentra registrado en la base de datos. Se considerara como 0.'
        depreciacion = getParameter(db.Operating_Parameters, "Depreciacion equipos y herramientas", pony.orm.core.ObjectNotFound, warning)
        
        warning = ' Aviso: el monto gastado en energia, luz, agua y otros en la fabrica no se encuentra registrado en la base de datos. Se considerara como 0.'
        energia_agua_otros = getParameter(db.Operating_Parameters, "Energia, luz, agua, otros", pony.orm.core.ObjectNotFound, warning)
        
        warning = ' Aviso: la venta (neta de IVA) mensual no se encuentra registrada en la base de datos. Se considerara como 0.'
        venta_neta_mensual = getParameter(db.Operating_Parameters, "Venta (neta de IVA) mensual", pony.orm.core.ObjectNotFound, warning)
    
    #parametros que deben obtenerse del archivo en cuestion, si el formato es suficientemente estandar
    metros_lineales = linearMeters(ws_read_manufacturing)
    with db_session:
        warning = ' Aviso: la cantidad de metros lineales vendidos en el mes no se encuentra registrada en la base de datos. Se considerara como la cantidad de metros lineales de este proyecto.'
        metros_lineales_mes = getParameter(db.Operating_Parameters, "Metros lineales vendidos en el mes", pony.orm.core.ObjectNotFound, warning)
        if metros_lineales_mes == 0:
            metros_lineales_mes = metros_lineales
    
    #ahora terminamos de escribir en el archivo
    writeInfo(db, project_cost, metros_lineales, metros_lineales_mes, remuneraciones_fijas, remuneraciones_variables, \
                porcentaje_venta, venta_neta_mensual, arriendo_fabrica, depreciacion, energia_agua_otros)
    
    
    
def linearMeters(ws_read_manufacturing):
    metros_lineales = 0
    width = ws_read_manufacturing.cell(row = 7, column = 4).value
    next_row = 8
    while(width > 0): #float(width.replace(',', '.')) en caso que width sea leido como string
        metros_lineales = metros_lineales + width/1000
        width = ws_read_manufacturing.cell(row = next_row, column = 4).value
        next_row = next_row + 1
    return metros_lineales
        
        
        
def writeInfo(db, project_cost, metros_lineales, metros_lineales_mes, remuneraciones_fijas, remuneraciones_variables, \
                porcentaje_venta, venta_neta_mensual, arriendo_fabrica, depreciacion, energia_agua_otros):
    
    #conseguimos algunos de los datos necesarios para escribir la informacion, partiendo por el los "materiales de fabricacion"
    materiales_fabricacion = porcentaje_venta * venta_neta_mensual
    
    #seguimos con el total de costos de fabricacion (entre todos los proyectos)
    total_costos_fabricacion = remuneraciones_fijas + remuneraciones_variables + materiales_fabricacion + \
                                    arriendo_fabrica + depreciacion + energia_agua_otros
                                    
    #seguimos con el costo de fabricacion por emtro lineal vendido
    costo_fabricacion_por_ml = total_costos_fabricacion / metros_lineales_mes
    
    #terminamos con el costo estandar de fabricacion
    costo_estandar_fabricacion = costo_fabricacion_por_ml * metros_lineales
    
    #escribimos efectivamente el costo de fabricacion en la base de datos
    with db_session:
        project_cost.standard_cost_fabrication = costo_estandar_fabricacion