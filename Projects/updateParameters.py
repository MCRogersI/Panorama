from pony.orm import *
from openpyxl import load_workbook
import convert

def updateFreightCosts(db, file_name):
    file_read = file_name + ".xlsx"
    wb_read = load_workbook(file_read, data_only=True)
    try:
        ws_read_freight = wb_read["Flete"]
    except KeyError:
        print(' Formato invalido del archivo. El archivo debe tener una hoja llamada Flete.')
        return
    
    next_row = 5
    comuna_to = ws_read_freight.cell(row = next_row, column = 1).value
    while(comuna_to != None):
        #parseamos el nombre de la comuna a formato estandar
        try:
            comuna_to = convert.get_fuzzy(comuna_to)
        except:
            print('\n Formato del archivo es invalido. Ha ocurrido un error al leer los nombres de las comunas.')
            return
        
        freight_cost = ws_read_freight.cell(row = next_row, column = 2).value
        with db_session:
            fc = db.Freight_Costs.get(comuna_to = comuna_to)
            #revisamos si el codigo leido esta ya en la base de datos. Si esta, actualizamos la informacion, si no esta, creamos el nuevo SKU con la informacion entregada
            if fc != None:
                fc.freight_cost = freight_cost
            else:
                db.Freight_Costs(comuna_to = comuna_to, freight_cost = freight_cost)
        next_row = next_row + 1
        comuna_to = ws_read_freight.cell(row = next_row, column = 1).value
    
    commit()
    return True


    
def updateOperatingCosts(db, file_name):
    file_read = file_name + ".xlsx"
    wb_read = load_workbook(file_read, data_only=True)
    try:
        ws_read_operating = wb_read["Operaciones"]
    except KeyError:
        print(' Formato invalido del archivo. El archivo debe tener una hoja llamada Operaciones.')
        return
    
    next_row = 5
    name = ws_read_operating.cell(row = next_row, column = 1).value
    while(name != None):
        value = ws_read_operating.cell(row = next_row, column = 2).value
        with db_session:
            oc = db.Operating_Parameters.get(name = name)
            #revisamos si el codigo leido esta ya en la base de datos. Si esta, actualizamos la informacion, si no esta, creamos el nuevo SKU con la informacion entregada
            if oc != None:
                oc.value = value
            else:
                db.Operating_Parameters(name = name, value = value)
        next_row = next_row + 1
        name = ws_read_operating.cell(row = next_row, column = 1).value
    
    commit()
    return True


    
def updateViaticCosts(db, file_name):
    file_read = file_name + ".xlsx"
    wb_read = load_workbook(file_read, data_only=True)
    try:
        ws_read_viatic = wb_read["Viaticos"]
    except KeyError:
        print(' Formato invalido del archivo. El archivo debe tener una hoja llamada Viaticos.')
        return
    
    next_row = 5
    comuna_from = ws_read_viatic.cell(row = next_row, column = 1).value
    while(comuna_from != None):
        comuna_to = ws_read_viatic.cell(row = next_row, column = 2).value
        #parseamos el nombre de la comuna a formato estandar
        try:
            comuna_from = convert.get_fuzzy(comuna_from)
            comuna_to = convert.get_fuzzy(comuna_to)
        except:
            print('\n Formato del archivo es invalido. Ha ocurrido un error al leer los nombres de las comunas.')
            return
        
        viatic_cost = ws_read_viatic.cell(row = next_row, column = 3).value
        with db_session:
            vc = db.Viatic_Costs.get(comuna_from = comuna_from, comuna_to = comuna_to)
            #revisamos si el codigo leido esta ya en la base de datos. Si esta, actualizamos la informacion, si no esta, creamos el nuevo SKU con la informacion entregada
            if vc != None:
                vc.viatic_cost = viatic_cost
            else:
                db.Viatic_Costs(comuna_from = comuna_from, comuna_to = comuna_to, viatic_cost = viatic_cost)
        next_row = next_row + 1
        comuna_from = ws_read_viatic.cell(row = next_row, column = 1).value
        
    commit()
    return True
        
        
        
def updateMovilizationCosts(db, file_name):
    file_read = file_name + ".xlsx"
    wb_read = load_workbook(file_read, data_only=True)
    try:
        ws_read_movilization = wb_read["Movilizacion"]
    except KeyError:
        print(' Formato invalido del archivo. El archivo debe tener una hoja llamada Movilizacion.')
        return
    
    next_row = 5
    comuna_from = ws_read_movilization.cell(row = next_row, column = 1).value
    while(comuna_from != None):
        comuna_to = ws_read_movilization.cell(row = next_row, column = 2).value
        #parseamos el nombre de la comuna a formato estandar
        try:
            comuna_from = convert.get_fuzzy(comuna_from)
            comuna_to = convert.get_fuzzy(comuna_to)
        except:
            print('\n Formato del archivo es invalido. Ha ocurrido un error al leer los nombres de las comunas.')
            return

        movilization_cost = ws_read_movilization.cell(row = next_row, column = 3).value
        with db_session:
            mc = db.Movilization_Costs.get(comuna_from = comuna_from, comuna_to = comuna_to)
            #revisamos si el codigo leido esta ya en la base de datos. Si esta, actualizamos la informacion, si no esta, creamos el nuevo SKU con la informacion entregada
            if mc != None:
                mc.movilization_cost = movilization_cost
            else:
                db.Movilization_Costs(comuna_from = comuna_from, comuna_to = comuna_to, movilization_cost = movilization_cost)
        next_row = next_row + 1
        comuna_from = ws_read_movilization.cell(row = next_row, column = 1).value
     
    commit()
    return True
        

        
def updateCrystalsParameters(db, file_name):
    file_read = file_name + ".xlsx"
    wb_read = load_workbook(file_read, data_only=True)
    try:
        ws_read_crystals = wb_read["Parametros cristales"]
    except KeyError:
        print(' Formato invalido del archivo. El archivo debe tener una hoja llamada Parametros cristales.')
        return
    
    next_row = 5
    thickness = str(ws_read_crystals.cell(row = next_row, column = 1).value)
    while(thickness != None):
        square_meter_cost = ws_read_crystals.cell(row = next_row, column = 2).value
        waste_factor = ws_read_crystals.cell(row = next_row, column = 3).value
        with db_session:
            cp = db.Crystals_Parameters.get(thickness = thickness)
            #revisamos si el codigo leido esta ya en la base de datos. Si esta, actualizamos la informacion, si no esta, creamos el nuevo SKU con la informacion entregada
            if cp != None:
                cp.square_meter_cost = square_meter_cost
                cp.waste_factor = waste_factor
            else:
                db.Crystals_Parameters(thickness = thickness, square_meter_cost = square_meter_cost, waste_factor = waste_factor)
        next_row = next_row + 1
        thickness = ws_read_crystals.cell(row = next_row, column = 1).value
    
    commit()
    return True
        

        
def updateProfilesParameters(db, file_name):
    file_read = file_name + ".xlsx"
    wb_read = load_workbook(file_read, data_only=True)
    try:
        ws_read_profiles = wb_read["Parametros perfiles y herrajes"]
    except KeyError:
        print(' Formato invalido del archivo. El archivo debe tener una hoja llamada Parametros perfiles y herrajes.')
        return
    
    next_row = 5
    name = ws_read_profiles.cell(row = next_row, column = 1).value
    while(name != None):
        waste_factor = ws_read_profiles.cell(row = next_row, column = 2).value
        with db_session:
            pp = db.Profiles_Fittings_Parameters.get(name = name)
            #revisamos si el codigo leido esta ya en la base de datos. Si esta, actualizamos la informacion, si no esta, creamos el nuevo SKU con la informacion entregada
            if pp != None:
                pp.waste_factor = waste_factor
            else:
                db.Profiles_Fittings_Parameters(name = name, waste_factor = waste_factor)
        next_row = next_row + 1
        name = ws_read_profiles.cell(row = next_row, column = 1).value
    
    commit()
    return True

    

