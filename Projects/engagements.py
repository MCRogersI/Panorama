from pony.orm import *
from openpyxl import Workbook, load_workbook
from math import ceil
from . import materialCost, instalationCost, fabricationCost
import convert

def createEngagements(db, contract_number, file_name):
    # primero, obtenemos el archivo de la hoja de corte
    file_read = file_name + ".xlsx"
    wb_read = load_workbook(file_read, data_only = True)
    #revisamos que el archivo tenga la hoja que necesitamos
    try:
        ws_read_manufacturing = wb_read["Manufacturing"]
    except KeyError:
        print(' Formato invalido del archivo. El archivo debe tener una hoja llamada Measures.')
        return
    with db_session:
        project = db.Projects[contract_number]
        
    #obtenemos la fecha de requerimiento, que se considera como el principio de la fabricación
    try:
        task = select(t for t in db.Tasks if t.project == project and t.skill == db.Skills[3] and t.failed == None).first()
        emp_task = select(et for et in db.Employees_Tasks if et.task == task).first()
        withdrawal_date = emp_task.planned_initial_date
    #si aún no se ha planificado, dejamos la fecha vacía
    except AttributeError:
        withdrawal_date = None
        
    createEngagementsProfiles(db, project, ws_read_manufacturing, withdrawal_date)
    createEngagementsComponents(db, project, ws_read_manufacturing, withdrawal_date)
    createEngagementsSealings(db, project, ws_read_manufacturing, withdrawal_date)
    
    
#################################################################################
# Desde aqui los metodos auxiliares que crean los Engagements para los Perfiles #
#################################################################################

def createEngagementsProfiles(db, project, ws_read_manufacturing, withdrawal_date):
    #primero, para el Perfil Superior
    with db_session:
        #recuperamos el factor de perdida
        warning = ' Aviso: el factor de perdida para Perfil Superior no se encuentra registrado en la base de datos. Se considerara como 0.'
        try:
            factor_perdida = db.Profiles_Fittings_Parameters.get(name = "Upper").waste_factor
        except AttributeError:
            print(warning)
            factor_perdida = 0
        #pasamos de porcentaje a fraccion
        factor_perdida = factor_perdida/100.0
        length_coords = [48, 2]
        code_coords = [46, 2]
        engageProfile(db, project, ws_read_manufacturing, factor_perdida, length_coords, code_coords, withdrawal_date)
    
    #ahora, el Perfil Inferior
        #recuperamos el factor de perdida
        warning = ' Aviso: el factor de perdida para Perfil Inferior no se encuentra registrado en la base de datos. Se considerara como 0.'
        try:
            factor_perdida = db.Profiles_Fittings_Parameters.get(name = "Lower").waste_factor
        except AttributeError:
            print(warning)
            factor_perdida = 0
        #pasamos de porcentaje a fraccion
        factor_perdida = factor_perdida/100.0
        length_coords = [60, 2]
        code_coords = [58, 2]
        engageProfile(db, project, ws_read_manufacturing, factor_perdida, length_coords, code_coords, withdrawal_date)
    
    #por ultimo, el Perfil Telescopico
        #recuperamos el factor de perdida
        warning = ' Aviso: el factor de perdida para Perfil Telescopico no se encuentra registrado en la base de datos. Se considerara como 0.'
        try:
            factor_perdida = db.Profiles_Fittings_Parameters.get(name = "Teleskopic").waste_factor
        except AttributeError:
            print(warning)
            factor_perdida = 0
        #pasamos de porcentaje a fraccion
        factor_perdida = factor_perdida/100.0
        length_coords = [72, 2]
        code_coords = [70, 2]
        engageProfile(db, project, ws_read_manufacturing, factor_perdida, length_coords, code_coords, withdrawal_date)
    
    #ahora, los Glassing Beads
        #recuperamos el factor de perdida
        warning = ' Aviso: el factor de perdida para Glassing Beads no se encuentra registrado en la base de datos. Se considerara como 0.'
        try:
            factor_perdida = db.Profiles_Fittings_Parameters.get(name = "Glassing beads").waste_factor
        except AttributeError:
            print(warning)
            factor_perdida = 0
        #pasamos de porcentaje a fraccion
        factor_perdida = factor_perdida/100.0
        length_coords = [47, 13]
        code_coords = [78, 14]
        engageProfile(db, project, ws_read_manufacturing, factor_perdida, length_coords, code_coords, withdrawal_date, True)
 
    #ahora, los Glassing Beads Covers
        #recuperamos el factor de perdida
        warning = ' Aviso: el factor de perdida para Glassing Bead Cover no se encuentra registrado en la base de datos. Se considerara como 0.'
        try:
            factor_perdida = db.Profiles_Fittings_Parameters.get(name = "Glassing bead cover").waste_factor
        except AttributeError:
            print(warning)
            factor_perdida = 0
        #pasamos de porcentaje a fraccion
        factor_perdida = factor_perdida/100.0
        length_coords = [47, 13]
        code_coords = [79, 14]
        engageProfile(db, project, ws_read_manufacturing, factor_perdida, length_coords, code_coords, withdrawal_date, True)
    
def engageProfile(db, project, ws_read_manufacturing, factor_perdida, length_coords, code_coords, withdrawal_date, bead = False):
    #primero, calculamos los metros lineales
    metros_lineales = 0
    length = ws_read_manufacturing.cell(row = length_coords[0], column = length_coords[1]).value
    next_row = length_coords[0] + 1
    while(length > 0):
        metros_lineales = metros_lineales + length/1000
        length = ws_read_manufacturing.cell(row = next_row, column = length_coords[1]).value
        next_row = next_row + 1
    if bead:
        metros_lineales = 2*metros_lineales
    metros_lineales_totales = metros_lineales/(1 - factor_perdida)
    
    #segundo, obtenemos el código y la cantidad requerida
    code = ws_read_manufacturing.cell(row = code_coords[0], column = code_coords[1]).value
    precio = getByCode(db, code)
    if precio == 0 and bead:
        code = 54220002
    
    #por último, generamos el Engagement
    with db_session:
        sku = db.Stock[code]
        db.Engagements(project = project, sku = sku, quantity = ceil(metros_lineales_totales), withdrawal_date = withdrawal_date)
        
        
        
        
        
#################################################################################
# Desde aqui los metodos auxiliares que crean los Engagements para los Herrajes #
#################################################################################

def createEngagementsComponents(db, project, ws_read_manufacturing, withdrawal_date):
    with db_session:
        #partimos por Components to Glass Panes
        #recuperamos el factor de perdida
        warning = ' Aviso: el factor de perdida para Components to Glass Panes no se encuentra registrado en la base de datos. Se considerara como 0.'
        try:
            factor_perdida = db.Profiles_Fittings_Parameters.get(name = "Components to glass panes").waste_factor
        except AttributeError:
            print(warning)
            factor_perdida = 0
        #pasamos de porcentaje a fraccion
        factor_perdida = factor_perdida/100.0
        rows_read = [126, 134]
        columns_read = [1, 6]
        rows = [22, 30]
        column = 2
        engageComponent(db, project, ws_read_manufacturing, factor_perdida, rows_read, columns_read, rows, withdrawal_date)
        
        #seguimos con Components to component box or profiles
        #recuperamos el factor de perdida
        warning = ' Aviso: el factor de perdida para Components to Component Box or Profiles no se encuentra registrado en la base de datos. Se considerara como 0.'
        try:
            factor_perdida = db.Profiles_Fittings_Parameters.get(name = "Components to component box or profiles").waste_factor
        except AttributeError:
            print(warning)
            factor_perdida = 0
        #pasamos de porcentaje a fraccion
        factor_perdida = factor_perdida/100.0
        rows_read = [140, 150]
        columns_read = [1, 6]
        rows = [36, 45]
        column = 2
        engageComponent(db, project, ws_read_manufacturing, factor_perdida, rows_read, columns_read, rows, withdrawal_date)
        
        #ahora Components to component box
        #recuperamos el factor de perdida
        warning = ' Aviso: el factor de perdida para Components to Component Box no se encuentra registrado en la base de datos. Se considerara como 0.'
        try:
            factor_perdida = db.Profiles_Fittings_Parameters.get(name = "Components to component box").waste_factor
        except AttributeError:
            print(warning)
            factor_perdida = 0
        #pasamos de porcentaje a fraccion
        factor_perdida = factor_perdida/100.0
        rows_read = [126, 148]
        columns_read = [8, 15]
        rows = [51, 73]
        column = 2
        engageComponent(db, project, ws_read_manufacturing, factor_perdida, rows_read, columns_read, rows, withdrawal_date)
        
def engageComponent(db, project, ws_read_manufacturing, factor_perdida, rows_read, columns_read, rows):
    precio_total = 0
    for i in range(0, rows_read[1] - rows_read[0] + 1):
        #vemos primero unidades totales requeridas del producto
        units = ws_read_manufacturing.cell(column = columns_read[1], row = rows_read[0] + i).value
        if units == None:
            units = 0
        if units > 500:
            units = units/1000
        units_total = units/(1 - factor_perdida)
        
        #ahora recuperamos el codigo y precio del producto
        code = ws_read_manufacturing.cell(column = columns_read[0], row = rows_read[0] + i).value
        if i > 19:
            code = 51220110
        
        #por último, generamos el Engagement
        with db_session:
            sku = db.Stock[code]
            db.Engagements(project = project, sku = sku, quantity = ceil(units_total), withdrawal_date = withdrawal_date)
            


            
            
#################################################################################
# Desde aqui los metodos auxiliares que crean los Engagements para los Herrajes #
#################################################################################

def createEngagementsSealings(db, project, ws_read_manufacturing, withdrawal_date):
    #primero, calculamos la cantidad de hojas
    hojas = 0
    width = ws_read_manufacturing.cell(row = 7, column = 4).value
    next_row = 8
    while(width > 0): #float(width.replace(',', '.')) en caso que width sea leido como string
        hojas = hojas + 1
        width = ws_read_manufacturing.cell(row = next_row, column = 4).value
        next_row = next_row + 1
        
    #ahora recuperamos el codigo del producto
    precio_total = 0
    for code in [54043034, 54043044, 54043064, 54043014, 54043024, 54043054, 54042014, 54042024, 54200204, 54200104, 11116200, 11116201]:
        if code in [54043034, 54043044, 54043064]:
            #por último, generamos el Engagement
            with db_session:
                sku = db.Stock[code]
                db.Engagements(project = project, sku = sku, quantity = hojas, withdrawal_date = withdrawal_date)
            
            


            
# Funcion auxiliar muy usada, bastante self-descriptive, auto-descriptiva        
def getByCode(db, code):
    with db_session:
        if code == None:
            return 0
        sku = db.Stock.get(id = code)
        if(sku != None):
            return sku.price
        return 0