from pony.orm import *
from openpyxl import load_workbook
from datetime import date, timedelta
import matplotlib.pyplot as plt
from operator import itemgetter
from matplotlib.pyplot import plot, show
from threading import Thread
import pandas
from tabulate import tabulate

def createSku(db,id, name, price, critical_level, real_quantity):
    ''' Este método crea una unidad nueva de stock, asigna automáticamente el ID de la misma.
        La cantidad estimada es la que se ve afectada por una planificación que podría cambiarse 
        en el futuro. Parte siendo igual a la cantidad real'''

    with db_session:
        if critical_level == None:
            critical_level = 0
        if real_quantity == None:
            real_quantity = 0
        s = db.Stock(id=id, name=name, price=price, critical_level=critical_level,
                     real_quantity=real_quantity, estimated_quantity=real_quantity)
        commit()



def editSku(db, id, name=None, price=None, critical_level=None, real_quantity=None,
            waste_factor = None):
    ''' Este método edita la unidad de stock, en cualquiera de sus características '''
    
    with db_session:
        try:
            s = db.Stock[id]
            if name != None:
                s.name = name
            if price != None:
                s.price = price
            if critical_level != None:
                s.critical_level = critical_level
            if real_quantity != None:
                s.real_quantity = real_quantity
                s.estimated_quantity = real_quantity
            if waste_factor != None:
                s.waste_factor = waste_factor
        except ObjectNotFound as e:
            print('Object not found: {}'.format(e))
        except ValueError as e:
            print('Value error: {}'.format(e))
        commit()


def deleteSku(db, id):
    ''' Este método elimina una de las entradas de SKU de la tabla de Stock'''
    with db_session:
        db.Stock[id].delete()
        commit()


def printStockConsole(db):
    ''' Este método elimina una de las entradas de SKU de la tabla de Stock '''
    with db_session:
        print('')
        st = db.Stock.select()
        data = [s.to_dict() for s in st]
        df = pandas.DataFrame(data, columns = ['id','name','price','critical_level','real_quantity','waste_factor','estimated_quantity'])
        df.columns = ['ID','Nombre','Precio','Punto Crítico','Cantidad Real en Bodega','Factor de Pérdida','Cantidad Estimada en Bodega']
        print( tabulate(df, headers='keys', tablefmt='psql'))


def createEngagement(db, contract_number, skus_list, withdrawal_date=None):
    ''' Este método crea una nueva entrada en la tabla de engagements a partir de los datos ingresados  '''
    # skus_list es una lista de tuplas con el id del SKU y la cantidad correspondiente.
    with db_session:
        if type(skus_list) == list:  # Caso en el que se ingresa un lista de tuplas.
            for sku_row in skus_list:
                try:
                    sku = db.Stock[sku_row[0]]
                    # IMPORTANTE: En 'project' se podría haber guardado simplemente el id del proyecto (contract_number),
                    # pero de esta forma el proyecto puede ser accedido de forma directa a través del engagement.
                    # Deberíamos instaurar una convención al respecto.
                    db.Engagements(project=db.Projects[contract_number], sku=sku, quantity=sku_row[1],
                                   withdrawal_date=withdrawal_date)
                except ObjectNotFound as e:
                    print('Object not found: {}'.format(e))
                except ValueError as e:
                    print('Value error: {}'.format(e))
        else:
            try:
                sku = db.Stock[skus_list[0]]
                db.Engagements(project=db.Projects[contract_number], sku=sku, quantity=skus_list[1],
                               withdrawal_date=withdrawal_date)

            except ObjectNotFound as e:
                print('Object not found: {}'.format(e))
            except ValueError as e:
                print('Value error: {}'.format(e))
            except TypeError as e:
                print('Type error: {}'.format(e))
        commit()


def createPurchases(db, skus_list, arrival_date):
    ''' Este método crea una nueva entrada en la tabla de purchases a partir de los datos ingresados  '''
    # skus_list es una lista de tuplas con el id del SKU y la cantidad correspondiente. Se DEBE ingresar la fecha de entrega.
    with db_session:
        if type(skus_list) == list:  # Caso en el que se ingresa un lista de tuplas.
            for sku_row in skus_list:
                try:
                    sku = db.Stock[sku_row[0]]
                    db.Purchases(sku=sku, quantity=sku_row[1], arrival_date=arrival_date)
                except ObjectNotFound as e:
                    print('Object not found: {}'.format(e))
                except ValueError as e:
                    print('Value error: {}'.format(e))
        else:
            try:
                sku = db.Stock[skus_list[0]]
                db.Purchases(sku=sku, quantity=skus_list[1], arrival_date=arrival_date)

            except ObjectNotFound as e:
                print('Object not found: {}'.format(e))
            except ValueError as e:
                print('Value error: {}'.format(e))
            except TypeError as e:
                print('Type error: {}'.format(e))
        commit()


def calculateStock(db, id_sku):
    ''' Este método retorna una tupla con los valores (fecha,cantidad) de stock (considerando las fechas en las que se presentan cambios)'''
    with db_session:
        try:
            engagements = select(en for en in db.Engagements if en.sku.id == id_sku).order_by(
                lambda en: en.withdrawal_date)
            purchases = select(pur for pur in db.Purchases if pur.sku.id == id_sku).order_by(
                lambda pur: pur.arrival_date)
            aux_engagements = []
            aux_purchases = []
            for en in engagements:
                aux_engagements.append((-en.quantity, en.withdrawal_date))
            for pur in purchases:
                aux_purchases.append((pur.quantity, pur.arrival_date))
            fluxes = aux_engagements + aux_purchases
            fluxes = sorted(fluxes, key=itemgetter(1))
            beginning_date = date.today()
            beginning_quantity = db.Stock[id_sku].real_quantity
            fluxes = [(0, beginning_date)] + fluxes
            values = [(beginning_quantity, beginning_date)]
            for i in range(1, len(fluxes)):
                values.append((values[i - 1][0] + fluxes[i][0], fluxes[i][1]))
            return values

        except ObjectNotFound as e:
            print('Object not found: {}'.format(e))
        except ValueError as e:
            print('Value error: {}'.format(e))

def calculateStockFix(db, id_sku,final_date): #Las fechas de engagements y purchases deben ser futuras (con una fecha mayor al día de hoy) !
    ''' Este método retorna una tupla con los valores (fecha,cantidad) de stock (considerando las fechas en las que se presentan cambios)'''
    with db_session:
        createEngagement(db, 1, [(id_sku, 0)], final_date) #Fix, engagement ficticio. Los Engagements (o Purchases) de la BD no deben tener una fecha superior a final_date
        try:
            engagements = select(
                en for en in db.Engagements if en.sku.id == id_sku).order_by(
                lambda en: en.withdrawal_date)
            purchases = select(
                pur for pur in db.Purchases if pur.sku.id == id_sku).order_by(
                lambda pur: pur.arrival_date)
            aux_engagements = []
            aux_purchases = []
            for en in engagements:
                aux_engagements.append((-en.quantity, en.withdrawal_date))
            for pur in purchases:
                aux_purchases.append((pur.quantity, pur.arrival_date))
            fluxes = aux_engagements + aux_purchases
            fluxes = sorted(fluxes, key=itemgetter(1))
            beginning_date = date.today()
            beginning_quantity = db.Stock[id_sku].real_quantity
            fluxes = [(0, beginning_date)] + fluxes
            values = [(beginning_quantity, beginning_date)]
            for i in range(1, len(fluxes)):
                values.append((values[i - 1][0] + fluxes[i][0], fluxes[i][1]))
            values = [v for v in values if v[1]>=beginning_date] #Para tomar en cuenta desde hoy
            return values

        except ObjectNotFound as e:
            print('Object not found: {}'.format(e))
        except ValueError as e:
            print('Value error: {}'.format(e))

#Función en progreso
# def calculateStockFixedPeriod(db, id_sku,last_day):
#     ''' Este método retorna una tupla con los valores (fecha,cantidad) de stock (considerando las fechas en las que se presentan cambios)'''
#     with db_session:
#         try:
#             engagements = select(en for en in db.Engagements if en.sku.id == id_sku).order_by(
#                 lambda en: en.withdrawal_date)
#             purchases = select(pur for pur in db.Purchases if pur.sku.id == id_sku).order_by(
#                 lambda pur: pur.arrival_date)
#             aux_engagements = []
#             aux_purchases = []
#             for en in engagements:
#                 aux_engagements.append((-en.quantity, en.withdrawal_date))
#             for pur in purchases:
#                 aux_purchases.append((pur.quantity, pur.arrival_date))
#             fluxes = aux_engagements + aux_purchases
#             fluxes = sorted(fluxes, key=itemgetter(1))
#             beginning_date = date.today()
#             beginning_quantity = db.Stock[id_sku].real_quantity
#             fluxes = [(0, beginning_date)] + fluxes
#             values = [(beginning_quantity, beginning_date)]
#             if (len(aux_engagements) < 1 and len(aux_purchases) < 1):
#                 for i in range(0,(beginning_date-last_day).days):
#                     values.append((beginning_quantity, beginning_date + timedelta(days=1)))
#                 return values
#             else:
#                 i = 1
#                 while(i<len(fluxes) and fluxes[i][1] <= last_day):
#                     values.append((values[i - 1][0] + fluxes[i][0], fluxes[i][1]))
#                     i+=1
#
#
#                 return values
#
#         except ObjectNotFound as e:
#             print('Object not found: {}'.format(e))
#         except ValueError as e:
#             print('Value error: {}'.format(e))

def updateEngagements(db, id_sku):
    '''Este método actualiza los engagements una vez que se ha hecho una planificación, asignando la fecha de inicio
        de la instalación '''
    with db_session:
        # installations = select(t for t in db.Tasks if t.skill.id == 4)
        assigned_inst = select(at for at in db.Employees_Tasks if at.task.skill.id == 4)
        engagements = select(e for e in db.Engagements if e.sku == db.Stock[id_sku])
        for e in engagements:
            for at in assigned_inst:
                if at.task.project == e.project:
                    e.withdrawal_date = at.planned_initial_date

def printStock(db, id_sku): #Revisar este método para el caso 3 gráfica una línea al principio que no parece función.
    '''Este método imprime el comportamiento de un SKU hasta el último de los movimientos registrados '''
    with db_session:
        movements = calculateStock(db, id_sku)
        quantities = []
        dates = []
        for l in movements:
            quantities.append(l[0])
            dates.append(l[1])

        #Lo siguiente le agrega 7 días de stock constante después del último movimiento (agrega un fecha 7 días después con el mismo valor)
        dates.append(dates[len(dates)-1]+timedelta(days = 7))
        quantities.append(quantities[len(quantities) - 1])

        def plot_graph(quantities, dates):
            min_quantity = min(quantities)
            max_quantity = max(quantities)
            span_quantities = max_quantity - min_quantity
            plt.figure()
            plt.step(dates, quantities, where='post')
            plt.xlim((date.today(), dates[len(dates) - 1] + timedelta(1)))
            plt.ylim(min_quantity - (span_quantities / 10), max_quantity + (span_quantities / 10))
            plt.xticks(dates, dates, rotation='vertical')
            plt.ylabel('Quantity')
            plt.xlabel('Date')
            plt.title('SKU (id = {}) quantities by date'.format(id_sku))
            plt.title(
                'Inventory prediction of unit ' + str(db.Stock[id_sku]) + ', code: ' + str(id_sku))
            plt.axhline(y=db.Stock[id_sku].critical_level, color='r', linestyle='-')
            plt.tight_layout()
            plt.show(block=False) #Esto debería estar definido con el valor de block = True.
            #Es un 'Workaround' para graficar múltiples gráficos.
        p = Thread(target=plot_graph(quantities=quantities, dates=dates))
        p.start()
        p.join()

def displayStock(db, id_sku):
    '''Este método grafica el comportamiento de un SKU hasta el último de los movimientos registrados '''
    with db_session:
        sku = db.Stock.get(id=id_sku)
        critical_level = sku.critical_level
        values = calculateStock(db, id_sku)
        quantities, dates = zip(*values)#<-- wooowowooo (que bonita función)

        def plot_graph(quantities, dates):
            plt.figure()
            plt.ylabel('Quantity')
            plt.xlabel('Date')
            plt.title(
                'Inventory prediction of unit ' + str(db.Stock[id_sku]) + ', code: ' + str(id_sku))
            min_date = min(dates)
            max_date = max(dates)
            min_quantity = min(quantities)
            max_quantity = max(quantities)
            span_quantities = max_quantity - min_quantity
            delta = max_date - min_date
            delta = delta.days
            aux_quantity = []
            aux_dates = []
            virtual_extra_days = 7

            def get_set_from_dates(quantities,dates): #Método auxiliar para generar un 'set' de los datos en values
                result = []
                for i in range(1,len(dates)):
                    if dates[i-1]!=dates[i]:
                        result.append((quantities[i-1],dates[i-1]))
                    if i == len(dates)-1:
                        result.append((quantities[i], dates[i]))
                return result

            set_of_quantities, set_of_dates = zip(*get_set_from_dates(quantities,dates)) #Revisar que sucede cuando no hay engagements y/o no hay purchases


            for i in range(0,delta + virtual_extra_days):
                aux_dates.append(min_date+timedelta(days=i))
                aux_quantity.append(0)

            d = 0
            for i in range(0,len(aux_dates)):
                if d < len(set_of_dates)-1:
                    if aux_dates[i] < set_of_dates[d+1]:
                        aux_quantity[i] = set_of_quantities[d]

                    elif d<len(set_of_dates)-1:
                        aux_quantity[i] = set_of_quantities[d+1]
                        d += 1

                else:
                    aux_quantity[i] = set_of_quantities[d]

            width = 1
            p1 = plt.bar(aux_dates, aux_quantity, width=width, color='#0000FF', align='center')

            for i in range(0,len(aux_dates)): #Para colorear los días con stock bajo el nivel crítico.
                if aux_quantity[i] <= critical_level:
                    p1[i].set_color('r')
                    # p1[i].set_linewidth(1)
                    # p1[i].set_edgecolor('k')
                p1[i].set_linewidth(1.1)
                p1[i].set_edgecolor('k')

            # plt.xticks(dates, dates, rotation='vertical') #Con esta configuración solo aparecen los labels de las fechas con cambios
            plt.xticks(aux_dates, aux_dates, rotation='vertical')  # Con esta configuración aparecen los labels de todos los días
            plt.grid()
            plt.axhline(y=db.Stock[id_sku].critical_level, color='r', linestyle='dashed',
                        linewidth=2.5)
            plt.axvline(x=max_date, color='k', linestyle='dashed', linewidth=1.5)
            plt.xlim(min_date, max_date + timedelta(
                days=7))  # Se agregan 7 días 'extras/ficticios' para un mejor display
            plt.ylim(min_quantity - (span_quantities / 10), max_quantity + (span_quantities / 10))
            # plt.ylim(0, max_quantity + (span_quantities / 10))
            # plt.savefig("SKU{}.png".format(id_sku))
            plt.tight_layout()
            plt.show(block=True)

        p = Thread(target=plot_graph(quantities=quantities, dates=dates))
        p.start()
        p.join()

def calculateStockForExcel(db, id_sku):
    '''Este método calcula el comportamiento de un SKU hasta el último de los movimientos registrados HECHO PARA EXCEL'''
    with db_session:
        sku = db.Stock.get(id=id_sku)
        critical_level = sku.critical_level
        values = calculateStockFix(db, id_sku,date(2017,10,1)) #Fecha final del display de los gráficos
        # values = calculateStock(db, id_sku) #Función sin el fix
        quantities, dates = zip(*values)#<-- wooowowooo (que bonita función)

        def intermediate_calculation(quantities, dates):
            # plt.figure()
            # plt.ylabel('Quantity')
            # plt.xlabel('Date')
            # plt.title(
            #     'Inventory prediction of unit ' + str(db.Stock[id_sku]) + ', code: ' + str(id_sku))
            min_date = min(dates)
            max_date = max(dates)
            min_quantity = min(quantities)
            max_quantity = max(quantities)
            span_quantities = max_quantity - min_quantity
            delta = max_date - min_date
            delta = delta.days
            aux_quantity = []
            aux_dates = []
            virtual_extra_days = 7

            def get_set_from_dates(quantities,dates): #Método auxiliar para generar un 'set' de los datos en values
                result = []
                for i in range(1,len(dates)):
                    if dates[i-1]!=dates[i]:
                        result.append((quantities[i-1],dates[i-1]))
                    if i == len(dates)-1:
                        result.append((quantities[i], dates[i]))
                return result

            set_of_quantities, set_of_dates = zip(*get_set_from_dates(quantities,dates)) #Revisar que sucede cuando no hay engagements y/o no hay purchases


            for i in range(0,delta + virtual_extra_days):
                aux_dates.append(min_date+timedelta(days=i))
                aux_quantity.append(0)

            d = 0
            for i in range(0,len(aux_dates)):
                if d < len(set_of_dates)-1:
                    if aux_dates[i] < set_of_dates[d+1]:
                        aux_quantity[i] = set_of_quantities[d]

                    elif d<len(set_of_dates)-1:
                        aux_quantity[i] = set_of_quantities[d+1]
                        d += 1

                else:
                    aux_quantity[i] = set_of_quantities[d]

            # width = 1
            # p1 = plt.bar(aux_dates, aux_quantity, width=width, color='#0000FF', align='center')

            # for i in range(0,len(aux_dates)): #Para colorear los días con stock bajo el nivel crítico.
            #     if aux_quantity[i] <= critical_level:
            #         p1[i].set_color('r')
                    ##p1[i].set_linewidth(1)
                    ##p1[i].set_edgecolor('k')
                # p1[i].set_linewidth(1.1)
                # p1[i].set_edgecolor('k')

            # plt.xticks(dates, dates, rotation='vertical') #Con esta configuración solo aparecen los labels de las fechas con cambios
            # plt.xticks(aux_dates, aux_dates, rotation='vertical')  # Con esta configuración aparecen los labels de todos los días
            # plt.grid()
            # plt.axhline(y=db.Stock[id_sku].critical_level, color='r', linestyle='dashed',
            #             linewidth=2.5)
            # plt.axvline(x=max_date, color='k', linestyle='dashed', linewidth=1.5)
            # plt.xlim(min_date, max_date + timedelta(
            #     days=7))  # Se agregan 7 días 'extras/ficticios' para un mejor display
            # plt.ylim(min_quantity - (span_quantities / 10), max_quantity + (span_quantities / 10))
            ## plt.ylim(0, max_quantity + (span_quantities / 10))
            ## plt.savefig("SKU{}.png".format(id_sku))
            # plt.tight_layout()
            # plt.show(block=True)

        # p = Thread(target=plot_graph(quantities=quantities, dates=dates))
        # p.start()
        # p.join()
            return (aux_dates, aux_quantity)
        return intermediate_calculation(quantities=quantities,dates=dates)
def displayStockForExcel(db, id_sku):
    '''Este método grafica el comportamiento de un SKU hasta el último de los movimientos registrados HECHO PARA EXCEL'''
    with db_session:
        sku = db.Stock.get(id=id_sku)
        critical_level = sku.critical_level
        values = calculateStock(db, id_sku)
        quantities, dates = zip(*values)  # <-- wooowowooo (que bonita función)

        def plot_graph(quantities, dates):
            plt.figure()
            plt.ylabel('Quantity')
            plt.xlabel('Date')
            plt.title(
                'Inventory prediction of unit ' + str(db.Stock[id_sku]) + ', code: ' + str(
                    id_sku))
            min_date = min(dates)
            max_date = max(dates)
            min_quantity = min(quantities)
            max_quantity = max(quantities)
            span_quantities = max_quantity - min_quantity
            delta = max_date - min_date
            delta = delta.days
            aux_quantity = []
            aux_dates = []
            virtual_extra_days = 7

            def get_set_from_dates(quantities,
                                   dates):  # Método auxiliar para generar un 'set' de los datos en values
                result = []
                for i in range(1, len(dates)):
                    if dates[i - 1] != dates[i]:
                        result.append((quantities[i - 1], dates[i - 1]))
                    if i == len(dates) - 1:
                        result.append((quantities[i], dates[i]))
                return result

            set_of_quantities, set_of_dates = zip(*get_set_from_dates(quantities, dates))

            for i in range(0, delta + virtual_extra_days):
                aux_dates.append(min_date + timedelta(days=i))
                aux_quantity.append(0)

            d = 0
            for i in range(0, len(aux_dates)):
                if d < len(set_of_dates) - 1:
                    if aux_dates[i] < set_of_dates[d + 1]:
                        aux_quantity[i] = set_of_quantities[d]

                    elif d < len(set_of_dates) - 1:
                        aux_quantity[i] = set_of_quantities[d + 1]
                        d += 1

                else:
                    aux_quantity[i] = set_of_quantities[d]

            width = 1
            return  aux_dates,aux_quantity
        plot_graph(quantities,dates)

#FUNCIÓN EN DESARROLLO
# def displayALlSKUs(db):
#     with db_session:
#         skus = select(sku for sku in db.Stock).order_by(lambda s: s.id)
#         for sku in skus:
#             displayStock(db,sku.id)

def checkStockAlarms(db):
    '''Este método revisa los niveles de stock para cada sku en el futuro y verifica si están bajo el nivel crítico.
    Retorna en una lista de tuplas, las cantidades bajo el nivel crítico detectadas y sus respectivas fechas.  '''
    alarms = []
    with db_session:
        skus = select(sku for sku in db.Stock)
        for sku in skus:
            sku_levels = calculateStock(db, sku.id)
            critical_level = sku.critical_level
            for sku_level in sku_levels:
                if sku_level[0] <= critical_level:
                    alarms.append(sku_level)
                    alarms = sorted(alarms, key=itemgetter(1))
                    print(
                        'SKU (id = {0}) quantity below critical level {1}, on the date {2}'.format(
                            sku.id, sku_level[0], sku_level[1]))
    return alarms
def getStockValue(db):
    '''Método que entrega el precio en euros del stock existente el día en que se llama el método '''
    with db_session:
        skus = select(sku for sku in db.Stock)
        total_sum = 0
        for s in skus:
            total_sum += s.price*s.real_quantity
        return total_sum
def updateStock(db):
    ''' Método que realiza los cambios efectivos de stock correspondientes a la fecha en que se llama al método,
    debería llamarse todos los días y/o cada vez que se ingrese algún movimiento efectivo '''
    with db_session:
        engagements = select (e for e in db.Engagements if e.withdrawal_date == date.today())
        purchases = select( pur for pur in db.Purchases if pur.arrival_date == date.today())

        
# para poder indicar un Purchase desde un Excel, en un formato cómodo       
def makePurchases(db, file_name):
    #primero cargamos la hoja donde esta la informacion
    file_read = file_name + ".xlsx"
    wb_read = load_workbook(file_read, data_only=True)
    ws_read_purchases = wb_read["OC"]
    
    #ahora creamos la lista que luego pasaremos al createPurchase, y obtenemos la fecha
    skus_list = []
    
    year = ws_read_purchases.cell(row = 3, column = 4).value
    month = ws_read_purchases.cell(row = 3, column = 5).value
    day = ws_read_purchases.cell(row = 3, column = 6).value
    arrival_date = date(year, month, day)
    
    #ahora recorremos la hoja cargada, llenando la informacion en las listas
    cell = ws_read_purchases.cell(row = 3, column = 2)
    next_row = 4
    while(cell.value):
        #si la celda correspondiente a Code no esta vacia, seguimos sacando informacion de la hoja cargada
        code = cell.value
        quantity = float(ws_read_purchases.cell(row = next_row - 1, column = 3).value)
        #agregamos la informacion a la lista
        skus_list.append([code, quantity])
        #pasamos a la siguiente fila y el While revisara si esta vacia, o si bien hay que seguir sacando informacion
        cell = ws_read_purchases.cell(row = next_row, column = 2)
        next_row = next_row + 1
    
    createPurchases(db, skus_list, arrival_date)
    
    

#para actualizar la lista de precios de una, a través de una hoja con el mismo formato de 
def editAllSkus(db, file_name):
    file_read = file_name + ".xlsx"
    wb_read = load_workbook(file_read, data_only=True)
    ws_read_skus = wb_read["Hoja2"]
    
    next_row = 13
    id = ws_read_skus.cell(row = next_row, column = 2).value
    while(id != None):
        name = ws_read_skus.cell(row = next_row, column = 4).value
        price = ws_read_skus.cell(row = next_row, column = 5).value
        critical_level = ws_read_skus.cell(row = next_row, column = 9).value
        real_quantity = ws_read_skus.cell(row = next_row, column = 10).value
        waste_factor = ws_read_skus.cell(row = next_row, column = 11).value
        
        with db_session:
            stock = db.Stock.get(id = id)
            #revisamos si el codigo leido esta ya en la base de datos. Si esta, actualizamos la informacion, si no esta, creamos el nuevo SKU con la informacion entregada
            if stock != None:
                editSku(db, id, name, price, critical_level, real_quantity)
            else:
                createSku(db, id, name, price, critical_level, real_quantity, waste_factor)
        
        next_row = next_row + 1
        id = ws_read_skus.cell(row = next_row, column = 2).value
