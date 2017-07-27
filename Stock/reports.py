from pony.orm import *
from datetime import date, timedelta
import pandas as pd
import numpy as np
from openpyxl import Workbook, load_workbook, drawing
from openpyxl.styles import Font
import openpyxl
from openpyxl import Workbook
from openpyxl.chart import BarChart, Series, Reference
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import Font,Alignment
from openpyxl.styles import PatternFill
from openpyxl.drawing.fill import PatternFillProperties, ColorChoice
from openpyxl.chart.marker import DataPoint
import os

def createStockReport(db):

    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))
    from Stock.features import displayStock, calculateStockForExcel
    with db_session:
        skus = select(s for s in db.Stock).order_by(lambda s : s.id)
        skus_ids = [sku.id for sku in skus]
        skus_critical_levels = [sku.critical_level for sku in skus]
        skus_names = [sku.name for sku in skus]
    bases = [calculateStockForExcel(db, id) for id in skus_ids]

    # wb = Workbook(write_only=True, guess_types=True)  # Atención con el guess_types y el write_only
    # wb = Workbook(write_only=False, guess_types=True)
    wb = Workbook(write_only=False)
    by_default_sheet = wb.get_sheet_by_name('Sheet')
    wb.remove_sheet(by_default_sheet)
    ws_raw = wb.create_sheet(title="Tablas de Stock")
    ws_plotted = wb.create_sheet(title="Gráficos de Stock", index=0)
    ws_raw.sheet_view.zoomScale = 30 #Para "alejar" el zoom de la hoja
    ws_plotted.sheet_view.zoomScale = 30 #Para "alejar" el zoom de la hoja
    counter = 0
    id_counter = 0 #Arreglar para que quede más compacto y limpio
    for b in bases:
        base = b
        dates = base[0]
        values = base[1]
        id_counter+=1
        current_sku_id = skus_ids[id_counter-1]
        current_sku_critical_level = skus_critical_levels[id_counter-1]
        current_sku_name = skus_names[id_counter-1]

        # rows = [
        #     ('Fechas', 'Cantidad', 'Batch 2'),
        #     (2, 10, 30),
        #     (3, 40, 60),
        #     (4, 50, 70),
        #     (5, 20, 10),
        #     (6, 10, 40),
        #     (7, 50, 30),
        # ]
        # vals = [(i,np.sqrt(i)) for i in range(0,20)]
        vals = zip(dates, values)
        rows = [('SKU', '{}'.format(current_sku_id)), ('Fecha', 'Cantidad')]
        # rows = rows
        # rows = rows.extend([('Fechas', 'Cantidad')])
        rows.extend(vals)

        row_counter = 0
        max_value = max(values)
        min_value = min(values)
        for row in rows:
            # ws_raw.append(row)
            cell1 = ws_raw.cell(row=1+row_counter,column=1 + counter*4)
            cell2 = ws_raw.cell(row=1+row_counter, column=2 + + counter*4)
            # ws_raw.cell(row=1+row_counter,column=1 + counter*4).value = row[0]
            cell1.value = row[0]
            cell1.font = Font(bold=True, )
            cell1.border = thin_border
            cell1.alignment = Alignment(horizontal='left')
            # ws_raw.cell(row=1+row_counter, column=2 + + counter*4).value = row[1]
            cell2.value = row[1]
            cell2.font = Font(bold=True, )
            cell2.border = thin_border
            cell2.alignment = Alignment(horizontal='left')

            row_counter+=1

        chart1 = BarChart()
        chart1.type = "col"
        chart1.style = 10
        # chart1.title = "Status del SKU: {0}\n{1}".format(current_sku_id,current_sku_name) #Nombre con salto de línea
        chart1.title = "Proyecciones para el SKU: {0}, {1}".format(current_sku_id, current_sku_name) #Nombre con flecha
        chart1.y_axis.title = 'Cantidad'
        chart1.x_axis.title = 'Fecha'

        # data = Reference(ws, min_col=2, min_row=1, max_row=7, max_col=3)
        data = Reference(ws_raw, min_col=2+ counter*4, min_row=2, max_row=len(rows))
        cats = Reference(ws_raw, min_col=1+ counter*4, min_row=3, max_row=len(rows))
        chart1.legend = None
        chart1.add_data(data, titles_from_data=True)
        chart1.set_categories(cats)
        chart1.shape = 4
        chart1.width = 100
        chart1.y_axis.scaling.min = -200# 200 ES UN BUEN VALOR PARA EL DISPLAY FIJO
        chart1.y_axis.scaling.min = min_value - 50  # VALOR MINIMO EN EL GRÁFICO
        chart1.y_axis.scaling.max = 500# 500 ES UN BUEN VALOR PARA EL DISPLAY FIJO
        chart1.y_axis.scaling.max = max_value + 50 # VALOR MAXIMO EN EL GRÁFICO
        # set a pattern for the whole series
        fill = PatternFillProperties(prst="ltUpDiag")
        fill.foreground = ColorChoice(prstClr="blue")
        fill.background = ColorChoice(prstClr="blue")
        # pat = PatternFill("solid", fgColor="ffff00")
        series = chart1.series[0]
        series.graphicalProperties.pattFill = fill

        #Obtener los índices de las alarmas
        alarm_indexes = []
        for i in range(0,len(values)):
            if values[i] <= current_sku_critical_level:
                alarm_indexes.append(i)
        # Pintar de rojo las barras críticas
        for alarm_index in alarm_indexes:
            pt = DataPoint(idx=alarm_index)
            alarm_fill = PatternFillProperties(prst="ltUpDiag")
            alarm_fill.foreground = ColorChoice(prstClr="red")
            alarm_fill.background = ColorChoice(prstClr="red")
            pt.graphicalProperties.pattFill = alarm_fill
            series.dPt.append(pt)

        row_counter = 0
        for alarm_index in alarm_indexes:
            cell1 = ws_raw.cell(row=1 + alarm_index +2, column=1 + counter * 4)
            cell2 = ws_raw.cell(row=1 + alarm_index +2, column=2 + + counter * 4)
            cell1.font = Font(bold=True)
            cell1.border = thin_border
            cell1.alignment = Alignment(horizontal='left')
            cell2.fill = PatternFill("solid", fgColor="ff0000")
            cell2.font = Font(bold=True)
            cell2.border = thin_border
            cell2.alignment = Alignment(horizontal='left')
            cell2.fill = PatternFill("solid", fgColor="ff0000")
        # if(len(alarm_indexes)>0): #Si hay algún día bajo el nivel crítico
        if (values[0] <= current_sku_critical_level):  # Si hoy el nivel está crítico
            cell3 = ws_raw.cell(row=1, column=1 + counter * 4 + 2)
            cell3.value = "COMPRAR"
            cell3.font = Font(bold=True)
            cell3.border = thin_border
            cell3.alignment = Alignment(horizontal='left')
            cell3.fill = PatternFill("solid", fgColor="ff0000")
            wrap_alignment = Alignment(wrap_text=True, horizontal="center",
                                       vertical="center")
            cell3.alignment = wrap_alignment


            row_counter += 1

        ws_plotted.add_chart(chart1, "{0}{1}".format("A",10 + counter*18))
        counter+=1

    # # Escribir la fecha en la que fue producida el reporte:
    # ws_plotted.column_dimensions["C"].width = 20
    # cell = ws_plotted.cell(row=3, column=3, value="Reporte producido el: ")
    # cell.font = Font(bold=True, )
    # cell.border = thin_border
    # cell.alignment = Alignment(horizontal='left')
    #
    # cell = ws_plotted.cell(row=3, column=4, value=date.today())
    # cell.font = Font(bold=True, )
    # cell.border = thin_border
    # cell.alignment = Alignment(horizontal='left')

    try:
        module_path = os.path.dirname(__file__)
        panorama_folder_path = os.path.abspath(os.path.join(module_path, os.pardir))
        report_folder_path = os.path.join(panorama_folder_path, "Reportes")
        if not os.path.exists(report_folder_path):
            os.makedirs(report_folder_path)
        report_file_name = "Informe de Stock {}.xlsx".format(date.today())
        fn = os.path.join(report_folder_path, report_file_name)
        wb.save(fn)
    except OSError as e:
        if e.args[0] != 13:
            raise
        input("\n Ha ocurrido un error porque el archivo Informe de Stock.xlsx está abierto. Por favor ciérrelo y presione cualquier tecla para que el programa pueda continuar.")


from database import db
# createStockReport(db)
