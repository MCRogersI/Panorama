from pony.orm import *
from datetime import date, timedelta
import pandas as pd
import numpy as np
from openpyxl import Workbook, load_workbook
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import Font,Alignment
from openpyxl.styles import PatternFill
#para borrar archivo Excel
import os
#para crear nombres de archivo al azar
import random
import string
import copy

#################
# Global report #
#################

def createGlobalReportModified(db):
    ''' Este método crea un informe en Excel compacto con una proción de la información de la base de datos. '''
    wb = Workbook()
    by_default_sheet = wb.get_sheet_by_name('Sheet')
    by_default_sheet.title = 'Introducción del informe'
    wb.remove_sheet(by_default_sheet)
    ws = wb.create_sheet(
        "Reporte",index=0)

    widths = {"A": 5, "B": 5, "C": 5,"D": 30, "E": 30, "F": 30,
              "G": 30, "H": 30, "I": 30,"J": 30, "K": 30, "L": 30,
              "M": 30, "N": 30, "O": 30,"P": 30, "Q": 30, "R": 30,
              "S": 50, "T": 50, "U": 50,"V": 50, "W": 50, "X": 50,
              "Y": 50, "Z": 50, "AA": 50, "AB": 50, "AC": 50,"AD":50,
              "AE": 50, "AF":50, "AG":50, "AH":50, "AI":50, "AJ":50, "AK":50, "AL":50,
              "AM":50,"AN":50,"AO":50,"AP":50,"AQ":50,"AR":50,"AS":50,"AT":50,"AU":50,
              "AV":50,"AW":50,"AX":50,"AY":50,"AZ":50,"BA":50,"BB":50,"BC":50,"BD":50,"BE":50,"BF":50,"BG":50,
              "BH":50,"BI":50,"BJ":50,"BK":50,"BL":50,"BM":50,"BN":50,"BO":50,"BP":50,"BQ":50,"BR":50}

    heights = {"A": 10, "B": 10, "C": 10, "D": 10, "E": 10, "F": 10,
              "G": 10, "H": 10, "I": 10, "J": 10, "K": 10, "L": 10,
              "M": 10, "N": 10, "O": 10,"P": 10, "Q": 10, "R": 10,
              "S": 10, "T": 10, "U": 10,"V": 10, "W": 10, "X": 10,
              "Y": 10, "Z": 10, "AA": 10, "AB": 10, "AC": 10,"AD":10,
               "AE":10, "AF":10, "AG":10, "AH":10, "AI":10, "AJ":10, "AK":10, "AL":10,
               "AM": 10, "AN": 10, "AO": 10, "AP": 10, "AQ": 10,"AR":10,"AS":10,"AT":10,"AU":10,
               "AV":10,"AW":10,"AX":10,"AY":10,"AZ":10,"BA":10,"BB":10,"BC":10,"BD":10,"BE":10,"BF":10,"BG":10,
               "BH":10,"BI":10,"BJ":10,"BK":10,"BL":10,"BM":10,"BN":10,"BO":10,"BP":10,"BQ":10,"BR":10}


    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))


    columns = ["A", "B", "C","D","E","F","G","H","I","J","K","L","M","N","O","P","Q","R","S","T","U","V","W","X","Y","Z","AA","AB",
               "AC","AD","AE", "AF", "AG", "AH", "AI", "AJ", "AK", "AL","AM","AN","AO","AP","AQ","AR","AS","AT","AU","AV","AW","AX","AY","AZ",
               "BA","BB","BC","BD","BE","BF","BG","BH","BI","BJ","BK","BL","BM","BN","BO","BP","BQ","BR"]

    for c in columns:
        ws.column_dimensions[c].width = widths[c]
        ws.column_dimensions[c].height = heights[c]


    num_columns, letter_columns = zip(*list(enumerate(columns)))
    num_columns = list(num_columns)
    num_columns = [x + 1 for x in num_columns] #Desplazamos los valores en 1 porque Excel está indexado desde el 1 y no desde el 0
    letter_columns = list(letter_columns)

    # escribir los títulos de las columnas, en negrita

    texts = ["","","","NUMERO DE CONTRATO", "NOMBRE CLIENTE", "PRIORIDAD",
             "FECHA LÍMITE","MAYOR PLAZO (ATRASO) TOTAL","ESTADO INICIAL","ESTADO FINAL","ORIGEN FALLA","COMUNA CLIENTE", "DIRECCIÓN CLIENTE","METROS LINEALES", "METROS LINEALES REALES", "COSTO ESTIMADO",
             "PRECIO DE VENTA", "FECHA DE VENTA",
             "FECHA ORIGINAL PLANIFICADA INICIO RECTIFICACIÓN","FECHA ORIGINAL PLANIFICADA TÉRMINO RECTIFICACIÓN",
             "FECHA PLANIFICADA INICIO RECTIFICACIÓN","FECHA PLANIFICADA TÉRMINO RECTIFICACIÓN",
             "FECHA EFECTIVA INICIO RECTIFICACIÓN","FECHA EFECTIVA TÉRMINO RECTIFICACIÓN","ID RECTIFICADOR",
             "MAYOR PLAZO (ATRASO) RECTIFICACION",
             "FECHA ORIGINAL PLANIFICADA INICIO DISEÑO","FECHA ORIGINAL PLANIFICADA TÉRMINO DISEÑO","FECHA PLANIFICADA INICIO DISEÑO",
             "FECHA PLANIFICADA TÉRMINO DISEÑO","FECHA EFECTIVA INICIO DISEÑO",
             "FECHA EFECTIVA TÉRMINO DISEÑO","ID DISEÑADOR","MAYOR PLAZO (ATRASO) DISEÑO",
             "FECHA ORIGINAL PLANIFICADA INICIO FABRICACIÓN","FECHA ORIGINAL PLANIFICADA TÉRMINO FABRICACIÓN",
             "FECHA PLANIFICADA INICIO FABRICACIÓN","FECHA PLANIFICADA TÉRMINO FABRICACIÓN","FECHA EFECTIVA INICIO FABRICACIÓN",
             "FECHA EFECTIVA TÉRMINO FABRICACIÓN","ID FABRICADOR","MAYOR PLAZO (ATRASO) FABRICACION",
             "FECHA ORIGINAL PLANIFICADA INICIO INSTALACIÓN","FECHA ORIGINAL PLANIFICADA TÉRMINO INSTALACIÓN","FECHA PLANIFICADA INICIO INSTALACIÓN",
             "FECHA PLANIFICADA TÉRMINO INSTALACIÓN",
             "FECHA EFECTIVA INICIO INSTALACIÓN","FECHA EFECTIVA TÉRMINO INSTALACIÓN","IDs INSTALADORES","MAYOR PLAZO (ATRASO) INSTALACION",
             "COSTO ESTANDAR PERFILES",
             "COSTO ESTANDAR HERRAJES","COSTO ESTANDAR CRISTALES","COSTO ESTANDAR M PRIMAS","COSTO ESTANDAR FABRICACION",
             "COSTO ESTANDAR INTALACION","COSTO ESTANDAR ADICIONALES","COSTO ESTANDAR TOTAL","COSTO EFECTIVO M PRIMAS","COSTO EFECTIVO FABRICACION",
             "COSTO EFECTIVO INSTALACION","COSTO EFECTIVO COMPLEMENTOS",
             "FECHA PLANIFICADA EMISION CRISTALES","FECHA EFECTIVA EMISION CRISTALES","FECHA PLANIFICADA LLEGADA CRISTALES",
             "FECHA EFECTIVA LLEGADA CRISTALES","ID EMISOR HC","ID PROVEEDOR CRISTALES","MAYOR PLAZO (ATRASO) EMISION HC",
             "MAYOR PLAZO (ATRASO) LLEGADA CRISTALES"]

    cellColorRect = "FFA200"
    cellColorDis = "2873FF"
    cellColorFab = "FFFF00"
    cellColorInst = "03FD24"
    cellColorOC = "E208F1"
    for i in range(4,len(num_columns)+1):
        cell = ws.cell(row=3, column=i, value=texts[i-1])
        cell.font = Font(bold=True,)
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')
        if(i>=19 and i<=26):
            cell.fill = PatternFill("solid", fgColor=cellColorRect)
        if (i >= 27 and i <= 34):
            cell.fill = PatternFill("solid", fgColor=cellColorDis)
        if (i >= 35 and i <= 42):
            cell.fill = PatternFill("solid", fgColor=cellColorFab)
        if (i >= 43 and i <= 50):
            cell.fill = PatternFill("solid", fgColor=cellColorInst)
        if (i >= 63 and i <= 70):
            cell.fill = PatternFill("solid", fgColor=cellColorOC)


    # A continuación llenamos con los datos

    with db_session:
        #El próximo bloque de código recupera los proyectos de la base de datos y los ordena por número de contrato y versión
        projects = select(p for p in db.Projects).order_by(lambda p: p.contract_number)
        aux_contract_numbers = []
        for p in projects:
            if p.contract_number not in aux_contract_numbers:
                aux_contract_numbers.append(p.contract_number)
        aux_projects_by_CN_group = []
        for aux_CN in aux_contract_numbers:
            aux_projects_by_CN_group.append(db.Projects.select(lambda p: p.contract_number == aux_CN).order_by(lambda p: p.version))
        aux_ordered_projects_list = []
        for CNgroup in aux_projects_by_CN_group:
            aux_ordered_projects_list.extend(list(CNgroup))

        projects = aux_ordered_projects_list

        r=4
        #A continuación se encuentra el loop que llena cada fila de la tabla del reporte (cada fila corresponde a un proyecto)
        for p in projects:


            project_tasks = select(t for t in db.Tasks if t.project == p)
            project_tasks = list(project_tasks)

            rectification = None
            rectification_et = None
            design = None
            design_et = None
            fabrication = None
            fabrication_et = None
            installation = None
            installation_et_first = None
            installation_et_last = None

            for project_task in project_tasks:
                if project_task.skill.id == 1:
                    rectification = project_task
                    rectification_et = db.Employees_Tasks.get(task=rectification)
                if project_task.skill.id == 2:
                    design = project_task
                    design_et = db.Employees_Tasks.get(task=design)
                if project_task.skill.id == 3:
                    fabrication = project_task
                    fabrication_et = db.Employees_Tasks.get(task=fabrication)
                if project_task.skill.id == 4:
                    installation = project_task
                    installation_et_first = list(select(et for et in db.Employees_Tasks if et.task == installation).order_by(lambda et: et.planned_initial_date))[0]
                    installation_et_last = list(select(et for et in db.Employees_Tasks if et.task == installation).order_by(lambda et: et.planned_end_date))[-1]

            # print("proyecto {0}, version {1}".format(p.contract_number,p.version))
            # print(rectification.skill.id)
            # print(design.skill.id)
            # print(fabrication.skill.id)
            # print(installation.skill.id)

            num_of_versions = 1
            project_contract_number = p.contract_number
            project_versions = select(v for v in db.Projects if v.contract_number ==project_contract_number)
            num_of_versions = len(project_versions)
            
            #Identificamos con color rojo las filas correspondientes a proyectos con reparos (con fallos).
            if p.version == num_of_versions:
                rowColorvalue = "E5DFEC"
            else:
                rowColorvalue = "F00909"

            # Obtenemos la fila de costos de la tabla de los costos del proyecto
            p_costs = db.Projects_Costs.get(project=p)

            # Obtenemos la fila correspondiente de la tabla de la información de las hojas de corte
            p_hc = db.Crystals_Sales_Order.get(project=p)

            # Escribe en qué etapa se originó el fallo
            id_tarea_origen_fallo = -1
            for tarea in project_tasks:
                if tarea.fail_cost != None:
                    id_tarea_origen_fallo = tarea.skill.id
            nombres_tareas = {}
            nombres_tareas[1] = "Rectificación"
            nombres_tareas[2] = "Diseño"
            nombres_tareas[3] = "Fabricación"
            nombres_tareas[4] = "Instalación"

            # Obtenemos información de los trabajadores para cada tarea
            # primero seleccionamos el Task asociado a este Project y a cada Skill, sabemos que es solo una con failed != True, así que tomamos first():

            task_rect = db.Tasks.get(project = p, skill = db.Skills[1])
            task_des = db.Tasks.get(project = p, skill = db.Skills[2])
            task_fab = db.Tasks.get(project = p, skill = db.Skills[3])
            task_inst = db.Tasks.get(project = p, skill = db.Skills[4])

            # sabemos que para las Skills 1,2,3 solamente puede haber un empleado encargado, así que tomamos el primero. Para la Skill 4 es distinto:
            employees_tasks_rect = select(
                et for et in db.Employees_Tasks if et.task == task_rect).first()
            employees_tasks_des = select(
                et for et in db.Employees_Tasks if et.task == task_des).first()
            employees_tasks_fab = select(
                et for et in db.Employees_Tasks if et.task == task_fab).first()
            employees_tasks_inst = select(
                et for et in db.Employees_Tasks if et.task == task_inst)

            instalation_employees = ""
            for et in employees_tasks_inst:
                instalation_employees = instalation_employees + str(et.employee) + " ;"

            #Definimos una variable col_pos para ir desplazando con mayor facilidad las posiciones de los datos en las columnas
            col_pos = 4

            # Escribe el número de contrato
            if p.contract_number != None:
                cell = ws.cell(row=r, column=col_pos, value=p.contract_number)
                cell.font = Font(bold=True)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center')
                cell.fill = PatternFill("solid", fgColor=rowColorvalue)
            else:
                cell = ws.cell(row=r, column=col_pos, value="Dato no disponible")
                cell.font = Font(bold=True)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center')
                cell.fill = PatternFill("solid", fgColor=rowColorvalue)
            col_pos+=1

            # Escribe el nombre del cliente
            if p.client_name != None:
                cell = ws.cell(row=r, column=col_pos, value=p.client_name)
                cell.font = Font(bold=True)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center')
                cell.fill = PatternFill("solid", fgColor=rowColorvalue)
            else:
                cell = ws.cell(row=r, column=col_pos, value="Dato no disponible")
                cell.font = Font(bold=True)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center')
                cell.fill = PatternFill("solid", fgColor=rowColorvalue)
            col_pos+=1

            # Escribe la prioridad del proyecto
            if p.priority != None:
                if p.priority >= 0:
                    cell = ws.cell(row=r, column=col_pos, value=p.priority)
                    cell.font = Font(bold=True)
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal='center')
                    cell.fill = PatternFill("solid", fgColor=rowColorvalue)
                else:
                    cell = ws.cell(row=r, column=col_pos, value="---")
                    cell.font = Font(bold=True)
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal='center')
                    cell.fill = PatternFill("solid", fgColor=rowColorvalue)
            else:
                cell = ws.cell(row=r, column=col_pos, value="Dato no disponible")
                cell.font = Font(bold=True)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center')
                cell.fill = PatternFill("solid", fgColor=rowColorvalue)
            col_pos+=1



            # Escribe la fecha límite del proyecto
            if p.deadline != None:
                cell = ws.cell(row=r, column=col_pos, value=p.deadline)
                cell.font = Font(bold=True)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center')
                cell.fill = PatternFill("solid", fgColor=rowColorvalue)
            else:
                cell = ws.cell(row=r, column=col_pos, value="Dato no disponible")
                cell.font = Font(bold=True)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center')
                cell.fill = PatternFill("solid", fgColor=rowColorvalue)
            col_pos+=1

            # Escribe el atraso, en caso de haberlo
            if installation_et_last != None and p.deadline < installation.original_end_date:
                cell = ws.cell(row=r, column=col_pos, value=(
                "{} días".format((installation_et_last.planned_end_date - p.deadline).days)))
                cell.font = Font(bold=True)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center')
                cell.fill = PatternFill("solid", fgColor=rowColorvalue)
            else:
                if installation != None:
                    cell = ws.cell(row=r, column=col_pos, value="A tiempo")
                    cell.font = Font(bold=True)
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal='center')
                    cell.fill = PatternFill("solid", fgColor=rowColorvalue)
                elif rectification != None or design != None or fabrication != None:
                    cell = ws.cell(row=r, column=col_pos, value="Término en versión posterior")
                    cell.font = Font(bold=True)
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal='center')
                    cell.fill = PatternFill("solid", fgColor=rowColorvalue)
                else:  # Este caso no es posible dado que sería una versión sin ninguna tarea
                    pass

            col_pos += 1

            # Escribe el estado inicial del proyecto
            if num_of_versions > 1:  # hay_un_fallo
                if p.version == 1:
                    cell = ws.cell(row=r, column=col_pos, value="NUEVO")
                    cell.font = Font(bold=True)
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal='center')
                    cell.fill = PatternFill("solid", fgColor=rowColorvalue)
                elif p.version < num_of_versions:
                    cell = ws.cell(row=r, column=col_pos, value="CON REPAROS")
                    cell.font = Font(bold=True)
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal='center')
                    cell.fill = PatternFill("solid", fgColor=rowColorvalue)
                else:  # última versión
                    cell = ws.cell(row=r, column=col_pos, value="CON REPAROS")
                    cell.font = Font(bold=True)
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal='center')
                    cell.fill = PatternFill("solid", fgColor=rowColorvalue)
            else:
                cell = ws.cell(row=r, column=col_pos, value="NUEVO")
                cell.font = Font(bold=True)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center')
                cell.fill = PatternFill("solid", fgColor=rowColorvalue)

            col_pos += 1

            # Escribe el estado final del proyecto
            if num_of_versions > 1:  # hay_un_fallo
                if p.version < num_of_versions:
                    cell = ws.cell(row=r, column=col_pos, value="CON REPAROS")
                    cell.font = Font(bold=True)
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal='center')
                    cell.fill = PatternFill("solid", fgColor=rowColorvalue)
                else:  # última versión
                    cell = ws.cell(row=r, column=col_pos, value="OK")
                    cell.font = Font(bold=True)
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal='center')
                    cell.fill = PatternFill("solid", fgColor=rowColorvalue)
            else:
                cell = ws.cell(row=r, column=col_pos, value="OK")
                cell.font = Font(bold=True)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center')
                cell.fill = PatternFill("solid", fgColor=rowColorvalue)

            col_pos += 1

            # Escribe el origen de la falla
            if id_tarea_origen_fallo >= 0:  # hay_un_fallo
                cell = ws.cell(row=r, column=col_pos, value=nombres_tareas[id_tarea_origen_fallo])
                cell.font = Font(bold=True)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center')
                cell.fill = PatternFill("solid", fgColor=rowColorvalue)
            else:
                cell = ws.cell(row=r, column=col_pos, value="No aplica")
                cell.font = Font(bold=True)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center')
                cell.fill = PatternFill("solid", fgColor=rowColorvalue)

            col_pos+=1

            # Escribe la comuna del cliente
            if p.client_comuna != None:
                cell = ws.cell(row=r, column=col_pos, value=p.client_comuna)
                cell.font = Font(bold=True)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center')
                cell.fill = PatternFill("solid", fgColor=rowColorvalue)
            else:
                cell = ws.cell(row=r, column=col_pos, value="Dato no disponible")
                cell.font = Font(bold=True)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center')
                cell.fill = PatternFill("solid", fgColor=rowColorvalue)
            col_pos+=1


            # Escribe la dirección del cliente
            if p.client_address != None:
                cell = ws.cell(row=r, column=col_pos, value=p.client_address)
                cell.font = Font(bold=True)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center')
                cell.fill = PatternFill("solid", fgColor=rowColorvalue)
            else:
                cell = ws.cell(row=r, column=col_pos, value="Dato no disponible")
                cell.font = Font(bold=True)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center')
                cell.fill = PatternFill("solid", fgColor=rowColorvalue)

            col_pos+=1

            # Escribe los metros lineales del proyecto
            if p.linear_meters != None:
                cell = ws.cell(row=r, column=col_pos, value=p.linear_meters)
                cell.font = Font(bold=True)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center')
                cell.fill = PatternFill("solid", fgColor=rowColorvalue)

            else:
                cell = ws.cell(row=r, column=col_pos, value="Dato no disponible")
                cell.font = Font(bold=True)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center')
                cell.fill = PatternFill("solid", fgColor=rowColorvalue)
            col_pos += 1

            # Escribe los metros lineales reales
            if p.real_linear_meters != None:
                cell = ws.cell(row=r, column=col_pos, value=p.real_linear_meters)
                cell.font = Font(bold=True)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center')
                cell.fill = PatternFill("solid", fgColor=rowColorvalue)
            else:
                cell = ws.cell(row=r, column=col_pos, value="Dato no disponible")
                cell.font = Font(bold=True)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center')
                cell.fill = PatternFill("solid", fgColor=rowColorvalue)

            col_pos+=1

            # Escribe el costo estimado del proyecto
            if p.estimated_cost != None:
                cell = ws.cell(row=r, column=col_pos, value=p.estimated_cost)
                cell.font = Font(bold=True)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center')
                cell.fill = PatternFill("solid", fgColor=rowColorvalue)
            else:
                cell = ws.cell(row=r, column=col_pos, value="Dato no disponible")
                cell.font = Font(bold=True)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center')
                cell.fill = PatternFill("solid", fgColor=rowColorvalue)

            col_pos+=1

            # Escribe el precio de venta del proyecto
            if p.sale_price != None:
                cell = ws.cell(row=r, column=col_pos, value=p.sale_price)
                cell.font = Font(bold=True)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center')
                cell.fill = PatternFill("solid", fgColor=rowColorvalue)
            else:
                cell = ws.cell(row=r, column=col_pos, value="Dato no disponible")
                cell.font = Font(bold=True)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center')
                cell.fill = PatternFill("solid", fgColor=rowColorvalue)

            col_pos+=1

            # Escribe la fecha de venta del proyecto
            if p.sale_date != None:
                cell = ws.cell(row=r, column=col_pos, value=p.sale_date)
                cell.font = Font(bold=True)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center')
                cell.fill = PatternFill("solid", fgColor=rowColorvalue)
            else:
                cell = ws.cell(row=r, column=col_pos, value="Dato no disponible")
                cell.font = Font(bold=True)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center')
                cell.fill = PatternFill("solid", fgColor=rowColorvalue)

            col_pos+=1

            # Escribe la fecha de original de inicio de rectificación
            if rectification != None and rectification.original_initial_date != None:
                cell = ws.cell(row=r, column=col_pos, value=rectification.original_initial_date)
                cell.font = Font(bold=True)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center')
                cell.fill = PatternFill("solid", fgColor=rowColorvalue)
            else:
                if rectification != None:
                    cell = ws.cell(row=r, column=col_pos, value="Dato no disponible")
                    cell.font = Font(bold=True)
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal='center')
                    cell.fill = PatternFill("solid", fgColor=rowColorvalue)
                else:
                    cell = ws.cell(row=r, column=col_pos, value="Realizada en versión anterior")
                    cell.font = Font(bold=True)
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal='center')
                    cell.fill = PatternFill("solid", fgColor=rowColorvalue)

            col_pos+=1

            # Escribe la fecha de original de término de rectificación
            if rectification != None and rectification.original_end_date != None:
                cell = ws.cell(row=r, column=col_pos, value=rectification.original_end_date)
                cell.font = Font(bold=True)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center')
                cell.fill = PatternFill("solid", fgColor=rowColorvalue)
            else:
                if rectification != None:
                    cell = ws.cell(row=r, column=col_pos, value="Dato no disponible")
                    cell.font = Font(bold=True)
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal='center')
                    cell.fill = PatternFill("solid", fgColor=rowColorvalue)
                else:
                    cell = ws.cell(row=r, column=col_pos, value="Realizada en versión anterior")
                    cell.font = Font(bold=True)
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal='center')
                    cell.fill = PatternFill("solid", fgColor=rowColorvalue)

            col_pos+=1

            # Escribe la fecha panificada de inicio de rectificación
            if rectification_et != None and rectification_et.planned_initial_date != None:
                cell = ws.cell(row=r, column=col_pos, value=rectification_et.planned_initial_date)
                cell.font = Font(bold=True)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center')
                cell.fill = PatternFill("solid", fgColor=rowColorvalue)
            else:
                if rectification != None:
                    cell = ws.cell(row=r, column=col_pos, value="Dato no disponible")
                    cell.font = Font(bold=True)
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal='center')
                    cell.fill = PatternFill("solid", fgColor=rowColorvalue)
                else:
                    cell = ws.cell(row=r, column=col_pos, value="Realizada en versión anterior")
                    cell.font = Font(bold=True)
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal='center')
                    cell.fill = PatternFill("solid", fgColor=rowColorvalue)

            col_pos += 1

            # Escribe la fecha panificada de término de rectificación
            if rectification_et != None and rectification_et.planned_end_date != None:
                cell = ws.cell(row=r, column=col_pos, value=rectification_et.planned_end_date)
                cell.font = Font(bold=True)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center')
                cell.fill = PatternFill("solid", fgColor=rowColorvalue)
            else:
                if rectification != None:
                    cell = ws.cell(row=r, column=col_pos, value="Dato no disponible")
                    cell.font = Font(bold=True)
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal='center')
                    cell.fill = PatternFill("solid", fgColor=rowColorvalue)
                else:
                    cell = ws.cell(row=r, column=col_pos, value="Realizada en versión anterior")
                    cell.font = Font(bold=True)
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal='center')
                    cell.fill = PatternFill("solid", fgColor=rowColorvalue)

            col_pos += 1


            # Escribe la fecha efectiva inicio de rectificación
            if rectification != None and rectification.effective_initial_date != None:
                cell = ws.cell(row=r, column=col_pos, value=rectification.effective_initial_date)
                cell.font = Font(bold=True)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center')
                cell.fill = PatternFill("solid", fgColor=rowColorvalue)
            else:
                if rectification != None:
                    cell = ws.cell(row=r, column=col_pos, value="Dato no disponible")
                    cell.font = Font(bold=True)
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal='center')
                    cell.fill = PatternFill("solid", fgColor=rowColorvalue)
                else:
                    cell = ws.cell(row=r, column=col_pos, value="Realizada en versión anterior")
                    cell.font = Font(bold=True)
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal='center')
                    cell.fill = PatternFill("solid", fgColor=rowColorvalue)

            col_pos+=1

            # Escribe la fecha efectiva de término de rectificación
            if rectification != None and rectification.effective_end_date != None:
                cell = ws.cell(row=r, column=col_pos, value=rectification.effective_end_date)
                cell.font = Font(bold=True)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center')
                cell.fill = PatternFill("solid", fgColor=rowColorvalue)
            else:
                if rectification != None:
                    cell = ws.cell(row=r, column=col_pos, value="Dato no disponible")
                    cell.font = Font(bold=True)
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal='center')
                    cell.fill = PatternFill("solid", fgColor=rowColorvalue)
                else:
                    cell = ws.cell(row=r, column=col_pos, value="Realizada en versión anterior")
                    cell.font = Font(bold=True)
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal='center')
                    cell.fill = PatternFill("solid", fgColor=rowColorvalue)

            col_pos+=1

            # Escribe el ID del rectificador, en caso de haberlos
            if employees_tasks_rect != None:
                cell = ws.cell(row=r, column=col_pos, value=employees_tasks_rect.employee.id)
                cell.font = Font(bold=True)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center')
                cell.fill = PatternFill("solid", fgColor=rowColorvalue)
            else:
                cell = ws.cell(row=r, column=col_pos, value="Dato no disponible")
                cell.font = Font(bold=True)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center')
                cell.fill = PatternFill("solid", fgColor=rowColorvalue)

            col_pos+=1

            # Escribe el atraso de la rectificación, en caso de haberlo
            if rectification != None and rectification.original_end_date != None and rectification.effective_end_date != None and rectification.original_end_date < rectification.effective_end_date:
                cell = ws.cell(row=r, column=col_pos, value=(
                    "{} días".format((
                                     rectification.effective_end_date - rectification.original_end_date).days)))
                cell.font = Font(bold=True)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center')
                cell.fill = PatternFill("solid", fgColor=rowColorvalue)
            else:
                if rectification != None:
                    cell = ws.cell(row=r, column=col_pos, value="A tiempo")
                    cell.font = Font(bold=True)
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal='center')
                    cell.fill = PatternFill("solid", fgColor=rowColorvalue)
                else:
                    cell = ws.cell(row=r, column=col_pos, value="Realizada en versión anterior")
                    cell.font = Font(bold=True)
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal='center')
                    cell.fill = PatternFill("solid", fgColor=rowColorvalue)

            col_pos+=1

            # Escribe la fecha de original de inicio de diseño
            if design != None and design.original_initial_date != None:
                cell = ws.cell(row=r, column=col_pos, value=design.original_initial_date)
                cell.font = Font(bold=True)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center')
                cell.fill = PatternFill("solid", fgColor=rowColorvalue)
            else:
                if design != None:
                    cell = ws.cell(row=r, column=col_pos, value="Dato no disponible")
                    cell.font = Font(bold=True)
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal='center')
                    cell.fill = PatternFill("solid", fgColor=rowColorvalue)
                elif rectification != None:
                    cell = ws.cell(row=r, column=col_pos, value="Realizada en versión posterior")
                    cell.font = Font(bold=True)
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal='center')
                    cell.fill = PatternFill("solid", fgColor=rowColorvalue)
                elif fabrication != None or installation != None:
                    cell = ws.cell(row=r, column=col_pos, value="Realizada en versión anterior")
                    cell.font = Font(bold=True)
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal='center')
                    cell.fill = PatternFill("solid", fgColor=rowColorvalue)

            col_pos+=1

            # Escribe la fecha de original de término de diseño
            if design != None and design.original_end_date != None:
                cell = ws.cell(row=r, column=col_pos, value=design.original_end_date)
                cell.font = Font(bold=True)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center')
                cell.fill = PatternFill("solid", fgColor=rowColorvalue)
            else:
                if design != None:
                    cell = ws.cell(row=r, column=col_pos, value="Dato no disponible")
                    cell.font = Font(bold=True)
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal='center')
                    cell.fill = PatternFill("solid", fgColor=rowColorvalue)
                elif rectification != None:
                    cell = ws.cell(row=r, column=col_pos, value="Realizada en versión posterior")
                    cell.font = Font(bold=True)
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal='center')
                    cell.fill = PatternFill("solid", fgColor=rowColorvalue)
                elif fabrication != None or installation != None:
                    cell = ws.cell(row=r, column=col_pos, value="Realizada en versión anterior")
                    cell.font = Font(bold=True)
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal='center')
                    cell.fill = PatternFill("solid", fgColor=rowColorvalue)

            col_pos+=1

            # Escribe la fecha planificada de inicio de diseño
            if design_et != None and design_et.planned_initial_date != None:
                cell = ws.cell(row=r, column=col_pos, value=design_et.planned_initial_date)
                cell.font = Font(bold=True)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center')
                cell.fill = PatternFill("solid", fgColor=rowColorvalue)
            else:
                if design != None:
                    cell = ws.cell(row=r, column=col_pos, value="Dato no disponible")
                    cell.font = Font(bold=True)
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal='center')
                    cell.fill = PatternFill("solid", fgColor=rowColorvalue)
                elif rectification != None:
                    cell = ws.cell(row=r, column=col_pos, value="Realizada en versión posterior")
                    cell.font = Font(bold=True)
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal='center')
                    cell.fill = PatternFill("solid", fgColor=rowColorvalue)
                elif fabrication != None or installation != None:
                    cell = ws.cell(row=r, column=col_pos, value="Realizada en versión anterior")
                    cell.font = Font(bold=True)
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal='center')
                    cell.fill = PatternFill("solid", fgColor=rowColorvalue)

            col_pos += 1

            # Escribe la fecha planificada de término de diseño
            if design_et != None and design_et.planned_end_date != None:
                cell = ws.cell(row=r, column=col_pos, value=design_et.planned_end_date)
                cell.font = Font(bold=True)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center')
                cell.fill = PatternFill("solid", fgColor=rowColorvalue)
            else:
                if design != None:
                    cell = ws.cell(row=r, column=col_pos, value="Dato no disponible")
                    cell.font = Font(bold=True)
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal='center')
                    cell.fill = PatternFill("solid", fgColor=rowColorvalue)
                elif rectification != None:
                    cell = ws.cell(row=r, column=col_pos, value="Realizada en versión posterior")
                    cell.font = Font(bold=True)
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal='center')
                    cell.fill = PatternFill("solid", fgColor=rowColorvalue)
                elif fabrication != None or installation != None:
                    cell = ws.cell(row=r, column=col_pos, value="Realizada en versión anterior")
                    cell.font = Font(bold=True)
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal='center')
                    cell.fill = PatternFill("solid", fgColor=rowColorvalue)

            col_pos += 1



            # Escribe la fecha efectiva de inicio de diseño
            if design != None and design.effective_initial_date != None:
                cell = ws.cell(row=r, column=col_pos, value=design.effective_initial_date)
                cell.font = Font(bold=True)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center')
                cell.fill = PatternFill("solid", fgColor=rowColorvalue)
            else:
                if design != None:
                    cell = ws.cell(row=r, column=col_pos, value="Dato no disponible")
                    cell.font = Font(bold=True)
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal='center')
                    cell.fill = PatternFill("solid", fgColor=rowColorvalue)
                elif rectification != None:
                    cell = ws.cell(row=r, column=col_pos, value="Realizada en versión posterior")
                    cell.font = Font(bold=True)
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal='center')
                    cell.fill = PatternFill("solid", fgColor=rowColorvalue)
                elif fabrication != None or installation != None:
                    cell = ws.cell(row=r, column=col_pos, value="Realizada en versión anterior")
                    cell.font = Font(bold=True)
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal='center')
                    cell.fill = PatternFill("solid", fgColor=rowColorvalue)

            col_pos+=1

            # Escribe la fecha efectiva de término de diseño
            if design != None and design.effective_end_date != None:
                cell = ws.cell(row=r, column=col_pos, value=design.effective_end_date)
                cell.font = Font(bold=True)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center')
                cell.fill = PatternFill("solid", fgColor=rowColorvalue)
            else:
                if design != None:
                    cell = ws.cell(row=r, column=col_pos, value="Dato no disponible")
                    cell.font = Font(bold=True)
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal='center')
                    cell.fill = PatternFill("solid", fgColor=rowColorvalue)
                elif rectification != None:
                    cell = ws.cell(row=r, column=col_pos, value="Realizada en versión posterior")
                    cell.font = Font(bold=True)
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal='center')
                    cell.fill = PatternFill("solid", fgColor=rowColorvalue)
                elif fabrication != None or installation != None:
                    cell = ws.cell(row=r, column=col_pos, value="Realizada en versión anterior")
                    cell.font = Font(bold=True)
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal='center')
                    cell.fill = PatternFill("solid", fgColor=rowColorvalue)

            col_pos+=1

            # Escribe el ID del diseñador, en caso de haberlos
            if employees_tasks_des != None:
                cell = ws.cell(row=r, column=col_pos, value=employees_tasks_des.employee.id)
                cell.font = Font(bold=True)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center')
                cell.fill = PatternFill("solid", fgColor=rowColorvalue)
            else:
                cell = ws.cell(row=r, column=col_pos, value="Dato no disponible")
                cell.font = Font(bold=True)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center')
                cell.fill = PatternFill("solid", fgColor=rowColorvalue)

            col_pos += 1

            # Escribe el atraso del diseño, en caso de haberlo
            if design != None and design.original_end_date != None and design.effective_end_date != None and design.original_end_date < design.effective_end_date:
                cell = ws.cell(row=r, column=col_pos, value=(
                    "{} días".format((
                                         design.effective_end_date - design.original_end_date).days)))
                cell.font = Font(bold=True)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center')
                cell.fill = PatternFill("solid", fgColor=rowColorvalue)
            else:
                if design != None:
                    cell = ws.cell(row=r, column=col_pos, value="A tiempo")
                    cell.font = Font(bold=True)
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal='center')
                    cell.fill = PatternFill("solid", fgColor=rowColorvalue)
                elif rectification != None:
                    cell = ws.cell(row=r, column=col_pos, value="Realizada en versión posterior")
                    cell.font = Font(bold=True)
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal='center')
                    cell.fill = PatternFill("solid", fgColor=rowColorvalue)
                elif fabrication != None or installation != None:
                    cell = ws.cell(row=r, column=col_pos, value="Realizada en versión anterior")
                    cell.font = Font(bold=True)
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal='center')
                    cell.fill = PatternFill("solid", fgColor=rowColorvalue)

            col_pos += 1

            # Escribe la fecha de original de inicio de fabricación
            if fabrication != None and fabrication.original_initial_date != None:
                cell = ws.cell(row=r, column=col_pos, value=fabrication.original_initial_date)
                cell.font = Font(bold=True)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center')
                cell.fill = PatternFill("solid", fgColor=rowColorvalue)
            else:
                if fabrication != None:
                    cell = ws.cell(row=r, column=col_pos, value="Dato no disponible")
                    cell.font = Font(bold=True)
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal='center')
                    cell.fill = PatternFill("solid", fgColor=rowColorvalue)
                elif rectification !=None or design != None:
                    cell = ws.cell(row=r, column=col_pos, value="Realizada en versión posterior")
                    cell.font = Font(bold=True)
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal='center')
                    cell.fill = PatternFill("solid", fgColor=rowColorvalue)
                elif installation != None:
                    cell = ws.cell(row=r, column=col_pos, value="Realizada en versión anterior")
                    cell.font = Font(bold=True)
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal='center')
                    cell.fill = PatternFill("solid", fgColor=rowColorvalue)

            col_pos+=1

            # Escribe la fecha de original de término de fabricación
            if fabrication != None and fabrication.original_end_date != None:
                cell = ws.cell(row=r, column=col_pos, value=fabrication.original_end_date)
                cell.font = Font(bold=True)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center')
                cell.fill = PatternFill("solid", fgColor=rowColorvalue)
            else:
                if fabrication != None:
                    cell = ws.cell(row=r, column=col_pos, value="Dato no disponible")
                    cell.font = Font(bold=True)
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal='center')
                    cell.fill = PatternFill("solid", fgColor=rowColorvalue)
                elif rectification !=None or design != None:
                    cell = ws.cell(row=r, column=col_pos, value="Realizada en versión posterior")
                    cell.font = Font(bold=True)
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal='center')
                    cell.fill = PatternFill("solid", fgColor=rowColorvalue)
                elif installation != None:
                    cell = ws.cell(row=r, column=col_pos, value="Realizada en versión anterior")
                    cell.font = Font(bold=True)
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal='center')
                    cell.fill = PatternFill("solid", fgColor=rowColorvalue)

            col_pos+=1

            # Escribe la fecha planificada de inicio de fabricación
            if fabrication_et != None and fabrication_et.planned_initial_date != None:
                cell = ws.cell(row=r, column=col_pos, value=fabrication_et.planned_initial_date)
                cell.font = Font(bold=True)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center')
                cell.fill = PatternFill("solid", fgColor=rowColorvalue)
            else:
                if fabrication != None:
                    cell = ws.cell(row=r, column=col_pos, value="Dato no disponible")
                    cell.font = Font(bold=True)
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal='center')
                    cell.fill = PatternFill("solid", fgColor=rowColorvalue)
                elif rectification != None or design != None:
                    cell = ws.cell(row=r, column=col_pos, value="Realizada en versión posterior")
                    cell.font = Font(bold=True)
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal='center')
                    cell.fill = PatternFill("solid", fgColor=rowColorvalue)
                elif installation != None:
                    cell = ws.cell(row=r, column=col_pos, value="Realizada en versión anterior")
                    cell.font = Font(bold=True)
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal='center')
                    cell.fill = PatternFill("solid", fgColor=rowColorvalue)

            col_pos += 1

            # Escribe la fecha planificada de término de fabricación
            if fabrication_et != None and fabrication_et.planned_end_date != None:
                cell = ws.cell(row=r, column=col_pos, value=fabrication.original_end_date)
                cell.font = Font(bold=True)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center')
                cell.fill = PatternFill("solid", fgColor=rowColorvalue)
            else:
                if fabrication != None:
                    cell = ws.cell(row=r, column=col_pos, value="Dato no disponible")
                    cell.font = Font(bold=True)
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal='center')
                    cell.fill = PatternFill("solid", fgColor=rowColorvalue)
                elif rectification != None or design != None:
                    cell = ws.cell(row=r, column=col_pos, value="Realizada en versión posterior")
                    cell.font = Font(bold=True)
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal='center')
                    cell.fill = PatternFill("solid", fgColor=rowColorvalue)
                elif installation != None:
                    cell = ws.cell(row=r, column=col_pos, value="Realizada en versión anterior")
                    cell.font = Font(bold=True)
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal='center')
                    cell.fill = PatternFill("solid", fgColor=rowColorvalue)

            col_pos += 1



            # Escribe la fecha efectiva de inicio de fabricación
            if fabrication != None and fabrication.effective_initial_date != None:
                cell = ws.cell(row=r, column=col_pos, value=fabrication.effective_initial_date)
                cell.font = Font(bold=True)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center')
                cell.fill = PatternFill("solid", fgColor=rowColorvalue)
            else:
                if fabrication != None:
                    cell = ws.cell(row=r, column=col_pos, value="Dato no disponible")
                    cell.font = Font(bold=True)
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal='center')
                    cell.fill = PatternFill("solid", fgColor=rowColorvalue)
                elif rectification !=None or design != None:
                    cell = ws.cell(row=r, column=col_pos, value="Realizada en versión posterior")
                    cell.font = Font(bold=True)
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal='center')
                    cell.fill = PatternFill("solid", fgColor=rowColorvalue)
                elif installation != None:
                    cell = ws.cell(row=r, column=col_pos, value="Realizada en versión anterior")
                    cell.font = Font(bold=True)
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal='center')
                    cell.fill = PatternFill("solid", fgColor=rowColorvalue)

            col_pos+=1

            # Escribe la fecha efectiva de término de fabricación
            if fabrication != None and fabrication.effective_end_date != None:
                cell = ws.cell(row=r, column=col_pos, value=fabrication.effective_end_date)
                cell.font = Font(bold=True)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center')
                cell.fill = PatternFill("solid", fgColor=rowColorvalue)
            else:
                if fabrication != None:
                    cell = ws.cell(row=r, column=col_pos, value="Dato no disponible")
                    cell.font = Font(bold=True)
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal='center')
                    cell.fill = PatternFill("solid", fgColor=rowColorvalue)
                elif rectification !=None or design != None:
                    cell = ws.cell(row=r, column=col_pos, value="Realizada en versión posterior")
                    cell.font = Font(bold=True)
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal='center')
                    cell.fill = PatternFill("solid", fgColor=rowColorvalue)
                elif installation != None:
                    cell = ws.cell(row=r, column=col_pos, value="Realizada en versión anterior")
                    cell.font = Font(bold=True)
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal='center')
                    cell.fill = PatternFill("solid", fgColor=rowColorvalue)

            col_pos+=1

            # Escribe el ID del fabricador, en caso de haberlo
            if employees_tasks_fab != None:
                cell = ws.cell(row=r, column=col_pos, value=employees_tasks_fab.employee.id)
                cell.font = Font(bold=True)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center')
                cell.fill = PatternFill("solid", fgColor=rowColorvalue)
            else:
                cell = ws.cell(row=r, column=col_pos, value="Dato no disponible")
                cell.font = Font(bold=True)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center')
                cell.fill = PatternFill("solid", fgColor=rowColorvalue)

            col_pos += 1

            # Escribe el atraso de la fabricación, en caso de haberlo
            if fabrication != None and fabrication.original_end_date != None and fabrication.effective_end_date != None and fabrication.original_end_date < fabrication.effective_end_date:
                cell = ws.cell(row=r, column=col_pos, value=("{} días".format(
                    (fabrication.effective_end_date - fabrication.original_end_date).days)))
                cell.font = Font(bold=True)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center')
                cell.fill = PatternFill("solid", fgColor=rowColorvalue)
            else:
                if fabrication != None:
                    cell = ws.cell(row=r, column=col_pos, value="A tiempo")
                    cell.font = Font(bold=True)
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal='center')
                    cell.fill = PatternFill("solid", fgColor=rowColorvalue)
                elif rectification != None or design != None:
                    cell = ws.cell(row=r, column=col_pos, value="Realizada en versión posterior")
                    cell.font = Font(bold=True)
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal='center')
                    cell.fill = PatternFill("solid", fgColor=rowColorvalue)
                elif installation != None:
                    cell = ws.cell(row=r, column=col_pos, value="Realizada en versión anterior")
                    cell.font = Font(bold=True)
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal='center')
                    cell.fill = PatternFill("solid", fgColor=rowColorvalue)

            col_pos += 1

            # Escribe la fecha de original de inicio de instalación
            if installation!= None and installation.original_initial_date != None:
                cell = ws.cell(row=r, column=col_pos, value=installation.original_initial_date)
                cell.font = Font(bold=True)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center')
                cell.fill = PatternFill("solid", fgColor=rowColorvalue)
            else:
                if installation != None:
                    cell = ws.cell(row=r, column=col_pos, value="Dato no disponible")
                    cell.font = Font(bold=True)
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal='center')
                    cell.fill = PatternFill("solid", fgColor=rowColorvalue)
                elif rectification !=None or design != None or fabrication!=None:
                    cell = ws.cell(row=r, column=col_pos, value="Realizada en versión posterior")
                    cell.font = Font(bold=True)
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal='center')
                    cell.fill = PatternFill("solid", fgColor=rowColorvalue)
                else: #Este caso no es posible dado que sería una versión sin ninguna tarea
                    pass

            col_pos+=1

            # Escribe la fecha de original de término de instalación
            if installation!= None and installation.original_end_date != None:
                cell = ws.cell(row=r, column=col_pos, value=installation.original_end_date)
                cell.font = Font(bold=True)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center')
                cell.fill = PatternFill("solid", fgColor=rowColorvalue)
            else:
                if installation != None:
                    cell = ws.cell(row=r, column=col_pos, value="Dato no disponible")
                    cell.font = Font(bold=True)
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal='center')
                    cell.fill = PatternFill("solid", fgColor=rowColorvalue)
                elif rectification !=None or design != None or fabrication!=None:
                    cell = ws.cell(row=r, column=col_pos, value="Realizada en versión posterior")
                    cell.font = Font(bold=True)
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal='center')
                    cell.fill = PatternFill("solid", fgColor=rowColorvalue)
                else: #Este caso no es posible dado que sería una versión sin ninguna tarea
                    pass

            col_pos+=1

            # Escribe la fecha planificada de inicio de instalación
            if installation_et_first != None and installation_et_first.planned_initial_date != None:
                cell = ws.cell(row=r, column=col_pos, value=installation_et_first.planned_initial_date)
                cell.font = Font(bold=True)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center')
                cell.fill = PatternFill("solid", fgColor=rowColorvalue)
            else:
                if installation != None:
                    cell = ws.cell(row=r, column=col_pos, value="Dato no disponible")
                    cell.font = Font(bold=True)
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal='center')
                    cell.fill = PatternFill("solid", fgColor=rowColorvalue)
                elif rectification != None or design != None or fabrication != None:
                    cell = ws.cell(row=r, column=col_pos, value="Realizada en versión posterior")
                    cell.font = Font(bold=True)
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal='center')
                    cell.fill = PatternFill("solid", fgColor=rowColorvalue)
                else:  # Este caso no es posible dado que sería una versión sin ninguna tarea
                    pass

            col_pos += 1

            # Escribe la fecha de original de término de instalación
            if installation_et_last != None and installation_et_last.planned_end_date != None:
                cell = ws.cell(row=r, column=col_pos, value=installation_et_last.planned_end_date)
                cell.font = Font(bold=True)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center')
                cell.fill = PatternFill("solid", fgColor=rowColorvalue)
            else:
                if installation != None:
                    cell = ws.cell(row=r, column=col_pos, value="Dato no disponible")
                    cell.font = Font(bold=True)
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal='center')
                    cell.fill = PatternFill("solid", fgColor=rowColorvalue)
                elif rectification != None or design != None or fabrication != None:
                    cell = ws.cell(row=r, column=col_pos, value="Realizada en versión posterior")
                    cell.font = Font(bold=True)
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal='center')
                    cell.fill = PatternFill("solid", fgColor=rowColorvalue)
                else:  # Este caso no es posible dado que sería una versión sin ninguna tarea
                    pass

            col_pos += 1

            # Escribe la fecha efectiva de inicio de instalación
            if installation!= None and installation.effective_initial_date != None:
                cell = ws.cell(row=r, column=col_pos, value=installation.effective_initial_date)
                cell.font = Font(bold=True)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center')
                cell.fill = PatternFill("solid", fgColor=rowColorvalue)
            else:
                if installation != None:
                    cell = ws.cell(row=r, column=col_pos, value="Dato no disponible")
                    cell.font = Font(bold=True)
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal='center')
                    cell.fill = PatternFill("solid", fgColor=rowColorvalue)
                elif rectification !=None or design != None or fabrication!=None:
                    cell = ws.cell(row=r, column=col_pos, value="Realizada en versión posterior")
                    cell.font = Font(bold=True)
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal='center')
                    cell.fill = PatternFill("solid", fgColor=rowColorvalue)
                else: #Este caso no es posible dado que sería una versión sin ninguna tarea
                    pass

            col_pos+=1

            # Escribe la fecha efectiva de término de instalación
            if installation!= None and installation.effective_end_date != None:
                cell = ws.cell(row=r, column=col_pos, value=installation.effective_end_date)
                cell.font = Font(bold=True)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center')
                cell.fill = PatternFill("solid", fgColor=rowColorvalue)
            else:
                if installation != None:
                    cell = ws.cell(row=r, column=col_pos, value="Dato no disponible")
                    cell.font = Font(bold=True)
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal='center')
                    cell.fill = PatternFill("solid", fgColor=rowColorvalue)
                elif rectification !=None or design != None or fabrication!=None:
                    cell = ws.cell(row=r, column=col_pos, value="Realizada en versión posterior")
                    cell.font = Font(bold=True)
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal='center')
                    cell.fill = PatternFill("solid", fgColor=rowColorvalue)
                else: #Este caso no es posible dado que sería una versión sin ninguna tarea
                    pass

            col_pos+=1

            # Escribe los IDs de los instaladores, en caso de haberlos
            if employees_tasks_inst != None and instalation_employees != "":  # Esto podría realizarse de forma más eficiente
                cell = ws.cell(row=r, column=col_pos, value=instalation_employees[0:-2])
                cell.font = Font(bold=True)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center')
                cell.fill = PatternFill("solid", fgColor=rowColorvalue)
            else:
                cell = ws.cell(row=r, column=col_pos, value="Dato no disponible")
                cell.font = Font(bold=True)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center')
                cell.fill = PatternFill("solid", fgColor=rowColorvalue)

            col_pos += 1

            # Escribe el atraso de la instalación, en caso de haberlo
            if installation != None and installation.original_end_date != None and installation.effective_end_date != None and installation.original_end_date < installation.effective_end_date:
                cell = ws.cell(row=r, column=col_pos, value=(
                    "{} días".format((
                                         installation.effective_end_date - installation.original_end_date).days)))
                cell.font = Font(bold=True)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center')
                cell.fill = PatternFill("solid", fgColor=rowColorvalue)
            else:
                if installation != None:
                    cell = ws.cell(row=r, column=col_pos, value="A tiempo")
                    cell.font = Font(bold=True)
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal='center')
                    cell.fill = PatternFill("solid", fgColor=rowColorvalue)
                elif rectification != None or design != None or fabrication != None:
                    cell = ws.cell(row=r, column=col_pos, value="Realizada en versión posterior")
                    cell.font = Font(bold=True)
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal='center')
                    cell.fill = PatternFill("solid", fgColor=rowColorvalue)
                else:  # Este caso no es posible dado que sería una versión sin ninguna tarea
                    pass

            col_pos += 1




            # Escribe el costo estandar de los perfiles, en caso de haberlo
            # Sería más ordenado separar lo siguiente en un gran if en lugar de usar "and"
            if p_costs!= None and p_costs.standard_cost_profiles != None:
                cell = ws.cell(row=r, column=col_pos, value=p.p_costs.standard_cost_profiles)
                cell.font = Font(bold=True)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center')
                cell.fill = PatternFill("solid", fgColor=rowColorvalue)
            else:
                cell = ws.cell(row=r, column=col_pos, value="Dato no disponible")
                cell.font = Font(bold=True)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center')
                cell.fill = PatternFill("solid", fgColor=rowColorvalue)

            col_pos+=1

            # Escribe el costo estandar de los herrajes, en caso de haberlo
            if p_costs!= None and p_costs.standard_cost_fittings != None:
                cell = ws.cell(row=r, column=col_pos, value=p.p_costs.standard_cost_fittings)
                cell.font = Font(bold=True)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center')
                cell.fill = PatternFill("solid", fgColor=rowColorvalue)
            else:
                cell = ws.cell(row=r, column=col_pos, value="Dato no disponible")
                cell.font = Font(bold=True)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center')
                cell.fill = PatternFill("solid", fgColor=rowColorvalue)

            col_pos+=1

            # Escribe el costo estandar de los cristales, en caso de haberlo
            if p_costs!= None and p_costs.standard_cost_crystals != None:
                cell = ws.cell(row=r, column=col_pos, value=p.p_costs.standard_cost_crystals)
                cell.font = Font(bold=True)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center')
                cell.fill = PatternFill("solid", fgColor=rowColorvalue)
            else:
                cell = ws.cell(row=r, column=col_pos, value="Dato no disponible")
                cell.font = Font(bold=True)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center')
                cell.fill = PatternFill("solid", fgColor=rowColorvalue)

            col_pos+=1

            # Escribe el costo estandar de las materias primas, en caso de haberlo
            if p_costs!= None and p_costs.standard_cost_material != None:
                cell = ws.cell(row=r, column=col_pos, value=p.p_costs.standard_cost_material)
                cell.font = Font(bold=True)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center')
                cell.fill = PatternFill("solid", fgColor=rowColorvalue)
            else:
                cell = ws.cell(row=r, column=col_pos, value="Dato no disponible")
                cell.font = Font(bold=True)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center')
                cell.fill = PatternFill("solid", fgColor=rowColorvalue)


            col_pos+=1

            # Escribe el costo estandar de la fabricación, en caso de haberlo
            if p_costs!= None and p_costs.standard_cost_fabrication != None:
                cell = ws.cell(row=r, column=col_pos, value=p.p_costs.standard_cost_fabrication)
                cell.font = Font(bold=True)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center')
                cell.fill = PatternFill("solid", fgColor=rowColorvalue)
            else:
                cell = ws.cell(row=r, column=col_pos, value="Dato no disponible")
                cell.font = Font(bold=True)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center')
                cell.fill = PatternFill("solid", fgColor=rowColorvalue)


            col_pos+=1

            # Escribe el costo estandar de la instalación, en caso de haberlo
            if p_costs!= None and p_costs.standard_cost_installation != None:
                cell = ws.cell(row=r, column=col_pos, value=p.p_costs.standard_cost_installation)
                cell.font = Font(bold=True)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center')
                cell.fill = PatternFill("solid", fgColor=rowColorvalue)
            else:
                cell = ws.cell(row=r, column=col_pos, value="Dato no disponible")
                cell.font = Font(bold=True)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center')
                cell.fill = PatternFill("solid", fgColor=rowColorvalue)

            col_pos+=1


            # Escribe los constos estandar "adicionales", en caso de haberlos
            if p_costs!= None and p_costs.standard_cost_additionals != None:
                cell = ws.cell(row=r, column=col_pos, value=p.p_costs.standard_cost_additionals)
                cell.font = Font(bold=True)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center')
                cell.fill = PatternFill("solid", fgColor=rowColorvalue)
            else:
                cell = ws.cell(row=r, column=col_pos, value="Dato no disponible")
                cell.font = Font(bold=True)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center')
                cell.fill = PatternFill("solid", fgColor=rowColorvalue)

            col_pos+=1

            # Escribe los costos estandar totales, en caso de haberlos
            if p_costs != None and p_costs.standard_cost_total != None:
                cell = ws.cell(row=r, column=col_pos, value=p.p_costs.standard_cost_total)
                cell.font = Font(bold=True)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center')
                cell.fill = PatternFill("solid", fgColor=rowColorvalue)
            else:
                cell = ws.cell(row=r, column=col_pos, value="Dato no disponible")
                cell.font = Font(bold=True)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center')
                cell.fill = PatternFill("solid", fgColor=rowColorvalue)

            col_pos+=1

            # Escribe los costos efectivos de materias primas, en caso de haberlos
            if p_costs != None and p_costs.effective_cost_material != None:
                cell = ws.cell(row=r, column=col_pos, value=p.p_costs.effective_cost_material)
                cell.font = Font(bold=True)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center')
                cell.fill = PatternFill("solid", fgColor=rowColorvalue)
            else:
                cell = ws.cell(row=r, column=col_pos, value="Dato no disponible")
                cell.font = Font(bold=True)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center')
                cell.fill = PatternFill("solid", fgColor=rowColorvalue)


            col_pos+=1

            # Escribe los costos efectivos de fabricación, en caso de haberlos
            if p_costs != None and p_costs.effective_cost_fabrication != None:
                cell = ws.cell(row=r, column=col_pos, value=p.p_costs.effective_cost_fabrication)
                cell.font = Font(bold=True)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center')
                cell.fill = PatternFill("solid", fgColor=rowColorvalue)
            else:
                cell = ws.cell(row=r, column=col_pos, value="Dato no disponible")
                cell.font = Font(bold=True)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center')
                cell.fill = PatternFill("solid", fgColor=rowColorvalue)


            col_pos+=1

            # Escribe los costos efectivos de instalación, en caso de haberlos
            if p_costs != None and p_costs.effective_cost_installation != None:
                cell = ws.cell(row=r, column=col_pos, value=p.p_costs.effective_cost_installation)
                cell.font = Font(bold=True)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center')
                cell.fill = PatternFill("solid", fgColor=rowColorvalue)
            else:
                cell = ws.cell(row=r, column=col_pos, value="Dato no disponible")
                cell.font = Font(bold=True)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center')
                cell.fill = PatternFill("solid", fgColor=rowColorvalue)

            col_pos+=1


            # Escribe los costos efectivos de complementos, en caso de haberlos
            if p_costs != None and p_costs.effective_cost_complements != None:
                cell = ws.cell(row=r, column=col_pos, value=p.p_costs.effective_cost_complements)
                cell.font = Font(bold=True)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center')
                cell.fill = PatternFill("solid", fgColor=rowColorvalue)
            else:
                cell = ws.cell(row=r, column=col_pos, value="Dato no disponible")
                cell.font = Font(bold=True)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center')
                cell.fill = PatternFill("solid", fgColor=rowColorvalue)

            col_pos+=1


            # Escribe la fecha planificada de emisión de la hoja de corte, en caso de haberlos (De la O.C. de los cristales)
            if p_hc!= None and p_hc.original_issuing_date != None:
                cell = ws.cell(row=r, column=col_pos, value=p_hc.original_issuing_date)
                cell.font = Font(bold=True)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center')
                cell.fill = PatternFill("solid", fgColor=rowColorvalue)
            else:
                cell = ws.cell(row=r, column=col_pos, value="Dato no disponible")
                cell.font = Font(bold=True)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center')
                cell.fill = PatternFill("solid", fgColor=rowColorvalue)

            col_pos+=1

            # Escribe la fecha efectiva de emisión de la hoja de corte, en caso de haberlos (De la O.C. de los cristales)
            if p_hc!= None and p_hc.effective_issuing_date != None:
                cell = ws.cell(row=r, column=col_pos, value=p_hc.effective_issuing_date)
                cell.font = Font(bold=True)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center')
                cell.fill = PatternFill("solid", fgColor=rowColorvalue)
            else:
                cell = ws.cell(row=r, column=col_pos, value="Dato no disponible")
                cell.font = Font(bold=True)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center')
                cell.fill = PatternFill("solid", fgColor=rowColorvalue)

            col_pos+=1


            # Escribe la fecha planificada de llegada de la hoja de corte, en caso de haberlos (De la O.C. de los cristales)
            if p_hc != None and p_hc.original_arrival_date != None:
                cell = ws.cell(row=r, column=col_pos, value=p_hc.original_arrival_date)
                cell.font = Font(bold=True)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center')
                cell.fill = PatternFill("solid", fgColor=rowColorvalue)
            else:
                cell = ws.cell(row=r, column=col_pos, value="Dato no disponible")
                cell.font = Font(bold=True)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center')
                cell.fill = PatternFill("solid", fgColor=rowColorvalue)

            col_pos+=1

            # Escribe la fecha efectiva de llegada de la hoja de corte, en caso de haberlos (De la O.C. de los cristales)
            if p_hc != None and p_hc.effective_arrival_date != None:
                cell = ws.cell(row=r, column=col_pos, value=p_hc.effective_arrival_date)
                cell.font = Font(bold=True)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center')
                cell.fill = PatternFill("solid", fgColor=rowColorvalue)
            else:
                cell = ws.cell(row=r, column=col_pos, value="Dato no disponible")
                cell.font = Font(bold=True)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center')
                cell.fill = PatternFill("solid", fgColor=rowColorvalue)

            col_pos+=1

            # Escribe el ID del emisor de la HC, en caso de haberlos
            if p_hc != None and p_hc.id_issuer_order != None and p_hc.id_issuer_order !="":
                cell = ws.cell(row=r, column=col_pos, value=p_hc.id_issuer_order)
                cell.font = Font(bold=True)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center')
                cell.fill = PatternFill("solid", fgColor=rowColorvalue)
            else:
                cell = ws.cell(row=r, column=col_pos, value="Dato no disponible")
                cell.font = Font(bold=True)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center')
                cell.fill = PatternFill("solid", fgColor=rowColorvalue)

            col_pos+=1


            # Escribe el ID del proveedor de cristales, en caso de haberlos
            if p_hc != None and p_hc.id_crystal_provider != None and p_hc.id_crystal_provider != "":
                cell = ws.cell(row=r, column=col_pos, value=p_hc.id_crystal_provider)
                cell.font = Font(bold=True)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center')
                cell.fill = PatternFill("solid", fgColor=rowColorvalue)
            else:
                cell = ws.cell(row=r, column=col_pos, value="Dato no disponible")
                cell.font = Font(bold=True)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center')
                cell.fill = PatternFill("solid", fgColor=rowColorvalue)

            col_pos+=1

            # Escribe el atraso en la emisión de la HC, en caso de haberlo
            if p_hc != None and p_hc.original_issuing_date!=None and p_hc.effective_issuing_date != None and p_hc.original_issuing_date < p_hc.effective_issuing_date:
                cell = ws.cell(row=r, column=col_pos, value=(
                "{} días".format((p_hc.effective_issuing_date - p_hc.original_issuing_date).days)))
                cell.font = Font(bold=True)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center')
                cell.fill = PatternFill("solid", fgColor=rowColorvalue)
            else:
                cell = ws.cell(row=r, column=col_pos, value="A tiempo")
                cell.font = Font(bold=True)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center')
                cell.fill = PatternFill("solid", fgColor=rowColorvalue)


            col_pos+=1

            # Escribe el atraso en la recepción de los cristales, en caso de haberlo
            if p_hc != None and p_hc.original_arrival_date!=None and p_hc.effective_arrival_date!= None and p_hc.original_arrival_date < p_hc.effective_arrival_date:
                cell = ws.cell(row=r, column=col_pos, value=(
                    "{} días".format(
                        (p_hc.effective_arrival_date - p_hc.original_arrival_date).days)))
                cell.font = Font(bold=True)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center')
                cell.fill = PatternFill("solid", fgColor=rowColorvalue)
            else:
                cell = ws.cell(row=r, column=col_pos, value="A tiempo")
                cell.font = Font(bold=True)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center')
                cell.fill = PatternFill("solid", fgColor=rowColorvalue)


            project_counter = 1
            for pt in project_tasks:
                if pt.fail_cost != None:
                    project_counter+=1
                    failed_task_id = pt.id
                else:
                    pass

            r+=1

    try:
        module_path = os.path.dirname(__file__)
        panorama_folder_path = os.path.abspath(os.path.join(module_path, os.pardir))
        report_folder_path = os.path.join(panorama_folder_path,"Reportes")
        if not os.path.exists(report_folder_path):
            os.makedirs(report_folder_path)
        fn = os.path.join(report_folder_path,"Reporte de Planificación generado el {}.xlsx".format(date.today()))
        wb.save(fn)
    except OSError as e:
        if e.args[0] != 13:
            raise
        input("\n Ha ocurrido un error porque el archivo Reporte global de planificación.xlsx está abierto. Por favor ciérrelo y presione cualquier tecla para que el programa pueda continuar.")

#--------------------------------------------------------------------------------
#--------------------------------------------------------------------------------
#--------------------------------------------------------------------------------
#--------------------------------------------------------------------------------
#--------------------------------------------------------------------------------
#--------------------------------------------------------------------------------
#--------------------------------------------------------------------------------
#--------------------------------------------------------------------------------
#--------------------------------------------------------------------------------
#--------------------------------------------------------------------------------
#--------------------------------------------------------------------------------
#--------------------------------------------------------------------------------
#--------------------------------------------------------------------------------
#--------------------------------------------------------------------------------
#--------------------------------------------------------------------------------
#--------------------------------------------------------------------------------
#--------------------------------------------------------------------------------
#--------------------------------------------------------------------------------
#--------------------------------------------------------------------------------
#--------------------------------------------------------------------------------
#--------------------------------------------------------------------------------
#--------------------------------------------------------------------------------
#--------------------------------------------------------------------------------
#--------------------------------------------------------------------------------
#--------------------------------------------------------------------------------
#--------------------------------------------------------------------------------
#--------------------------------------------------------------------------------
#--------------------------------------------------------------------------------
#--------------------------------------------------------------------------------
#--------------------------------------------------------------------------------
#--------------------------------------------------------------------------------
#--------------------------------------------------------------------------------
#--------------------------------------------------------------------------------
#--------------------------------------------------------------------------------
#--------------------------------------------------------------------------------
#--------------------------------------------------------------------------------
#--------------------------------------------------------------------------------
#--------------------------------------------------------------------------------
#--------------------------------------------------------------------------------
#--------------------------------------------------------------------------------
#--------------------------------------------------------------------------------
#--------------------------------------------------------------------------------
#--------------------------------------------------------------------------------
#--------------------------------------------------------------------------------
#--------------------------------------------------------------------------------







##########################################################
# Métodos relacionados a los informes post-planificación #
##########################################################

'''caused_by_delay es un boolean: si es True significa que estamos usando el método solo para verificar factibilidad de planificacion.
Ese caso significa que createReport() esta siendo llamado desde el metodo createDelay(), y por lo tanto el archivo Excel que creemos aca 
sera solamente temporal y lo borraremos al final. Si es False, significa que creareReport() esta siendo llamado desde el metodo 
doPlanning(), y por lo tanto el archivo Excel no es solo auxiliar: el usuario podrá interferir con el.'''
def createReport(db, Delayed, caused_by_delay):
    if caused_by_delay:
        wb = Workbook()
        by_default_sheet = wb['Sheet']
        wb.remove_sheet(by_default_sheet)
        createPlanningReport(db, wb)
        #creamos un Excel con nombre cualquiera, sera borrado despues de todas maneras
        N = random.randint(20, 40)
        file_name = ''.join(random.choice(string.ascii_uppercase + string.digits) for _ in range(N)) + ".xlsx"
        
        wb.save(file_name)
        changes_plausible = planningChangesPlausible(db, wb)
        os.remove(file_name)
        return changes_plausible[0]
        
    
    with db_session:
        wb = Workbook()
        createDelayedReport(wb, Delayed)
        createPlanningReport(db, wb)
        
        file_open = True
        while(file_open):
            try:
                #para guardar en la carpeta Reportes
                module_path = os.path.dirname(__file__)
                panorama_folder_path = os.path.abspath(os.path.join(module_path, os.pardir))
                report_folder_path = os.path.join(panorama_folder_path,"Reportes")
                if not os.path.exists(report_folder_path):
                    os.makedirs(report_folder_path)
                fn = os.path.join(report_folder_path, 'Reporte planificación propuesta.xlsx')
                wb.save(fn)
                file_open = False
            except OSError as e:
                if e.args[0] != 13:
                    raise
                input("\n Ha ocurrido un error porque el archivo ReportePlanificacion.xlsx está abierto. Por favor ciérrelo y presione cualquier tecla para que el programa pueda continuar.")
        
        
        #entramos en el siguiente ciclo para ver si el usuario acepta la planificación o quiere cambiar datos de los empleados
        while(True):
            print('\n El reporte de la última planificación se encuentra en el archivo Reporte planificación propuesta.xlsx.')
            opt = input(" Marque una de las siguientes opciones:\n - 1: Si acepta la planificación propuesta. \
                                                                \n - 2: Si desea cambiar la asignación de empleados. \
                                                                \n Ingrese la alternativa elegida: ")
            if opt == '1':
                break
            elif opt == '2':
                input(" Realice los cambios en el archivo Reporte planificación propuesta.xlsx, y presione cualquier tecla al terminar.")
                changes_plausible = planningChangesPlausible(db, wb)
                if not changes_plausible[0]:
                    print(changes_plausible[1] + " Los cambios a la planificación no serán aplicados.")
                else:
                    print(changes_plausible[1])
                    implementChanges(db, wb)


def createDelayedReport(wb, Delayed):
    ws = wb.create_sheet(
        "Reporte atrasos")  # aquí se pueden agregar otras restricciones no cumplidas

    # cambiar ancho de columnas
    widths = {"A": 20, "B": 30, "C": 30}
    columns = ["A", "B", "C"]
    for c in columns:
        ws.column_dimensions[c].width = widths[c]

    # escribir texto en algunas celdas, en negrita
    rows = [1, 3, 3, 3]
    columns = [1, 1, 2, 3]
    texts = ["Reporte de atrasos", "Número de contrato", "Fecha de entrega comprometida",
             "Fecha de entrega planificada"]
    for i in range(0, len(rows)):
        cell = ws.cell(row=rows[i], column=columns[i], value=texts[i])
        cell.font = Font(bold=True)

    # llenar con los datos de Delayed
    next_row = 4
    for index, row in Delayed.iterrows():
        ws.cell(row=next_row, column=1, value=row['contract number'])
        ws.cell(row=next_row, column=2, value=row['deadline'])
        ws.cell(row=next_row, column=3, value=row['ending date'])
        next_row = next_row + 1


def createPlanningReport(db, wb):
    ws = wb.create_sheet(
        "Reporte planificación")  # aquí se pueden agregar otras restricciones no cumplidas

    # cambiar ancho de columnas
    widths = {"A": 20, "B": 30, "C": 30, "D": 30, "E": 30, "F": 30, "G": 30, "H": 30, "I": 30,
              "J": 30, "K": 30, "L": 30, "M": 30}
    columns = ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M"]
    for c in columns:
        ws.column_dimensions[c].width = widths[c]

    # escribir texto en algunas celdas, en negrita
    rows = [1, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3]
    columns = [1, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13]
    texts = ["Reporte de planificación", "Número de contrato",
             "Fecha inicio rectificación", "Fecha término rectificación",
             "Empleado encargado rectificación",
             "Fecha inicio diseño", "Fecha término diseño", "Empleado encargado diseño",
             "Fecha inicio fabricación", "Fecha término fabricación",
             "Empleado encargado fabricación",
             "Fecha inicio instalación", "Fecha término instalación",
             "Empleados encargados instalación"]
    for i in range(0, len(rows)):
        cell = ws.cell(row=rows[i], column=columns[i], value=texts[i])
        cell.font = Font(bold=True)

    # llenar con los datos de Delayed
    next_row = 4
    with db_session:
        projects = select(p for p in db.Projects if p.finished == None).order_by(lambda p: p.contract_number)
        for p in projects:
            # primero seleccionamos el Task asociado a este Project y a cada Skill, sabemos que es solo una con failed != True, así que tomamos first():
            task_rect = select(t for t in db.Tasks if t.project == p and t.skill == db.Skills[1] and (t.failed == None or t.failed == False)).first()
            task_des = select(t for t in db.Tasks if t.project == p and t.skill == db.Skills[2] and (t.failed == None or t.failed == False)).first()
            task_fab = select(t for t in db.Tasks if t.project == p and t.skill == db.Skills[3] and (t.failed == None or t.failed == False)).first()
            task_inst = select(t for t in db.Tasks if t.project == p and t.skill == db.Skills[4] and (t.failed == None or t.failed == False)).first()

            # sabemos que para las Skills 1,2,3 solamente puede haber un empleado encargado, así que tomamos el primero. Para la Skill 4 es distinto:
            employees_tasks_rect = select(et for et in db.Employees_Tasks if et.task == task_rect).first()
            employees_tasks_des = select(et for et in db.Employees_Tasks if et.task == task_des).first()
            employees_tasks_fab = select(et for et in db.Employees_Tasks if et.task == task_fab).first()
            employees_tasks_inst = select(et for et in db.Employees_Tasks if et.task == task_inst)
            
            # para el Skill 4, pueden ser varios empleados:
            instalation_planned_initial_date = employees_tasks_inst.first().planned_initial_date
            instalation_planned_end_date = employees_tasks_inst.first().planned_end_date
            instalation_employees = ""
            for et in employees_tasks_inst:
                instalation_employees = instalation_employees + str(et.employee) + " ;"

            # finalmente, escribimos en el Excel el resumen de la planificación:
            values = []
            values.append(p.contract_number)
            for et in [employees_tasks_rect, employees_tasks_des, employees_tasks_fab]:
                # algunas tareas pueden haber sido completadas en una versión anterior, en ese caso llenamos con un mensaje apropiado
                if et != None:
                    values.append(et.planned_initial_date)
                    values.append(et.planned_end_date)
                    values.append(et.employee.id)
                else:
                    values.append('Tarea lista en versión anterior.')
                    values.append('Tarea lista en versión anterior.')
                    values.append('Tarea lista en versión anterior.')
            
            values.append(instalation_planned_initial_date)
            values.append(instalation_planned_end_date)
            values.append(instalation_employees[0:-2])
            
            for c in range(1, len(columns)):
                ws.cell(row=next_row, column=columns[c], value=values[c - 1])
            next_row = next_row + 1



#############################################################################
# Métodos relacionados con cambiar empleados manualmente post-planificación #
#############################################################################

#método que revisa si las asignaciones de empleados que hizo el usuario entregan una planificación factible
def planningChangesPlausible(db, wb):
    #primero cargamos la hoja de planificación y calculamos cuál es la máxima fila
    # wb = load_workbook('ReportePlanificacion.xlsx')
    ws = wb["Reporte planificación"]
    max_row = 3
    while(True):
        if ws.cell(row = max_row + 1, column = 1).value == None:
            break
        max_row = max_row + 1
    #ahora revisamos que la planificación sea factible en distintos sentidos
    if not employeesSkillsPlausible(db, ws, max_row):
        return False, " Uno de los empleados fue asignado a una tarea para la cual no está capacitado, o bien la cantidad de isntaladores senior y junior para un proyecto no coincide."
    if not employeesActivitiesPlausible(db, ws, max_row):
        return False, " Uno de los empleados fue asignado a una tarea en fecha que coincide con sus vacaciones o alguna licencia."
    if not employeesTasksPlausible(db, ws, max_row):
        return False, " Uno de los empleados fue asignado a demasiadas tareas en la misma fecha."
    if not employeesRestrictionsPlausible(db, ws, max_row):
        return False, " No se está respetando la restricción de empleados fijos/vetados en la planificación."
    return True, " Los cambios hechos a la planificación son válidos, por lo tanto, serán aplicados."

#revisa factibilidad en cuanto a que a un empleado no se le asigne una tarea que requiera una Skill que no maneje
def employeesSkillsPlausible(db, ws, max_row):
    with db_session:
        for next_row in range(4, max_row + 1):
            #revisamos primero el caso de las Skills 1, 2, 3 que es el más sencillo
            rectifier = ws.cell(row = next_row, column = 4).value
            designer = ws.cell(row = next_row, column = 7).value
            fabricator = ws.cell(row = next_row, column = 10).value
            
            rectifier_skill = db.Employees_Skills.get(employee = db.Employees[rectifier], skill = db.Skills[1])
            designer_skill = db.Employees_Skills.get(employee = db.Employees[designer], skill = db.Skills[2])
            fabricator_skill = db.Employees_Skills.get(employee = db.Employees[fabricator], skill = db.Skills[3])
            if rectifier_skill == None or rectifier_skill.performance == 0 or \
                designer_skill == None or designer_skill.performance == 0 or \
                fabricator_skill == None or fabricator_skill.performance == 0:
                return False
            
            #ahora revisamos para el Skill 4
            installers = str(ws.cell(row = next_row, column = 13).value).split(';')
            num_seniors = 0
            num_juniors = 0
            for i in installers:
                id = int(i)
                installer_skill = db.Employees_Skills.get(employee = db.Employees[id], skill = db.Skills[4])
                if installer_skill == None or installer_skill.performance == 0:
                    return False
                #revisamos también que por cada Installer Senior haya un Installer Junior
                if db.Employees[id].senior:
                    num_seniors = num_seniors + 1
                else:
                    num_juniors = num_juniors + 1
            if num_seniors != num_juniors:
                return False
        return True
        
#revisa factibilidad en cuanto a que una tarea no tope con actividades (licencia/vacaciones)
def employeesActivitiesPlausible(db, ws, max_row):
    with db_session:
        for next_row in range(4, max_row + 1):
            rectifier = ws.cell(row = next_row, column = 4).value
            designer = ws.cell(row = next_row, column = 7).value
            fabricator = ws.cell(row = next_row, column = 10).value
            installers = str(ws.cell(row = next_row, column = 13).value).split(';')
            employees = [[rectifier], [designer], [fabricator], installers]
            for i in range(0, 4):
                initial_date = ws.cell(row = next_row, column = 3*i+2).value.date()
                end_date = ws.cell(row = next_row, column = 3*i+3).value.date()
                if not employeesAvailableActivities(db, employees[i], initial_date, end_date, True):
                    return False
    return True

#revisa factibilidad en cuanto a que dos tareas del mismo empleado no se topen 
def employeesTasksPlausible(db, ws, max_row):
    from Planning.features import datesOverlap, fillCommitments
    with db_session:
        #recorremos para cada empleado el reporte de planificación, guardando todas las veces en que aparece asignado para una tarea
        employees = select(e for e in db.Employees)
        for e in employees:
            id = e.id
            initial_date_rows = []
            initial_date_columns = []
            skills = []
            for next_row in range(4, max_row + 1):
                if id == ws.cell(row = next_row, column = 4).value:
                    initial_date_rows.append(next_row)
                    initial_date_columns.append(2)
                    skills.append(1)
                if id == ws.cell(row = next_row, column = 7).value:
                    initial_date_rows.append(next_row)
                    initial_date_columns.append(5)
                    skills.append(2)
                if id == ws.cell(row = next_row, column = 10).value:
                    initial_date_rows.append(next_row)
                    initial_date_columns.append(8)
                    skills.append(3)
                if str(id) in str(ws.cell(row = next_row, column = 13).value).split(';'):
                    initial_date_rows.append(next_row)
                    initial_date_columns.append(11)
                    skills.append(4)
            #teniendo esa información guardada, ahora revisamos si calzan entre ellas
            dates = len(initial_date_columns)
            for i in range(0, dates):
                #si el Skill es 3 o 4, entonces su fecha no puede calzar con ninguna otra
                if skills[i] == 3 or skills[i] == 4:
                    for j in (k for k in range(0, dates) if k != i):
                        initial_date_1 = ws.cell(row = initial_date_rows[i], column = initial_date_columns[i]).value.date()
                        end_date_1 = ws.cell(row = initial_date_rows[i], column = initial_date_columns[i] + 1).value.date()
                        initial_date_2 = ws.cell(row = initial_date_rows[j], column = initial_date_columns[j]).value.date()
                        end_date_2 = ws.cell(row = initial_date_rows[j], column = initial_date_columns[j] + 1).value.date()
                        if datesOverlap(initial_date_1, end_date_1, initial_date_2, end_date_2):
                            return False
                #si el Skill es 1 o 2, nos fijamos solamente en las otras fechas con Skill 1 o 2, los otros Skills estan cubiertos por el IF de arriba
                else:
                    initial_date_1 = ws.cell(row = initial_date_rows[i], column = initial_date_columns[i]).value.date()
                    end_date_1 = ws.cell(row = initial_date_rows[i], column = initial_date_columns[i] + 1).value.date()
                    # creamos un arreglo de puros 0's de largo el lapso de tiempo entre initial_date y end_date, y lo vamos llenando con las tareas que tienen
                    commitments_skill_1 = np.zeros(abs((end_date_1 - initial_date_1).days) + 1 )
                    commitments_skill_2 = np.zeros(abs((end_date_1 - initial_date_1).days) + 1 )
                    for j in (k for k in range(0, dates) if skills[k] == 1):
                        initial_date_2 = ws.cell(row = initial_date_rows[j], column = initial_date_columns[j]).value.date()
                        end_date_2 = ws.cell(row = initial_date_rows[j], column = initial_date_columns[j] + 1).value.date()
                        commitments_skill_1 = fillCommitments(commitments_skill_1, initial_date_1, end_date_1, initial_date_2, end_date_2)
                    for j in (k for k in range(0, dates) if skills[k] == 2):
                        initial_date_2 = ws.cell(row = initial_date_rows[j], column = initial_date_columns[j]).value.date()
                        end_date_2 = ws.cell(row = initial_date_rows[j], column = initial_date_columns[j] + 1).value.date()
                        commitments_skill_2 = fillCommitments(commitments_skill_2, initial_date_1, end_date_1, initial_date_2, end_date_2)    
                    # vemos cuánto es lo máximo que puede hacer por día, entre ambos Skills (si alguno es 0, solo entre un Skill)
                    es_skill_1 = db.Employees_Skills.get(employee = db.Employees[id], skill = db.Skills[1])
                    es_skill_2 = db.Employees_Skills.get(employee = db.Employees[id], skill = db.Skills[2])
                    # en este caso nos fijamos solo en los Tasks del Skill 2
                    if es_skill_1 == None:
                        limit_skill_2 = np.floor(es_skill_2.performance)
                        for c in commitments_skill_2:
                            if c > limit_skill_2:
                                return False
                    # en este caso nos fijamos solo en los Tasks del Skill 1
                    elif es_skill_2 == None:
                        limit_skill_1 = np.floor(es_skill_1.performance)
                        for c in commitments_skill_1:
                            if c > limit_skill_1:
                                return False
                    # en este caso nos fijamos en los Tasks del Skill 1 y tambien del Skill 2
                    else:                        
                        limit_skill_1 = np.floor(es_skill_1.performance)
                        limit_skill_2 = np.floor(es_skill_2.performance)
                        for i in range(0, len(commitments_skill_1)):
                            proportion = commitments_skill_1[i]/limit_skill_1 + commitments_skill_2[i]/limit_skill_2
                            if proportion > 1:
                                return False
    return True

#revisa factibilidad en cuanto a empleados fijos/baneados de proyectos
def employeesRestrictionsPlausible(db, ws, max_row):
    projects = []
    employees_assigned = {}
    #primero recuperamos qué empleados fueron asignados a cada proyecto
    for next_row in range(4, max_row + 1):
        project = ws.cell(row = next_row, column = 1).value
        projects.append(project)
        employees = []
        
        employees.append(ws.cell(row = next_row, column = 4).value)
        employees.append(ws.cell(row = next_row, column = 7).value)
        employees.append(ws.cell(row = next_row, column = 10).value)
        installers = str(ws.cell(row = next_row, column = 4).value).split(';')
        
        for i in installers:
            employees.append(int(i))
        
        employees_assigned[project] = employees
    #ahora revisamos que NINGUN empleado vetado en un proyecto haya sido asignado a el, y que los fijos a un proyecto SI hayan sido asignados
    for p in projects:
        #primero revisamos respecto a los vetados
        with db_session:
            emp_rests = select(er for er in db.Employees_Restrictions if er.project.contract_number == p and er.fixed == False)
            employees_vetoed = list(er.employee.id for er in emp_rests)
            for e in employees_assigned[p]:
                if e in employees_vetoed:
                    return False
        #ahora revisamos respecto a los fijos
        with db_session:
            emp_rests = select(er for er in db.Employees_Restrictions if er.project.contract_number == p and er.fixed == True)
            employees_fixed = list(er.employee.id for er in emp_rests)
            for e in employees_fixed:
                if e not in employees_assigned[p]:
                    return False
    return True
    
#método auxiliar para ver si un empleado está disponible según sus Activities
def employeesAvailableActivities(db, ids_employees, initial_date, end_date, activities):
    from Planning.features import datesOverlap
    with db_session:
        emp_acts = select(ea for ea in db.Employees_Activities if ea.employee.id in ids_employees)
        for ea in emp_acts:
            if datesOverlap(initial_date, end_date, ea.initial_date, ea.end_date):
                return False
        return True

#método que aplica los cambios si es que son factibles
def implementChanges(db, wb):
    from Planning.features import assignTask
    with db_session:
        #primero cargamos la hoja de planificación y calculamos cuál es la máxima fila
        # wb = load_workbook('ReportePlanificacion.xlsx')
        ws = wb["Reporte planificación"]
        max_row = 3
        while(True):
            if ws.cell(row = max_row + 1, column = 1).value == None:
                break
            max_row = max_row + 1
        #implementamos primero el caso de las Skills 1, 2, 3 que es el más sencillo
        columns = [4, 7, 10]
        skills = [1, 2, 3]
        for next_row in range(4, max_row + 1):
            for i in range(0, len(columns)):
                #recuperamos el Task, sabiendo que solo puede haber uno para un par Project, Skill que no tenga failed = True, usamos el .first()
                contract_number = ws.cell(row = next_row, column = 1).value
                skill = skills[i]
                task = select(t for t in db.Tasks if t.project == db.Projects.get(contract_number = contract_number, finished = None) and t.skill == db.Skills[skill] and \
                                    (t.failed == None or t.failed == False)).first()
                #luego recuperamos el Employee_Tasks, para cambiar el empleado asociado
                c = columns[i]
                initial_date = ws.cell(row = next_row, column = c - 2).value.date()
                end_date = ws.cell(row = next_row, column = c - 1).value.date()
                new_employee = ws.cell(row = next_row, column = c).value
                employee_task = db.Employees_Tasks.get(task = task, planned_initial_date = initial_date, planned_end_date = end_date)
                if employee_task != None:
                    employee_task.delete()
                    #el commit() es clave, si no, el empleado no se considera borrado aún, y la línea de assignTask() tira un error
                    commit()
                    assignTask(db, [new_employee], task.id, initial_date, end_date)
        #implementamos ahora para el caso de la Skill 4
        for next_row in range(4, max_row + 1):
            #recuperamos el Task, sabiendo que solo puede haber uno para un par Project, Skill que no tenga failed = True, usamos el .first()
            contract_number = ws.cell(row = next_row, column = 1).value
            skill = 4
            task = select(t for t in db.Tasks if t.project == db.Projects.get(contract_number = contract_number, finished = None) and t.skill == db.Skills[skill] and \
                                    (t.failed == None or t.failed == False)).first()
            #luego recuperamos los Employee_Tasks, para cambiar el empleado asociado
            c = 13
            initial_date = ws.cell(row = next_row, column = c - 2).value.date()
            end_date = ws.cell(row = next_row, column = c - 1).value.date()
            new_employees = str(ws.cell(row = next_row, column = c).value).split(';')
            emp_tasks = select(et for et in db.Employees_Tasks if et.task == task and et.planned_initial_date == initial_date and et.planned_end_date == end_date)
            #eliminamos los Employees_Tasks antiguos
            for et in emp_tasks:
                et.delete()
            #el commit() es clave, si no, los Employees_Tasks no se consideran borrado aún, y la línea de assignTask() tira un error
            commit()
            #asignamos los Employees_Tasks nuevos
            assignTask(db, new_employees, task.id, initial_date, end_date)
            