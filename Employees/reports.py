from pony.orm import *
from datetime import date, timedelta
import pandas as pd
import numpy as np
from openpyxl import Workbook, load_workbook
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import Font,Alignment
from openpyxl.styles import PatternFill
from openpyxl.styles.colors import Color
import os

# # from pony.orm import *
from database import db

def createPersonalEmployeeReport(db,id_employee):
    ''' Este método crea un informe en Excel compacto con una proción de la información de la base de datos. '''
    with db_session:
        e = db.Employees.get(id=id_employee)
        es = db.Employees_Skills.get(employee=e)
        e_activities = select(ea for ea in db.Employees_Activities if ea.employee == e) #Para ver lo de las vacaciones
        from Planning.features import isHoliday, isNotWorkday  # Es realmente necesario importar isHoliday
        tareas = select(et for et in db.Employees_Tasks if et.employee == e)

        wb = Workbook()
        by_default_sheet = wb.get_sheet_by_name('Sheet')
        by_default_sheet.title = 'Introducción del informe'
        wb.remove_sheet(by_default_sheet)
        ws = wb.create_sheet(
            "Informe trabajador",index=0)

        widths = {"A": 20, "B": 20, "C": 20,"D": 20, "E": 20, "F": 20,
                  "G": 20, "H": 20, "I": 20,"J": 20, "K": 20, "L": 20,
                  "M": 20, "N": 20}

        heights = {"A": 10, "B": 10, "C": 10, "D": 10, "E": 10, "F": 10,
                  "G": 10, "H": 10, "I": 10, "J": 10, "K": 10, "L": 10,
                  "M": 10, "N": 10}

        thin_border = Border(left=Side(style='thin'),
                             right=Side(style='thin'),
                             top=Side(style='thin'),
                             bottom=Side(style='thin'))

        if (False): #(es.skill.id == 4) <-- introducir esta condición si es que se quiere mostrar la información de "senior"
            # columns = ["A", "B", "C","D","E","F","G","H","I"] #Para un instalador y con el rendimiento
            columns = ["A", "B", "C", "D", "E", "F", "G", "H"]  # Para un instalador sin el rendimiento
            # texts = ["","","","ID", "NOMBRE", "ZONA", "COMPETENCIA",
            #          "SENIOR","RENDIMIENTO"] #Para un instalador y con el rendimiento
            texts = ["","","","ID", "NOMBRE", "ZONA", "COMPETENCIA",
                     "SENIOR"] #Para un instalador sin el rendimiento
        else:
            columns = ["A", "B", "C", "D", "E", "F", "G"]
            texts = ["", "", "", "ID", "NOMBRE", "ZONA", "COMPETENCIA"]

        aux_columns = ["A", "B", "C", "D", "E", "F", "G", "H"] #Solo para mantener el ancho de las celdas en el caso que no sea un instalador

        for c in aux_columns:
            ws.column_dimensions[c].width = widths[c]
            ws.column_dimensions[c].height = heights[c]


        num_columns, letter_columns = zip(*list(enumerate(columns)))
        num_columns = list(num_columns)
        num_columns = [x + 1 for x in num_columns] #Desplazamos los valores en 1 porque Excel está indexado desde el 1 y no desde el 0
        letter_columns = list(letter_columns)

        # escribir los títulos de las columnas, en negrita

        for i in range(4,len(num_columns)+1):
            cell = ws.cell(row=i, column=3, value=texts[i-1])
            cell.font = Font(bold=True,)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='left')


        #Escribir la fecha en la que fue producida el reporte:
        cell = ws.cell(row=3, column=3, value="Reporte producido el: ")
        cell.font = Font(bold=True, )
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='left')

        cell = ws.cell(row=3, column=4, value=date.today())
        cell.font = Font(bold=True, )
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='left')

        # A continuación llenamos con los datos


        r=4
        #Escribe el ID del empleado
        if e.id != None:
            cell = ws.cell(row=r, column=4, value=e.id)
            cell.font = Font(bold=True)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='left')
        else:
            cell = ws.cell(row=r, column=4, value="Dato no disponible")
            cell.font = Font(bold=True)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='left')

        # Escribe el nombre del empleado
        if e.name != None:
            cell = ws.cell(row=r+1, column=4, value=e.name)
            cell.font = Font(bold=True)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='left')
        else:
            cell = ws.cell(row=r+1, column=4, value="Dato no disponible")
            cell.font = Font(bold=True)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='left')

        # Escribe la zona geográfica del empleado
        if e.zone != None:
            cell = ws.cell(row=r+2, column=4, value=e.zone)
            cell.font = Font(bold=True)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='left')
        else:
            cell = ws.cell(row=r+2, column=4, value="Dato no disponible")
            cell.font = Font(bold=True)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='left')

        # Escribe la competencia del empleado (por el momento se asume que solo es una)

        if es.skill != None:
            skill_name = es.skill.name
            cell = ws.cell(row=r+3, column=4, value=skill_name)
            cell.font = Font(bold=True)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='left')
        else:
            cell = ws.cell(row=r+3, column=4, value="Dato no disponible")
            cell.font = Font(bold=True)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='left')

        #Información "delicada"
        # Escribe si el empleado es senior o no (Sí o No)
        # if(es.skill.id == 4):
        #     if e.senior != None:
        #         if e.senior:
        #             value = "Sí"
        #         else:
        #             value = "No"
        #         cell = ws.cell(row=r+4, column=4, value = value)
        #         cell.font = Font(bold=True)
        #         cell.border = thin_border
        #         cell.alignment = Alignment(horizontal='left')
        #     else:
        #         cell = ws.cell(row=r+4, column=4, value="Dato no disponible")
        #         cell.font = Font(bold=True)
        #         cell.border = thin_border
        #         cell.alignment = Alignment(horizontal='left')

        # # Escribe el rendimiento del trabajador
        # if es.performance != None:
        #     cell = ws.cell(row=r+5, column=4, value = es.performance)
        #     cell.font = Font(bold=True)
        #     cell.border = thin_border
        #     cell.alignment = Alignment(horizontal='left')
        # else:
        #     cell = ws.cell(row=r+5, column=4, value="Dato no disponible")
        #     cell.font = Font(bold=True)
        #     cell.border = thin_border
        #     cell.alignment = Alignment(horizontal='left')

        #Preparar el formato del horario/calendario del trabajador
        cell = ws.cell(row=r+9, column=3, value="SEMANA")
        cell.font = Font(bold=True)
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')
        aux_c = 1
        for day in ["Lunes","Martes", "Miercoles","Jueves","Viernes"]:
            cell = ws.cell(row=r + 9, column=3 + aux_c, value="{0}".format(day))
            cell.font = Font(bold=True)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')
            aux_c+=1
        span_of_weeks = 14
        dates = []
        today = date.today()
        initial_week_date = today + timedelta(days=-today.weekday())
        calendar_r = r + 10 #Esta variable para la fila podría definirse se otra forma
        for i in range (0,span_of_weeks):
            dates.append(initial_week_date+i*timedelta(days=7))
        for d in dates:
            cell = ws.cell(row=calendar_r, column=3, value=d)
            cell.font = Font(bold=True)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')
            calendar_r+=1

        for f in [i for i in range(14, 14+len(dates))]:#Revisar los limites de indexación
            ws.row_dimensions[f].height = 40
            for c in [i for i in range(4,9)]:
                cell = ws.cell(row=f, column=c)
                cell.font = Font(bold=True)
                cell.border = thin_border
                # cell.alignment = Alignment(horizontal='center')
                wrap_alignment = Alignment(wrap_text=True, horizontal="center",vertical="center")
                cell.alignment = wrap_alignment



        #Imprimir las tareas del trabajador
        #Se definen las fechas de trabajo del trabajador

        tasks = []
        tasks_dates_blocks = []
        for tarea in tareas:
            working_dates = []
            task_initial_d = tarea.planned_initial_date
            task_final_d = tarea.planned_end_date
            task = db.Tasks.get(id=tarea.task.id)
            tasks.append(task)
            # Ajustar para que solo tome fechas de trabajo en días hábiles
            working_days = (task_final_d-task_initial_d).days
            working_days_span = working_days

            # working_dates = [task_initial_d + i*timedelta(days=1) for i in range(0,working_days)]
            working_days_counter = 0
            while(working_days_counter<working_days_span+1): #Ojo con el +1 que es para tomar el intervalo cerrado de la fecha de término
                if(not isNotWorkday(task_initial_d + working_days_counter*timedelta(days=1))):
                    working_dates.append(task_initial_d + working_days_counter*timedelta(days=1))
                else:
                    # working_days_span+=1 #Ya está considerado en el Planning
                    pass
                working_days_counter+=1
            tasks_dates_blocks.append(working_dates)

        #Funcionalidad para el display de las vacaciones y licencias
        activities = []
        activities_dates_blocks = []
        for e_act in e_activities:
            ea_dates = []
            ea_initial_d = e_act.initial_date
            ea_final_d = e_act.end_date
            activity = db.Activities.get(id=e_act.activity.id)
            activities.append(activity)
            ea_days = (ea_final_d - ea_initial_d).days
            ea_days_span = ea_days
            ea_days_counter = 0
            while (ea_days_counter < ea_days_span + 1):  # Ojo con el +1 que es para tomar el intervalo cerrado de la fecha de término
                if (not isNotWorkday(ea_initial_d + ea_days_counter * timedelta(days=1))):
                    ea_dates.append(ea_initial_d + ea_days_counter * timedelta(days=1))
                else:
                    # working_days_span+=1 #Ya está considerado en el Planning
                    pass
                ea_days_counter += 1
            activities_dates_blocks.append(ea_dates)
        # Funcionalidad para el display de las vacaciones y licencias

        #Se imprimen los días de trabajo
        for i in range(0,len(dates)-1):
            task_counter = 0
            for task_date_block in tasks_dates_blocks:
                for d in task_date_block:
                    if (dates[i] <= d < dates[i + 1]):
                        # cell = ws.cell(row=r +10 + i, column=4+task_initial_d.weekday(), value="{0} \n Proyecto: {1}".format(task_initial_d, task.project))
                        cell = ws.cell(row=r + 10 + i, column=4 + d.weekday(),
                                       value="Proyecto {0}\n{1}".format(tasks[task_counter].project,tasks[task_counter].project.client_address))
                        cell.font = Font(bold=True)
                        cell.border = thin_border
                        cell.alignment = Alignment(horizontal='center')
                        # cell.style.alignment.wrap_text = True #Para autoajustar el tamaño de la celca. Revisar su correcto funcionamiento
                        # create alignment style
                        wrap_alignment = Alignment(wrap_text=True, horizontal="center",
                                                   vertical="center")
                        # assign
                        cell.alignment = wrap_alignment
                        cell.fill = PatternFill("solid", fgColor="ffff00")
                    elif (dates[len(dates) - 1] <= d):
                        # cell = ws.cell(row=r +10 + i, column=4+task_initial_d.weekday(), value="{0} \n Proyecto: {1}".format(task_initial_d, task.project))
                        cell = ws.cell(row=r + 10 + i + 1, column=4 + d.weekday(),
                                       value="Proyecto {0}\n{1}".format(tasks[task_counter].project,
                                                                        tasks[task_counter].project.client_address))
                        cell.font = Font(bold=True)
                        cell.border = thin_border
                        cell.alignment = Alignment(horizontal='center')
                        # cell.style.alignment.wrap_text = True #Para autoajustar el tamaño de la celca. Revisar su correcto funcionamiento
                        # create alignment style
                        wrap_alignment = Alignment(wrap_text=True, horizontal="center",
                                                   vertical="center")
                        # assign
                        cell.alignment = wrap_alignment
                        cell.fill = PatternFill("solid", fgColor="ffff00")
                task_counter+=1

            # Funcionalidad para el display de las vacaciones y licencias
            activity_counter = 0
            for activity_date_block in activities_dates_blocks:
                for d in activity_date_block:
                    if (dates[i] <= d < dates[i + 1]):
                        # cell = ws.cell(row=r +10 + i, column=4+task_initial_d.weekday(), value="{0} \n Proyecto: {1}".format(task_initial_d, task.project))
                        cell = ws.cell(row=r + 10 + i, column=4 + d.weekday(),
                                       value=activities[activity_counter].description)
                        cell.font = Font(bold=True)
                        cell.border = thin_border
                        cell.alignment = Alignment(horizontal='center')
                        # cell.style.alignment.wrap_text = True #Para autoajustar el tamaño de la celca. Revisar su correcto funcionamiento
                        # create alignment style
                        wrap_alignment = Alignment(wrap_text=True, horizontal="center",
                                                   vertical="center")
                        # assign
                        cell.alignment = wrap_alignment
                        if activities[activity_counter].id == 1:
                            cell.fill = PatternFill("solid", fgColor="E60404")
                        elif activities[activity_counter].id == 2:
                            cell.fill = PatternFill("solid", fgColor="810777")
                    elif (dates[len(dates) - 1] <= d):
                        # cell = ws.cell(row=r +10 + i, column=4+task_initial_d.weekday(), value="{0} \n Proyecto: {1}".format(task_initial_d, task.project))
                        cell = ws.cell(row=r + 10 + i + 1, column=4 + d.weekday(),
                                       value=activities[activity_counter].description)
                        cell.font = Font(bold=True)
                        cell.border = thin_border
                        cell.alignment = Alignment(horizontal='center')
                        # cell.style.alignment.wrap_text = True #Para autoajustar el tamaño de la celca. Revisar su correcto funcionamiento
                        # create alignment style
                        wrap_alignment = Alignment(wrap_text=True, horizontal="center",
                                                   vertical="center")
                        # assign
                        cell.alignment = wrap_alignment
                        if activities[activity_counter].id == 1:
                            cell.fill = PatternFill("solid", fgColor="E60404")
                        elif activities[activity_counter].id == 2:
                            cell.fill = PatternFill("solid", fgColor="810777")

                activity_counter += 1
            # Adición de funcionalidad para las vacaciones y licencias (Atención)

        # Imprimir los días feriados
        all_the_days_on_display = []
        for i in range(0,span_of_weeks*7):
            all_the_days_on_display.append(initial_week_date+timedelta(days=i))
        for i in range(0, len(dates)-1):
            for d in all_the_days_on_display:
                if (dates[i] <= d < dates[i + 1]):
                    if(isHoliday(d) and d.weekday() < 5):
                        cell = ws.cell(row=r + 10 + i, column=4 + d.weekday(),value="Feriado")
                        cell.font = Font(bold=True)
                        cell.border = thin_border
                        cell.alignment = Alignment(horizontal='center')
                        wrap_alignment = Alignment(wrap_text=True, horizontal="center",vertical="center")
                        cell.alignment = wrap_alignment
                        cell.fill = PatternFill("solid", fgColor="00ffff")
                    elif(ws.cell(row=r + 10 + i, column=4 + d.weekday()).value==None and d.weekday() < 5):
                        ws.cell(row=r + 10 + i, column=4 + d.weekday()).value = "Disponible"
                        ws.cell(row=r + 10 + i, column=4 + d.weekday()).fill = PatternFill("solid", fgColor="00ff00")
                elif(dates[len(dates)-1] <= d):
                    if (isHoliday(d) and d.weekday() < 5):
                        cell = ws.cell(row=r + 10 + i + 1, column=4 + d.weekday(), value="Feriado")
                        cell.font = Font(bold=True)
                        cell.border = thin_border
                        cell.alignment = Alignment(horizontal='center')
                        wrap_alignment = Alignment(wrap_text=True, horizontal="center",
                                                   vertical="center")
                        cell.alignment = wrap_alignment
                        cell.fill = PatternFill("solid", fgColor="00ffff")
                    elif (ws.cell(row=r + 10 + i + 1, column=4 + d.weekday()).value == None and d.weekday() < 5):
                        ws.cell(row=r + 10 + i + 1, column=4 + d.weekday()).value = "Disponible"
                        ws.cell(row=r + 10 + i + 1, column=4 + d.weekday()).fill = PatternFill("solid", fgColor="00ff00")


        # for f in [i for i in range(14, 14+len(dates))]:
        #     ws.row_dimensions[f].height = 40
        #     for c in [i for i in range(4,9)]:
        #         cell = ws.cell(row=f, column=c)
        #         cell.font = Font(bold=True)
        #         cell.border = thin_border
        #         cell.alignment = Alignment(horizontal='center')

    try:
        module_path = os.path.dirname(__file__)
        panorama_folder_path = os.path.abspath(os.path.join(module_path, os.pardir))
        report_folder_path = os.path.join(panorama_folder_path,"Reportes")
        if not os.path.exists(report_folder_path):
            os.makedirs(report_folder_path)
        report_file_name = "Reporte empleado {0} {1}.xlsx".format(id_employee,date.today())
        fn = os.path.join(report_folder_path,report_file_name)
        wb.save(fn)
    except OSError as e:
        if e.args[0] != 13:
            raise
        input("\n Ha ocurrido un error porque el archivo Reporte empleado {}.xlsx está abierto. Por favor ciérrelo y presione cualquier tecla para que el programa pueda continuar.".format(id_employee))
####################################################################



def createRectificatorsReport(db):
    ''' Este método crea un informe en Excel compacto con una proción de la información de la base de datos. '''
    id_skill = 1
    with db_session:
        employees_skills = select(es for es in db.Employees_Skills if es.skill.id == id_skill)
        employees = []
        for es in employees_skills:
            employee = db.Employees.get(id=es.employee.id)
            employees.append(employee)
        sheet_index = 0
        wb = Workbook()
        by_default_sheet = wb.get_sheet_by_name('Sheet')
        by_default_sheet.title = 'Introducción del informe'
        wb.remove_sheet(by_default_sheet)
        ws = wb.create_sheet("Informe trabajadores",index=sheet_index)
        sheet_index+=1
        employee_row_counter = 0
        for e in employees:
            e_activities = select(ea for ea in db.Employees_Activities if
                                  ea.employee == e)  # Para ver lo de las vacaciones
            from Planning.features import isHoliday, \
                isNotWorkday  # Es realmente necesario importar isHoliday
            tareas = select(et for et in db.Employees_Tasks if et.employee == e)


            widths = {"A": 20, "B": 20, "C": 20, "D": 20, "E": 20, "F": 20,
                      "G": 20, "H": 20, "I": 20, "J": 20, "K": 20, "L": 20,
                      "M": 20, "N": 20, "O": 20, "P": 20, "Q": 20, "R": 20,
                      "S": 20, "T": 20, "U": 20, "V": 20, "W": 20, "X": 20,
                      "Y": 20, "Z": 20, "AA": 20, "AB": 20, "AC": 20, "AD": 20,
                      "AE": 20, "AF": 20, "AG": 20, "AH": 20, "AI": 20, "AJ": 20,
                      "AK": 20, "AL": 20, "AM": 20, "AN": 20, "AO": 20, "AP": 20,
                      "AQ": 20, "AR": 20, "AS": 20, "AT": 20, "AU": 20, "AV": 20,
                      "AW": 20, "AX": 20, "AY": 20, "AZ": 20, "BA": 20, "BB": 20,
                      "BC": 20, "BD": 20, "BE": 20, "BF": 20, "BG": 20, "BH": 20,
                      "BI": 20, "BJ": 20, "BK": 20, "BL": 20, "BM": 20,"BN": 20,
                      "BO": 20, "BP": 20, "BQ": 20, "BR": 20,"BS": 20,"BT": 20,
                      "BU": 20, "BV": 20, "BW": 20, "BX": 20, "BY": 20, "BZ": 20,
                      "CA": 20, "CB": 20, "CC": 20, "CD": 20, "CE": 20, "CF": 20,
                      "CG": 20, "CH": 20, "CI": 20, "CJ": 20, "CK": 20, "CL": 20,
                      "CM": 20, "CN": 20, "CO": 20, "CP": 20, "CQ": 20, "CR": 20,
                      "CS": 20, "CT": 20, "CU": 20, "CV": 20, "CW": 20, "CX": 20,
                      "CY": 20, "CZ": 20, "DA": 20, "DB": 20, "DC": 20, "DD": 20,
                      "DE": 20, "DF": 20, "DG": 20, "DH": 20, "DI": 20, "DJ": 20,
                      "DK": 20, "DL": 20, "DM": 20, "DN": 20, "DO": 20, "DP": 20,
                      "DQ": 20, "DR": 20, "DS": 20, "DT": 20, "DU": 20, "DV": 20,
                      "DW": 20, "DX": 20, "DY": 20, "DZ": 20}

            heights = {"A": 10, "B": 10, "C": 10, "D": 10, "E": 10, "F": 10,
                       "G": 10, "H": 10, "I": 10, "J": 10, "K": 10, "L": 10,
                       "M": 10, "N": 10, "O": 10, "P": 10, "Q": 10, "R": 10,
                       "S": 10, "T": 10, "U": 10, "V": 10, "W": 10, "X": 10,
                       "Y": 10, "Z": 10, "AA": 10, "AB": 10, "AC": 10, "AD": 10,
                       "AE": 10, "AF": 10, "AG": 10, "AH": 10, "AI": 10, "AJ": 10,
                       "AK": 10, "AL": 10, "AM": 10, "AN": 10, "AO": 10, "AP": 10,
                       "AQ": 10, "AR": 10, "AS": 10, "AT": 10, "AU": 10, "AV": 10,
                       "AW": 10, "AX": 10, "AY": 10, "AZ": 10, "BA": 10, "BB": 10,
                       "BC": 10, "BD": 10, "BE": 10, "BF": 10, "BG": 10, "BH": 10,
                       "BI": 10, "BJ": 10, "BK": 10, "BL": 10, "BM": 10,"BN": 20,
                      "BO": 10, "BP": 10, "BQ": 10, "BR": 10,"BS": 10,"BT": 10,
                      "BU": 10, "BV": 10, "BW": 10, "BX": 10, "BY": 10, "BZ": 10,
                    "CA": 10, "CB": 10,"CC": 10, "CD": 10, "CE": 10, "CF": 10,
                       "CG": 10, "CH": 10,"CI": 10, "CJ": 10, "CK": 10, "CL": 10,
                       "CM": 10,"CN": 10,"CO": 10, "CP": 10, "CQ": 10, "CR": 10,
                       "CS": 10,"CT": 10,"CU": 10, "CV": 10, "CW": 10, "CX": 10,
                       "CY": 10, "CZ": 10, "DA": 10, "DB": 10,"DC": 10, "DD": 10,
                       "DE": 10, "DF": 10,"DG": 10, "DH": 10,"DI": 10, "DJ": 10,
                       "DK": 10, "DL": 10,"DM": 10,"DN": 10,"DO": 10, "DP": 10,
                       "DQ": 10, "DR": 10,"DS": 10,"DT": 10,"DU": 10, "DV": 10,
                       "DW": 10, "DX": 10,"DY": 10, "DZ": 10}

            thin_border = Border(left=Side(style='thin'),
                                 right=Side(style='thin'),
                                 top=Side(style='thin'),
                                 bottom=Side(style='thin'))

            columns = ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M", "N",
                       "O", "P", "Q", "R", "S", "T", "U", "V", "W", "X", "Y", "Z", "AA", "AB",
                       "AC", "AD", "AE", "AF", "AG", "AH", "AI", "AJ", "AK", "AL", "AM", "AN",
                       "AO", "AP", "AQ", "AR", "AS", "AT", "AU", "AV", "AW", "AX", "AY", "AZ",
                       "BA", "BB", "BC", "BD", "BE", "BF", "BG", "BH", "BI", "BJ", "BK", "BL", "BM",
                       "BN","BO","BP","BQ","BR","BS","BT","BU","BV","BW","BX","BY","BZ",
                       "CA", "CB", "CC", "CD", "CE", "CF", "CG", "CH", "CI", "CJ", "CK", "CL", "CM",
                       "CN", "CO", "CP", "CQ", "CR", "CS", "CT", "CU", "CV", "CW", "CX", "CY", "CZ",
                       "DA", "DB", "DC", "DD", "DE", "DF", "DG", "DH", "DI", "DJ", "DK", "DL", "DM",
                       "DN", "DO", "DP", "DQ", "DR", "DS", "DT", "DU", "DV", "DW", "DX", "DY", "DZ"]
            for c in columns:
                ws.column_dimensions[c].width = widths[c]
                ws.column_dimensions[c].height = heights[c]

            thin_border = Border(left=Side(style='thin'),
                                 right=Side(style='thin'),
                                 top=Side(style='thin'),
                                 bottom=Side(style='thin'))

            if (
            False):  # (es.skill.id == 4) <-- introducir esta condición si es que se quiere mostrar la información de "senior"
                # columns = ["A", "B", "C","D","E","F","G","H","I"] #Para un instalador y con el rendimiento
                columns = ["A", "B", "C", "D", "E", "F", "G",
                           "H"]  # Para un instalador sin el rendimiento
                # texts = ["","","","ID", "NOMBRE", "ZONA", "COMPETENCIA",
                #          "SENIOR","RENDIMIENTO"] #Para un instalador y con el rendimiento
                texts = ["", "", "", "ID", "NOMBRE", "ZONA", "COMPETENCIA",
                         "SENIOR"]  # Para un instalador sin el rendimiento
            else:
                columns = ["A", "B", "C", "D", "E", "F", "G"]
                texts = ["", "", "", "ID", "NOMBRE", "ZONA", "COMPETENCIA"]

            aux_columns = ["A", "B", "C", "D", "E", "F", "G",
                           "H"]  # Solo para mantener el ancho de las celdas en el caso que no sea un instalador

            for c in aux_columns:
                ws.column_dimensions[c].width = widths[c]
                ws.column_dimensions[c].height = heights[c]

            num_columns, letter_columns = zip(*list(enumerate(columns)))
            num_columns = list(num_columns)
            num_columns = [x + 1 for x in
                           num_columns]  # Desplazamos los valores en 1 porque Excel está indexado desde el 1 y no desde el 0
            letter_columns = list(letter_columns)


            # Escribir la fecha en la que fue producida el reporte:
            cell = ws.cell(row=6, column=3, value="Reporte producido el: ")
            cell.font = Font(bold=True, )
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='left')

            cell = ws.cell(row=6, column=4, value=date.today())
            cell.font = Font(bold=True, )
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='left')

            # A continuación llenamos con los datos


            r = 4
            r += employee_row_counter
            # Escribe el ID del empleado
            if e.id != None:
                cell = ws.cell(row=r+10, column=3, value="id: {0}\nNombre:{1}".format(e.id,e.name))
                cell.font = Font(bold=True)
                cell.border = thin_border
                wrap_alignment = Alignment(wrap_text=True, horizontal="center",
                                           vertical="center")
                # assign
                cell.alignment = wrap_alignment

            # Preparar el formato del horario/calendario del trabajador
            cell = ws.cell(row=4 + 9, column=3, value="EMPLEADO")
            cell.font = Font(bold=True)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')

            span_of_weeks = 15
            week_counter = 0
            aux_c = 0
            for week in range(0,span_of_weeks):
                for day in ["Lunes", "Martes", "Miercoles", "Jueves", "Viernes", "Sábado", "Domingo"]:
                    cell = ws.cell(row=4 + 9, column=4 + aux_c, value="{0} {1}".format(day,date.today()+timedelta(days=aux_c-date.today().weekday())))
                    cell.font = Font(bold=True)
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal='center')
                    aux_c += 1
                week_counter+=1

            dates = []
            today = date.today()
            for i in range(0, span_of_weeks*7):
                # dates.append(today + i * timedelta(days=1))
                dates.append(date.today()+timedelta(days=i-date.today().weekday()))
            # Imprimir las tareas del trabajador



            # Se definen las fechas de trabajo del trabajador

            tasks = []
            tasks_dates_blocks = []
            for tarea in tareas:
                working_dates = []
                task_initial_d = tarea.planned_initial_date
                task_final_d = tarea.planned_end_date
                task = db.Tasks.get(id=tarea.task.id)
                tasks.append(task)
                # Ajustar para que solo tome fechas de trabajo en días hábiles
                working_days = (task_final_d - task_initial_d).days
                working_days_span = working_days

                # working_dates = [task_initial_d + i*timedelta(days=1) for i in range(0,working_days)]
                working_days_counter = 0
                while (working_days_counter < working_days_span + 1):  # Ojo con el +1 que es para tomar el intervalo cerrado de la fecha de término
                    if (not isNotWorkday(task_initial_d + working_days_counter * timedelta(days=1))):
                        working_dates.append(task_initial_d + working_days_counter * timedelta(days=1))
                    else:
                        # working_days_span+=1 #Ya está considerado en el Planning
                        pass
                    working_days_counter += 1
                tasks_dates_blocks.append(working_dates)

            # # ########Adición de funcionalidad para las vacaciones y licencias (Atención)
            activities = []
            activities_dates_blocks = []
            for e_act in e_activities:
                ea_dates = []
                ea_initial_d = e_act.initial_date
                ea_final_d = e_act.end_date
                activity = db.Activities.get(id=e_act.activity.id)
                activities.append(activity)
                # Ajustar para que solo tome fechas de trabajo en días hábiles
                ea_days = (ea_final_d - ea_initial_d).days
                ea_days_span = ea_days

                # working_dates = [task_initial_d + i*timedelta(days=1) for i in range(0,working_days)]
                ea_days_counter = 0
                while (
                    ea_days_counter < ea_days_span + 1):  # Ojo con el +1 que es para tomar el intervalo cerrado de la fecha de término
                    if (not isNotWorkday(ea_initial_d + ea_days_counter * timedelta(days=1))):
                        ea_dates.append(ea_initial_d + ea_days_counter * timedelta(days=1))
                    else:
                        # working_days_span+=1 #Ya está considerado en el Planning
                        pass
                    ea_days_counter += 1
                    # # #########Adición de funcionalidad para las vacaciones y licencias (Atención)
                activities_dates_blocks.append(ea_dates)
            # print(ea_dates)

            # Se imprimen los días de trabajo
            for i in range(0,len(dates)):
                task_counter = 0
                for tasks_dates_block in tasks_dates_blocks:
                    for d in tasks_dates_block:
                        if d == dates[i]:
                            cell = ws.cell(row=r + 10, column=4 + i,
                                           value="Proyecto {0}\n{1}".format(
                                               tasks[task_counter].project,
                                               tasks[task_counter].project.client_address))
                            cell.font = Font(bold=True)
                            cell.border = thin_border
                            cell.alignment = Alignment(horizontal='center')
                            # cell.style.alignment.wrap_text = True #Para autoajustar el tamaño de la celca. Revisar su correcto funcionamiento
                            # create alignment style
                            wrap_alignment = Alignment(wrap_text=True, horizontal="center",
                                                       vertical="center")
                            # assign
                            cell.alignment = wrap_alignment
                            cell.fill = PatternFill("solid", fgColor="ffff00")
                    task_counter = 0
                activity_counter = 0
                for activity_date_block in activities_dates_blocks:
                    for d in activity_date_block:
                        if d == dates[i]:
                            cell = ws.cell(row=r + 10 , column=4 + i,
                                           value=activities[activity_counter].description)
                            cell.font = Font(bold=True)
                            cell.border = thin_border
                            wrap_alignment = Alignment(wrap_text=True, horizontal="center",
                                                       vertical="center")
                            cell.alignment = wrap_alignment
                            # cell.style.alignment.wrap_text = True #Para autoajustar el tamaño de la celca. Revisar su correcto funcionamiento
                            # create alignment style
                            wrap_alignment = Alignment(wrap_text=True, horizontal="center",
                                                       vertical="center")
                            if activities[activity_counter].id == 1:
                                cell.fill = PatternFill("solid", fgColor="E60404")
                            elif activities[activity_counter].id == 2:
                                cell.fill = PatternFill("solid", fgColor="810777")
                    activity_counter = 0
                if isHoliday(dates[i]):
                    cell = ws.cell(row=r + 10, column=4 + i, value="Feriado")
                    cell.font = Font(bold=True)
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal='center')
                    wrap_alignment = Alignment(wrap_text=True, horizontal="center",
                                               vertical="center")
                    cell.alignment = wrap_alignment
                    cell.fill = PatternFill("solid", fgColor="00ffff")
                if dates[i].weekday()>=5:
                    cell = ws.cell(row=r + 10, column=4 + i, value="Fin de semana")
                    cell.font = Font(bold=True)
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal='center')
                    wrap_alignment = Alignment(wrap_text=True, horizontal="center",
                                               vertical="center")
                    cell.alignment = wrap_alignment
                    cell.fill = PatternFill("solid", fgColor="00ffff")
                if (ws.cell(row=r + 10, column=4 + i).value == None and dates[i].weekday() < 5): #ws.cell(row=r + 10, column=4 + i).value == None and
                    cell = ws.cell(row=r + 10, column=4 + i, value="Disponible")
                    cell.font = Font(bold=True)
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal='center')
                    wrap_alignment = Alignment(wrap_text=True, horizontal="center",
                                               vertical="center")
                    cell.alignment = wrap_alignment
                    cell.fill = PatternFill("solid", fgColor="00ff00")

            employee_row_counter+=1
        for e in employees:
            e_activities = select(ea for ea in db.Employees_Activities if ea.employee == e) #Para ver lo de las vacaciones
            from Planning.features import isHoliday, isNotWorkday  # Es realmente necesario importar isHoliday
            tareas = select(et for et in db.Employees_Tasks if et.employee == e)
            ws = wb.create_sheet(
                "Informe trabajador {}".format(e.id),index=sheet_index)

            widths = {"A": 20, "B": 20, "C": 20,"D": 20, "E": 20, "F": 20,
                      "G": 20, "H": 20, "I": 20,"J": 20, "K": 20, "L": 20,
                      "M": 20, "N": 20}

            heights = {"A": 10, "B": 10, "C": 10, "D": 10, "E": 10, "F": 10,
                      "G": 10, "H": 10, "I": 10, "J": 10, "K": 10, "L": 10,
                      "M": 10, "N": 10}

            thin_border = Border(left=Side(style='thin'),
                                 right=Side(style='thin'),
                                 top=Side(style='thin'),
                                 bottom=Side(style='thin'))

            if (False): #(es.skill.id == 4) <-- introducir esta condición si es que se quiere mostrar la información de "senior"
                # columns = ["A", "B", "C","D","E","F","G","H","I"] #Para un instalador y con el rendimiento
                columns = ["A", "B", "C", "D", "E", "F", "G", "H"]  # Para un instalador sin el rendimiento
                # texts = ["","","","ID", "NOMBRE", "ZONA", "COMPETENCIA",
                #          "SENIOR","RENDIMIENTO"] #Para un instalador y con el rendimiento
                texts = ["","","","ID", "NOMBRE", "ZONA", "COMPETENCIA",
                         "SENIOR"] #Para un instalador sin el rendimiento
            else:
                columns = ["A", "B", "C", "D", "E", "F", "G"]
                texts = ["", "", "", "ID", "NOMBRE", "ZONA", "COMPETENCIA"]

            aux_columns = ["A", "B", "C", "D", "E", "F", "G", "H"] #Solo para mantener el ancho de las celdas en el caso que no sea un instalador

            for c in aux_columns:
                ws.column_dimensions[c].width = widths[c]
                ws.column_dimensions[c].height = heights[c]


            num_columns, letter_columns = zip(*list(enumerate(columns)))
            num_columns = list(num_columns)
            num_columns = [x + 1 for x in num_columns] #Desplazamos los valores en 1 porque Excel está indexado desde el 1 y no desde el 0
            letter_columns = list(letter_columns)

            # escribir los títulos de las columnas, en negrita

            for i in range(4,len(num_columns)+1):
                cell = ws.cell(row=i, column=3, value=texts[i-1])
                cell.font = Font(bold=True,)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='left')


            #Escribir la fecha en la que fue producida el reporte:
            cell = ws.cell(row=3, column=3, value="Reporte producido el: ")
            cell.font = Font(bold=True, )
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='left')

            cell = ws.cell(row=3, column=4, value=date.today())
            cell.font = Font(bold=True, )
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='left')

            # A continuación llenamos con los datos


            r=4
            #Escribe el ID del empleado
            if e.id != None:
                cell = ws.cell(row=r, column=4, value=e.id)
                cell.font = Font(bold=True)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='left')
            else:
                cell = ws.cell(row=r, column=4, value="Dato no disponible")
                cell.font = Font(bold=True)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='left')

            # Escribe el nombre del empleado
            if e.name != None:
                cell = ws.cell(row=r+1, column=4, value=e.name)
                cell.font = Font(bold=True)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='left')
            else:
                cell = ws.cell(row=r+1, column=4, value="Dato no disponible")
                cell.font = Font(bold=True)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='left')

            # Escribe la zona geográfica del empleado
            if e.zone != None:
                cell = ws.cell(row=r+2, column=4, value=e.zone)
                cell.font = Font(bold=True)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='left')
            else:
                cell = ws.cell(row=r+2, column=4, value="Dato no disponible")
                cell.font = Font(bold=True)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='left')

            # Escribe la competencia del empleado (por el momento se asume que solo es una)

            if es.skill != None:
                skill_name = es.skill.name
                cell = ws.cell(row=r+3, column=4, value=skill_name)
                cell.font = Font(bold=True)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='left')
            else:
                cell = ws.cell(row=r+3, column=4, value="Dato no disponible")
                cell.font = Font(bold=True)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='left')

            # Escribe si el empleado es senior o no (Sí o No)
            # if(es.skill.id == 4):
            #     if e.senior != None:
            #         if e.senior:
            #             value = "Sí"
            #         else:
            #             value = "No"
            #         cell = ws.cell(row=r+4, column=4, value = value)
            #         cell.font = Font(bold=True)
            #         cell.border = thin_border
            #         cell.alignment = Alignment(horizontal='left')
            #     else:
            #         cell = ws.cell(row=r+4, column=4, value="Dato no disponible")
            #         cell.font = Font(bold=True)
            #         cell.border = thin_border
            #         cell.alignment = Alignment(horizontal='left')

            # # Escribe el rendimiento del trabajador
            # if es.performance != None:
            #     cell = ws.cell(row=r+5, column=4, value = es.performance)
            #     cell.font = Font(bold=True)
            #     cell.border = thin_border
            #     cell.alignment = Alignment(horizontal='left')
            # else:
            #     cell = ws.cell(row=r+5, column=4, value="Dato no disponible")
            #     cell.font = Font(bold=True)
            #     cell.border = thin_border
            #     cell.alignment = Alignment(horizontal='left')

            #Preparar el formato del horario/calendario del trabajador
            cell = ws.cell(row=r+9, column=3, value="SEMANA")
            cell.font = Font(bold=True)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')
            aux_c = 1
            for day in ["Lunes","Martes", "Miercoles","Jueves","Viernes"]:
                cell = ws.cell(row=r + 9, column=3 + aux_c, value="{0}".format(day))
                cell.font = Font(bold=True)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center')
                aux_c+=1
            span_of_weeks = 14
            dates = []
            today = date.today()
            initial_week_date = today + timedelta(days=-today.weekday())
            calendar_r = r + 10 #Esta variable para la fila podría definirse se otra forma
            for i in range (0,span_of_weeks):
                dates.append(initial_week_date+i*timedelta(days=7))
            for d in dates:
                cell = ws.cell(row=calendar_r, column=3, value=d)
                cell.font = Font(bold=True)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center')
                calendar_r+=1

            for f in [i for i in range(14, 14+len(dates))]:#Revisar los limites de indexación
                ws.row_dimensions[f].height = 40
                for c in [i for i in range(4,9)]:
                    cell = ws.cell(row=f, column=c)
                    cell.font = Font(bold=True)
                    cell.border = thin_border
                    # cell.alignment = Alignment(horizontal='center')
                    wrap_alignment = Alignment(wrap_text=True, horizontal="center",vertical="center")
                    cell.alignment = wrap_alignment



            #Imprimir las tareas del trabajador



            #Se definen las fechas de trabajo del trabajador

            tasks = []
            tasks_dates_blocks = []
            for tarea in tareas:
                working_dates = []
                task_initial_d = tarea.planned_initial_date
                task_final_d = tarea.planned_end_date
                task = db.Tasks.get(id=tarea.task.id)
                tasks.append(task)
                # Ajustar para que solo tome fechas de trabajo en días hábiles
                working_days = (task_final_d-task_initial_d).days
                working_days_span = working_days

                # working_dates = [task_initial_d + i*timedelta(days=1) for i in range(0,working_days)]
                working_days_counter = 0
                while(working_days_counter<working_days_span+1): #Ojo con el +1 que es para tomar el intervalo cerrado de la fecha de término
                    if(not isNotWorkday(task_initial_d + working_days_counter*timedelta(days=1))):
                        working_dates.append(task_initial_d + working_days_counter*timedelta(days=1))
                    else:
                        # working_days_span+=1 #Ya está considerado en el Planning
                        pass
                    working_days_counter+=1
                tasks_dates_blocks.append(working_dates)

            # # ########Adición de funcionalidad para las vacaciones y licencias (Atención)
            activities = []
            activities_dates_blocks = []
            for e_act in e_activities:
                ea_dates = []
                ea_initial_d = e_act.initial_date
                ea_final_d = e_act.end_date
                activity = db.Activities.get(id=e_act.activity.id)
                activities.append(activity)
                # Ajustar para que solo tome fechas de trabajo en días hábiles
                ea_days = (ea_final_d - ea_initial_d).days
                ea_days_span = ea_days

                # working_dates = [task_initial_d + i*timedelta(days=1) for i in range(0,working_days)]
                ea_days_counter = 0
                while (ea_days_counter < ea_days_span + 1):  # Ojo con el +1 que es para tomar el intervalo cerrado de la fecha de término
                    if (not isNotWorkday(ea_initial_d + ea_days_counter * timedelta(days=1))):
                        ea_dates.append(ea_initial_d + ea_days_counter * timedelta(days=1))
                    else:
                        # working_days_span+=1 #Ya está considerado en el Planning
                        pass
                    ea_days_counter += 1
                    # # #########Adición de funcionalidad para las vacaciones y licencias (Atención)
                activities_dates_blocks.append(ea_dates)
            # print(ea_dates)

            #Se imprimen los días de trabajo
            for i in range(0,len(dates)-1):
                task_counter = 0
                for task_date_block in tasks_dates_blocks:
                    for d in task_date_block:
                        if (dates[i] <= d < dates[i + 1]):
                            # cell = ws.cell(row=r +10 + i, column=4+task_initial_d.weekday(), value="{0} \n Proyecto: {1}".format(task_initial_d, task.project))
                            cell = ws.cell(row=r + 10 + i, column=4 + d.weekday(),
                                           value="Proyecto {0}\n{1}".format(tasks[task_counter].project,tasks[task_counter].project.client_address))
                            cell.font = Font(bold=True)
                            cell.border = thin_border
                            cell.alignment = Alignment(horizontal='center')
                            # cell.style.alignment.wrap_text = True #Para autoajustar el tamaño de la celca. Revisar su correcto funcionamiento
                            # create alignment style
                            wrap_alignment = Alignment(wrap_text=True, horizontal="center",
                                                       vertical="center")
                            # assign
                            cell.alignment = wrap_alignment
                            cell.fill = PatternFill("solid", fgColor="ffff00")
                        elif (dates[len(dates) - 1] <= d):
                            # cell = ws.cell(row=r +10 + i, column=4+task_initial_d.weekday(), value="{0} \n Proyecto: {1}".format(task_initial_d, task.project))
                            cell = ws.cell(row=r + 10 + i + 1, column=4 + d.weekday(),
                                           value="Proyecto {0}\n{1}".format(tasks[task_counter].project,
                                                                            tasks[task_counter].project.client_address))
                            cell.font = Font(bold=True)
                            cell.border = thin_border
                            cell.alignment = Alignment(horizontal='center')
                            # cell.style.alignment.wrap_text = True #Para autoajustar el tamaño de la celca. Revisar su correcto funcionamiento
                            # create alignment style
                            wrap_alignment = Alignment(wrap_text=True, horizontal="center",
                                                       vertical="center")
                            # assign
                            cell.alignment = wrap_alignment
                            cell.fill = PatternFill("solid", fgColor="ffff00")
                    task_counter+=1

                # Adición de funcionalidad para las vacaciones y licencias (Atención)
                # print(len(ea_dates))
                activity_counter = 0
                for activity_date_block in activities_dates_blocks:
                    for d in activity_date_block:
                        if (dates[i] <= d < dates[i + 1]):
                            # cell = ws.cell(row=r +10 + i, column=4+task_initial_d.weekday(), value="{0} \n Proyecto: {1}".format(task_initial_d, task.project))
                            cell = ws.cell(row=r + 10 + i, column=4 + d.weekday(),
                                           value=activities[activity_counter].description)
                            cell.font = Font(bold=True)
                            cell.border = thin_border
                            cell.alignment = Alignment(horizontal='center')
                            # cell.style.alignment.wrap_text = True #Para autoajustar el tamaño de la celca. Revisar su correcto funcionamiento
                            # create alignment style
                            wrap_alignment = Alignment(wrap_text=True, horizontal="center",
                                                       vertical="center")
                            # assign
                            cell.alignment = wrap_alignment
                            if activities[activity_counter].id == 1:
                                cell.fill = PatternFill("solid", fgColor="E60404")
                            elif activities[activity_counter].id == 2:
                                cell.fill = PatternFill("solid", fgColor="810777")
                        elif (dates[len(dates) - 1] <= d):
                            # cell = ws.cell(row=r +10 + i, column=4+task_initial_d.weekday(), value="{0} \n Proyecto: {1}".format(task_initial_d, task.project))
                            cell = ws.cell(row=r + 10 + i + 1, column=4 + d.weekday(),
                                           value=activities[activity_counter].description)
                            cell.font = Font(bold=True)
                            cell.border = thin_border
                            cell.alignment = Alignment(horizontal='center')
                            # cell.style.alignment.wrap_text = True #Para autoajustar el tamaño de la celca. Revisar su correcto funcionamiento
                            # create alignment style
                            wrap_alignment = Alignment(wrap_text=True, horizontal="center",
                                                       vertical="center")
                            # assign
                            cell.alignment = wrap_alignment
                            if activities[activity_counter].id == 1:
                                cell.fill = PatternFill("solid", fgColor="E60404")
                            elif activities[activity_counter].id == 2:
                                cell.fill = PatternFill("solid", fgColor="810777")

                    activity_counter += 1
                # Adición de funcionalidad para las vacaciones y licencias (Atención)

            # Imprimir los días feriados
            all_the_days_on_display = []
            for i in range(0,span_of_weeks*7):
                all_the_days_on_display.append(initial_week_date+timedelta(days=i))
            for i in range(0, len(dates)-1):
                for d in all_the_days_on_display:
                    if (dates[i] <= d < dates[i + 1]):
                        if(isHoliday(d) and d.weekday() < 5):
                            cell = ws.cell(row=r + 10 + i, column=4 + d.weekday(),value="Feriado")
                            cell.font = Font(bold=True)
                            cell.border = thin_border
                            cell.alignment = Alignment(horizontal='center')
                            wrap_alignment = Alignment(wrap_text=True, horizontal="center",vertical="center")
                            cell.alignment = wrap_alignment
                            cell.fill = PatternFill("solid", fgColor="00ffff")
                        elif(ws.cell(row=r + 10 + i, column=4 + d.weekday()).value==None and d.weekday() < 5):
                            ws.cell(row=r + 10 + i, column=4 + d.weekday()).value = "Disponible"
                            ws.cell(row=r + 10 + i, column=4 + d.weekday()).fill = PatternFill("solid", fgColor="00ff00")
                    elif(dates[len(dates)-1] <= d):
                        if (isHoliday(d) and d.weekday() < 5):
                            cell = ws.cell(row=r + 10 + i + 1, column=4 + d.weekday(), value="Feriado")
                            cell.font = Font(bold=True)
                            cell.border = thin_border
                            cell.alignment = Alignment(horizontal='center')
                            wrap_alignment = Alignment(wrap_text=True, horizontal="center",
                                                       vertical="center")
                            cell.alignment = wrap_alignment
                            cell.fill = PatternFill("solid", fgColor="00ffff")
                        elif (ws.cell(row=r + 10 + i + 1, column=4 + d.weekday()).value == None and d.weekday() < 5):
                            ws.cell(row=r + 10 + i + 1, column=4 + d.weekday()).value = "Disponible"
                            ws.cell(row=r + 10 + i + 1, column=4 + d.weekday()).fill = PatternFill("solid", fgColor="00ff00")

        sheet_index+=1
    try:
        module_path = os.path.dirname(__file__)
        panorama_folder_path = os.path.abspath(os.path.join(module_path, os.pardir))
        report_folder_path = os.path.join(panorama_folder_path,"Reportes")
        if not os.path.exists(report_folder_path):
            os.makedirs(report_folder_path)
        report_file_name = "Reporte empleados {0} {1}.xlsx".format("Rectificación",date.today())
        fn = os.path.join(report_folder_path,report_file_name)
        wb.save(fn)
    except OSError as e:
        if e.args[0] != 13:
            raise
        input("\n Ha ocurrido un error porque el archivo Reporte empleados {0} {1}.xlsx está abierto. Por favor ciérrelo y presione cualquier tecla para que el programa pueda continuar.".format("Rectificación",date.today()))

def createDesignersReport(db):
    ''' Este método crea un informe en Excel compacto con una proción de la información de la base de datos. '''
    id_skill = 2
    with db_session:
        employees_skills = select(es for es in db.Employees_Skills if es.skill.id == id_skill)
        employees = []
        for es in employees_skills:
            employee = db.Employees.get(id=es.employee.id)
            employees.append(employee)
        sheet_index = 0
        wb = Workbook()
        by_default_sheet = wb.get_sheet_by_name('Sheet')
        by_default_sheet.title = 'Introducción del informe'
        wb.remove_sheet(by_default_sheet)
        employee_row_counter = 0
        ws = wb.create_sheet("Informe trabajadores",index=sheet_index)
        sheet_index+=1
        for e in employees:
            e_activities = select(ea for ea in db.Employees_Activities if
                                  ea.employee == e)  # Para ver lo de las vacaciones
            from Planning.features import isHoliday, \
                isNotWorkday  # Es realmente necesario importar isHoliday
            tareas = select(et for et in db.Employees_Tasks if et.employee == e)

            widths = {"A": 20, "B": 20, "C": 20, "D": 20, "E": 20, "F": 20,
                      "G": 20, "H": 20, "I": 20, "J": 20, "K": 20, "L": 20,
                      "M": 20, "N": 20, "O": 20, "P": 20, "Q": 20, "R": 20,
                      "S": 20, "T": 20, "U": 20, "V": 20, "W": 20, "X": 20,
                      "Y": 20, "Z": 20, "AA": 20, "AB": 20, "AC": 20, "AD": 20,
                      "AE": 20, "AF": 20, "AG": 20, "AH": 20, "AI": 20, "AJ": 20,
                      "AK": 20, "AL": 20, "AM": 20, "AN": 20, "AO": 20, "AP": 20,
                      "AQ": 20, "AR": 20, "AS": 20, "AT": 20, "AU": 20, "AV": 20,
                      "AW": 20, "AX": 20, "AY": 20, "AZ": 20, "BA": 20, "BB": 20,
                      "BC": 20, "BD": 20, "BE": 20, "BF": 20, "BG": 20, "BH": 20,
                      "BI": 20, "BJ": 20, "BK": 20, "BL": 20, "BM": 20, "BN": 20,
                      "BO": 20, "BP": 20, "BQ": 20, "BR": 20, "BS": 20, "BT": 20,
                      "BU": 20, "BV": 20, "BW": 20, "BX": 20, "BY": 20, "BZ": 20,
                      "CA": 20, "CB": 20, "CC": 20, "CD": 20, "CE": 20, "CF": 20,
                      "CG": 20, "CH": 20, "CI": 20, "CJ": 20, "CK": 20, "CL": 20,
                      "CM": 20, "CN": 20, "CO": 20, "CP": 20, "CQ": 20, "CR": 20,
                      "CS": 20, "CT": 20, "CU": 20, "CV": 20, "CW": 20, "CX": 20,
                      "CY": 20, "CZ": 20, "DA": 20, "DB": 20, "DC": 20, "DD": 20,
                      "DE": 20, "DF": 20, "DG": 20, "DH": 20, "DI": 20, "DJ": 20,
                      "DK": 20, "DL": 20, "DM": 20, "DN": 20, "DO": 20, "DP": 20,
                      "DQ": 20, "DR": 20, "DS": 20, "DT": 20, "DU": 20, "DV": 20,
                      "DW": 20, "DX": 20, "DY": 20, "DZ": 20}

            heights = {"A": 10, "B": 10, "C": 10, "D": 10, "E": 10, "F": 10,
                       "G": 10, "H": 10, "I": 10, "J": 10, "K": 10, "L": 10,
                       "M": 10, "N": 10, "O": 10, "P": 10, "Q": 10, "R": 10,
                       "S": 10, "T": 10, "U": 10, "V": 10, "W": 10, "X": 10,
                       "Y": 10, "Z": 10, "AA": 10, "AB": 10, "AC": 10, "AD": 10,
                       "AE": 10, "AF": 10, "AG": 10, "AH": 10, "AI": 10, "AJ": 10,
                       "AK": 10, "AL": 10, "AM": 10, "AN": 10, "AO": 10, "AP": 10,
                       "AQ": 10, "AR": 10, "AS": 10, "AT": 10, "AU": 10, "AV": 10,
                       "AW": 10, "AX": 10, "AY": 10, "AZ": 10, "BA": 10, "BB": 10,
                       "BC": 10, "BD": 10, "BE": 10, "BF": 10, "BG": 10, "BH": 10,
                       "BI": 10, "BJ": 10, "BK": 10, "BL": 10, "BM": 10, "BN": 20,
                       "BO": 10, "BP": 10, "BQ": 10, "BR": 10, "BS": 10, "BT": 10,
                       "BU": 10, "BV": 10, "BW": 10, "BX": 10, "BY": 10, "BZ": 10,
                       "CA": 10, "CB": 10, "CC": 10, "CD": 10, "CE": 10, "CF": 10,
                       "CG": 10, "CH": 10, "CI": 10, "CJ": 10, "CK": 10, "CL": 10,
                       "CM": 10, "CN": 10, "CO": 10, "CP": 10, "CQ": 10, "CR": 10,
                       "CS": 10, "CT": 10, "CU": 10, "CV": 10, "CW": 10, "CX": 10,
                       "CY": 10, "CZ": 10, "DA": 10, "DB": 10, "DC": 10, "DD": 10,
                       "DE": 10, "DF": 10, "DG": 10, "DH": 10, "DI": 10, "DJ": 10,
                       "DK": 10, "DL": 10, "DM": 10, "DN": 10, "DO": 10, "DP": 10,
                       "DQ": 10, "DR": 10, "DS": 10, "DT": 10, "DU": 10, "DV": 10,
                       "DW": 10, "DX": 10, "DY": 10, "DZ": 10}

            thin_border = Border(left=Side(style='thin'),
                                 right=Side(style='thin'),
                                 top=Side(style='thin'),
                                 bottom=Side(style='thin'))

            columns = ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M", "N",
                       "O", "P", "Q", "R", "S", "T", "U", "V", "W", "X", "Y", "Z", "AA", "AB",
                       "AC", "AD", "AE", "AF", "AG", "AH", "AI", "AJ", "AK", "AL", "AM", "AN",
                       "AO", "AP", "AQ", "AR", "AS", "AT", "AU", "AV", "AW", "AX", "AY", "AZ",
                       "BA", "BB", "BC", "BD", "BE", "BF", "BG", "BH", "BI", "BJ", "BK", "BL", "BM",
                       "BN", "BO", "BP", "BQ", "BR", "BS", "BT", "BU", "BV", "BW", "BX", "BY", "BZ",
                       "CA", "CB", "CC", "CD", "CE", "CF", "CG", "CH", "CI", "CJ", "CK", "CL", "CM",
                       "CN", "CO", "CP", "CQ", "CR", "CS", "CT", "CU", "CV", "CW", "CX", "CY", "CZ",
                       "DA", "DB", "DC", "DD", "DE", "DF", "DG", "DH", "DI", "DJ", "DK", "DL", "DM",
                       "DN", "DO", "DP", "DQ", "DR", "DS", "DT", "DU", "DV", "DW", "DX", "DY", "DZ"]
            for c in columns:
                ws.column_dimensions[c].width = widths[c]
                ws.column_dimensions[c].height = heights[c]

            thin_border = Border(left=Side(style='thin'),
                                 right=Side(style='thin'),
                                 top=Side(style='thin'),
                                 bottom=Side(style='thin'))

            if (
                    False):  # (es.skill.id == 4) <-- introducir esta condición si es que se quiere mostrar la información de "senior"
                # columns = ["A", "B", "C","D","E","F","G","H","I"] #Para un instalador y con el rendimiento
                columns = ["A", "B", "C", "D", "E", "F", "G",
                           "H"]  # Para un instalador sin el rendimiento
                # texts = ["","","","ID", "NOMBRE", "ZONA", "COMPETENCIA",
                #          "SENIOR","RENDIMIENTO"] #Para un instalador y con el rendimiento
                texts = ["", "", "", "ID", "NOMBRE", "ZONA", "COMPETENCIA",
                         "SENIOR"]  # Para un instalador sin el rendimiento
            else:
                columns = ["A", "B", "C", "D", "E", "F", "G"]
                texts = ["", "", "", "ID", "NOMBRE", "ZONA", "COMPETENCIA"]

            aux_columns = ["A", "B", "C", "D", "E", "F", "G",
                           "H"]  # Solo para mantener el ancho de las celdas en el caso que no sea un instalador

            for c in aux_columns:
                ws.column_dimensions[c].width = widths[c]
                ws.column_dimensions[c].height = heights[c]

            num_columns, letter_columns = zip(*list(enumerate(columns)))
            num_columns = list(num_columns)
            num_columns = [x + 1 for x in
                           num_columns]  # Desplazamos los valores en 1 porque Excel está indexado desde el 1 y no desde el 0
            letter_columns = list(letter_columns)

            # Escribir la fecha en la que fue producida el reporte:
            cell = ws.cell(row=6, column=3, value="Reporte producido el: ")
            cell.font = Font(bold=True, )
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='left')

            cell = ws.cell(row=6, column=4, value=date.today())
            cell.font = Font(bold=True, )
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='left')

            # A continuación llenamos con los datos


            r = 4
            r += employee_row_counter
            # Escribe el ID del empleado
            if e.id != None:
                cell = ws.cell(row=r + 10, column=3,
                               value="id: {0}\nNombre:{1}".format(e.id, e.name))
                cell.font = Font(bold=True)
                cell.border = thin_border
                wrap_alignment = Alignment(wrap_text=True, horizontal="center",
                                           vertical="center")
                # assign
                cell.alignment = wrap_alignment

            # Preparar el formato del horario/calendario del trabajador
            cell = ws.cell(row=4 + 9, column=3, value="EMPLEADO")
            cell.font = Font(bold=True)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')

            span_of_weeks = 15
            week_counter = 0
            aux_c = 0
            for week in range(0, span_of_weeks):
                for day in ["Lunes", "Martes", "Miercoles", "Jueves", "Viernes", "Sábado",
                            "Domingo"]:
                    cell = ws.cell(row=4 + 9, column=4 + aux_c, value="{0} {1}".format(day,
                                                                                       date.today() + timedelta(
                                                                                           days=aux_c - date.today().weekday())))
                    cell.font = Font(bold=True)
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal='center')
                    aux_c += 1
                week_counter += 1

            dates = []
            today = date.today()
            for i in range(0, span_of_weeks * 7):
                # dates.append(today + i * timedelta(days=1))
                dates.append(date.today() + timedelta(days=i - date.today().weekday()))
            # Imprimir las tareas del trabajador



            # Se definen las fechas de trabajo del trabajador

            tasks = []
            tasks_dates_blocks = []
            for tarea in tareas:
                working_dates = []
                task_initial_d = tarea.planned_initial_date
                task_final_d = tarea.planned_end_date
                task = db.Tasks.get(id=tarea.task.id)
                tasks.append(task)
                # Ajustar para que solo tome fechas de trabajo en días hábiles
                working_days = (task_final_d - task_initial_d).days
                working_days_span = working_days

                # working_dates = [task_initial_d + i*timedelta(days=1) for i in range(0,working_days)]
                working_days_counter = 0
                while (
                    working_days_counter < working_days_span + 1):  # Ojo con el +1 que es para tomar el intervalo cerrado de la fecha de término
                    if (
                    not isNotWorkday(task_initial_d + working_days_counter * timedelta(days=1))):
                        working_dates.append(
                            task_initial_d + working_days_counter * timedelta(days=1))
                    else:
                        # working_days_span+=1 #Ya está considerado en el Planning
                        pass
                    working_days_counter += 1
                tasks_dates_blocks.append(working_dates)

            # # ########Adición de funcionalidad para las vacaciones y licencias (Atención)
            activities = []
            activities_dates_blocks = []
            for e_act in e_activities:
                ea_dates = []
                ea_initial_d = e_act.initial_date
                ea_final_d = e_act.end_date
                activity = db.Activities.get(id=e_act.activity.id)
                activities.append(activity)
                # Ajustar para que solo tome fechas de trabajo en días hábiles
                ea_days = (ea_final_d - ea_initial_d).days
                ea_days_span = ea_days

                # working_dates = [task_initial_d + i*timedelta(days=1) for i in range(0,working_days)]
                ea_days_counter = 0
                while (
                            ea_days_counter < ea_days_span + 1):  # Ojo con el +1 que es para tomar el intervalo cerrado de la fecha de término
                    if (not isNotWorkday(ea_initial_d + ea_days_counter * timedelta(days=1))):
                        ea_dates.append(ea_initial_d + ea_days_counter * timedelta(days=1))
                    else:
                        # working_days_span+=1 #Ya está considerado en el Planning
                        pass
                    ea_days_counter += 1
                    # # #########Adición de funcionalidad para las vacaciones y licencias (Atención)
                activities_dates_blocks.append(ea_dates)
            # print(ea_dates)

            # Se imprimen los días de trabajo
            for i in range(0, len(dates)):
                task_counter = 0
                for tasks_dates_block in tasks_dates_blocks:
                    for d in tasks_dates_block:
                        if d == dates[i]:
                            cell = ws.cell(row=r + 10, column=4 + i,
                                           value="Proyecto {0}\n{1}".format(
                                               tasks[task_counter].project,
                                               tasks[task_counter].project.client_address))
                            cell.font = Font(bold=True)
                            cell.border = thin_border
                            cell.alignment = Alignment(horizontal='center')
                            # cell.style.alignment.wrap_text = True #Para autoajustar el tamaño de la celca. Revisar su correcto funcionamiento
                            # create alignment style
                            wrap_alignment = Alignment(wrap_text=True, horizontal="center",
                                                       vertical="center")
                            # assign
                            cell.alignment = wrap_alignment
                            cell.fill = PatternFill("solid", fgColor="ffff00")
                    task_counter = 0
                activity_counter = 0
                for activity_date_block in activities_dates_blocks:
                    for d in activity_date_block:
                        if d == dates[i]:
                            cell = ws.cell(row=r + 10, column=4 + i,
                                           value=activities[activity_counter].description)
                            cell.font = Font(bold=True)
                            cell.border = thin_border
                            wrap_alignment = Alignment(wrap_text=True, horizontal="center",
                                                       vertical="center")
                            cell.alignment = wrap_alignment
                            # cell.style.alignment.wrap_text = True #Para autoajustar el tamaño de la celca. Revisar su correcto funcionamiento
                            # create alignment style
                            wrap_alignment = Alignment(wrap_text=True, horizontal="center",
                                                       vertical="center")
                            if activities[activity_counter].id == 1:
                                cell.fill = PatternFill("solid", fgColor="E60404")
                            elif activities[activity_counter].id == 2:
                                cell.fill = PatternFill("solid", fgColor="810777")
                    activity_counter = 0
                if isHoliday(dates[i]):
                    cell = ws.cell(row=r + 10, column=4 + i, value="Feriado")
                    cell.font = Font(bold=True)
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal='center')
                    wrap_alignment = Alignment(wrap_text=True, horizontal="center",
                                               vertical="center")
                    cell.alignment = wrap_alignment
                    cell.fill = PatternFill("solid", fgColor="00ffff")
                if dates[i].weekday() >= 5:
                    cell = ws.cell(row=r + 10, column=4 + i, value="Fin de semana")
                    cell.font = Font(bold=True)
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal='center')
                    wrap_alignment = Alignment(wrap_text=True, horizontal="center",
                                               vertical="center")
                    cell.alignment = wrap_alignment
                    cell.fill = PatternFill("solid", fgColor="00ffff")
                if (ws.cell(row=r + 10, column=4 + i).value == None and dates[
                    i].weekday() < 5):  # ws.cell(row=r + 10, column=4 + i).value == None and
                    cell = ws.cell(row=r + 10, column=4 + i, value="Disponible")
                    cell.font = Font(bold=True)
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal='center')
                    wrap_alignment = Alignment(wrap_text=True, horizontal="center",
                                               vertical="center")
                    cell.alignment = wrap_alignment
                    cell.fill = PatternFill("solid", fgColor="00ff00")

            employee_row_counter += 1
        for e in employees:
            e_activities = select(ea for ea in db.Employees_Activities if ea.employee == e) #Para ver lo de las vacaciones
            from Planning.features import isHoliday, isNotWorkday  # Es realmente necesario importar isHoliday
            tareas = select(et for et in db.Employees_Tasks if et.employee == e)
            ws = wb.create_sheet(
                "Informe trabajador {}".format(e.id),index=sheet_index)

            widths = {"A": 20, "B": 20, "C": 20,"D": 20, "E": 20, "F": 20,
                      "G": 20, "H": 20, "I": 20,"J": 20, "K": 20, "L": 20,
                      "M": 20, "N": 20}

            heights = {"A": 10, "B": 10, "C": 10, "D": 10, "E": 10, "F": 10,
                      "G": 10, "H": 10, "I": 10, "J": 10, "K": 10, "L": 10,
                      "M": 10, "N": 10}

            thin_border = Border(left=Side(style='thin'),
                                 right=Side(style='thin'),
                                 top=Side(style='thin'),
                                 bottom=Side(style='thin'))

            if (False): #(es.skill.id == 4) <-- introducir esta condición si es que se quiere mostrar la información de "senior"
                # columns = ["A", "B", "C","D","E","F","G","H","I"] #Para un instalador y con el rendimiento
                columns = ["A", "B", "C", "D", "E", "F", "G", "H"]  # Para un instalador sin el rendimiento
                # texts = ["","","","ID", "NOMBRE", "ZONA", "COMPETENCIA",
                #          "SENIOR","RENDIMIENTO"] #Para un instalador y con el rendimiento
                texts = ["","","","ID", "NOMBRE", "ZONA", "COMPETENCIA",
                         "SENIOR"] #Para un instalador sin el rendimiento
            else:
                columns = ["A", "B", "C", "D", "E", "F", "G"]
                texts = ["", "", "", "ID", "NOMBRE", "ZONA", "COMPETENCIA"]

            aux_columns = ["A", "B", "C", "D", "E", "F", "G", "H"] #Solo para mantener el ancho de las celdas en el caso que no sea un instalador

            for c in aux_columns:
                ws.column_dimensions[c].width = widths[c]
                ws.column_dimensions[c].height = heights[c]


            num_columns, letter_columns = zip(*list(enumerate(columns)))
            num_columns = list(num_columns)
            num_columns = [x + 1 for x in num_columns] #Desplazamos los valores en 1 porque Excel está indexado desde el 1 y no desde el 0
            letter_columns = list(letter_columns)

            # escribir los títulos de las columnas, en negrita

            for i in range(4,len(num_columns)+1):
                cell = ws.cell(row=i, column=3, value=texts[i-1])
                cell.font = Font(bold=True,)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='left')


            #Escribir la fecha en la que fue producida el reporte:
            cell = ws.cell(row=3, column=3, value="Reporte producido el: ")
            cell.font = Font(bold=True, )
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='left')

            cell = ws.cell(row=3, column=4, value=date.today())
            cell.font = Font(bold=True, )
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='left')

            # A continuación llenamos con los datos


            r=4
            #Escribe el ID del empleado
            if e.id != None:
                cell = ws.cell(row=r, column=4, value=e.id)
                cell.font = Font(bold=True)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='left')
            else:
                cell = ws.cell(row=r, column=4, value="Dato no disponible")
                cell.font = Font(bold=True)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='left')

            # Escribe el nombre del empleado
            if e.name != None:
                cell = ws.cell(row=r+1, column=4, value=e.name)
                cell.font = Font(bold=True)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='left')
            else:
                cell = ws.cell(row=r+1, column=4, value="Dato no disponible")
                cell.font = Font(bold=True)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='left')

            # Escribe la zona geográfica del empleado
            if e.zone != None:
                cell = ws.cell(row=r+2, column=4, value=e.zone)
                cell.font = Font(bold=True)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='left')
            else:
                cell = ws.cell(row=r+2, column=4, value="Dato no disponible")
                cell.font = Font(bold=True)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='left')

            # Escribe la competencia del empleado (por el momento se asume que solo es una)

            if es.skill != None:
                skill_name = es.skill.name
                cell = ws.cell(row=r+3, column=4, value=skill_name)
                cell.font = Font(bold=True)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='left')
            else:
                cell = ws.cell(row=r+3, column=4, value="Dato no disponible")
                cell.font = Font(bold=True)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='left')

            # Escribe si el empleado es senior o no (Sí o No)
            # if(es.skill.id == 4):
            #     if e.senior != None:
            #         if e.senior:
            #             value = "Sí"
            #         else:
            #             value = "No"
            #         cell = ws.cell(row=r+4, column=4, value = value)
            #         cell.font = Font(bold=True)
            #         cell.border = thin_border
            #         cell.alignment = Alignment(horizontal='left')
            #     else:
            #         cell = ws.cell(row=r+4, column=4, value="Dato no disponible")
            #         cell.font = Font(bold=True)
            #         cell.border = thin_border
            #         cell.alignment = Alignment(horizontal='left')

            # # Escribe el rendimiento del trabajador
            # if es.performance != None:
            #     cell = ws.cell(row=r+5, column=4, value = es.performance)
            #     cell.font = Font(bold=True)
            #     cell.border = thin_border
            #     cell.alignment = Alignment(horizontal='left')
            # else:
            #     cell = ws.cell(row=r+5, column=4, value="Dato no disponible")
            #     cell.font = Font(bold=True)
            #     cell.border = thin_border
            #     cell.alignment = Alignment(horizontal='left')

            #Preparar el formato del horario/calendario del trabajador
            cell = ws.cell(row=r+9, column=3, value="SEMANA")
            cell.font = Font(bold=True)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')
            aux_c = 1
            for day in ["Lunes","Martes", "Miercoles","Jueves","Viernes"]:
                cell = ws.cell(row=r + 9, column=3 + aux_c, value="{0}".format(day))
                cell.font = Font(bold=True)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center')
                aux_c+=1
            span_of_weeks = 14
            dates = []
            today = date.today()
            initial_week_date = today + timedelta(days=-today.weekday())
            calendar_r = r + 10 #Esta variable para la fila podría definirse se otra forma
            for i in range (0,span_of_weeks):
                dates.append(initial_week_date+i*timedelta(days=7))
            for d in dates:
                cell = ws.cell(row=calendar_r, column=3, value=d)
                cell.font = Font(bold=True)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center')
                calendar_r+=1

            for f in [i for i in range(14, 14+len(dates))]:#Revisar los limites de indexación
                ws.row_dimensions[f].height = 40
                for c in [i for i in range(4,9)]:
                    cell = ws.cell(row=f, column=c)
                    cell.font = Font(bold=True)
                    cell.border = thin_border
                    # cell.alignment = Alignment(horizontal='center')
                    wrap_alignment = Alignment(wrap_text=True, horizontal="center",vertical="center")
                    cell.alignment = wrap_alignment



            #Imprimir las tareas del trabajador



            #Se definen las fechas de trabajo del trabajador

            tasks = []
            tasks_dates_blocks = []
            for tarea in tareas:
                working_dates = []
                task_initial_d = tarea.planned_initial_date
                task_final_d = tarea.planned_end_date
                task = db.Tasks.get(id=tarea.task.id)
                tasks.append(task)
                # Ajustar para que solo tome fechas de trabajo en días hábiles
                working_days = (task_final_d-task_initial_d).days
                working_days_span = working_days

                # working_dates = [task_initial_d + i*timedelta(days=1) for i in range(0,working_days)]
                working_days_counter = 0
                while(working_days_counter<working_days_span+1): #Ojo con el +1 que es para tomar el intervalo cerrado de la fecha de término
                    if(not isNotWorkday(task_initial_d + working_days_counter*timedelta(days=1))):
                        working_dates.append(task_initial_d + working_days_counter*timedelta(days=1))
                    else:
                        # working_days_span+=1 #Ya está considerado en el Planning
                        pass
                    working_days_counter+=1
                tasks_dates_blocks.append(working_dates)

            # # ########Adición de funcionalidad para las vacaciones y licencias (Atención)
            activities = []
            activities_dates_blocks = []
            for e_act in e_activities:
                ea_dates = []
                ea_initial_d = e_act.initial_date
                ea_final_d = e_act.end_date
                activity = db.Activities.get(id=e_act.activity.id)
                activities.append(activity)
                # Ajustar para que solo tome fechas de trabajo en días hábiles
                ea_days = (ea_final_d - ea_initial_d).days
                ea_days_span = ea_days

                # working_dates = [task_initial_d + i*timedelta(days=1) for i in range(0,working_days)]
                ea_days_counter = 0
                while (ea_days_counter < ea_days_span + 1):  # Ojo con el +1 que es para tomar el intervalo cerrado de la fecha de término
                    if (not isNotWorkday(ea_initial_d + ea_days_counter * timedelta(days=1))):
                        ea_dates.append(ea_initial_d + ea_days_counter * timedelta(days=1))
                    else:
                        # working_days_span+=1 #Ya está considerado en el Planning
                        pass
                    ea_days_counter += 1
                    # # #########Adición de funcionalidad para las vacaciones y licencias (Atención)
                activities_dates_blocks.append(ea_dates)
            # print(ea_dates)

            #Se imprimen los días de trabajo
            for i in range(0,len(dates)-1):
                task_counter = 0
                for task_date_block in tasks_dates_blocks:
                    for d in task_date_block:
                        if (dates[i] <= d < dates[i + 1]):
                            # cell = ws.cell(row=r +10 + i, column=4+task_initial_d.weekday(), value="{0} \n Proyecto: {1}".format(task_initial_d, task.project))
                            cell = ws.cell(row=r + 10 + i, column=4 + d.weekday(),
                                           value="Proyecto {0}\n{1}".format(tasks[task_counter].project,tasks[task_counter].project.client_address))
                            cell.font = Font(bold=True)
                            cell.border = thin_border
                            cell.alignment = Alignment(horizontal='center')
                            # cell.style.alignment.wrap_text = True #Para autoajustar el tamaño de la celca. Revisar su correcto funcionamiento
                            # create alignment style
                            wrap_alignment = Alignment(wrap_text=True, horizontal="center",
                                                       vertical="center")
                            # assign
                            cell.alignment = wrap_alignment
                            cell.fill = PatternFill("solid", fgColor="ffff00")
                        elif (dates[len(dates) - 1] <= d):
                            # cell = ws.cell(row=r +10 + i, column=4+task_initial_d.weekday(), value="{0} \n Proyecto: {1}".format(task_initial_d, task.project))
                            cell = ws.cell(row=r + 10 + i + 1, column=4 + d.weekday(),
                                           value="Proyecto {0}\n{1}".format(tasks[task_counter].project,
                                                                            tasks[task_counter].project.client_address))
                            cell.font = Font(bold=True)
                            cell.border = thin_border
                            cell.alignment = Alignment(horizontal='center')
                            # cell.style.alignment.wrap_text = True #Para autoajustar el tamaño de la celca. Revisar su correcto funcionamiento
                            # create alignment style
                            wrap_alignment = Alignment(wrap_text=True, horizontal="center",
                                                       vertical="center")
                            # assign
                            cell.alignment = wrap_alignment
                            cell.fill = PatternFill("solid", fgColor="ffff00")
                    task_counter+=1

                # Adición de funcionalidad para las vacaciones y licencias (Atención)
                # print(len(ea_dates))
                activity_counter = 0
                for activity_date_block in activities_dates_blocks:
                    for d in activity_date_block:
                        if (dates[i] <= d < dates[i + 1]):
                            # cell = ws.cell(row=r +10 + i, column=4+task_initial_d.weekday(), value="{0} \n Proyecto: {1}".format(task_initial_d, task.project))
                            cell = ws.cell(row=r + 10 + i, column=4 + d.weekday(),
                                           value=activities[activity_counter].description)
                            cell.font = Font(bold=True)
                            cell.border = thin_border
                            cell.alignment = Alignment(horizontal='center')
                            # cell.style.alignment.wrap_text = True #Para autoajustar el tamaño de la celca. Revisar su correcto funcionamiento
                            # create alignment style
                            wrap_alignment = Alignment(wrap_text=True, horizontal="center",
                                                       vertical="center")
                            # assign
                            cell.alignment = wrap_alignment
                            if activities[activity_counter].id == 1:
                                cell.fill = PatternFill("solid", fgColor="E60404")
                            elif activities[activity_counter].id == 2:
                                cell.fill = PatternFill("solid", fgColor="810777")
                        elif (dates[len(dates) - 1] <= d):
                            # cell = ws.cell(row=r +10 + i, column=4+task_initial_d.weekday(), value="{0} \n Proyecto: {1}".format(task_initial_d, task.project))
                            cell = ws.cell(row=r + 10 + i + 1, column=4 + d.weekday(),
                                           value=activities[activity_counter].description)
                            cell.font = Font(bold=True)
                            cell.border = thin_border
                            cell.alignment = Alignment(horizontal='center')
                            # cell.style.alignment.wrap_text = True #Para autoajustar el tamaño de la celca. Revisar su correcto funcionamiento
                            # create alignment style
                            wrap_alignment = Alignment(wrap_text=True, horizontal="center",
                                                       vertical="center")
                            # assign
                            cell.alignment = wrap_alignment
                            if activities[activity_counter].id == 1:
                                cell.fill = PatternFill("solid", fgColor="E60404")
                            elif activities[activity_counter].id == 2:
                                cell.fill = PatternFill("solid", fgColor="810777")

                    activity_counter += 1
                # Adición de funcionalidad para las vacaciones y licencias (Atención)

            # Imprimir los días feriados
            all_the_days_on_display = []
            for i in range(0,span_of_weeks*7):
                all_the_days_on_display.append(initial_week_date+timedelta(days=i))
            for i in range(0, len(dates)-1):
                for d in all_the_days_on_display:
                    if (dates[i] <= d < dates[i + 1]):
                        if(isHoliday(d) and d.weekday() < 5):
                            cell = ws.cell(row=r + 10 + i, column=4 + d.weekday(),value="Feriado")
                            cell.font = Font(bold=True)
                            cell.border = thin_border
                            cell.alignment = Alignment(horizontal='center')
                            wrap_alignment = Alignment(wrap_text=True, horizontal="center",vertical="center")
                            cell.alignment = wrap_alignment
                            cell.fill = PatternFill("solid", fgColor="00ffff")
                        elif(ws.cell(row=r + 10 + i, column=4 + d.weekday()).value==None and d.weekday() < 5):
                            ws.cell(row=r + 10 + i, column=4 + d.weekday()).value = "Disponible"
                            ws.cell(row=r + 10 + i, column=4 + d.weekday()).fill = PatternFill("solid", fgColor="00ff00")
                    elif(dates[len(dates)-1] <= d):
                        if (isHoliday(d) and d.weekday() < 5):
                            cell = ws.cell(row=r + 10 + i + 1, column=4 + d.weekday(), value="Feriado")
                            cell.font = Font(bold=True)
                            cell.border = thin_border
                            cell.alignment = Alignment(horizontal='center')
                            wrap_alignment = Alignment(wrap_text=True, horizontal="center",
                                                       vertical="center")
                            cell.alignment = wrap_alignment
                            cell.fill = PatternFill("solid", fgColor="00ffff")
                        elif (ws.cell(row=r + 10 + i + 1, column=4 + d.weekday()).value == None and d.weekday() < 5):
                            ws.cell(row=r + 10 + i + 1, column=4 + d.weekday()).value = "Disponible"
                            ws.cell(row=r + 10 + i + 1, column=4 + d.weekday()).fill = PatternFill("solid", fgColor="00ff00")

        sheet_index+=1
    try:
        module_path = os.path.dirname(__file__)
        panorama_folder_path = os.path.abspath(os.path.join(module_path, os.pardir))
        report_folder_path = os.path.join(panorama_folder_path,"Reportes")
        if not os.path.exists(report_folder_path):
            os.makedirs(report_folder_path)
        report_file_name = "Reporte empleados {0} {1}.xlsx".format("Diseño",date.today())
        fn = os.path.join(report_folder_path,report_file_name)
        wb.save(fn)
    except OSError as e:
        if e.args[0] != 13:
            raise
        input("\n Ha ocurrido un error porque el archivo Reporte empleados {0} {1}.xlsx está abierto. Por favor ciérrelo y presione cualquier tecla para que el programa pueda continuar.".format("Diseño",date.today()))

def createFabricatorsReport(db):
    ''' Este método crea un informe en Excel compacto con una proción de la información de la base de datos. '''
    id_skill = 3
    with db_session:
        employees_skills = select(es for es in db.Employees_Skills if es.skill.id == id_skill)
        employees = []
        for es in employees_skills:
            employee = db.Employees.get(id=es.employee.id)
            employees.append(employee)
        sheet_index = 0
        wb = Workbook()
        by_default_sheet = wb.get_sheet_by_name('Sheet')
        by_default_sheet.title = 'Introducción del informe'
        wb.remove_sheet(by_default_sheet)
        employee_row_counter = 0
        ws = wb.create_sheet("Informe trabajadores",index = sheet_index)
        sheet_index+=1
        for e in employees:
            e_activities = select(ea for ea in db.Employees_Activities if
                                  ea.employee == e)  # Para ver lo de las vacaciones
            from Planning.features import isHoliday, \
                isNotWorkday  # Es realmente necesario importar isHoliday
            tareas = select(et for et in db.Employees_Tasks if et.employee == e)

            widths = {"A": 20, "B": 20, "C": 20, "D": 20, "E": 20, "F": 20,
                      "G": 20, "H": 20, "I": 20, "J": 20, "K": 20, "L": 20,
                      "M": 20, "N": 20, "O": 20, "P": 20, "Q": 20, "R": 20,
                      "S": 20, "T": 20, "U": 20, "V": 20, "W": 20, "X": 20,
                      "Y": 20, "Z": 20, "AA": 20, "AB": 20, "AC": 20, "AD": 20,
                      "AE": 20, "AF": 20, "AG": 20, "AH": 20, "AI": 20, "AJ": 20,
                      "AK": 20, "AL": 20, "AM": 20, "AN": 20, "AO": 20, "AP": 20,
                      "AQ": 20, "AR": 20, "AS": 20, "AT": 20, "AU": 20, "AV": 20,
                      "AW": 20, "AX": 20, "AY": 20, "AZ": 20, "BA": 20, "BB": 20,
                      "BC": 20, "BD": 20, "BE": 20, "BF": 20, "BG": 20, "BH": 20,
                      "BI": 20, "BJ": 20, "BK": 20, "BL": 20, "BM": 20, "BN": 20,
                      "BO": 20, "BP": 20, "BQ": 20, "BR": 20, "BS": 20, "BT": 20,
                      "BU": 20, "BV": 20, "BW": 20, "BX": 20, "BY": 20, "BZ": 20,
                      "CA": 20, "CB": 20, "CC": 20, "CD": 20, "CE": 20, "CF": 20,
                      "CG": 20, "CH": 20, "CI": 20, "CJ": 20, "CK": 20, "CL": 20,
                      "CM": 20, "CN": 20, "CO": 20, "CP": 20, "CQ": 20, "CR": 20,
                      "CS": 20, "CT": 20, "CU": 20, "CV": 20, "CW": 20, "CX": 20,
                      "CY": 20, "CZ": 20, "DA": 20, "DB": 20, "DC": 20, "DD": 20,
                      "DE": 20, "DF": 20, "DG": 20, "DH": 20, "DI": 20, "DJ": 20,
                      "DK": 20, "DL": 20, "DM": 20, "DN": 20, "DO": 20, "DP": 20,
                      "DQ": 20, "DR": 20, "DS": 20, "DT": 20, "DU": 20, "DV": 20,
                      "DW": 20, "DX": 20, "DY": 20, "DZ": 20}

            heights = {"A": 10, "B": 10, "C": 10, "D": 10, "E": 10, "F": 10,
                       "G": 10, "H": 10, "I": 10, "J": 10, "K": 10, "L": 10,
                       "M": 10, "N": 10, "O": 10, "P": 10, "Q": 10, "R": 10,
                       "S": 10, "T": 10, "U": 10, "V": 10, "W": 10, "X": 10,
                       "Y": 10, "Z": 10, "AA": 10, "AB": 10, "AC": 10, "AD": 10,
                       "AE": 10, "AF": 10, "AG": 10, "AH": 10, "AI": 10, "AJ": 10,
                       "AK": 10, "AL": 10, "AM": 10, "AN": 10, "AO": 10, "AP": 10,
                       "AQ": 10, "AR": 10, "AS": 10, "AT": 10, "AU": 10, "AV": 10,
                       "AW": 10, "AX": 10, "AY": 10, "AZ": 10, "BA": 10, "BB": 10,
                       "BC": 10, "BD": 10, "BE": 10, "BF": 10, "BG": 10, "BH": 10,
                       "BI": 10, "BJ": 10, "BK": 10, "BL": 10, "BM": 10, "BN": 20,
                       "BO": 10, "BP": 10, "BQ": 10, "BR": 10, "BS": 10, "BT": 10,
                       "BU": 10, "BV": 10, "BW": 10, "BX": 10, "BY": 10, "BZ": 10,
                       "CA": 10, "CB": 10, "CC": 10, "CD": 10, "CE": 10, "CF": 10,
                       "CG": 10, "CH": 10, "CI": 10, "CJ": 10, "CK": 10, "CL": 10,
                       "CM": 10, "CN": 10, "CO": 10, "CP": 10, "CQ": 10, "CR": 10,
                       "CS": 10, "CT": 10, "CU": 10, "CV": 10, "CW": 10, "CX": 10,
                       "CY": 10, "CZ": 10, "DA": 10, "DB": 10, "DC": 10, "DD": 10,
                       "DE": 10, "DF": 10, "DG": 10, "DH": 10, "DI": 10, "DJ": 10,
                       "DK": 10, "DL": 10, "DM": 10, "DN": 10, "DO": 10, "DP": 10,
                       "DQ": 10, "DR": 10, "DS": 10, "DT": 10, "DU": 10, "DV": 10,
                       "DW": 10, "DX": 10, "DY": 10, "DZ": 10}

            thin_border = Border(left=Side(style='thin'),
                                 right=Side(style='thin'),
                                 top=Side(style='thin'),
                                 bottom=Side(style='thin'))

            columns = ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M", "N",
                       "O", "P", "Q", "R", "S", "T", "U", "V", "W", "X", "Y", "Z", "AA", "AB",
                       "AC", "AD", "AE", "AF", "AG", "AH", "AI", "AJ", "AK", "AL", "AM", "AN",
                       "AO", "AP", "AQ", "AR", "AS", "AT", "AU", "AV", "AW", "AX", "AY", "AZ",
                       "BA", "BB", "BC", "BD", "BE", "BF", "BG", "BH", "BI", "BJ", "BK", "BL", "BM",
                       "BN", "BO", "BP", "BQ", "BR", "BS", "BT", "BU", "BV", "BW", "BX", "BY", "BZ",
                       "CA", "CB", "CC", "CD", "CE", "CF", "CG", "CH", "CI", "CJ", "CK", "CL", "CM",
                       "CN", "CO", "CP", "CQ", "CR", "CS", "CT", "CU", "CV", "CW", "CX", "CY", "CZ",
                       "DA", "DB", "DC", "DD", "DE", "DF", "DG", "DH", "DI", "DJ", "DK", "DL", "DM",
                       "DN", "DO", "DP", "DQ", "DR", "DS", "DT", "DU", "DV", "DW", "DX", "DY", "DZ"]
            for c in columns:
                ws.column_dimensions[c].width = widths[c]
                ws.column_dimensions[c].height = heights[c]

            thin_border = Border(left=Side(style='thin'),
                                 right=Side(style='thin'),
                                 top=Side(style='thin'),
                                 bottom=Side(style='thin'))

            if (
                    False):  # (es.skill.id == 4) <-- introducir esta condición si es que se quiere mostrar la información de "senior"
                # columns = ["A", "B", "C","D","E","F","G","H","I"] #Para un instalador y con el rendimiento
                columns = ["A", "B", "C", "D", "E", "F", "G",
                           "H"]  # Para un instalador sin el rendimiento
                # texts = ["","","","ID", "NOMBRE", "ZONA", "COMPETENCIA",
                #          "SENIOR","RENDIMIENTO"] #Para un instalador y con el rendimiento
                texts = ["", "", "", "ID", "NOMBRE", "ZONA", "COMPETENCIA",
                         "SENIOR"]  # Para un instalador sin el rendimiento
            else:
                columns = ["A", "B", "C", "D", "E", "F", "G"]
                texts = ["", "", "", "ID", "NOMBRE", "ZONA", "COMPETENCIA"]

            aux_columns = ["A", "B", "C", "D", "E", "F", "G",
                           "H"]  # Solo para mantener el ancho de las celdas en el caso que no sea un instalador

            for c in aux_columns:
                ws.column_dimensions[c].width = widths[c]
                ws.column_dimensions[c].height = heights[c]

            num_columns, letter_columns = zip(*list(enumerate(columns)))
            num_columns = list(num_columns)
            num_columns = [x + 1 for x in
                           num_columns]  # Desplazamos los valores en 1 porque Excel está indexado desde el 1 y no desde el 0
            letter_columns = list(letter_columns)

            # Escribir la fecha en la que fue producida el reporte:
            cell = ws.cell(row=6, column=3, value="Reporte producido el: ")
            cell.font = Font(bold=True, )
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='left')

            cell = ws.cell(row=6, column=4, value=date.today())
            cell.font = Font(bold=True, )
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='left')

            # A continuación llenamos con los datos


            r = 4
            r += employee_row_counter
            # Escribe el ID del empleado
            if e.id != None:
                cell = ws.cell(row=r + 10, column=3,
                               value="id: {0}\nNombre:{1}".format(e.id, e.name))
                cell.font = Font(bold=True)
                cell.border = thin_border
                wrap_alignment = Alignment(wrap_text=True, horizontal="center",
                                           vertical="center")
                # assign
                cell.alignment = wrap_alignment

            # Preparar el formato del horario/calendario del trabajador
            cell = ws.cell(row=4 + 9, column=3, value="EMPLEADO")
            cell.font = Font(bold=True)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')

            span_of_weeks = 15
            week_counter = 0
            aux_c = 0
            for week in range(0, span_of_weeks):
                for day in ["Lunes", "Martes", "Miercoles", "Jueves", "Viernes", "Sábado",
                            "Domingo"]:
                    cell = ws.cell(row=4 + 9, column=4 + aux_c, value="{0} {1}".format(day,
                                                                                       date.today() + timedelta(
                                                                                           days=aux_c - date.today().weekday())))
                    cell.font = Font(bold=True)
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal='center')
                    aux_c += 1
                week_counter += 1

            dates = []
            today = date.today()
            for i in range(0, span_of_weeks * 7):
                # dates.append(today + i * timedelta(days=1))
                dates.append(date.today() + timedelta(days=i - date.today().weekday()))
            # Imprimir las tareas del trabajador



            # Se definen las fechas de trabajo del trabajador

            tasks = []
            tasks_dates_blocks = []
            for tarea in tareas:
                working_dates = []
                task_initial_d = tarea.planned_initial_date
                task_final_d = tarea.planned_end_date
                task = db.Tasks.get(id=tarea.task.id)
                tasks.append(task)
                # Ajustar para que solo tome fechas de trabajo en días hábiles
                working_days = (task_final_d - task_initial_d).days
                working_days_span = working_days

                # working_dates = [task_initial_d + i*timedelta(days=1) for i in range(0,working_days)]
                working_days_counter = 0
                while (
                    working_days_counter < working_days_span + 1):  # Ojo con el +1 que es para tomar el intervalo cerrado de la fecha de término
                    if (
                    not isNotWorkday(task_initial_d + working_days_counter * timedelta(days=1))):
                        working_dates.append(
                            task_initial_d + working_days_counter * timedelta(days=1))
                    else:
                        # working_days_span+=1 #Ya está considerado en el Planning
                        pass
                    working_days_counter += 1
                tasks_dates_blocks.append(working_dates)

            # # ########Adición de funcionalidad para las vacaciones y licencias (Atención)
            activities = []
            activities_dates_blocks = []
            for e_act in e_activities:
                ea_dates = []
                ea_initial_d = e_act.initial_date
                ea_final_d = e_act.end_date
                activity = db.Activities.get(id=e_act.activity.id)
                activities.append(activity)
                # Ajustar para que solo tome fechas de trabajo en días hábiles
                ea_days = (ea_final_d - ea_initial_d).days
                ea_days_span = ea_days

                # working_dates = [task_initial_d + i*timedelta(days=1) for i in range(0,working_days)]
                ea_days_counter = 0
                while (
                            ea_days_counter < ea_days_span + 1):  # Ojo con el +1 que es para tomar el intervalo cerrado de la fecha de término
                    if (not isNotWorkday(ea_initial_d + ea_days_counter * timedelta(days=1))):
                        ea_dates.append(ea_initial_d + ea_days_counter * timedelta(days=1))
                    else:
                        # working_days_span+=1 #Ya está considerado en el Planning
                        pass
                    ea_days_counter += 1
                    # # #########Adición de funcionalidad para las vacaciones y licencias (Atención)
                activities_dates_blocks.append(ea_dates)
            # print(ea_dates)

            # Se imprimen los días de trabajo
            for i in range(0, len(dates)):
                task_counter = 0
                for tasks_dates_block in tasks_dates_blocks:
                    for d in tasks_dates_block:
                        if d == dates[i]:
                            cell = ws.cell(row=r + 10, column=4 + i,
                                           value="Proyecto {0}\n{1}".format(
                                               tasks[task_counter].project,
                                               tasks[task_counter].project.client_address))
                            cell.font = Font(bold=True)
                            cell.border = thin_border
                            cell.alignment = Alignment(horizontal='center')
                            # cell.style.alignment.wrap_text = True #Para autoajustar el tamaño de la celca. Revisar su correcto funcionamiento
                            # create alignment style
                            wrap_alignment = Alignment(wrap_text=True, horizontal="center",
                                                       vertical="center")
                            # assign
                            cell.alignment = wrap_alignment
                            cell.fill = PatternFill("solid", fgColor="ffff00")
                    task_counter = 0
                activity_counter = 0
                for activity_date_block in activities_dates_blocks:
                    for d in activity_date_block:
                        if d == dates[i]:
                            cell = ws.cell(row=r + 10, column=4 + i,
                                           value=activities[activity_counter].description)
                            cell.font = Font(bold=True)
                            cell.border = thin_border
                            wrap_alignment = Alignment(wrap_text=True, horizontal="center",
                                                       vertical="center")
                            cell.alignment = wrap_alignment
                            # cell.style.alignment.wrap_text = True #Para autoajustar el tamaño de la celca. Revisar su correcto funcionamiento
                            # create alignment style
                            wrap_alignment = Alignment(wrap_text=True, horizontal="center",
                                                       vertical="center")
                            if activities[activity_counter].id == 1:
                                cell.fill = PatternFill("solid", fgColor="E60404")
                            elif activities[activity_counter].id == 2:
                                cell.fill = PatternFill("solid", fgColor="810777")
                    activity_counter = 0
                if isHoliday(dates[i]):
                    cell = ws.cell(row=r + 10, column=4 + i, value="Feriado")
                    cell.font = Font(bold=True)
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal='center')
                    wrap_alignment = Alignment(wrap_text=True, horizontal="center",
                                               vertical="center")
                    cell.alignment = wrap_alignment
                    cell.fill = PatternFill("solid", fgColor="00ffff")
                if dates[i].weekday() >= 5:
                    cell = ws.cell(row=r + 10, column=4 + i, value="Fin de semana")
                    cell.font = Font(bold=True)
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal='center')
                    wrap_alignment = Alignment(wrap_text=True, horizontal="center",
                                               vertical="center")
                    cell.alignment = wrap_alignment
                    cell.fill = PatternFill("solid", fgColor="00ffff")
                if (ws.cell(row=r + 10, column=4 + i).value == None and dates[
                    i].weekday() < 5):  # ws.cell(row=r + 10, column=4 + i).value == None and
                    cell = ws.cell(row=r + 10, column=4 + i, value="Disponible")
                    cell.font = Font(bold=True)
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal='center')
                    wrap_alignment = Alignment(wrap_text=True, horizontal="center",
                                               vertical="center")
                    cell.alignment = wrap_alignment
                    cell.fill = PatternFill("solid", fgColor="00ff00")

            employee_row_counter += 1
        for e in employees:
            e_activities = select(ea for ea in db.Employees_Activities if ea.employee == e) #Para ver lo de las vacaciones
            from Planning.features import isHoliday, isNotWorkday  # Es realmente necesario importar isHoliday
            tareas = select(et for et in db.Employees_Tasks if et.employee == e)
            ws = wb.create_sheet(
                "Informe trabajador {}".format(e.id),index=sheet_index)

            widths = {"A": 20, "B": 20, "C": 20,"D": 20, "E": 20, "F": 20,
                      "G": 20, "H": 20, "I": 20,"J": 20, "K": 20, "L": 20,
                      "M": 20, "N": 20}

            heights = {"A": 10, "B": 10, "C": 10, "D": 10, "E": 10, "F": 10,
                      "G": 10, "H": 10, "I": 10, "J": 10, "K": 10, "L": 10,
                      "M": 10, "N": 10}

            thin_border = Border(left=Side(style='thin'),
                                 right=Side(style='thin'),
                                 top=Side(style='thin'),
                                 bottom=Side(style='thin'))

            if (False): #(es.skill.id == 4) <-- introducir esta condición si es que se quiere mostrar la información de "senior"
                # columns = ["A", "B", "C","D","E","F","G","H","I"] #Para un instalador y con el rendimiento
                columns = ["A", "B", "C", "D", "E", "F", "G", "H"]  # Para un instalador sin el rendimiento
                # texts = ["","","","ID", "NOMBRE", "ZONA", "COMPETENCIA",
                #          "SENIOR","RENDIMIENTO"] #Para un instalador y con el rendimiento
                texts = ["","","","ID", "NOMBRE", "ZONA", "COMPETENCIA",
                         "SENIOR"] #Para un instalador sin el rendimiento
            else:
                columns = ["A", "B", "C", "D", "E", "F", "G"]
                texts = ["", "", "", "ID", "NOMBRE", "ZONA", "COMPETENCIA"]

            aux_columns = ["A", "B", "C", "D", "E", "F", "G", "H"] #Solo para mantener el ancho de las celdas en el caso que no sea un instalador

            for c in aux_columns:
                ws.column_dimensions[c].width = widths[c]
                ws.column_dimensions[c].height = heights[c]


            num_columns, letter_columns = zip(*list(enumerate(columns)))
            num_columns = list(num_columns)
            num_columns = [x + 1 for x in num_columns] #Desplazamos los valores en 1 porque Excel está indexado desde el 1 y no desde el 0
            letter_columns = list(letter_columns)

            # escribir los títulos de las columnas, en negrita

            for i in range(4,len(num_columns)+1):
                cell = ws.cell(row=i, column=3, value=texts[i-1])
                cell.font = Font(bold=True,)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='left')


            #Escribir la fecha en la que fue producida el reporte:
            cell = ws.cell(row=3, column=3, value="Reporte producido el: ")
            cell.font = Font(bold=True, )
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='left')

            cell = ws.cell(row=3, column=4, value=date.today())
            cell.font = Font(bold=True, )
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='left')

            # A continuación llenamos con los datos


            r=4
            #Escribe el ID del empleado
            if e.id != None:
                cell = ws.cell(row=r, column=4, value=e.id)
                cell.font = Font(bold=True)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='left')
            else:
                cell = ws.cell(row=r, column=4, value="Dato no disponible")
                cell.font = Font(bold=True)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='left')

            # Escribe el nombre del empleado
            if e.name != None:
                cell = ws.cell(row=r+1, column=4, value=e.name)
                cell.font = Font(bold=True)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='left')
            else:
                cell = ws.cell(row=r+1, column=4, value="Dato no disponible")
                cell.font = Font(bold=True)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='left')

            # Escribe la zona geográfica del empleado
            if e.zone != None:
                cell = ws.cell(row=r+2, column=4, value=e.zone)
                cell.font = Font(bold=True)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='left')
            else:
                cell = ws.cell(row=r+2, column=4, value="Dato no disponible")
                cell.font = Font(bold=True)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='left')

            # Escribe la competencia del empleado (por el momento se asume que solo es una)

            if es.skill != None:
                skill_name = es.skill.name
                cell = ws.cell(row=r+3, column=4, value=skill_name)
                cell.font = Font(bold=True)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='left')
            else:
                cell = ws.cell(row=r+3, column=4, value="Dato no disponible")
                cell.font = Font(bold=True)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='left')

            # Escribe si el empleado es senior o no (Sí o No)
            # if(es.skill.id == 4):
            #     if e.senior != None:
            #         if e.senior:
            #             value = "Sí"
            #         else:
            #             value = "No"
            #         cell = ws.cell(row=r+4, column=4, value = value)
            #         cell.font = Font(bold=True)
            #         cell.border = thin_border
            #         cell.alignment = Alignment(horizontal='left')
            #     else:
            #         cell = ws.cell(row=r+4, column=4, value="Dato no disponible")
            #         cell.font = Font(bold=True)
            #         cell.border = thin_border
            #         cell.alignment = Alignment(horizontal='left')

            # # Escribe el rendimiento del trabajador
            # if es.performance != None:
            #     cell = ws.cell(row=r+5, column=4, value = es.performance)
            #     cell.font = Font(bold=True)
            #     cell.border = thin_border
            #     cell.alignment = Alignment(horizontal='left')
            # else:
            #     cell = ws.cell(row=r+5, column=4, value="Dato no disponible")
            #     cell.font = Font(bold=True)
            #     cell.border = thin_border
            #     cell.alignment = Alignment(horizontal='left')

            #Preparar el formato del horario/calendario del trabajador
            cell = ws.cell(row=r+9, column=3, value="SEMANA")
            cell.font = Font(bold=True)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')
            aux_c = 1
            for day in ["Lunes","Martes", "Miercoles","Jueves","Viernes"]:
                cell = ws.cell(row=r + 9, column=3 + aux_c, value="{0}".format(day))
                cell.font = Font(bold=True)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center')
                aux_c+=1
            span_of_weeks = 14
            dates = []
            today = date.today()
            initial_week_date = today + timedelta(days=-today.weekday())
            calendar_r = r + 10 #Esta variable para la fila podría definirse se otra forma
            for i in range (0,span_of_weeks):
                dates.append(initial_week_date+i*timedelta(days=7))
            for d in dates:
                cell = ws.cell(row=calendar_r, column=3, value=d)
                cell.font = Font(bold=True)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center')
                calendar_r+=1

            for f in [i for i in range(14, 14+len(dates))]:#Revisar los limites de indexación
                ws.row_dimensions[f].height = 40
                for c in [i for i in range(4,9)]:
                    cell = ws.cell(row=f, column=c)
                    cell.font = Font(bold=True)
                    cell.border = thin_border
                    # cell.alignment = Alignment(horizontal='center')
                    wrap_alignment = Alignment(wrap_text=True, horizontal="center",vertical="center")
                    cell.alignment = wrap_alignment



            #Imprimir las tareas del trabajador



            #Se definen las fechas de trabajo del trabajador

            tasks = []
            tasks_dates_blocks = []
            for tarea in tareas:
                working_dates = []
                task_initial_d = tarea.planned_initial_date
                task_final_d = tarea.planned_end_date
                task = db.Tasks.get(id=tarea.task.id)
                tasks.append(task)
                # Ajustar para que solo tome fechas de trabajo en días hábiles
                working_days = (task_final_d-task_initial_d).days
                working_days_span = working_days

                # working_dates = [task_initial_d + i*timedelta(days=1) for i in range(0,working_days)]
                working_days_counter = 0
                while(working_days_counter<working_days_span+1): #Ojo con el +1 que es para tomar el intervalo cerrado de la fecha de término
                    if(not isNotWorkday(task_initial_d + working_days_counter*timedelta(days=1))):
                        working_dates.append(task_initial_d + working_days_counter*timedelta(days=1))
                    else:
                        # working_days_span+=1 #Ya está considerado en el Planning
                        pass
                    working_days_counter+=1
                tasks_dates_blocks.append(working_dates)

            # # ########Adición de funcionalidad para las vacaciones y licencias (Atención)
            activities = []
            activities_dates_blocks = []
            for e_act in e_activities:
                ea_dates = []
                ea_initial_d = e_act.initial_date
                ea_final_d = e_act.end_date
                activity = db.Activities.get(id=e_act.activity.id)
                activities.append(activity)
                # Ajustar para que solo tome fechas de trabajo en días hábiles
                ea_days = (ea_final_d - ea_initial_d).days
                ea_days_span = ea_days

                # working_dates = [task_initial_d + i*timedelta(days=1) for i in range(0,working_days)]
                ea_days_counter = 0
                while (ea_days_counter < ea_days_span + 1):  # Ojo con el +1 que es para tomar el intervalo cerrado de la fecha de término
                    if (not isNotWorkday(ea_initial_d + ea_days_counter * timedelta(days=1))):
                        ea_dates.append(ea_initial_d + ea_days_counter * timedelta(days=1))
                    else:
                        # working_days_span+=1 #Ya está considerado en el Planning
                        pass
                    ea_days_counter += 1
                    # # #########Adición de funcionalidad para las vacaciones y licencias (Atención)
                activities_dates_blocks.append(ea_dates)
            # print(ea_dates)

            #Se imprimen los días de trabajo
            for i in range(0,len(dates)-1):
                task_counter = 0
                for task_date_block in tasks_dates_blocks:
                    for d in task_date_block:
                        if (dates[i] <= d < dates[i + 1]):
                            # cell = ws.cell(row=r +10 + i, column=4+task_initial_d.weekday(), value="{0} \n Proyecto: {1}".format(task_initial_d, task.project))
                            cell = ws.cell(row=r + 10 + i, column=4 + d.weekday(),
                                           value="Proyecto {0}\n{1}".format(tasks[task_counter].project,tasks[task_counter].project.client_address))
                            cell.font = Font(bold=True)
                            cell.border = thin_border
                            cell.alignment = Alignment(horizontal='center')
                            # cell.style.alignment.wrap_text = True #Para autoajustar el tamaño de la celca. Revisar su correcto funcionamiento
                            # create alignment style
                            wrap_alignment = Alignment(wrap_text=True, horizontal="center",
                                                       vertical="center")
                            # assign
                            cell.alignment = wrap_alignment
                            cell.fill = PatternFill("solid", fgColor="ffff00")
                        elif (dates[len(dates) - 1] <= d):
                            # cell = ws.cell(row=r +10 + i, column=4+task_initial_d.weekday(), value="{0} \n Proyecto: {1}".format(task_initial_d, task.project))
                            cell = ws.cell(row=r + 10 + i + 1, column=4 + d.weekday(),
                                           value="Proyecto {0}\n{1}".format(tasks[task_counter].project,
                                                                            tasks[task_counter].project.client_address))
                            cell.font = Font(bold=True)
                            cell.border = thin_border
                            cell.alignment = Alignment(horizontal='center')
                            # cell.style.alignment.wrap_text = True #Para autoajustar el tamaño de la celca. Revisar su correcto funcionamiento
                            # create alignment style
                            wrap_alignment = Alignment(wrap_text=True, horizontal="center",
                                                       vertical="center")
                            # assign
                            cell.alignment = wrap_alignment
                            cell.fill = PatternFill("solid", fgColor="ffff00")
                    task_counter+=1

                # Adición de funcionalidad para las vacaciones y licencias (Atención)
                # print(len(ea_dates))
                activity_counter = 0
                for activity_date_block in activities_dates_blocks:
                    for d in activity_date_block:
                        if (dates[i] <= d < dates[i + 1]):
                            # cell = ws.cell(row=r +10 + i, column=4+task_initial_d.weekday(), value="{0} \n Proyecto: {1}".format(task_initial_d, task.project))
                            cell = ws.cell(row=r + 10 + i, column=4 + d.weekday(),
                                           value=activities[activity_counter].description)
                            cell.font = Font(bold=True)
                            cell.border = thin_border
                            cell.alignment = Alignment(horizontal='center')
                            # cell.style.alignment.wrap_text = True #Para autoajustar el tamaño de la celca. Revisar su correcto funcionamiento
                            # create alignment style
                            wrap_alignment = Alignment(wrap_text=True, horizontal="center",
                                                       vertical="center")
                            # assign
                            cell.alignment = wrap_alignment
                            if activities[activity_counter].id == 1:
                                cell.fill = PatternFill("solid", fgColor="E60404")
                            elif activities[activity_counter].id == 2:
                                cell.fill = PatternFill("solid", fgColor="810777")
                        elif (dates[len(dates) - 1] <= d):
                            # cell = ws.cell(row=r +10 + i, column=4+task_initial_d.weekday(), value="{0} \n Proyecto: {1}".format(task_initial_d, task.project))
                            cell = ws.cell(row=r + 10 + i + 1, column=4 + d.weekday(),
                                           value=activities[activity_counter].description)
                            cell.font = Font(bold=True)
                            cell.border = thin_border
                            cell.alignment = Alignment(horizontal='center')
                            # cell.style.alignment.wrap_text = True #Para autoajustar el tamaño de la celca. Revisar su correcto funcionamiento
                            # create alignment style
                            wrap_alignment = Alignment(wrap_text=True, horizontal="center",
                                                       vertical="center")
                            # assign
                            cell.alignment = wrap_alignment
                            if activities[activity_counter].id == 1:
                                cell.fill = PatternFill("solid", fgColor="E60404")
                            elif activities[activity_counter].id == 2:
                                cell.fill = PatternFill("solid", fgColor="810777")

                    activity_counter += 1
                # Adición de funcionalidad para las vacaciones y licencias (Atención)

            # Imprimir los días feriados
            all_the_days_on_display = []
            for i in range(0,span_of_weeks*7):
                all_the_days_on_display.append(initial_week_date+timedelta(days=i))
            for i in range(0, len(dates)-1):
                for d in all_the_days_on_display:
                    if (dates[i] <= d < dates[i + 1]):
                        if(isHoliday(d) and d.weekday() < 5):
                            cell = ws.cell(row=r + 10 + i, column=4 + d.weekday(),value="Feriado")
                            cell.font = Font(bold=True)
                            cell.border = thin_border
                            cell.alignment = Alignment(horizontal='center')
                            wrap_alignment = Alignment(wrap_text=True, horizontal="center",vertical="center")
                            cell.alignment = wrap_alignment
                            cell.fill = PatternFill("solid", fgColor="00ffff")
                        elif(ws.cell(row=r + 10 + i, column=4 + d.weekday()).value==None and d.weekday() < 5):
                            ws.cell(row=r + 10 + i, column=4 + d.weekday()).value = "Disponible"
                            ws.cell(row=r + 10 + i, column=4 + d.weekday()).fill = PatternFill("solid", fgColor="00ff00")
                    elif(dates[len(dates)-1] <= d):
                        if (isHoliday(d) and d.weekday() < 5):
                            cell = ws.cell(row=r + 10 + i + 1, column=4 + d.weekday(), value="Feriado")
                            cell.font = Font(bold=True)
                            cell.border = thin_border
                            cell.alignment = Alignment(horizontal='center')
                            wrap_alignment = Alignment(wrap_text=True, horizontal="center",
                                                       vertical="center")
                            cell.alignment = wrap_alignment
                            cell.fill = PatternFill("solid", fgColor="00ffff")
                        elif (ws.cell(row=r + 10 + i + 1, column=4 + d.weekday()).value == None and d.weekday() < 5):
                            ws.cell(row=r + 10 + i + 1, column=4 + d.weekday()).value = "Disponible"
                            ws.cell(row=r + 10 + i + 1, column=4 + d.weekday()).fill = PatternFill("solid", fgColor="00ff00")

        sheet_index+=1
    try:
        module_path = os.path.dirname(__file__)
        panorama_folder_path = os.path.abspath(os.path.join(module_path, os.pardir))
        report_folder_path = os.path.join(panorama_folder_path,"Reportes")
        if not os.path.exists(report_folder_path):
            os.makedirs(report_folder_path)
        report_file_name = "Reporte empleados {0} {1}.xlsx".format("Fabricación",date.today())
        fn = os.path.join(report_folder_path,report_file_name)
        wb.save(fn)
    except OSError as e:
        if e.args[0] != 13:
            raise
        input("\n Ha ocurrido un error porque el archivo Reporte empleados {0} {1}.xlsx está abierto. Por favor ciérrelo y presione cualquier tecla para que el programa pueda continuar.".format("Fabricación",date.today()))

def createInstallersReport(db):
    ''' Este método crea un informe en Excel compacto con una proción de la información de la base de datos. '''
    id_skill = 4
    with db_session:
        employees_skills = select(es for es in db.Employees_Skills if es.skill.id == id_skill)
        employees = []
        for es in employees_skills:
            employee = db.Employees.get(id=es.employee.id)
            employees.append(employee)
        sheet_index = 0
        wb = Workbook()
        by_default_sheet = wb.get_sheet_by_name('Sheet')
        by_default_sheet.title = 'Introducción del informe'
        wb.remove_sheet(by_default_sheet)
        employee_row_counter = 0
        ws = wb.create_sheet("Informe trabajadores",index=sheet_index)
        sheet_index+=1
        for e in employees:
            e_activities = select(ea for ea in db.Employees_Activities if
                                  ea.employee == e)  # Para ver lo de las vacaciones
            from Planning.features import isHoliday, \
                isNotWorkday  # Es realmente necesario importar isHoliday
            tareas = select(et for et in db.Employees_Tasks if et.employee == e)

            widths = {"A": 20, "B": 20, "C": 20, "D": 20, "E": 20, "F": 20,
                      "G": 20, "H": 20, "I": 20, "J": 20, "K": 20, "L": 20,
                      "M": 20, "N": 20, "O": 20, "P": 20, "Q": 20, "R": 20,
                      "S": 20, "T": 20, "U": 20, "V": 20, "W": 20, "X": 20,
                      "Y": 20, "Z": 20, "AA": 20, "AB": 20, "AC": 20, "AD": 20,
                      "AE": 20, "AF": 20, "AG": 20, "AH": 20, "AI": 20, "AJ": 20,
                      "AK": 20, "AL": 20, "AM": 20, "AN": 20, "AO": 20, "AP": 20,
                      "AQ": 20, "AR": 20, "AS": 20, "AT": 20, "AU": 20, "AV": 20,
                      "AW": 20, "AX": 20, "AY": 20, "AZ": 20, "BA": 20, "BB": 20,
                      "BC": 20, "BD": 20, "BE": 20, "BF": 20, "BG": 20, "BH": 20,
                      "BI": 20, "BJ": 20, "BK": 20, "BL": 20, "BM": 20, "BN": 20,
                      "BO": 20, "BP": 20, "BQ": 20, "BR": 20, "BS": 20, "BT": 20,
                      "BU": 20, "BV": 20, "BW": 20, "BX": 20, "BY": 20, "BZ": 20,
                      "CA": 20, "CB": 20, "CC": 20, "CD": 20, "CE": 20, "CF": 20,
                      "CG": 20, "CH": 20, "CI": 20, "CJ": 20, "CK": 20, "CL": 20,
                      "CM": 20, "CN": 20, "CO": 20, "CP": 20, "CQ": 20, "CR": 20,
                      "CS": 20, "CT": 20, "CU": 20, "CV": 20, "CW": 20, "CX": 20,
                      "CY": 20, "CZ": 20, "DA": 20, "DB": 20, "DC": 20, "DD": 20,
                      "DE": 20, "DF": 20, "DG": 20, "DH": 20, "DI": 20, "DJ": 20,
                      "DK": 20, "DL": 20, "DM": 20, "DN": 20, "DO": 20, "DP": 20,
                      "DQ": 20, "DR": 20, "DS": 20, "DT": 20, "DU": 20, "DV": 20,
                      "DW": 20, "DX": 20, "DY": 20, "DZ": 20}

            heights = {"A": 10, "B": 10, "C": 10, "D": 10, "E": 10, "F": 10,
                       "G": 10, "H": 10, "I": 10, "J": 10, "K": 10, "L": 10,
                       "M": 10, "N": 10, "O": 10, "P": 10, "Q": 10, "R": 10,
                       "S": 10, "T": 10, "U": 10, "V": 10, "W": 10, "X": 10,
                       "Y": 10, "Z": 10, "AA": 10, "AB": 10, "AC": 10, "AD": 10,
                       "AE": 10, "AF": 10, "AG": 10, "AH": 10, "AI": 10, "AJ": 10,
                       "AK": 10, "AL": 10, "AM": 10, "AN": 10, "AO": 10, "AP": 10,
                       "AQ": 10, "AR": 10, "AS": 10, "AT": 10, "AU": 10, "AV": 10,
                       "AW": 10, "AX": 10, "AY": 10, "AZ": 10, "BA": 10, "BB": 10,
                       "BC": 10, "BD": 10, "BE": 10, "BF": 10, "BG": 10, "BH": 10,
                       "BI": 10, "BJ": 10, "BK": 10, "BL": 10, "BM": 10, "BN": 20,
                       "BO": 10, "BP": 10, "BQ": 10, "BR": 10, "BS": 10, "BT": 10,
                       "BU": 10, "BV": 10, "BW": 10, "BX": 10, "BY": 10, "BZ": 10,
                       "CA": 10, "CB": 10, "CC": 10, "CD": 10, "CE": 10, "CF": 10,
                       "CG": 10, "CH": 10, "CI": 10, "CJ": 10, "CK": 10, "CL": 10,
                       "CM": 10, "CN": 10, "CO": 10, "CP": 10, "CQ": 10, "CR": 10,
                       "CS": 10, "CT": 10, "CU": 10, "CV": 10, "CW": 10, "CX": 10,
                       "CY": 10, "CZ": 10, "DA": 10, "DB": 10, "DC": 10, "DD": 10,
                       "DE": 10, "DF": 10, "DG": 10, "DH": 10, "DI": 10, "DJ": 10,
                       "DK": 10, "DL": 10, "DM": 10, "DN": 10, "DO": 10, "DP": 10,
                       "DQ": 10, "DR": 10, "DS": 10, "DT": 10, "DU": 10, "DV": 10,
                       "DW": 10, "DX": 10, "DY": 10, "DZ": 10}

            thin_border = Border(left=Side(style='thin'),
                                 right=Side(style='thin'),
                                 top=Side(style='thin'),
                                 bottom=Side(style='thin'))

            columns = ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M", "N",
                       "O", "P", "Q", "R", "S", "T", "U", "V", "W", "X", "Y", "Z", "AA", "AB",
                       "AC", "AD", "AE", "AF", "AG", "AH", "AI", "AJ", "AK", "AL", "AM", "AN",
                       "AO", "AP", "AQ", "AR", "AS", "AT", "AU", "AV", "AW", "AX", "AY", "AZ",
                       "BA", "BB", "BC", "BD", "BE", "BF", "BG", "BH", "BI", "BJ", "BK", "BL", "BM",
                       "BN", "BO", "BP", "BQ", "BR", "BS", "BT", "BU", "BV", "BW", "BX", "BY", "BZ",
                       "CA", "CB", "CC", "CD", "CE", "CF", "CG", "CH", "CI", "CJ", "CK", "CL", "CM",
                       "CN", "CO", "CP", "CQ", "CR", "CS", "CT", "CU", "CV", "CW", "CX", "CY", "CZ",
                       "DA", "DB", "DC", "DD", "DE", "DF", "DG", "DH", "DI", "DJ", "DK", "DL", "DM",
                       "DN", "DO", "DP", "DQ", "DR", "DS", "DT", "DU", "DV", "DW", "DX", "DY", "DZ"]
            for c in columns:
                ws.column_dimensions[c].width = widths[c]
                ws.column_dimensions[c].height = heights[c]

            thin_border = Border(left=Side(style='thin'),
                                 right=Side(style='thin'),
                                 top=Side(style='thin'),
                                 bottom=Side(style='thin'))

            if (
                    False):  # (es.skill.id == 4) <-- introducir esta condición si es que se quiere mostrar la información de "senior"
                # columns = ["A", "B", "C","D","E","F","G","H","I"] #Para un instalador y con el rendimiento
                columns = ["A", "B", "C", "D", "E", "F", "G",
                           "H"]  # Para un instalador sin el rendimiento
                # texts = ["","","","ID", "NOMBRE", "ZONA", "COMPETENCIA",
                #          "SENIOR","RENDIMIENTO"] #Para un instalador y con el rendimiento
                texts = ["", "", "", "ID", "NOMBRE", "ZONA", "COMPETENCIA",
                         "SENIOR"]  # Para un instalador sin el rendimiento
            else:
                columns = ["A", "B", "C", "D", "E", "F", "G"]
                texts = ["", "", "", "ID", "NOMBRE", "ZONA", "COMPETENCIA"]

            aux_columns = ["A", "B", "C", "D", "E", "F", "G",
                           "H"]  # Solo para mantener el ancho de las celdas en el caso que no sea un instalador

            for c in aux_columns:
                ws.column_dimensions[c].width = widths[c]
                ws.column_dimensions[c].height = heights[c]

            num_columns, letter_columns = zip(*list(enumerate(columns)))
            num_columns = list(num_columns)
            num_columns = [x + 1 for x in
                           num_columns]  # Desplazamos los valores en 1 porque Excel está indexado desde el 1 y no desde el 0
            letter_columns = list(letter_columns)

            # Escribir la fecha en la que fue producida el reporte:
            cell = ws.cell(row=6, column=3, value="Reporte producido el: ")
            cell.font = Font(bold=True, )
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='left')

            cell = ws.cell(row=6, column=4, value=date.today())
            cell.font = Font(bold=True, )
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='left')

            # A continuación llenamos con los datos


            r = 4
            r += employee_row_counter
            # Escribe el ID del empleado
            if e.id != None:
                cell = ws.cell(row=r + 10, column=3,
                               value="id: {0}\nNombre:{1}".format(e.id, e.name))
                cell.font = Font(bold=True)
                cell.border = thin_border
                wrap_alignment = Alignment(wrap_text=True, horizontal="center",
                                           vertical="center")
                # assign
                cell.alignment = wrap_alignment

            # Preparar el formato del horario/calendario del trabajador
            cell = ws.cell(row=4 + 9, column=3, value="EMPLEADO")
            cell.font = Font(bold=True)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')

            span_of_weeks = 15
            week_counter = 0
            aux_c = 0
            for week in range(0, span_of_weeks):
                for day in ["Lunes", "Martes", "Miercoles", "Jueves", "Viernes", "Sábado",
                            "Domingo"]:
                    cell = ws.cell(row=4 + 9, column=4 + aux_c, value="{0} {1}".format(day,
                                                                                       date.today() + timedelta(
                                                                                           days=aux_c - date.today().weekday())))
                    cell.font = Font(bold=True)
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal='center')
                    aux_c += 1
                week_counter += 1

            dates = []
            today = date.today()
            for i in range(0, span_of_weeks * 7):
                # dates.append(today + i * timedelta(days=1))
                dates.append(date.today() + timedelta(days=i - date.today().weekday()))
            # Imprimir las tareas del trabajador



            # Se definen las fechas de trabajo del trabajador

            tasks = []
            tasks_dates_blocks = []
            for tarea in tareas:
                working_dates = []
                task_initial_d = tarea.planned_initial_date
                task_final_d = tarea.planned_end_date
                task = db.Tasks.get(id=tarea.task.id)
                tasks.append(task)
                # Ajustar para que solo tome fechas de trabajo en días hábiles
                working_days = (task_final_d - task_initial_d).days
                working_days_span = working_days

                # working_dates = [task_initial_d + i*timedelta(days=1) for i in range(0,working_days)]
                working_days_counter = 0
                while (
                    working_days_counter < working_days_span + 1):  # Ojo con el +1 que es para tomar el intervalo cerrado de la fecha de término
                    if (
                    not isNotWorkday(task_initial_d + working_days_counter * timedelta(days=1))):
                        working_dates.append(
                            task_initial_d + working_days_counter * timedelta(days=1))
                    else:
                        # working_days_span+=1 #Ya está considerado en el Planning
                        pass
                    working_days_counter += 1
                tasks_dates_blocks.append(working_dates)

            # # ########Adición de funcionalidad para las vacaciones y licencias (Atención)
            activities = []
            activities_dates_blocks = []
            for e_act in e_activities:
                ea_dates = []
                ea_initial_d = e_act.initial_date
                ea_final_d = e_act.end_date
                activity = db.Activities.get(id=e_act.activity.id)
                activities.append(activity)
                # Ajustar para que solo tome fechas de trabajo en días hábiles
                ea_days = (ea_final_d - ea_initial_d).days
                ea_days_span = ea_days

                # working_dates = [task_initial_d + i*timedelta(days=1) for i in range(0,working_days)]
                ea_days_counter = 0
                while (
                            ea_days_counter < ea_days_span + 1):  # Ojo con el +1 que es para tomar el intervalo cerrado de la fecha de término
                    if (not isNotWorkday(ea_initial_d + ea_days_counter * timedelta(days=1))):
                        ea_dates.append(ea_initial_d + ea_days_counter * timedelta(days=1))
                    else:
                        # working_days_span+=1 #Ya está considerado en el Planning
                        pass
                    ea_days_counter += 1
                    # # #########Adición de funcionalidad para las vacaciones y licencias (Atención)
                activities_dates_blocks.append(ea_dates)
            # print(ea_dates)

            # Se imprimen los días de trabajo
            for i in range(0, len(dates)):
                task_counter = 0
                for tasks_dates_block in tasks_dates_blocks:
                    for d in tasks_dates_block:
                        if d == dates[i]:
                            cell = ws.cell(row=r + 10, column=4 + i,
                                           value="Proyecto {0}\n{1}".format(
                                               tasks[task_counter].project,
                                               tasks[task_counter].project.client_address))
                            cell.font = Font(bold=True)
                            cell.border = thin_border
                            cell.alignment = Alignment(horizontal='center')
                            # cell.style.alignment.wrap_text = True #Para autoajustar el tamaño de la celca. Revisar su correcto funcionamiento
                            # create alignment style
                            wrap_alignment = Alignment(wrap_text=True, horizontal="center",
                                                       vertical="center")
                            # assign
                            cell.alignment = wrap_alignment
                            cell.fill = PatternFill("solid", fgColor="ffff00")
                    task_counter = 0
                activity_counter = 0
                for activity_date_block in activities_dates_blocks:
                    for d in activity_date_block:
                        if d == dates[i]:
                            cell = ws.cell(row=r + 10, column=4 + i,
                                           value=activities[activity_counter].description)
                            cell.font = Font(bold=True)
                            cell.border = thin_border
                            wrap_alignment = Alignment(wrap_text=True, horizontal="center",
                                                       vertical="center")
                            cell.alignment = wrap_alignment
                            # cell.style.alignment.wrap_text = True #Para autoajustar el tamaño de la celca. Revisar su correcto funcionamiento
                            # create alignment style
                            wrap_alignment = Alignment(wrap_text=True, horizontal="center",
                                                       vertical="center")
                            if activities[activity_counter].id == 1:
                                cell.fill = PatternFill("solid", fgColor="E60404")
                            elif activities[activity_counter].id == 2:
                                cell.fill = PatternFill("solid", fgColor="810777")
                    activity_counter = 0
                if isHoliday(dates[i]):
                    cell = ws.cell(row=r + 10, column=4 + i, value="Feriado")
                    cell.font = Font(bold=True)
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal='center')
                    wrap_alignment = Alignment(wrap_text=True, horizontal="center",
                                               vertical="center")
                    cell.alignment = wrap_alignment
                    cell.fill = PatternFill("solid", fgColor="00ffff")
                if dates[i].weekday() >= 5:
                    cell = ws.cell(row=r + 10, column=4 + i, value="Fin de semana")
                    cell.font = Font(bold=True)
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal='center')
                    wrap_alignment = Alignment(wrap_text=True, horizontal="center",
                                               vertical="center")
                    cell.alignment = wrap_alignment
                    cell.fill = PatternFill("solid", fgColor="00ffff")
                if (ws.cell(row=r + 10, column=4 + i).value == None and dates[
                    i].weekday() < 5):  # ws.cell(row=r + 10, column=4 + i).value == None and
                    cell = ws.cell(row=r + 10, column=4 + i, value="Disponible")
                    cell.font = Font(bold=True)
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal='center')
                    wrap_alignment = Alignment(wrap_text=True, horizontal="center",
                                               vertical="center")
                    cell.alignment = wrap_alignment
                    cell.fill = PatternFill("solid", fgColor="00ff00")

            employee_row_counter += 1
        for e in employees:
            e_activities = select(ea for ea in db.Employees_Activities if ea.employee == e) #Para ver lo de las vacaciones
            from Planning.features import isHoliday, isNotWorkday  # Es realmente necesario importar isHoliday
            tareas = select(et for et in db.Employees_Tasks if et.employee == e)
            ws = wb.create_sheet(
                "Informe trabajador {}".format(e.id),index=sheet_index)

            widths = {"A": 20, "B": 20, "C": 20,"D": 20, "E": 20, "F": 20,
                      "G": 20, "H": 20, "I": 20,"J": 20, "K": 20, "L": 20,
                      "M": 20, "N": 20}

            heights = {"A": 10, "B": 10, "C": 10, "D": 10, "E": 10, "F": 10,
                      "G": 10, "H": 10, "I": 10, "J": 10, "K": 10, "L": 10,
                      "M": 10, "N": 10}

            thin_border = Border(left=Side(style='thin'),
                                 right=Side(style='thin'),
                                 top=Side(style='thin'),
                                 bottom=Side(style='thin'))

            if (False): #(es.skill.id == 4) <-- introducir esta condición si es que se quiere mostrar la información de "senior"
                # columns = ["A", "B", "C","D","E","F","G","H","I"] #Para un instalador y con el rendimiento
                columns = ["A", "B", "C", "D", "E", "F", "G", "H"]  # Para un instalador sin el rendimiento
                # texts = ["","","","ID", "NOMBRE", "ZONA", "COMPETENCIA",
                #          "SENIOR","RENDIMIENTO"] #Para un instalador y con el rendimiento
                texts = ["","","","ID", "NOMBRE", "ZONA", "COMPETENCIA",
                         "SENIOR"] #Para un instalador sin el rendimiento
            else:
                columns = ["A", "B", "C", "D", "E", "F", "G"]
                texts = ["", "", "", "ID", "NOMBRE", "ZONA", "COMPETENCIA"]

            aux_columns = ["A", "B", "C", "D", "E", "F", "G", "H"] #Solo para mantener el ancho de las celdas en el caso que no sea un instalador

            for c in aux_columns:
                ws.column_dimensions[c].width = widths[c]
                ws.column_dimensions[c].height = heights[c]


            num_columns, letter_columns = zip(*list(enumerate(columns)))
            num_columns = list(num_columns)
            num_columns = [x + 1 for x in num_columns] #Desplazamos los valores en 1 porque Excel está indexado desde el 1 y no desde el 0
            letter_columns = list(letter_columns)

            # escribir los títulos de las columnas, en negrita

            for i in range(4,len(num_columns)+1):
                cell = ws.cell(row=i, column=3, value=texts[i-1])
                cell.font = Font(bold=True,)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='left')


            #Escribir la fecha en la que fue producida el reporte:
            cell = ws.cell(row=3, column=3, value="Reporte producido el: ")
            cell.font = Font(bold=True, )
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='left')

            cell = ws.cell(row=3, column=4, value=date.today())
            cell.font = Font(bold=True, )
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='left')

            # A continuación llenamos con los datos


            r=4
            #Escribe el ID del empleado
            if e.id != None:
                cell = ws.cell(row=r, column=4, value=e.id)
                cell.font = Font(bold=True)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='left')
            else:
                cell = ws.cell(row=r, column=4, value="Dato no disponible")
                cell.font = Font(bold=True)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='left')

            # Escribe el nombre del empleado
            if e.name != None:
                cell = ws.cell(row=r+1, column=4, value=e.name)
                cell.font = Font(bold=True)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='left')
            else:
                cell = ws.cell(row=r+1, column=4, value="Dato no disponible")
                cell.font = Font(bold=True)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='left')

            # Escribe la zona geográfica del empleado
            if e.zone != None:
                cell = ws.cell(row=r+2, column=4, value=e.zone)
                cell.font = Font(bold=True)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='left')
            else:
                cell = ws.cell(row=r+2, column=4, value="Dato no disponible")
                cell.font = Font(bold=True)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='left')

            # Escribe la competencia del empleado (por el momento se asume que solo es una)

            if es.skill != None:
                skill_name = es.skill.name
                cell = ws.cell(row=r+3, column=4, value=skill_name)
                cell.font = Font(bold=True)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='left')
            else:
                cell = ws.cell(row=r+3, column=4, value="Dato no disponible")
                cell.font = Font(bold=True)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='left')

            # Escribe si el empleado es senior o no (Sí o No)
            # if(es.skill.id == 4):
            #     if e.senior != None:
            #         if e.senior:
            #             value = "Sí"
            #         else:
            #             value = "No"
            #         cell = ws.cell(row=r+4, column=4, value = value)
            #         cell.font = Font(bold=True)
            #         cell.border = thin_border
            #         cell.alignment = Alignment(horizontal='left')
            #     else:
            #         cell = ws.cell(row=r+4, column=4, value="Dato no disponible")
            #         cell.font = Font(bold=True)
            #         cell.border = thin_border
            #         cell.alignment = Alignment(horizontal='left')

            # # Escribe el rendimiento del trabajador
            # if es.performance != None:
            #     cell = ws.cell(row=r+5, column=4, value = es.performance)
            #     cell.font = Font(bold=True)
            #     cell.border = thin_border
            #     cell.alignment = Alignment(horizontal='left')
            # else:
            #     cell = ws.cell(row=r+5, column=4, value="Dato no disponible")
            #     cell.font = Font(bold=True)
            #     cell.border = thin_border
            #     cell.alignment = Alignment(horizontal='left')

            #Preparar el formato del horario/calendario del trabajador
            cell = ws.cell(row=r+9, column=3, value="SEMANA")
            cell.font = Font(bold=True)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')
            aux_c = 1
            for day in ["Lunes","Martes", "Miercoles","Jueves","Viernes"]:
                cell = ws.cell(row=r + 9, column=3 + aux_c, value="{0}".format(day))
                cell.font = Font(bold=True)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center')
                aux_c+=1
            span_of_weeks = 14
            dates = []
            today = date.today()
            initial_week_date = today + timedelta(days=-today.weekday())
            calendar_r = r + 10 #Esta variable para la fila podría definirse se otra forma
            for i in range (0,span_of_weeks):
                dates.append(initial_week_date+i*timedelta(days=7))
            for d in dates:
                cell = ws.cell(row=calendar_r, column=3, value=d)
                cell.font = Font(bold=True)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center')
                calendar_r+=1

            for f in [i for i in range(14, 14+len(dates))]:#Revisar los limites de indexación
                ws.row_dimensions[f].height = 40
                for c in [i for i in range(4,9)]:
                    cell = ws.cell(row=f, column=c)
                    cell.font = Font(bold=True)
                    cell.border = thin_border
                    # cell.alignment = Alignment(horizontal='center')
                    wrap_alignment = Alignment(wrap_text=True, horizontal="center",vertical="center")
                    cell.alignment = wrap_alignment



            #Imprimir las tareas del trabajador



            #Se definen las fechas de trabajo del trabajador

            tasks = []
            tasks_dates_blocks = []
            for tarea in tareas:
                working_dates = []
                task_initial_d = tarea.planned_initial_date
                task_final_d = tarea.planned_end_date
                task = db.Tasks.get(id=tarea.task.id)
                tasks.append(task)
                # Ajustar para que solo tome fechas de trabajo en días hábiles
                working_days = (task_final_d-task_initial_d).days
                working_days_span = working_days

                # working_dates = [task_initial_d + i*timedelta(days=1) for i in range(0,working_days)]
                working_days_counter = 0
                while(working_days_counter<working_days_span+1): #Ojo con el +1 que es para tomar el intervalo cerrado de la fecha de término
                    if(not isNotWorkday(task_initial_d + working_days_counter*timedelta(days=1))):
                        working_dates.append(task_initial_d + working_days_counter*timedelta(days=1))
                    else:
                        # working_days_span+=1 #Ya está considerado en el Planning
                        pass
                    working_days_counter+=1
                tasks_dates_blocks.append(working_dates)

            # # ########Adición de funcionalidad para las vacaciones y licencias (Atención)
            activities = []
            activities_dates_blocks = []
            for e_act in e_activities:
                ea_dates = []
                ea_initial_d = e_act.initial_date
                ea_final_d = e_act.end_date
                activity = db.Activities.get(id=e_act.activity.id)
                activities.append(activity)
                # Ajustar para que solo tome fechas de trabajo en días hábiles
                ea_days = (ea_final_d - ea_initial_d).days
                ea_days_span = ea_days

                # working_dates = [task_initial_d + i*timedelta(days=1) for i in range(0,working_days)]
                ea_days_counter = 0
                while (ea_days_counter < ea_days_span + 1):  # Ojo con el +1 que es para tomar el intervalo cerrado de la fecha de término
                    if (not isNotWorkday(ea_initial_d + ea_days_counter * timedelta(days=1))):
                        ea_dates.append(ea_initial_d + ea_days_counter * timedelta(days=1))
                    else:
                        # working_days_span+=1 #Ya está considerado en el Planning
                        pass
                    ea_days_counter += 1
                    # # #########Adición de funcionalidad para las vacaciones y licencias (Atención)
                activities_dates_blocks.append(ea_dates)
            # print(ea_dates)

            #Se imprimen los días de trabajo
            for i in range(0,len(dates)-1):
                task_counter = 0
                for task_date_block in tasks_dates_blocks:
                    for d in task_date_block:
                        if (dates[i] <= d < dates[i + 1]):
                            # cell = ws.cell(row=r +10 + i, column=4+task_initial_d.weekday(), value="{0} \n Proyecto: {1}".format(task_initial_d, task.project))
                            cell = ws.cell(row=r + 10 + i, column=4 + d.weekday(),
                                           value="Proyecto {0}\n{1}".format(tasks[task_counter].project,tasks[task_counter].project.client_address))
                            cell.font = Font(bold=True)
                            cell.border = thin_border
                            cell.alignment = Alignment(horizontal='center')
                            # cell.style.alignment.wrap_text = True #Para autoajustar el tamaño de la celca. Revisar su correcto funcionamiento
                            # create alignment style
                            wrap_alignment = Alignment(wrap_text=True, horizontal="center",
                                                       vertical="center")
                            # assign
                            cell.alignment = wrap_alignment
                            cell.fill = PatternFill("solid", fgColor="ffff00")
                        elif (dates[len(dates) - 1] <= d):
                            # cell = ws.cell(row=r +10 + i, column=4+task_initial_d.weekday(), value="{0} \n Proyecto: {1}".format(task_initial_d, task.project))
                            cell = ws.cell(row=r + 10 + i + 1, column=4 + d.weekday(),
                                           value="Proyecto {0}\n{1}".format(tasks[task_counter].project,
                                                                            tasks[task_counter].project.client_address))
                            cell.font = Font(bold=True)
                            cell.border = thin_border
                            cell.alignment = Alignment(horizontal='center')
                            # cell.style.alignment.wrap_text = True #Para autoajustar el tamaño de la celca. Revisar su correcto funcionamiento
                            # create alignment style
                            wrap_alignment = Alignment(wrap_text=True, horizontal="center",
                                                       vertical="center")
                            # assign
                            cell.alignment = wrap_alignment
                            cell.fill = PatternFill("solid", fgColor="ffff00")
                    task_counter+=1

                # Adición de funcionalidad para las vacaciones y licencias (Atención)
                # print(len(ea_dates))
                activity_counter = 0
                for activity_date_block in activities_dates_blocks:
                    for d in activity_date_block:
                        if (dates[i] <= d < dates[i + 1]):
                            # cell = ws.cell(row=r +10 + i, column=4+task_initial_d.weekday(), value="{0} \n Proyecto: {1}".format(task_initial_d, task.project))
                            cell = ws.cell(row=r + 10 + i, column=4 + d.weekday(),
                                           value=activities[activity_counter].description)
                            cell.font = Font(bold=True)
                            cell.border = thin_border
                            cell.alignment = Alignment(horizontal='center')
                            # cell.style.alignment.wrap_text = True #Para autoajustar el tamaño de la celca. Revisar su correcto funcionamiento
                            # create alignment style
                            wrap_alignment = Alignment(wrap_text=True, horizontal="center",
                                                       vertical="center")
                            # assign
                            cell.alignment = wrap_alignment
                            if activities[activity_counter].id == 1:
                                cell.fill = PatternFill("solid", fgColor="E60404")
                            elif activities[activity_counter].id == 2:
                                cell.fill = PatternFill("solid", fgColor="810777")
                        elif (dates[len(dates) - 1] <= d):
                            # cell = ws.cell(row=r +10 + i, column=4+task_initial_d.weekday(), value="{0} \n Proyecto: {1}".format(task_initial_d, task.project))
                            cell = ws.cell(row=r + 10 + i + 1, column=4 + d.weekday(),
                                           value=activities[activity_counter].description)
                            cell.font = Font(bold=True)
                            cell.border = thin_border
                            cell.alignment = Alignment(horizontal='center')
                            # cell.style.alignment.wrap_text = True #Para autoajustar el tamaño de la celca. Revisar su correcto funcionamiento
                            # create alignment style
                            wrap_alignment = Alignment(wrap_text=True, horizontal="center",
                                                       vertical="center")
                            # assign
                            cell.alignment = wrap_alignment
                            if activities[activity_counter].id == 1:
                                cell.fill = PatternFill("solid", fgColor="E60404")
                            elif activities[activity_counter].id == 2:
                                cell.fill = PatternFill("solid", fgColor="810777")

                    activity_counter += 1
                # Adición de funcionalidad para las vacaciones y licencias (Atención)

            # Imprimir los días feriados
            all_the_days_on_display = []
            for i in range(0,span_of_weeks*7):
                all_the_days_on_display.append(initial_week_date+timedelta(days=i))
            for i in range(0, len(dates)-1):
                for d in all_the_days_on_display:
                    if (dates[i] <= d < dates[i + 1]):
                        if(isHoliday(d) and d.weekday() < 5):
                            cell = ws.cell(row=r + 10 + i, column=4 + d.weekday(),value="Feriado")
                            cell.font = Font(bold=True)
                            cell.border = thin_border
                            cell.alignment = Alignment(horizontal='center')
                            wrap_alignment = Alignment(wrap_text=True, horizontal="center",vertical="center")
                            cell.alignment = wrap_alignment
                            cell.fill = PatternFill("solid", fgColor="00ffff")
                        elif(ws.cell(row=r + 10 + i, column=4 + d.weekday()).value==None and d.weekday() < 5):
                            ws.cell(row=r + 10 + i, column=4 + d.weekday()).value = "Disponible"
                            ws.cell(row=r + 10 + i, column=4 + d.weekday()).fill = PatternFill("solid", fgColor="00ff00")
                    elif(dates[len(dates)-1] <= d):
                        if (isHoliday(d) and d.weekday() < 5):
                            cell = ws.cell(row=r + 10 + i + 1, column=4 + d.weekday(), value="Feriado")
                            cell.font = Font(bold=True)
                            cell.border = thin_border
                            cell.alignment = Alignment(horizontal='center')
                            wrap_alignment = Alignment(wrap_text=True, horizontal="center",
                                                       vertical="center")
                            cell.alignment = wrap_alignment
                            cell.fill = PatternFill("solid", fgColor="00ffff")
                        elif (ws.cell(row=r + 10 + i + 1, column=4 + d.weekday()).value == None and d.weekday() < 5):
                            ws.cell(row=r + 10 + i + 1, column=4 + d.weekday()).value = "Disponible"
                            ws.cell(row=r + 10 + i + 1, column=4 + d.weekday()).fill = PatternFill("solid", fgColor="00ff00")

        sheet_index+=1
    try:
        module_path = os.path.dirname(__file__)
        panorama_folder_path = os.path.abspath(os.path.join(module_path, os.pardir))
        report_folder_path = os.path.join(panorama_folder_path,"Reportes")
        if not os.path.exists(report_folder_path):
            os.makedirs(report_folder_path)
        report_file_name = "Reporte empleados {0} {1}.xlsx".format("Instalación",date.today())
        fn = os.path.join(report_folder_path,report_file_name)
        wb.save(fn)
    except OSError as e:
        if e.args[0] != 13:
            raise
        input("\n Ha ocurrido un error porque el archivo Reporte empleados {0} {1}.xlsx está abierto. Por favor ciérrelo y presione cualquier tecla para que el programa pueda continuar.".format("Instalación",date.today()))

def createRectificatorsReportWide(db):
    ''' Este método crea un informe en Excel compacto con una proción de la información de la base de datos. '''
    id_skill = 1
    with db_session:
        employees_skills = select(es for es in db.Employees_Skills if es.skill.id == id_skill)
        employees = []
        for es in employees_skills:
            employee = db.Employees.get(id=es.employee.id)
            employees.append(employee)
        sheet_index = 0
        wb = Workbook()
        by_default_sheet = wb.get_sheet_by_name('Sheet')
        by_default_sheet.title = 'Introducción del informe'
        wb.remove_sheet(by_default_sheet)
        employee_row_counter = 0
        ws = wb.create_sheet("Informe trabajadores")
        for e in employees:
            e_activities = select(ea for ea in db.Employees_Activities if
                                  ea.employee == e)  # Para ver lo de las vacaciones
            from Planning.features import isHoliday, \
                isNotWorkday  # Es realmente necesario importar isHoliday
            tareas = select(et for et in db.Employees_Tasks if et.employee == e)


            widths = {"A": 20, "B": 20, "C": 20, "D": 20, "E": 20, "F": 20,
                      "G": 20, "H": 20, "I": 20, "J": 20, "K": 20, "L": 20,
                      "M": 20, "N": 20, "O": 20, "P": 20, "Q": 20, "R": 20,
                      "S": 20, "T": 20, "U": 20, "V": 20, "W": 20, "X": 20,
                      "Y": 20, "Z": 20, "AA": 20, "AB": 20, "AC": 20, "AD": 20,
                      "AE": 20, "AF": 20, "AG": 20, "AH": 20, "AI": 20, "AJ": 20,
                      "AK": 20, "AL": 20, "AM": 20, "AN": 20, "AO": 20, "AP": 20,
                      "AQ": 20, "AR": 20, "AS": 20, "AT": 20, "AU": 20, "AV": 20,
                      "AW": 20, "AX": 20, "AY": 20, "AZ": 20, "BA": 20, "BB": 20,
                      "BC": 20, "BD": 20, "BE": 20, "BF": 20, "BG": 20, "BH": 20,
                      "BI": 20, "BJ": 20, "BK": 20, "BL": 20, "BM": 20,"BN": 20,
                      "BO": 20, "BP": 20, "BQ": 20, "BR": 20,"BS": 20,"BT": 20,
                      "BU": 20, "BV": 20, "BW": 20, "BX": 20, "BY": 20, "BZ": 20,
                      "CA": 20, "CB": 20, "CC": 20, "CD": 20, "CE": 20, "CF": 20,
                      "CG": 20, "CH": 20, "CI": 20, "CJ": 20, "CK": 20, "CL": 20,
                      "CM": 20, "CN": 20, "CO": 20, "CP": 20, "CQ": 20, "CR": 20,
                      "CS": 20, "CT": 20, "CU": 20, "CV": 20, "CW": 20, "CX": 20,
                      "CY": 20, "CZ": 20, "DA": 20, "DB": 20, "DC": 20, "DD": 20,
                      "DE": 20, "DF": 20, "DG": 20, "DH": 20, "DI": 20, "DJ": 20,
                      "DK": 20, "DL": 20, "DM": 20, "DN": 20, "DO": 20, "DP": 20,
                      "DQ": 20, "DR": 20, "DS": 20, "DT": 20, "DU": 20, "DV": 20,
                      "DW": 20, "DX": 20, "DY": 20, "DZ": 20}

            heights = {"A": 10, "B": 10, "C": 10, "D": 10, "E": 10, "F": 10,
                       "G": 10, "H": 10, "I": 10, "J": 10, "K": 10, "L": 10,
                       "M": 10, "N": 10, "O": 10, "P": 10, "Q": 10, "R": 10,
                       "S": 10, "T": 10, "U": 10, "V": 10, "W": 10, "X": 10,
                       "Y": 10, "Z": 10, "AA": 10, "AB": 10, "AC": 10, "AD": 10,
                       "AE": 10, "AF": 10, "AG": 10, "AH": 10, "AI": 10, "AJ": 10,
                       "AK": 10, "AL": 10, "AM": 10, "AN": 10, "AO": 10, "AP": 10,
                       "AQ": 10, "AR": 10, "AS": 10, "AT": 10, "AU": 10, "AV": 10,
                       "AW": 10, "AX": 10, "AY": 10, "AZ": 10, "BA": 10, "BB": 10,
                       "BC": 10, "BD": 10, "BE": 10, "BF": 10, "BG": 10, "BH": 10,
                       "BI": 10, "BJ": 10, "BK": 10, "BL": 10, "BM": 10,"BN": 20,
                      "BO": 10, "BP": 10, "BQ": 10, "BR": 10,"BS": 10,"BT": 10,
                      "BU": 10, "BV": 10, "BW": 10, "BX": 10, "BY": 10, "BZ": 10,
                    "CA": 10, "CB": 10,"CC": 10, "CD": 10, "CE": 10, "CF": 10,
                       "CG": 10, "CH": 10,"CI": 10, "CJ": 10, "CK": 10, "CL": 10,
                       "CM": 10,"CN": 10,"CO": 10, "CP": 10, "CQ": 10, "CR": 10,
                       "CS": 10,"CT": 10,"CU": 10, "CV": 10, "CW": 10, "CX": 10,
                       "CY": 10, "CZ": 10, "DA": 10, "DB": 10,"DC": 10, "DD": 10,
                       "DE": 10, "DF": 10,"DG": 10, "DH": 10,"DI": 10, "DJ": 10,
                       "DK": 10, "DL": 10,"DM": 10,"DN": 10,"DO": 10, "DP": 10,
                       "DQ": 10, "DR": 10,"DS": 10,"DT": 10,"DU": 10, "DV": 10,
                       "DW": 10, "DX": 10,"DY": 10, "DZ": 10}

            thin_border = Border(left=Side(style='thin'),
                                 right=Side(style='thin'),
                                 top=Side(style='thin'),
                                 bottom=Side(style='thin'))

            columns = ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M", "N",
                       "O", "P", "Q", "R", "S", "T", "U", "V", "W", "X", "Y", "Z", "AA", "AB",
                       "AC", "AD", "AE", "AF", "AG", "AH", "AI", "AJ", "AK", "AL", "AM", "AN",
                       "AO", "AP", "AQ", "AR", "AS", "AT", "AU", "AV", "AW", "AX", "AY", "AZ",
                       "BA", "BB", "BC", "BD", "BE", "BF", "BG", "BH", "BI", "BJ", "BK", "BL", "BM",
                       "BN","BO","BP","BQ","BR","BS","BT","BU","BV","BW","BX","BY","BZ",
                       "CA", "CB", "CC", "CD", "CE", "CF", "CG", "CH", "CI", "CJ", "CK", "CL", "CM",
                       "CN", "CO", "CP", "CQ", "CR", "CS", "CT", "CU", "CV", "CW", "CX", "CY", "CZ",
                       "DA", "DB", "DC", "DD", "DE", "DF", "DG", "DH", "DI", "DJ", "DK", "DL", "DM",
                       "DN", "DO", "DP", "DQ", "DR", "DS", "DT", "DU", "DV", "DW", "DX", "DY", "DZ"]
            for c in columns:
                ws.column_dimensions[c].width = widths[c]
                ws.column_dimensions[c].height = heights[c]

            thin_border = Border(left=Side(style='thin'),
                                 right=Side(style='thin'),
                                 top=Side(style='thin'),
                                 bottom=Side(style='thin'))

            if (
            False):  # (es.skill.id == 4) <-- introducir esta condición si es que se quiere mostrar la información de "senior"
                # columns = ["A", "B", "C","D","E","F","G","H","I"] #Para un instalador y con el rendimiento
                columns = ["A", "B", "C", "D", "E", "F", "G",
                           "H"]  # Para un instalador sin el rendimiento
                # texts = ["","","","ID", "NOMBRE", "ZONA", "COMPETENCIA",
                #          "SENIOR","RENDIMIENTO"] #Para un instalador y con el rendimiento
                texts = ["", "", "", "ID", "NOMBRE", "ZONA", "COMPETENCIA",
                         "SENIOR"]  # Para un instalador sin el rendimiento
            else:
                columns = ["A", "B", "C", "D", "E", "F", "G"]
                texts = ["", "", "", "ID", "NOMBRE", "ZONA", "COMPETENCIA"]

            aux_columns = ["A", "B", "C", "D", "E", "F", "G",
                           "H"]  # Solo para mantener el ancho de las celdas en el caso que no sea un instalador

            for c in aux_columns:
                ws.column_dimensions[c].width = widths[c]
                ws.column_dimensions[c].height = heights[c]

            num_columns, letter_columns = zip(*list(enumerate(columns)))
            num_columns = list(num_columns)
            num_columns = [x + 1 for x in
                           num_columns]  # Desplazamos los valores en 1 porque Excel está indexado desde el 1 y no desde el 0
            letter_columns = list(letter_columns)


            # Escribir la fecha en la que fue producida el reporte:
            cell = ws.cell(row=6, column=3, value="Reporte producido el: ")
            cell.font = Font(bold=True, )
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='left')

            cell = ws.cell(row=6, column=4, value=date.today())
            cell.font = Font(bold=True, )
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='left')

            # A continuación llenamos con los datos


            r = 4
            r += employee_row_counter
            # Escribe el ID del empleado
            if e.id != None:
                cell = ws.cell(row=r+10, column=3, value="id: {0}\nNombre:{1}".format(e.id,e.name))
                cell.font = Font(bold=True)
                cell.border = thin_border
                wrap_alignment = Alignment(wrap_text=True, horizontal="center",
                                           vertical="center")
                # assign
                cell.alignment = wrap_alignment

            # Preparar el formato del horario/calendario del trabajador
            cell = ws.cell(row=4 + 9, column=3, value="EMPLEADO")
            cell.font = Font(bold=True)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')

            span_of_weeks = 15
            week_counter = 0
            aux_c = 0
            for week in range(0,span_of_weeks):
                for day in ["Lunes", "Martes", "Miercoles", "Jueves", "Viernes", "Sábado", "Domingo"]:
                    cell = ws.cell(row=4 + 9, column=4 + aux_c, value="{0} {1}".format(day,date.today()+timedelta(days=aux_c-date.today().weekday())))
                    cell.font = Font(bold=True)
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal='center')
                    aux_c += 1
                week_counter+=1

            dates = []
            today = date.today()
            for i in range(0, span_of_weeks*7):
                # dates.append(today + i * timedelta(days=1))
                dates.append(date.today()+timedelta(days=i-date.today().weekday()))
            # Imprimir las tareas del trabajador



            # Se definen las fechas de trabajo del trabajador

            tasks = []
            tasks_dates_blocks = []
            for tarea in tareas:
                working_dates = []
                task_initial_d = tarea.planned_initial_date
                task_final_d = tarea.planned_end_date
                task = db.Tasks.get(id=tarea.task.id)
                tasks.append(task)
                # Ajustar para que solo tome fechas de trabajo en días hábiles
                working_days = (task_final_d - task_initial_d).days
                working_days_span = working_days

                # working_dates = [task_initial_d + i*timedelta(days=1) for i in range(0,working_days)]
                working_days_counter = 0
                while (working_days_counter < working_days_span + 1):  # Ojo con el +1 que es para tomar el intervalo cerrado de la fecha de término
                    if (not isNotWorkday(task_initial_d + working_days_counter * timedelta(days=1))):
                        working_dates.append(task_initial_d + working_days_counter * timedelta(days=1))
                    else:
                        # working_days_span+=1 #Ya está considerado en el Planning
                        pass
                    working_days_counter += 1
                tasks_dates_blocks.append(working_dates)

            # # ########Adición de funcionalidad para las vacaciones y licencias (Atención)
            activities = []
            activities_dates_blocks = []
            for e_act in e_activities:
                ea_dates = []
                ea_initial_d = e_act.initial_date
                ea_final_d = e_act.end_date
                activity = db.Activities.get(id=e_act.activity.id)
                activities.append(activity)
                # Ajustar para que solo tome fechas de trabajo en días hábiles
                ea_days = (ea_final_d - ea_initial_d).days
                ea_days_span = ea_days

                # working_dates = [task_initial_d + i*timedelta(days=1) for i in range(0,working_days)]
                ea_days_counter = 0
                while (
                    ea_days_counter < ea_days_span + 1):  # Ojo con el +1 que es para tomar el intervalo cerrado de la fecha de término
                    if (not isNotWorkday(ea_initial_d + ea_days_counter * timedelta(days=1))):
                        ea_dates.append(ea_initial_d + ea_days_counter * timedelta(days=1))
                    else:
                        # working_days_span+=1 #Ya está considerado en el Planning
                        pass
                    ea_days_counter += 1
                    # # #########Adición de funcionalidad para las vacaciones y licencias (Atención)
                activities_dates_blocks.append(ea_dates)
            # print(ea_dates)

            # Se imprimen los días de trabajo
            for i in range(0,len(dates)):
                task_counter = 0
                for tasks_dates_block in tasks_dates_blocks:
                    for d in tasks_dates_block:
                        if d == dates[i]:
                            cell = ws.cell(row=r + 10, column=4 + i,
                                           value="Proyecto {0}\n{1}".format(
                                               tasks[task_counter].project,
                                               tasks[task_counter].project.client_address))
                            cell.font = Font(bold=True)
                            cell.border = thin_border
                            cell.alignment = Alignment(horizontal='center')
                            # cell.style.alignment.wrap_text = True #Para autoajustar el tamaño de la celca. Revisar su correcto funcionamiento
                            # create alignment style
                            wrap_alignment = Alignment(wrap_text=True, horizontal="center",
                                                       vertical="center")
                            # assign
                            cell.alignment = wrap_alignment
                            cell.fill = PatternFill("solid", fgColor="ffff00")
                    task_counter = 0
                activity_counter = 0
                for activity_date_block in activities_dates_blocks:
                    for d in activity_date_block:
                        if d == dates[i]:
                            cell = ws.cell(row=r + 10 , column=4 + i,
                                           value=activities[activity_counter].description)
                            cell.font = Font(bold=True)
                            cell.border = thin_border
                            wrap_alignment = Alignment(wrap_text=True, horizontal="center",
                                                       vertical="center")
                            cell.alignment = wrap_alignment
                            # cell.style.alignment.wrap_text = True #Para autoajustar el tamaño de la celca. Revisar su correcto funcionamiento
                            # create alignment style
                            wrap_alignment = Alignment(wrap_text=True, horizontal="center",
                                                       vertical="center")
                            if activities[activity_counter].id == 1:
                                cell.fill = PatternFill("solid", fgColor="E60404")
                            elif activities[activity_counter].id == 2:
                                cell.fill = PatternFill("solid", fgColor="810777")
                    activity_counter = 0
                if isHoliday(dates[i]):
                    cell = ws.cell(row=r + 10, column=4 + i, value="Feriado")
                    cell.font = Font(bold=True)
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal='center')
                    wrap_alignment = Alignment(wrap_text=True, horizontal="center",
                                               vertical="center")
                    cell.alignment = wrap_alignment
                    cell.fill = PatternFill("solid", fgColor="00ffff")
                if dates[i].weekday()>=5:
                    cell = ws.cell(row=r + 10, column=4 + i, value="Fin de semana")
                    cell.font = Font(bold=True)
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal='center')
                    wrap_alignment = Alignment(wrap_text=True, horizontal="center",
                                               vertical="center")
                    cell.alignment = wrap_alignment
                    cell.fill = PatternFill("solid", fgColor="00ffff")
                if (ws.cell(row=r + 10, column=4 + i).value == None and dates[i].weekday() < 5): #ws.cell(row=r + 10, column=4 + i).value == None and
                    cell = ws.cell(row=r + 10, column=4 + i, value="Disponible")
                    cell.font = Font(bold=True)
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal='center')
                    wrap_alignment = Alignment(wrap_text=True, horizontal="center",
                                               vertical="center")
                    cell.alignment = wrap_alignment
                    cell.fill = PatternFill("solid", fgColor="00ff00")

            employee_row_counter+=1
    try:
        module_path = os.path.dirname(__file__)
        panorama_folder_path = os.path.abspath(os.path.join(module_path, os.pardir))
        report_folder_path = os.path.join(panorama_folder_path, "Reportes")
        if not os.path.exists(report_folder_path):
            os.makedirs(report_folder_path)
        report_file_name = "Reporte empleados {0} {1}.xlsx".format("Rectificación", date.today())
        fn = os.path.join(report_folder_path, report_file_name)
        wb.save(fn)
    except OSError as e:
        if e.args[0] != 13:
            raise
        input(
            "\n Ha ocurrido un error porque el archivo Reporte empleados {0} {1}.xlsx está abierto. Por favor ciérrelo y presione cualquier tecla para que el programa pueda continuar.".format(
                "Rectificación", date.today()))


def createDesignersReportWide(db):
    ''' Este método crea un informe en Excel compacto con una proción de la información de la base de datos. '''
    id_skill = 2
    with db_session:
        employees_skills = select(es for es in db.Employees_Skills if es.skill.id == id_skill)
        employees = []
        for es in employees_skills:
            employee = db.Employees.get(id=es.employee.id)
            employees.append(employee)
        sheet_index = 0
        wb = Workbook()
        by_default_sheet = wb.get_sheet_by_name('Sheet')
        by_default_sheet.title = 'Introducción del informe'
        wb.remove_sheet(by_default_sheet)
        employee_row_counter = 0
        ws = wb.create_sheet("Informe trabajadores")
        for e in employees:
            e_activities = select(ea for ea in db.Employees_Activities if
                                  ea.employee == e)  # Para ver lo de las vacaciones
            from Planning.features import isHoliday, \
                isNotWorkday  # Es realmente necesario importar isHoliday
            tareas = select(et for et in db.Employees_Tasks if et.employee == e)


            widths = {"A": 20, "B": 20, "C": 20, "D": 20, "E": 20, "F": 20,
                      "G": 20, "H": 20, "I": 20, "J": 20, "K": 20, "L": 20,
                      "M": 20, "N": 20, "O": 20, "P": 20, "Q": 20, "R": 20,
                      "S": 20, "T": 20, "U": 20, "V": 20, "W": 20, "X": 20,
                      "Y": 20, "Z": 20, "AA": 20, "AB": 20, "AC": 20, "AD": 20,
                      "AE": 20, "AF": 20, "AG": 20, "AH": 20, "AI": 20, "AJ": 20,
                      "AK": 20, "AL": 20, "AM": 20, "AN": 20, "AO": 20, "AP": 20,
                      "AQ": 20, "AR": 20, "AS": 20, "AT": 20, "AU": 20, "AV": 20,
                      "AW": 20, "AX": 20, "AY": 20, "AZ": 20, "BA": 20, "BB": 20,
                      "BC": 20, "BD": 20, "BE": 20, "BF": 20, "BG": 20, "BH": 20,
                      "BI": 20, "BJ": 20, "BK": 20, "BL": 20, "BM": 20,"BN": 20,
                      "BO": 20, "BP": 20, "BQ": 20, "BR": 20,"BS": 20,"BT": 20,
                      "BU": 20, "BV": 20, "BW": 20, "BX": 20, "BY": 20, "BZ": 20,
                      "CA": 20, "CB": 20, "CC": 20, "CD": 20, "CE": 20, "CF": 20,
                      "CG": 20, "CH": 20, "CI": 20, "CJ": 20, "CK": 20, "CL": 20,
                      "CM": 20, "CN": 20, "CO": 20, "CP": 20, "CQ": 20, "CR": 20,
                      "CS": 20, "CT": 20, "CU": 20, "CV": 20, "CW": 20, "CX": 20,
                      "CY": 20, "CZ": 20, "DA": 20, "DB": 20, "DC": 20, "DD": 20,
                      "DE": 20, "DF": 20, "DG": 20, "DH": 20, "DI": 20, "DJ": 20,
                      "DK": 20, "DL": 20, "DM": 20, "DN": 20, "DO": 20, "DP": 20,
                      "DQ": 20, "DR": 20, "DS": 20, "DT": 20, "DU": 20, "DV": 20,
                      "DW": 20, "DX": 20, "DY": 20, "DZ": 20}

            heights = {"A": 10, "B": 10, "C": 10, "D": 10, "E": 10, "F": 10,
                       "G": 10, "H": 10, "I": 10, "J": 10, "K": 10, "L": 10,
                       "M": 10, "N": 10, "O": 10, "P": 10, "Q": 10, "R": 10,
                       "S": 10, "T": 10, "U": 10, "V": 10, "W": 10, "X": 10,
                       "Y": 10, "Z": 10, "AA": 10, "AB": 10, "AC": 10, "AD": 10,
                       "AE": 10, "AF": 10, "AG": 10, "AH": 10, "AI": 10, "AJ": 10,
                       "AK": 10, "AL": 10, "AM": 10, "AN": 10, "AO": 10, "AP": 10,
                       "AQ": 10, "AR": 10, "AS": 10, "AT": 10, "AU": 10, "AV": 10,
                       "AW": 10, "AX": 10, "AY": 10, "AZ": 10, "BA": 10, "BB": 10,
                       "BC": 10, "BD": 10, "BE": 10, "BF": 10, "BG": 10, "BH": 10,
                       "BI": 10, "BJ": 10, "BK": 10, "BL": 10, "BM": 10,"BN": 20,
                      "BO": 10, "BP": 10, "BQ": 10, "BR": 10,"BS": 10,"BT": 10,
                      "BU": 10, "BV": 10, "BW": 10, "BX": 10, "BY": 10, "BZ": 10,
                    "CA": 10, "CB": 10,"CC": 10, "CD": 10, "CE": 10, "CF": 10,
                       "CG": 10, "CH": 10,"CI": 10, "CJ": 10, "CK": 10, "CL": 10,
                       "CM": 10,"CN": 10,"CO": 10, "CP": 10, "CQ": 10, "CR": 10,
                       "CS": 10,"CT": 10,"CU": 10, "CV": 10, "CW": 10, "CX": 10,
                       "CY": 10, "CZ": 10, "DA": 10, "DB": 10,"DC": 10, "DD": 10,
                       "DE": 10, "DF": 10,"DG": 10, "DH": 10,"DI": 10, "DJ": 10,
                       "DK": 10, "DL": 10,"DM": 10,"DN": 10,"DO": 10, "DP": 10,
                       "DQ": 10, "DR": 10,"DS": 10,"DT": 10,"DU": 10, "DV": 10,
                       "DW": 10, "DX": 10,"DY": 10, "DZ": 10}

            thin_border = Border(left=Side(style='thin'),
                                 right=Side(style='thin'),
                                 top=Side(style='thin'),
                                 bottom=Side(style='thin'))

            columns = ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M", "N",
                       "O", "P", "Q", "R", "S", "T", "U", "V", "W", "X", "Y", "Z", "AA", "AB",
                       "AC", "AD", "AE", "AF", "AG", "AH", "AI", "AJ", "AK", "AL", "AM", "AN",
                       "AO", "AP", "AQ", "AR", "AS", "AT", "AU", "AV", "AW", "AX", "AY", "AZ",
                       "BA", "BB", "BC", "BD", "BE", "BF", "BG", "BH", "BI", "BJ", "BK", "BL", "BM",
                       "BN","BO","BP","BQ","BR","BS","BT","BU","BV","BW","BX","BY","BZ",
                       "CA", "CB", "CC", "CD", "CE", "CF", "CG", "CH", "CI", "CJ", "CK", "CL", "CM",
                       "CN", "CO", "CP", "CQ", "CR", "CS", "CT", "CU", "CV", "CW", "CX", "CY", "CZ",
                       "DA", "DB", "DC", "DD", "DE", "DF", "DG", "DH", "DI", "DJ", "DK", "DL", "DM",
                       "DN", "DO", "DP", "DQ", "DR", "DS", "DT", "DU", "DV", "DW", "DX", "DY", "DZ"]
            for c in columns:
                ws.column_dimensions[c].width = widths[c]
                ws.column_dimensions[c].height = heights[c]

            thin_border = Border(left=Side(style='thin'),
                                 right=Side(style='thin'),
                                 top=Side(style='thin'),
                                 bottom=Side(style='thin'))

            if (
            False):  # (es.skill.id == 4) <-- introducir esta condición si es que se quiere mostrar la información de "senior"
                # columns = ["A", "B", "C","D","E","F","G","H","I"] #Para un instalador y con el rendimiento
                columns = ["A", "B", "C", "D", "E", "F", "G",
                           "H"]  # Para un instalador sin el rendimiento
                # texts = ["","","","ID", "NOMBRE", "ZONA", "COMPETENCIA",
                #          "SENIOR","RENDIMIENTO"] #Para un instalador y con el rendimiento
                texts = ["", "", "", "ID", "NOMBRE", "ZONA", "COMPETENCIA",
                         "SENIOR"]  # Para un instalador sin el rendimiento
            else:
                columns = ["A", "B", "C", "D", "E", "F", "G"]
                texts = ["", "", "", "ID", "NOMBRE", "ZONA", "COMPETENCIA"]

            aux_columns = ["A", "B", "C", "D", "E", "F", "G",
                           "H"]  # Solo para mantener el ancho de las celdas en el caso que no sea un instalador

            for c in aux_columns:
                ws.column_dimensions[c].width = widths[c]
                ws.column_dimensions[c].height = heights[c]

            num_columns, letter_columns = zip(*list(enumerate(columns)))
            num_columns = list(num_columns)
            num_columns = [x + 1 for x in
                           num_columns]  # Desplazamos los valores en 1 porque Excel está indexado desde el 1 y no desde el 0
            letter_columns = list(letter_columns)


            # Escribir la fecha en la que fue producida el reporte:
            cell = ws.cell(row=6, column=3, value="Reporte producido el: ")
            cell.font = Font(bold=True, )
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='left')

            cell = ws.cell(row=6, column=4, value=date.today())
            cell.font = Font(bold=True, )
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='left')

            # A continuación llenamos con los datos


            r = 4
            r += employee_row_counter
            # Escribe el ID del empleado
            if e.id != None:
                cell = ws.cell(row=r+10, column=3, value="id: {0}\nNombre:{1}".format(e.id,e.name))
                cell.font = Font(bold=True)
                cell.border = thin_border
                wrap_alignment = Alignment(wrap_text=True, horizontal="center",
                                           vertical="center")
                # assign
                cell.alignment = wrap_alignment

            # Preparar el formato del horario/calendario del trabajador
            cell = ws.cell(row=4 + 9, column=3, value="EMPLEADO")
            cell.font = Font(bold=True)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')

            span_of_weeks = 15
            week_counter = 0
            aux_c = 0
            for week in range(0,span_of_weeks):
                for day in ["Lunes", "Martes", "Miercoles", "Jueves", "Viernes", "Sábado", "Domingo"]:
                    cell = ws.cell(row=4 + 9, column=4 + aux_c, value="{0} {1}".format(day,date.today()+timedelta(days=aux_c-date.today().weekday())))
                    cell.font = Font(bold=True)
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal='center')
                    aux_c += 1
                week_counter+=1

            dates = []
            today = date.today()
            for i in range(0, span_of_weeks*7):
                # dates.append(today + i * timedelta(days=1))
                dates.append(date.today()+timedelta(days=i-date.today().weekday()))
            # Imprimir las tareas del trabajador



            # Se definen las fechas de trabajo del trabajador

            tasks = []
            tasks_dates_blocks = []
            for tarea in tareas:
                working_dates = []
                task_initial_d = tarea.planned_initial_date
                task_final_d = tarea.planned_end_date
                task = db.Tasks.get(id=tarea.task.id)
                tasks.append(task)
                # Ajustar para que solo tome fechas de trabajo en días hábiles
                working_days = (task_final_d - task_initial_d).days
                working_days_span = working_days

                # working_dates = [task_initial_d + i*timedelta(days=1) for i in range(0,working_days)]
                working_days_counter = 0
                while (working_days_counter < working_days_span + 1):  # Ojo con el +1 que es para tomar el intervalo cerrado de la fecha de término
                    if (not isNotWorkday(task_initial_d + working_days_counter * timedelta(days=1))):
                        working_dates.append(task_initial_d + working_days_counter * timedelta(days=1))
                    else:
                        # working_days_span+=1 #Ya está considerado en el Planning
                        pass
                    working_days_counter += 1
                tasks_dates_blocks.append(working_dates)

            # # ########Adición de funcionalidad para las vacaciones y licencias (Atención)
            activities = []
            activities_dates_blocks = []
            for e_act in e_activities:
                ea_dates = []
                ea_initial_d = e_act.initial_date
                ea_final_d = e_act.end_date
                activity = db.Activities.get(id=e_act.activity.id)
                activities.append(activity)
                # Ajustar para que solo tome fechas de trabajo en días hábiles
                ea_days = (ea_final_d - ea_initial_d).days
                ea_days_span = ea_days

                # working_dates = [task_initial_d + i*timedelta(days=1) for i in range(0,working_days)]
                ea_days_counter = 0
                while (
                    ea_days_counter < ea_days_span + 1):  # Ojo con el +1 que es para tomar el intervalo cerrado de la fecha de término
                    if (not isNotWorkday(ea_initial_d + ea_days_counter * timedelta(days=1))):
                        ea_dates.append(ea_initial_d + ea_days_counter * timedelta(days=1))
                    else:
                        # working_days_span+=1 #Ya está considerado en el Planning
                        pass
                    ea_days_counter += 1
                    # # #########Adición de funcionalidad para las vacaciones y licencias (Atención)
                activities_dates_blocks.append(ea_dates)
            # print(ea_dates)

            # Se imprimen los días de trabajo
            for i in range(0,len(dates)):
                task_counter = 0
                for tasks_dates_block in tasks_dates_blocks:
                    for d in tasks_dates_block:
                        if d == dates[i]:
                            cell = ws.cell(row=r + 10, column=4 + i,
                                           value="Proyecto {0}\n{1}".format(
                                               tasks[task_counter].project,
                                               tasks[task_counter].project.client_address))
                            cell.font = Font(bold=True)
                            cell.border = thin_border
                            cell.alignment = Alignment(horizontal='center')
                            # cell.style.alignment.wrap_text = True #Para autoajustar el tamaño de la celca. Revisar su correcto funcionamiento
                            # create alignment style
                            wrap_alignment = Alignment(wrap_text=True, horizontal="center",
                                                       vertical="center")
                            # assign
                            cell.alignment = wrap_alignment
                            cell.fill = PatternFill("solid", fgColor="ffff00")
                    task_counter = 0
                activity_counter = 0
                for activity_date_block in activities_dates_blocks:
                    for d in activity_date_block:
                        if d == dates[i]:
                            cell = ws.cell(row=r + 10 , column=4 + i,
                                           value=activities[activity_counter].description)
                            cell.font = Font(bold=True)
                            cell.border = thin_border
                            wrap_alignment = Alignment(wrap_text=True, horizontal="center",
                                                       vertical="center")
                            cell.alignment = wrap_alignment
                            # cell.style.alignment.wrap_text = True #Para autoajustar el tamaño de la celca. Revisar su correcto funcionamiento
                            # create alignment style
                            wrap_alignment = Alignment(wrap_text=True, horizontal="center",
                                                       vertical="center")
                            if activities[activity_counter].id == 1:
                                cell.fill = PatternFill("solid", fgColor="E60404")
                            elif activities[activity_counter].id == 2:
                                cell.fill = PatternFill("solid", fgColor="810777")
                    activity_counter = 0
                if isHoliday(dates[i]):
                    cell = ws.cell(row=r + 10, column=4 + i, value="Feriado")
                    cell.font = Font(bold=True)
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal='center')
                    wrap_alignment = Alignment(wrap_text=True, horizontal="center",
                                               vertical="center")
                    cell.alignment = wrap_alignment
                    cell.fill = PatternFill("solid", fgColor="00ffff")
                if dates[i].weekday()>=5:
                    cell = ws.cell(row=r + 10, column=4 + i, value="Fin de semana")
                    cell.font = Font(bold=True)
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal='center')
                    wrap_alignment = Alignment(wrap_text=True, horizontal="center",
                                               vertical="center")
                    cell.alignment = wrap_alignment
                    cell.fill = PatternFill("solid", fgColor="00ffff")
                if (ws.cell(row=r + 10, column=4 + i).value == None and dates[i].weekday() < 5): #ws.cell(row=r + 10, column=4 + i).value == None and
                    cell = ws.cell(row=r + 10, column=4 + i, value="Disponible")
                    cell.font = Font(bold=True)
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal='center')
                    wrap_alignment = Alignment(wrap_text=True, horizontal="center",
                                               vertical="center")
                    cell.alignment = wrap_alignment
                    cell.fill = PatternFill("solid", fgColor="00ff00")

            employee_row_counter+=1
    try:
        module_path = os.path.dirname(__file__)
        panorama_folder_path = os.path.abspath(os.path.join(module_path, os.pardir))
        report_folder_path = os.path.join(panorama_folder_path, "Reportes")
        if not os.path.exists(report_folder_path):
            os.makedirs(report_folder_path)
        report_file_name = "Reporte empleados {0} {1}.xlsx".format("Diseño", date.today())
        fn = os.path.join(report_folder_path, report_file_name)
        wb.save(fn)
    except OSError as e:
        if e.args[0] != 13:
            raise
        input(
            "\n Ha ocurrido un error porque el archivo Reporte empleados {0} {1}.xlsx está abierto. Por favor ciérrelo y presione cualquier tecla para que el programa pueda continuar.".format(
                "Diseño", date.today()))


def createFabricatorsReportWide(db):
    ''' Este método crea un informe en Excel compacto con una proción de la información de la base de datos. '''
    id_skill = 3
    with db_session:
        employees_skills = select(es for es in db.Employees_Skills if es.skill.id == id_skill)
        employees = []
        for es in employees_skills:
            employee = db.Employees.get(id=es.employee.id)
            employees.append(employee)
        sheet_index = 0
        wb = Workbook()
        by_default_sheet = wb.get_sheet_by_name('Sheet')
        by_default_sheet.title = 'Introducción del informe'
        wb.remove_sheet(by_default_sheet)
        employee_row_counter = 0
        ws = wb.create_sheet("Informe trabajadores")
        for e in employees:
            e_activities = select(ea for ea in db.Employees_Activities if
                                  ea.employee == e)  # Para ver lo de las vacaciones
            from Planning.features import isHoliday, \
                isNotWorkday  # Es realmente necesario importar isHoliday
            tareas = select(et for et in db.Employees_Tasks if et.employee == e)


            widths = {"A": 20, "B": 20, "C": 20, "D": 20, "E": 20, "F": 20,
                      "G": 20, "H": 20, "I": 20, "J": 20, "K": 20, "L": 20,
                      "M": 20, "N": 20, "O": 20, "P": 20, "Q": 20, "R": 20,
                      "S": 20, "T": 20, "U": 20, "V": 20, "W": 20, "X": 20,
                      "Y": 20, "Z": 20, "AA": 20, "AB": 20, "AC": 20, "AD": 20,
                      "AE": 20, "AF": 20, "AG": 20, "AH": 20, "AI": 20, "AJ": 20,
                      "AK": 20, "AL": 20, "AM": 20, "AN": 20, "AO": 20, "AP": 20,
                      "AQ": 20, "AR": 20, "AS": 20, "AT": 20, "AU": 20, "AV": 20,
                      "AW": 20, "AX": 20, "AY": 20, "AZ": 20, "BA": 20, "BB": 20,
                      "BC": 20, "BD": 20, "BE": 20, "BF": 20, "BG": 20, "BH": 20,
                      "BI": 20, "BJ": 20, "BK": 20, "BL": 20, "BM": 20,"BN": 20,
                      "BO": 20, "BP": 20, "BQ": 20, "BR": 20,"BS": 20,"BT": 20,
                      "BU": 20, "BV": 20, "BW": 20, "BX": 20, "BY": 20, "BZ": 20,
                      "CA": 20, "CB": 20, "CC": 20, "CD": 20, "CE": 20, "CF": 20,
                      "CG": 20, "CH": 20, "CI": 20, "CJ": 20, "CK": 20, "CL": 20,
                      "CM": 20, "CN": 20, "CO": 20, "CP": 20, "CQ": 20, "CR": 20,
                      "CS": 20, "CT": 20, "CU": 20, "CV": 20, "CW": 20, "CX": 20,
                      "CY": 20, "CZ": 20, "DA": 20, "DB": 20, "DC": 20, "DD": 20,
                      "DE": 20, "DF": 20, "DG": 20, "DH": 20, "DI": 20, "DJ": 20,
                      "DK": 20, "DL": 20, "DM": 20, "DN": 20, "DO": 20, "DP": 20,
                      "DQ": 20, "DR": 20, "DS": 20, "DT": 20, "DU": 20, "DV": 20,
                      "DW": 20, "DX": 20, "DY": 20, "DZ": 20}

            heights = {"A": 10, "B": 10, "C": 10, "D": 10, "E": 10, "F": 10,
                       "G": 10, "H": 10, "I": 10, "J": 10, "K": 10, "L": 10,
                       "M": 10, "N": 10, "O": 10, "P": 10, "Q": 10, "R": 10,
                       "S": 10, "T": 10, "U": 10, "V": 10, "W": 10, "X": 10,
                       "Y": 10, "Z": 10, "AA": 10, "AB": 10, "AC": 10, "AD": 10,
                       "AE": 10, "AF": 10, "AG": 10, "AH": 10, "AI": 10, "AJ": 10,
                       "AK": 10, "AL": 10, "AM": 10, "AN": 10, "AO": 10, "AP": 10,
                       "AQ": 10, "AR": 10, "AS": 10, "AT": 10, "AU": 10, "AV": 10,
                       "AW": 10, "AX": 10, "AY": 10, "AZ": 10, "BA": 10, "BB": 10,
                       "BC": 10, "BD": 10, "BE": 10, "BF": 10, "BG": 10, "BH": 10,
                       "BI": 10, "BJ": 10, "BK": 10, "BL": 10, "BM": 10,"BN": 20,
                      "BO": 10, "BP": 10, "BQ": 10, "BR": 10,"BS": 10,"BT": 10,
                      "BU": 10, "BV": 10, "BW": 10, "BX": 10, "BY": 10, "BZ": 10,
                    "CA": 10, "CB": 10,"CC": 10, "CD": 10, "CE": 10, "CF": 10,
                       "CG": 10, "CH": 10,"CI": 10, "CJ": 10, "CK": 10, "CL": 10,
                       "CM": 10,"CN": 10,"CO": 10, "CP": 10, "CQ": 10, "CR": 10,
                       "CS": 10,"CT": 10,"CU": 10, "CV": 10, "CW": 10, "CX": 10,
                       "CY": 10, "CZ": 10, "DA": 10, "DB": 10,"DC": 10, "DD": 10,
                       "DE": 10, "DF": 10,"DG": 10, "DH": 10,"DI": 10, "DJ": 10,
                       "DK": 10, "DL": 10,"DM": 10,"DN": 10,"DO": 10, "DP": 10,
                       "DQ": 10, "DR": 10,"DS": 10,"DT": 10,"DU": 10, "DV": 10,
                       "DW": 10, "DX": 10,"DY": 10, "DZ": 10}

            thin_border = Border(left=Side(style='thin'),
                                 right=Side(style='thin'),
                                 top=Side(style='thin'),
                                 bottom=Side(style='thin'))

            columns = ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M", "N",
                       "O", "P", "Q", "R", "S", "T", "U", "V", "W", "X", "Y", "Z", "AA", "AB",
                       "AC", "AD", "AE", "AF", "AG", "AH", "AI", "AJ", "AK", "AL", "AM", "AN",
                       "AO", "AP", "AQ", "AR", "AS", "AT", "AU", "AV", "AW", "AX", "AY", "AZ",
                       "BA", "BB", "BC", "BD", "BE", "BF", "BG", "BH", "BI", "BJ", "BK", "BL", "BM",
                       "BN","BO","BP","BQ","BR","BS","BT","BU","BV","BW","BX","BY","BZ",
                       "CA", "CB", "CC", "CD", "CE", "CF", "CG", "CH", "CI", "CJ", "CK", "CL", "CM",
                       "CN", "CO", "CP", "CQ", "CR", "CS", "CT", "CU", "CV", "CW", "CX", "CY", "CZ",
                       "DA", "DB", "DC", "DD", "DE", "DF", "DG", "DH", "DI", "DJ", "DK", "DL", "DM",
                       "DN", "DO", "DP", "DQ", "DR", "DS", "DT", "DU", "DV", "DW", "DX", "DY", "DZ"]
            for c in columns:
                ws.column_dimensions[c].width = widths[c]
                ws.column_dimensions[c].height = heights[c]

            thin_border = Border(left=Side(style='thin'),
                                 right=Side(style='thin'),
                                 top=Side(style='thin'),
                                 bottom=Side(style='thin'))

            if (
            False):  # (es.skill.id == 4) <-- introducir esta condición si es que se quiere mostrar la información de "senior"
                # columns = ["A", "B", "C","D","E","F","G","H","I"] #Para un instalador y con el rendimiento
                columns = ["A", "B", "C", "D", "E", "F", "G",
                           "H"]  # Para un instalador sin el rendimiento
                # texts = ["","","","ID", "NOMBRE", "ZONA", "COMPETENCIA",
                #          "SENIOR","RENDIMIENTO"] #Para un instalador y con el rendimiento
                texts = ["", "", "", "ID", "NOMBRE", "ZONA", "COMPETENCIA",
                         "SENIOR"]  # Para un instalador sin el rendimiento
            else:
                columns = ["A", "B", "C", "D", "E", "F", "G"]
                texts = ["", "", "", "ID", "NOMBRE", "ZONA", "COMPETENCIA"]

            aux_columns = ["A", "B", "C", "D", "E", "F", "G",
                           "H"]  # Solo para mantener el ancho de las celdas en el caso que no sea un instalador

            for c in aux_columns:
                ws.column_dimensions[c].width = widths[c]
                ws.column_dimensions[c].height = heights[c]

            num_columns, letter_columns = zip(*list(enumerate(columns)))
            num_columns = list(num_columns)
            num_columns = [x + 1 for x in
                           num_columns]  # Desplazamos los valores en 1 porque Excel está indexado desde el 1 y no desde el 0
            letter_columns = list(letter_columns)


            # Escribir la fecha en la que fue producida el reporte:
            cell = ws.cell(row=6, column=3, value="Reporte producido el: ")
            cell.font = Font(bold=True, )
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='left')

            cell = ws.cell(row=6, column=4, value=date.today())
            cell.font = Font(bold=True, )
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='left')

            # A continuación llenamos con los datos


            r = 4
            r += employee_row_counter
            # Escribe el ID del empleado
            if e.id != None:
                cell = ws.cell(row=r+10, column=3, value="id: {0}\nNombre:{1}".format(e.id,e.name))
                cell.font = Font(bold=True)
                cell.border = thin_border
                wrap_alignment = Alignment(wrap_text=True, horizontal="center",
                                           vertical="center")
                # assign
                cell.alignment = wrap_alignment

            # Preparar el formato del horario/calendario del trabajador
            cell = ws.cell(row=4 + 9, column=3, value="EMPLEADO")
            cell.font = Font(bold=True)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')

            span_of_weeks = 15
            week_counter = 0
            aux_c = 0
            for week in range(0,span_of_weeks):
                for day in ["Lunes", "Martes", "Miercoles", "Jueves", "Viernes", "Sábado", "Domingo"]:
                    cell = ws.cell(row=4 + 9, column=4 + aux_c, value="{0} {1}".format(day,date.today()+timedelta(days=aux_c-date.today().weekday())))
                    cell.font = Font(bold=True)
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal='center')
                    aux_c += 1
                week_counter+=1

            dates = []
            today = date.today()
            for i in range(0, span_of_weeks*7):
                # dates.append(today + i * timedelta(days=1))
                dates.append(date.today()+timedelta(days=i-date.today().weekday()))
            # Imprimir las tareas del trabajador



            # Se definen las fechas de trabajo del trabajador

            tasks = []
            tasks_dates_blocks = []
            for tarea in tareas:
                working_dates = []
                task_initial_d = tarea.planned_initial_date
                task_final_d = tarea.planned_end_date
                task = db.Tasks.get(id=tarea.task.id)
                tasks.append(task)
                # Ajustar para que solo tome fechas de trabajo en días hábiles
                working_days = (task_final_d - task_initial_d).days
                working_days_span = working_days

                # working_dates = [task_initial_d + i*timedelta(days=1) for i in range(0,working_days)]
                working_days_counter = 0
                while (working_days_counter < working_days_span + 1):  # Ojo con el +1 que es para tomar el intervalo cerrado de la fecha de término
                    if (not isNotWorkday(task_initial_d + working_days_counter * timedelta(days=1))):
                        working_dates.append(task_initial_d + working_days_counter * timedelta(days=1))
                    else:
                        # working_days_span+=1 #Ya está considerado en el Planning
                        pass
                    working_days_counter += 1
                tasks_dates_blocks.append(working_dates)

            # # ########Adición de funcionalidad para las vacaciones y licencias (Atención)
            activities = []
            activities_dates_blocks = []
            for e_act in e_activities:
                ea_dates = []
                ea_initial_d = e_act.initial_date
                ea_final_d = e_act.end_date
                activity = db.Activities.get(id=e_act.activity.id)
                activities.append(activity)
                # Ajustar para que solo tome fechas de trabajo en días hábiles
                ea_days = (ea_final_d - ea_initial_d).days
                ea_days_span = ea_days

                # working_dates = [task_initial_d + i*timedelta(days=1) for i in range(0,working_days)]
                ea_days_counter = 0
                while (
                    ea_days_counter < ea_days_span + 1):  # Ojo con el +1 que es para tomar el intervalo cerrado de la fecha de término
                    if (not isNotWorkday(ea_initial_d + ea_days_counter * timedelta(days=1))):
                        ea_dates.append(ea_initial_d + ea_days_counter * timedelta(days=1))
                    else:
                        # working_days_span+=1 #Ya está considerado en el Planning
                        pass
                    ea_days_counter += 1
                    # # #########Adición de funcionalidad para las vacaciones y licencias (Atención)
                activities_dates_blocks.append(ea_dates)
            # print(ea_dates)

            # Se imprimen los días de trabajo
            for i in range(0,len(dates)):
                task_counter = 0
                for tasks_dates_block in tasks_dates_blocks:
                    for d in tasks_dates_block:
                        if d == dates[i]:
                            cell = ws.cell(row=r + 10, column=4 + i,
                                           value="Proyecto {0}\n{1}".format(
                                               tasks[task_counter].project,
                                               tasks[task_counter].project.client_address))
                            cell.font = Font(bold=True)
                            cell.border = thin_border
                            cell.alignment = Alignment(horizontal='center')
                            # cell.style.alignment.wrap_text = True #Para autoajustar el tamaño de la celca. Revisar su correcto funcionamiento
                            # create alignment style
                            wrap_alignment = Alignment(wrap_text=True, horizontal="center",
                                                       vertical="center")
                            # assign
                            cell.alignment = wrap_alignment
                            cell.fill = PatternFill("solid", fgColor="ffff00")
                    task_counter = 0
                activity_counter = 0
                for activity_date_block in activities_dates_blocks:
                    for d in activity_date_block:
                        if d == dates[i]:
                            cell = ws.cell(row=r + 10 , column=4 + i,
                                           value=activities[activity_counter].description)
                            cell.font = Font(bold=True)
                            cell.border = thin_border
                            wrap_alignment = Alignment(wrap_text=True, horizontal="center",
                                                       vertical="center")
                            cell.alignment = wrap_alignment
                            # cell.style.alignment.wrap_text = True #Para autoajustar el tamaño de la celca. Revisar su correcto funcionamiento
                            # create alignment style
                            wrap_alignment = Alignment(wrap_text=True, horizontal="center",
                                                       vertical="center")
                            if activities[activity_counter].id == 1:
                                cell.fill = PatternFill("solid", fgColor="E60404")
                            elif activities[activity_counter].id == 2:
                                cell.fill = PatternFill("solid", fgColor="810777")
                    activity_counter = 0
                if isHoliday(dates[i]):
                    cell = ws.cell(row=r + 10, column=4 + i, value="Feriado")
                    cell.font = Font(bold=True)
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal='center')
                    wrap_alignment = Alignment(wrap_text=True, horizontal="center",
                                               vertical="center")
                    cell.alignment = wrap_alignment
                    cell.fill = PatternFill("solid", fgColor="00ffff")
                if dates[i].weekday()>=5:
                    cell = ws.cell(row=r + 10, column=4 + i, value="Fin de semana")
                    cell.font = Font(bold=True)
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal='center')
                    wrap_alignment = Alignment(wrap_text=True, horizontal="center",
                                               vertical="center")
                    cell.alignment = wrap_alignment
                    cell.fill = PatternFill("solid", fgColor="00ffff")
                if (ws.cell(row=r + 10, column=4 + i).value == None and dates[i].weekday() < 5): #ws.cell(row=r + 10, column=4 + i).value == None and
                    cell = ws.cell(row=r + 10, column=4 + i, value="Disponible")
                    cell.font = Font(bold=True)
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal='center')
                    wrap_alignment = Alignment(wrap_text=True, horizontal="center",
                                               vertical="center")
                    cell.alignment = wrap_alignment
                    cell.fill = PatternFill("solid", fgColor="00ff00")

            employee_row_counter+=1
    try:
        module_path = os.path.dirname(__file__)
        panorama_folder_path = os.path.abspath(os.path.join(module_path, os.pardir))
        report_folder_path = os.path.join(panorama_folder_path, "Reportes")
        if not os.path.exists(report_folder_path):
            os.makedirs(report_folder_path)
        report_file_name = "Reporte empleados {0} {1}.xlsx".format("Fabricación", date.today())
        fn = os.path.join(report_folder_path, report_file_name)
        wb.save(fn)
    except OSError as e:
        if e.args[0] != 13:
            raise
        input(
            "\n Ha ocurrido un error porque el archivo Reporte empleados {0} {1}.xlsx está abierto. Por favor ciérrelo y presione cualquier tecla para que el programa pueda continuar.".format(
                "Fabricación", date.today()))


def createInstallatorsReportWide(db):
    ''' Este método crea un informe en Excel compacto con una proción de la información de la base de datos. '''
    id_skill = 4
    with db_session:
        employees_skills = select(es for es in db.Employees_Skills if es.skill.id == id_skill)
        employees = []
        employees_skills_aux = []
        for es in employees_skills:
            employee = db.Employees.get(id=es.employee.id)
            employees.append(employee)
            employees_skills_aux.append(es.skill.name)
        sheet_index = 0
        wb = Workbook()
        by_default_sheet = wb.get_sheet_by_name('Sheet')
        by_default_sheet.title = 'Introducción del informe'
        wb.remove_sheet(by_default_sheet)
        employee_row_counter = 0
        ws = wb.create_sheet("Informe trabajadores")
        for e in employees:
            e_activities = select(ea for ea in db.Employees_Activities if
                                  ea.employee == e)  # Para ver lo de las vacaciones
            from Planning.features import isHoliday, \
                isNotWorkday  # Es realmente necesario importar isHoliday
            tareas = select(et for et in db.Employees_Tasks if et.employee == e)


            widths = {"A": 20, "B": 20, "C": 20, "D": 20, "E": 20, "F": 20,
                      "G": 20, "H": 20, "I": 20, "J": 20, "K": 20, "L": 20,
                      "M": 20, "N": 20, "O": 20, "P": 20, "Q": 20, "R": 20,
                      "S": 20, "T": 20, "U": 20, "V": 20, "W": 20, "X": 20,
                      "Y": 20, "Z": 20, "AA": 20, "AB": 20, "AC": 20, "AD": 20,
                      "AE": 20, "AF": 20, "AG": 20, "AH": 20, "AI": 20, "AJ": 20,
                      "AK": 20, "AL": 20, "AM": 20, "AN": 20, "AO": 20, "AP": 20,
                      "AQ": 20, "AR": 20, "AS": 20, "AT": 20, "AU": 20, "AV": 20,
                      "AW": 20, "AX": 20, "AY": 20, "AZ": 20, "BA": 20, "BB": 20,
                      "BC": 20, "BD": 20, "BE": 20, "BF": 20, "BG": 20, "BH": 20,
                      "BI": 20, "BJ": 20, "BK": 20, "BL": 20, "BM": 20,"BN": 20,
                      "BO": 20, "BP": 20, "BQ": 20, "BR": 20,"BS": 20,"BT": 20,
                      "BU": 20, "BV": 20, "BW": 20, "BX": 20, "BY": 20, "BZ": 20,
                      "CA": 20, "CB": 20, "CC": 20, "CD": 20, "CE": 20, "CF": 20,
                      "CG": 20, "CH": 20, "CI": 20, "CJ": 20, "CK": 20, "CL": 20,
                      "CM": 20, "CN": 20, "CO": 20, "CP": 20, "CQ": 20, "CR": 20,
                      "CS": 20, "CT": 20, "CU": 20, "CV": 20, "CW": 20, "CX": 20,
                      "CY": 20, "CZ": 20, "DA": 20, "DB": 20, "DC": 20, "DD": 20,
                      "DE": 20, "DF": 20, "DG": 20, "DH": 20, "DI": 20, "DJ": 20,
                      "DK": 20, "DL": 20, "DM": 20, "DN": 20, "DO": 20, "DP": 20,
                      "DQ": 20, "DR": 20, "DS": 20, "DT": 20, "DU": 20, "DV": 20,
                      "DW": 20, "DX": 20, "DY": 20, "DZ": 20}

            heights = {"A": 10, "B": 10, "C": 10, "D": 10, "E": 10, "F": 10,
                       "G": 10, "H": 10, "I": 10, "J": 10, "K": 10, "L": 10,
                       "M": 10, "N": 10, "O": 10, "P": 10, "Q": 10, "R": 10,
                       "S": 10, "T": 10, "U": 10, "V": 10, "W": 10, "X": 10,
                       "Y": 10, "Z": 10, "AA": 10, "AB": 10, "AC": 10, "AD": 10,
                       "AE": 10, "AF": 10, "AG": 10, "AH": 10, "AI": 10, "AJ": 10,
                       "AK": 10, "AL": 10, "AM": 10, "AN": 10, "AO": 10, "AP": 10,
                       "AQ": 10, "AR": 10, "AS": 10, "AT": 10, "AU": 10, "AV": 10,
                       "AW": 10, "AX": 10, "AY": 10, "AZ": 10, "BA": 10, "BB": 10,
                       "BC": 10, "BD": 10, "BE": 10, "BF": 10, "BG": 10, "BH": 10,
                       "BI": 10, "BJ": 10, "BK": 10, "BL": 10, "BM": 10,"BN": 20,
                      "BO": 10, "BP": 10, "BQ": 10, "BR": 10,"BS": 10,"BT": 10,
                      "BU": 10, "BV": 10, "BW": 10, "BX": 10, "BY": 10, "BZ": 10,
                    "CA": 10, "CB": 10,"CC": 10, "CD": 10, "CE": 10, "CF": 10,
                       "CG": 10, "CH": 10,"CI": 10, "CJ": 10, "CK": 10, "CL": 10,
                       "CM": 10,"CN": 10,"CO": 10, "CP": 10, "CQ": 10, "CR": 10,
                       "CS": 10,"CT": 10,"CU": 10, "CV": 10, "CW": 10, "CX": 10,
                       "CY": 10, "CZ": 10, "DA": 10, "DB": 10,"DC": 10, "DD": 10,
                       "DE": 10, "DF": 10,"DG": 10, "DH": 10,"DI": 10, "DJ": 10,
                       "DK": 10, "DL": 10,"DM": 10,"DN": 10,"DO": 10, "DP": 10,
                       "DQ": 10, "DR": 10,"DS": 10,"DT": 10,"DU": 10, "DV": 10,
                       "DW": 10, "DX": 10,"DY": 10, "DZ": 10}

            thin_border = Border(left=Side(style='thin'),
                                 right=Side(style='thin'),
                                 top=Side(style='thin'),
                                 bottom=Side(style='thin'))

            columns = ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M", "N",
                       "O", "P", "Q", "R", "S", "T", "U", "V", "W", "X", "Y", "Z", "AA", "AB",
                       "AC", "AD", "AE", "AF", "AG", "AH", "AI", "AJ", "AK", "AL", "AM", "AN",
                       "AO", "AP", "AQ", "AR", "AS", "AT", "AU", "AV", "AW", "AX", "AY", "AZ",
                       "BA", "BB", "BC", "BD", "BE", "BF", "BG", "BH", "BI", "BJ", "BK", "BL", "BM",
                       "BN","BO","BP","BQ","BR","BS","BT","BU","BV","BW","BX","BY","BZ",
                       "CA", "CB", "CC", "CD", "CE", "CF", "CG", "CH", "CI", "CJ", "CK", "CL", "CM",
                       "CN", "CO", "CP", "CQ", "CR", "CS", "CT", "CU", "CV", "CW", "CX", "CY", "CZ",
                       "DA", "DB", "DC", "DD", "DE", "DF", "DG", "DH", "DI", "DJ", "DK", "DL", "DM",
                       "DN", "DO", "DP", "DQ", "DR", "DS", "DT", "DU", "DV", "DW", "DX", "DY", "DZ"]
            for c in columns:
                ws.column_dimensions[c].width = widths[c]
                ws.column_dimensions[c].height = heights[c]

            thin_border = Border(left=Side(style='thin'),
                                 right=Side(style='thin'),
                                 top=Side(style='thin'),
                                 bottom=Side(style='thin'))

            if (
            False):  # (es.skill.id == 4) <-- introducir esta condición si es que se quiere mostrar la información de "senior"
                # columns = ["A", "B", "C","D","E","F","G","H","I"] #Para un instalador y con el rendimiento
                columns = ["A", "B", "C", "D", "E", "F", "G",
                           "H"]  # Para un instalador sin el rendimiento
                # texts = ["","","","ID", "NOMBRE", "ZONA", "COMPETENCIA",
                #          "SENIOR","RENDIMIENTO"] #Para un instalador y con el rendimiento
                texts = ["", "", "", "ID", "NOMBRE", "ZONA", "COMPETENCIA",
                         "SENIOR"]  # Para un instalador sin el rendimiento
            else:
                columns = ["A", "B", "C", "D", "E", "F", "G"]
                texts = ["", "", "", "ID", "NOMBRE", "ZONA", "COMPETENCIA"]

            aux_columns = ["A", "B", "C", "D", "E", "F", "G",
                           "H"]  # Solo para mantener el ancho de las celdas en el caso que no sea un instalador

            for c in aux_columns:
                ws.column_dimensions[c].width = widths[c]
                ws.column_dimensions[c].height = heights[c]

            num_columns, letter_columns = zip(*list(enumerate(columns)))
            num_columns = list(num_columns)
            num_columns = [x + 1 for x in
                           num_columns]  # Desplazamos los valores en 1 porque Excel está indexado desde el 1 y no desde el 0
            letter_columns = list(letter_columns)


            # Escribir la fecha en la que fue producida el reporte:
            cell = ws.cell(row=6, column=3, value="Reporte producido el: ")
            cell.font = Font(bold=True, )
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='left')

            cell = ws.cell(row=6, column=4, value=date.today())
            cell.font = Font(bold=True, )
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='left')

            # A continuación llenamos con los datos


            r = 4
            r += employee_row_counter
            # Escribe el ID del empleado
            if e.id != None:
                cell = ws.cell(row=r+10, column=3, value="id: {0}\nNombre:{1}".format(e.id,e.name))
                cell.font = Font(bold=True)
                cell.border = thin_border
                wrap_alignment = Alignment(wrap_text=True, horizontal="center",
                                           vertical="center")
                # assign
                cell.alignment = wrap_alignment

            # Preparar el formato del horario/calendario del trabajador
            cell = ws.cell(row=4 + 9, column=3, value="EMPLEADO")
            cell.font = Font(bold=True)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')

            span_of_weeks = 15
            week_counter = 0
            aux_c = 0
            for week in range(0,span_of_weeks):
                for day in ["Lunes", "Martes", "Miercoles", "Jueves", "Viernes", "Sábado", "Domingo"]:
                    cell = ws.cell(row=4 + 9, column=4 + aux_c, value="{0} {1}".format(day,date.today()+timedelta(days=aux_c-date.today().weekday())))
                    cell.font = Font(bold=True)
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal='center')
                    aux_c += 1
                week_counter+=1

            dates = []
            today = date.today()
            for i in range(0, span_of_weeks*7):
                # dates.append(today + i * timedelta(days=1))
                dates.append(date.today()+timedelta(days=i-date.today().weekday()))
            # Imprimir las tareas del trabajador



            # Se definen las fechas de trabajo del trabajador

            tasks = []
            tasks_dates_blocks = []
            for tarea in tareas:
                working_dates = []
                task_initial_d = tarea.planned_initial_date
                task_final_d = tarea.planned_end_date
                task = db.Tasks.get(id=tarea.task.id)
                tasks.append(task)
                # Ajustar para que solo tome fechas de trabajo en días hábiles
                working_days = (task_final_d - task_initial_d).days
                working_days_span = working_days

                # working_dates = [task_initial_d + i*timedelta(days=1) for i in range(0,working_days)]
                working_days_counter = 0
                while (working_days_counter < working_days_span + 1):  # Ojo con el +1 que es para tomar el intervalo cerrado de la fecha de término
                    if (not isNotWorkday(task_initial_d + working_days_counter * timedelta(days=1))):
                        working_dates.append(task_initial_d + working_days_counter * timedelta(days=1))
                    else:
                        # working_days_span+=1 #Ya está considerado en el Planning
                        pass
                    working_days_counter += 1
                tasks_dates_blocks.append(working_dates)

            # # ########Adición de funcionalidad para las vacaciones y licencias (Atención)
            activities = []
            activities_dates_blocks = []
            for e_act in e_activities:
                ea_dates = []
                ea_initial_d = e_act.initial_date
                ea_final_d = e_act.end_date
                activity = db.Activities.get(id=e_act.activity.id)
                activities.append(activity)
                # Ajustar para que solo tome fechas de trabajo en días hábiles
                ea_days = (ea_final_d - ea_initial_d).days
                ea_days_span = ea_days

                # working_dates = [task_initial_d + i*timedelta(days=1) for i in range(0,working_days)]
                ea_days_counter = 0
                while (
                    ea_days_counter < ea_days_span + 1):  # Ojo con el +1 que es para tomar el intervalo cerrado de la fecha de término
                    if (not isNotWorkday(ea_initial_d + ea_days_counter * timedelta(days=1))):
                        ea_dates.append(ea_initial_d + ea_days_counter * timedelta(days=1))
                    else:
                        # working_days_span+=1 #Ya está considerado en el Planning
                        pass
                    ea_days_counter += 1
                    # # #########Adición de funcionalidad para las vacaciones y licencias (Atención)
                activities_dates_blocks.append(ea_dates)
            # print(ea_dates)

            # Se imprimen los días de trabajo
            for i in range(0,len(dates)):
                task_counter = 0
                for tasks_dates_block in tasks_dates_blocks:
                    for d in tasks_dates_block:
                        if d == dates[i]:
                            cell = ws.cell(row=r + 10, column=4 + i,
                                           value="Proyecto {0}\n{1}".format(
                                               tasks[task_counter].project,
                                               tasks[task_counter].project.client_address))
                            cell.font = Font(bold=True)
                            cell.border = thin_border
                            cell.alignment = Alignment(horizontal='center')
                            # cell.style.alignment.wrap_text = True #Para autoajustar el tamaño de la celca. Revisar su correcto funcionamiento
                            # create alignment style
                            wrap_alignment = Alignment(wrap_text=True, horizontal="center",
                                                       vertical="center")
                            # assign
                            cell.alignment = wrap_alignment
                            cell.fill = PatternFill("solid", fgColor="ffff00")
                    task_counter = 0
                activity_counter = 0
                for activity_date_block in activities_dates_blocks:
                    for d in activity_date_block:
                        if d == dates[i]:
                            cell = ws.cell(row=r + 10 , column=4 + i,
                                           value=activities[activity_counter].description)
                            cell.font = Font(bold=True)
                            cell.border = thin_border
                            wrap_alignment = Alignment(wrap_text=True, horizontal="center",
                                                       vertical="center")
                            cell.alignment = wrap_alignment
                            # cell.style.alignment.wrap_text = True #Para autoajustar el tamaño de la celca. Revisar su correcto funcionamiento
                            # create alignment style
                            wrap_alignment = Alignment(wrap_text=True, horizontal="center",
                                                       vertical="center")
                            if activities[activity_counter].id == 1:
                                cell.fill = PatternFill("solid", fgColor="E60404")
                            elif activities[activity_counter].id == 2:
                                cell.fill = PatternFill("solid", fgColor="810777")
                    activity_counter = 0
                if isHoliday(dates[i]):
                    cell = ws.cell(row=r + 10, column=4 + i, value="Feriado")
                    cell.font = Font(bold=True)
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal='center')
                    wrap_alignment = Alignment(wrap_text=True, horizontal="center",
                                               vertical="center")
                    cell.alignment = wrap_alignment
                    cell.fill = PatternFill("solid", fgColor="00ffff")
                if dates[i].weekday()>=5:
                    cell = ws.cell(row=r + 10, column=4 + i, value="Fin de semana")
                    cell.font = Font(bold=True)
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal='center')
                    wrap_alignment = Alignment(wrap_text=True, horizontal="center",
                                               vertical="center")
                    cell.alignment = wrap_alignment
                    cell.fill = PatternFill("solid", fgColor="00ffff")
                if (ws.cell(row=r + 10, column=4 + i).value == None and dates[i].weekday() < 5): #ws.cell(row=r + 10, column=4 + i).value == None and
                    cell = ws.cell(row=r + 10, column=4 + i, value="Disponible")
                    cell.font = Font(bold=True)
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal='center')
                    wrap_alignment = Alignment(wrap_text=True, horizontal="center",
                                               vertical="center")
                    cell.alignment = wrap_alignment
                    cell.fill = PatternFill("solid", fgColor="00ff00")

            employee_row_counter+=1

    try:
        module_path = os.path.dirname(__file__)
        panorama_folder_path = os.path.abspath(os.path.join(module_path, os.pardir))
        report_folder_path = os.path.join(panorama_folder_path, "Reportes")
        if not os.path.exists(report_folder_path):
            os.makedirs(report_folder_path)
        report_file_name = "Reporte empleados {0} {1}.xlsx".format("Instalación", date.today())
        fn = os.path.join(report_folder_path, report_file_name)
        wb.save(fn)
    except OSError as e:
        if e.args[0] != 13:
            raise
        input(
            "\n Ha ocurrido un error porque el archivo Reporte empleados {0} {1}.xlsx está abierto. Por favor ciérrelo y presione cualquier tecla para que el programa pueda continuar.".format(
                "Instalación", date.today()))

def createWorkersReportWide(db):
    ''' Este método crea un informe en Excel compacto con una proción de la información de la base de datos. '''
    with db_session:
        employees_skills_rect = select(es for es in db.Employees_Skills if es.skill.id == 1)
        employees_skills_rect = list(employees_skills_rect)
        employees_skills_dis = select(es for es in db.Employees_Skills if es.skill.id == 2)
        employees_skills_dis = list(employees_skills_dis)
        employees_skills_fab = select(es for es in db.Employees_Skills if es.skill.id == 3)
        employees_skills_fab = list(employees_skills_fab)
        employees_skills_inst = select(es for es in db.Employees_Skills if es.skill.id == 4)
        employees_skills_inst = list(employees_skills_inst)
        employees_skills = []
        employees_skills.append(employees_skills_rect)
        employees_skills.append(employees_skills_dis)
        employees_skills.append(employees_skills_fab)
        employees_skills.append(employees_skills_inst)

        employees = []

        employees_skills_aux = []
        for es in employees_skills:
            for es_s in es:
                employee = db.Employees.get(id=es_s.employee.id)
                employees.append(employee)
                employees_skills_aux.append(es_s.skill.name)
        sheet_index = 0
        wb = Workbook()
        by_default_sheet = wb.get_sheet_by_name('Sheet')
        by_default_sheet.title = 'Introducción del informe'
        wb.remove_sheet(by_default_sheet)
        employee_row_counter = 0
        ws = wb.create_sheet("Informe trabajadores")
        employee_current_index = 0
        for e in employees:
            e_activities = select(ea for ea in db.Employees_Activities if
                                  ea.employee == e)  # Para ver lo de las vacaciones
            from Planning.features import isHoliday, \
                isNotWorkday  # Es realmente necesario importar isHoliday
            tareas = select(et for et in db.Employees_Tasks if et.employee == e)

            widths = {"A": 20, "B": 20, "C": 20, "D": 20, "E": 20, "F": 20,
                      "G": 20, "H": 20, "I": 20, "J": 20, "K": 20, "L": 20,
                      "M": 20, "N": 20, "O": 20, "P": 20, "Q": 20, "R": 20,
                      "S": 20, "T": 20, "U": 20, "V": 20, "W": 20, "X": 20,
                      "Y": 20, "Z": 20, "AA": 20, "AB": 20, "AC": 20, "AD": 20,
                      "AE": 20, "AF": 20, "AG": 20, "AH": 20, "AI": 20, "AJ": 20,
                      "AK": 20, "AL": 20, "AM": 20, "AN": 20, "AO": 20, "AP": 20,
                      "AQ": 20, "AR": 20, "AS": 20, "AT": 20, "AU": 20, "AV": 20,
                      "AW": 20, "AX": 20, "AY": 20, "AZ": 20, "BA": 20, "BB": 20,
                      "BC": 20, "BD": 20, "BE": 20, "BF": 20, "BG": 20, "BH": 20,
                      "BI": 20, "BJ": 20, "BK": 20, "BL": 20, "BM": 20, "BN": 20,
                      "BO": 20, "BP": 20, "BQ": 20, "BR": 20, "BS": 20, "BT": 20,
                      "BU": 20, "BV": 20, "BW": 20, "BX": 20, "BY": 20, "BZ": 20,
                      "CA": 20, "CB": 20, "CC": 20, "CD": 20, "CE": 20, "CF": 20,
                      "CG": 20, "CH": 20, "CI": 20, "CJ": 20, "CK": 20, "CL": 20,
                      "CM": 20, "CN": 20, "CO": 20, "CP": 20, "CQ": 20, "CR": 20,
                      "CS": 20, "CT": 20, "CU": 20, "CV": 20, "CW": 20, "CX": 20,
                      "CY": 20, "CZ": 20, "DA": 20, "DB": 20, "DC": 20, "DD": 20,
                      "DE": 20, "DF": 20, "DG": 20, "DH": 20, "DI": 20, "DJ": 20,
                      "DK": 20, "DL": 20, "DM": 20, "DN": 20, "DO": 20, "DP": 20,
                      "DQ": 20, "DR": 20, "DS": 20, "DT": 20, "DU": 20, "DV": 20,
                      "DW": 20, "DX": 20, "DY": 20, "DZ": 20}

            heights = {"A": 10, "B": 10, "C": 10, "D": 10, "E": 10, "F": 10,
                       "G": 10, "H": 10, "I": 10, "J": 10, "K": 10, "L": 10,
                       "M": 10, "N": 10, "O": 10, "P": 10, "Q": 10, "R": 10,
                       "S": 10, "T": 10, "U": 10, "V": 10, "W": 10, "X": 10,
                       "Y": 10, "Z": 10, "AA": 10, "AB": 10, "AC": 10, "AD": 10,
                       "AE": 10, "AF": 10, "AG": 10, "AH": 10, "AI": 10, "AJ": 10,
                       "AK": 10, "AL": 10, "AM": 10, "AN": 10, "AO": 10, "AP": 10,
                       "AQ": 10, "AR": 10, "AS": 10, "AT": 10, "AU": 10, "AV": 10,
                       "AW": 10, "AX": 10, "AY": 10, "AZ": 10, "BA": 10, "BB": 10,
                       "BC": 10, "BD": 10, "BE": 10, "BF": 10, "BG": 10, "BH": 10,
                       "BI": 10, "BJ": 10, "BK": 10, "BL": 10, "BM": 10, "BN": 20,
                       "BO": 10, "BP": 10, "BQ": 10, "BR": 10, "BS": 10, "BT": 10,
                       "BU": 10, "BV": 10, "BW": 10, "BX": 10, "BY": 10, "BZ": 10,
                       "CA": 10, "CB": 10, "CC": 10, "CD": 10, "CE": 10, "CF": 10,
                       "CG": 10, "CH": 10, "CI": 10, "CJ": 10, "CK": 10, "CL": 10,
                       "CM": 10, "CN": 10, "CO": 10, "CP": 10, "CQ": 10, "CR": 10,
                       "CS": 10, "CT": 10, "CU": 10, "CV": 10, "CW": 10, "CX": 10,
                       "CY": 10, "CZ": 10, "DA": 10, "DB": 10, "DC": 10, "DD": 10,
                       "DE": 10, "DF": 10, "DG": 10, "DH": 10, "DI": 10, "DJ": 10,
                       "DK": 10, "DL": 10, "DM": 10, "DN": 10, "DO": 10, "DP": 10,
                       "DQ": 10, "DR": 10, "DS": 10, "DT": 10, "DU": 10, "DV": 10,
                       "DW": 10, "DX": 10, "DY": 10, "DZ": 10}

            thin_border = Border(left=Side(style='thin'),
                                 right=Side(style='thin'),
                                 top=Side(style='thin'),
                                 bottom=Side(style='thin'))

            columns = ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M", "N",
                       "O", "P", "Q", "R", "S", "T", "U", "V", "W", "X", "Y", "Z", "AA", "AB",
                       "AC", "AD", "AE", "AF", "AG", "AH", "AI", "AJ", "AK", "AL", "AM", "AN",
                       "AO", "AP", "AQ", "AR", "AS", "AT", "AU", "AV", "AW", "AX", "AY", "AZ",
                       "BA", "BB", "BC", "BD", "BE", "BF", "BG", "BH", "BI", "BJ", "BK", "BL",
                       "BM",
                       "BN", "BO", "BP", "BQ", "BR", "BS", "BT", "BU", "BV", "BW", "BX", "BY",
                       "BZ",
                       "CA", "CB", "CC", "CD", "CE", "CF", "CG", "CH", "CI", "CJ", "CK", "CL",
                       "CM",
                       "CN", "CO", "CP", "CQ", "CR", "CS", "CT", "CU", "CV", "CW", "CX", "CY",
                       "CZ",
                       "DA", "DB", "DC", "DD", "DE", "DF", "DG", "DH", "DI", "DJ", "DK", "DL",
                       "DM",
                       "DN", "DO", "DP", "DQ", "DR", "DS", "DT", "DU", "DV", "DW", "DX", "DY",
                       "DZ"]
            for c in columns:
                ws.column_dimensions[c].width = widths[c]
                ws.column_dimensions[c].height = heights[c]

            thin_border = Border(left=Side(style='thin'),
                                 right=Side(style='thin'),
                                 top=Side(style='thin'),
                                 bottom=Side(style='thin'))

            if (
                    False):  # (es.skill.id == 4) <-- introducir esta condición si es que se quiere mostrar la información de "senior"
                # columns = ["A", "B", "C","D","E","F","G","H","I"] #Para un instalador y con el rendimiento
                columns = ["A", "B", "C", "D", "E", "F", "G",
                           "H"]  # Para un instalador sin el rendimiento
                # texts = ["","","","ID", "NOMBRE", "ZONA", "COMPETENCIA",
                #          "SENIOR","RENDIMIENTO"] #Para un instalador y con el rendimiento
                texts = ["", "", "", "ID", "NOMBRE", "ZONA", "COMPETENCIA",
                         "SENIOR"]  # Para un instalador sin el rendimiento
            else:
                columns = ["A", "B", "C", "D", "E", "F", "G"]
                texts = ["", "", "", "ID", "NOMBRE", "ZONA", "COMPETENCIA"]

            aux_columns = ["A", "B", "C", "D", "E", "F", "G",
                           "H"]  # Solo para mantener el ancho de las celdas en el caso que no sea un instalador

            for c in aux_columns:
                ws.column_dimensions[c].width = widths[c]
                ws.column_dimensions[c].height = heights[c]

            num_columns, letter_columns = zip(*list(enumerate(columns)))
            num_columns = list(num_columns)
            num_columns = [x + 1 for x in
                           num_columns]  # Desplazamos los valores en 1 porque Excel está indexado desde el 1 y no desde el 0
            letter_columns = list(letter_columns)

            # Escribir la fecha en la que fue producida el reporte:
            cell = ws.cell(row=6, column=3, value="Reporte producido el: ")
            cell.font = Font(bold=True, )
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='left')

            cell = ws.cell(row=6, column=4, value=date.today())
            cell.font = Font(bold=True, )
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='left')

            # A continuación llenamos con los datos


            r = 4
            r += employee_row_counter
            # Escribe el ID del empleado
            if e.id != None:
                cell = ws.cell(row=r + 10, column=3,
                               value="id: {0}\n{1}\n{2}".format(e.id, e.name,employees_skills_aux[employee_current_index]))
                cell.font = Font(bold=True)
                cell.border = thin_border
                wrap_alignment = Alignment(wrap_text=True, horizontal="center",
                                           vertical="center")
                # assign
                cell.alignment = wrap_alignment

            # Preparar el formato del horario/calendario del trabajador
            cell = ws.cell(row=4 + 9, column=3, value="EMPLEADO")
            cell.font = Font(bold=True)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')

            span_of_weeks = 15
            week_counter = 0
            aux_c = 0
            for week in range(0, span_of_weeks):
                for day in ["Lunes", "Martes", "Miercoles", "Jueves", "Viernes", "Sábado",
                            "Domingo"]:
                    cell = ws.cell(row=4 + 9, column=4 + aux_c, value="{0} {1}".format(day,
                                                                                       date.today() + timedelta(
                                                                                           days=aux_c - date.today().weekday())))
                    cell.font = Font(bold=True)
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal='center')
                    aux_c += 1
                week_counter += 1

            dates = []
            today = date.today()
            for i in range(0, span_of_weeks * 7):
                # dates.append(today + i * timedelta(days=1))
                dates.append(date.today() + timedelta(days=i - date.today().weekday()))
            # Imprimir las tareas del trabajador



            # Se definen las fechas de trabajo del trabajador

            tasks = []
            tasks_dates_blocks = []
            for tarea in tareas:
                working_dates = []
                task_initial_d = tarea.planned_initial_date
                task_final_d = tarea.planned_end_date
                task = db.Tasks.get(id=tarea.task.id)
                tasks.append(task)
                # Ajustar para que solo tome fechas de trabajo en días hábiles
                working_days = (task_final_d - task_initial_d).days
                working_days_span = working_days

                # working_dates = [task_initial_d + i*timedelta(days=1) for i in range(0,working_days)]
                working_days_counter = 0
                while (
                    working_days_counter < working_days_span + 1):  # Ojo con el +1 que es para tomar el intervalo cerrado de la fecha de término
                    if (not isNotWorkday(
                                task_initial_d + working_days_counter * timedelta(days=1))):
                        working_dates.append(
                            task_initial_d + working_days_counter * timedelta(days=1))
                    else:
                        # working_days_span+=1 #Ya está considerado en el Planning
                        pass
                    working_days_counter += 1
                tasks_dates_blocks.append(working_dates)

            # # ########Adición de funcionalidad para las vacaciones y licencias (Atención)
            activities = []
            activities_dates_blocks = []
            for e_act in e_activities:
                ea_dates = []
                ea_initial_d = e_act.initial_date
                ea_final_d = e_act.end_date
                activity = db.Activities.get(id=e_act.activity.id)
                activities.append(activity)
                # Ajustar para que solo tome fechas de trabajo en días hábiles
                ea_days = (ea_final_d - ea_initial_d).days
                ea_days_span = ea_days

                # working_dates = [task_initial_d + i*timedelta(days=1) for i in range(0,working_days)]
                ea_days_counter = 0
                while (
                            ea_days_counter < ea_days_span + 1):  # Ojo con el +1 que es para tomar el intervalo cerrado de la fecha de término
                    if (not isNotWorkday(ea_initial_d + ea_days_counter * timedelta(days=1))):
                        ea_dates.append(ea_initial_d + ea_days_counter * timedelta(days=1))
                    else:
                        # working_days_span+=1 #Ya está considerado en el Planning
                        pass
                    ea_days_counter += 1
                    # # #########Adición de funcionalidad para las vacaciones y licencias (Atención)
                activities_dates_blocks.append(ea_dates)
            # print(ea_dates)

            # Se imprimen los días de trabajo
            for i in range(0, len(dates)):
                task_counter = 0
                for tasks_dates_block in tasks_dates_blocks:
                    for d in tasks_dates_block:
                        if d == dates[i]:
                            cell = ws.cell(row=r + 10, column=4 + i,
                                           value="Proyecto {0}\n{1}".format(
                                               tasks[task_counter].project,
                                               tasks[task_counter].project.client_address))
                            cell.font = Font(bold=True)
                            cell.border = thin_border
                            cell.alignment = Alignment(horizontal='center')
                            # cell.style.alignment.wrap_text = True #Para autoajustar el tamaño de la celca. Revisar su correcto funcionamiento
                            # create alignment style
                            wrap_alignment = Alignment(wrap_text=True, horizontal="center",
                                                       vertical="center")
                            # assign
                            cell.alignment = wrap_alignment
                            cell.fill = PatternFill("solid", fgColor="ffff00")
                    task_counter = 0
                activity_counter = 0
                for activity_date_block in activities_dates_blocks:
                    for d in activity_date_block:
                        if d == dates[i]:
                            cell = ws.cell(row=r + 10, column=4 + i,
                                           value=activities[activity_counter].description)
                            cell.font = Font(bold=True)
                            cell.border = thin_border
                            wrap_alignment = Alignment(wrap_text=True, horizontal="center",
                                                       vertical="center")
                            cell.alignment = wrap_alignment
                            # cell.style.alignment.wrap_text = True #Para autoajustar el tamaño de la celca. Revisar su correcto funcionamiento
                            # create alignment style
                            wrap_alignment = Alignment(wrap_text=True, horizontal="center",
                                                       vertical="center")
                            if activities[activity_counter].id == 1:
                                cell.fill = PatternFill("solid", fgColor="E60404")
                            elif activities[activity_counter].id == 2:
                                cell.fill = PatternFill("solid", fgColor="810777")
                    activity_counter = 0
                if isHoliday(dates[i]):
                    cell = ws.cell(row=r + 10, column=4 + i, value="Feriado")
                    cell.font = Font(bold=True)
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal='center')
                    wrap_alignment = Alignment(wrap_text=True, horizontal="center",
                                               vertical="center")
                    cell.alignment = wrap_alignment
                    cell.fill = PatternFill("solid", fgColor="00ffff")
                if dates[i].weekday() >= 5:
                    cell = ws.cell(row=r + 10, column=4 + i, value="Fin de semana")
                    cell.font = Font(bold=True)
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal='center')
                    wrap_alignment = Alignment(wrap_text=True, horizontal="center",
                                               vertical="center")
                    cell.alignment = wrap_alignment
                    cell.fill = PatternFill("solid", fgColor="00ffff")
                if (ws.cell(row=r + 10, column=4 + i).value == None and dates[
                    i].weekday() < 5):  # ws.cell(row=r + 10, column=4 + i).value == None and
                    cell = ws.cell(row=r + 10, column=4 + i, value="Disponible")
                    cell.font = Font(bold=True)
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal='center')
                    wrap_alignment = Alignment(wrap_text=True, horizontal="center",
                                               vertical="center")
                    cell.alignment = wrap_alignment
                    cell.fill = PatternFill("solid", fgColor="00ff00")

            employee_row_counter += 1
            employee_current_index += 1

    try:
        module_path = os.path.dirname(__file__)
        panorama_folder_path = os.path.abspath(os.path.join(module_path, os.pardir))
        report_folder_path = os.path.join(panorama_folder_path, "Reportes")
        if not os.path.exists(report_folder_path):
            os.makedirs(report_folder_path)
        report_file_name = "Reporte empleados {}.xlsx".format(date.today())
        fn = os.path.join(report_folder_path, report_file_name)
        wb.save(fn)
    except OSError as e:
        if e.args[0] != 13:
            raise
        input(
            "\n Ha ocurrido un error porque el archivo Reporte empleados {}.xlsx está abierto. Por favor ciérrelo y presione cualquier tecla para que el programa pueda continuar.".format(date.today()))

createWorkersReportWide(db)
# createRectificatorsReportWide(db)
# createDesignersReportWide(db)
# createFabricatorsReportWide(db)
# createInstallatorsReportWide(db)
createRectificatorsReport(db)
createDesignersReport(db)
createFabricatorsReport(db)
createInstallersReport(db)

# createPersonalEmployeeReport(db,1)
# createPersonalEmployeeReport(db,2)
# createPersonalEmployeeReport(db,3)
# createPersonalEmployeeReport(db,4)
# createPersonalEmployeeReport(db,5)
# createPersonalEmployeeReport(db,6)
# createPersonalEmployeeReport(db,7)
# createPersonalEmployeeReport(db,8)
# createPersonalEmployeeReport(db,9)
# createPersonalEmployeeReport(db,10)
# createPersonalEmployeeReport(db,11)
# createPersonalEmployeeReport(db,12)
# createPersonalEmployeeReport(db,13)
